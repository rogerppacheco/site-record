"""APIs do pedido de ajuda/socorro ao GC da Nio."""

from __future__ import annotations

from rest_framework import generics, permissions, status
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from crm_app.models import EtapaErroAjudaGc, PedidoAjudaGc, Venda
from crm_app.serializers import EtapaErroAjudaGcSerializer
from crm_app.services.pedido_ajuda_gc_service import (
    PDV_PADRAO,
    contato_do_usuario,
    cpf_cnpj_da_venda,
    numero_pedido_da_venda,
    obter_config_gc,
    processar_pedido_abrir_chamado_ti,
    protocolo_registro_auditoria,
)


class PedidoAjudaGcExportarView(APIView):
    """GET: exporta acionamentos (pedidos de ajuda GC) em Excel. Query: data_inicio, data_fim, origem."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from datetime import datetime, time
        from io import BytesIO

        import openpyxl
        from django.http import HttpResponse
        from django.utils import timezone
        from django.utils.dateparse import parse_date
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        di_raw = (request.query_params.get('data_inicio') or '').strip()
        df_raw = (request.query_params.get('data_fim') or '').strip()
        origem = (request.query_params.get('origem') or '').strip().lower()

        di = parse_date(di_raw) if di_raw else None
        df = parse_date(df_raw) if df_raw else None
        if not di or not df:
            return Response(
                {'detail': 'Informe data_inicio e data_fim (YYYY-MM-DD).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if df < di:
            return Response(
                {'detail': 'data_fim deve ser maior ou igual a data_inicio.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tz = timezone.get_current_timezone()
        inicio = timezone.make_aware(datetime.combine(di, time.min), tz)
        fim = timezone.make_aware(datetime.combine(df, time.max), tz)

        qs = (
            PedidoAjudaGc.objects.select_related('usuario', 'venda', 'venda__cliente')
            .filter(criado_em__gte=inicio, criado_em__lte=fim)
            .order_by('-criado_em')
        )
        if origem in (PedidoAjudaGc.ORIGEM_AUDITORIA, PedidoAjudaGc.ORIGEM_ESTEIRA):
            qs = qs.filter(origem=origem)

        limite = 5000
        lista = list(qs[:limite])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Acionamentos'
        headers = [
            'Data/hora',
            'Origem',
            'Tipo',
            'Solicitante',
            'Nome GC',
            'E-mail GC',
            'WhatsApp GC',
            'PDV',
            'Login BO',
            'Login vendedor',
            'CPF/CNPJ',
            'Nº pedido (O.S)',
            'Contato',
            'Etapa do erro',
            'Cenário reportado',
            'Nº registro atendimento',
            'Enviado e-mail',
            'Enviado WhatsApp',
            'Erros',
            'ID venda',
            'Mensagem enviada',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4E73DF', end_color='4E73DF', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        tipo_labels = dict(PedidoAjudaGc.TIPO_CHOICES)
        origem_labels = dict(PedidoAjudaGc.ORIGEM_CHOICES)
        for p in lista:
            criado = timezone.localtime(p.criado_em).strftime('%d/%m/%Y %H:%M') if p.criado_em else ''
            solicitante = ''
            if p.usuario:
                solicitante = (
                    p.usuario.get_full_name() or p.usuario.username or ''
                ).strip()
            erros = ''
            if isinstance(p.erros, list):
                erros = '; '.join(str(x) for x in p.erros if x)
            elif p.erros:
                erros = str(p.erros)
            ws.append([
                criado,
                origem_labels.get(p.origem, p.origem),
                tipo_labels.get(p.tipo, p.tipo),
                solicitante,
                p.nome_gc or '',
                p.email_gc or '',
                p.telefone_gc or '',
                p.pdv or '',
                p.login_bo or '',
                p.login_vendedor or '',
                p.cpf_cnpj_cliente or '',
                p.numero_pedido or '',
                p.contato or '',
                p.etapa_erro or '',
                p.detalhe_cenario or '',
                p.numero_registro or '',
                'Sim' if p.enviado_email else 'Não',
                'Sim' if p.enviado_whatsapp else 'Não',
                erros,
                p.venda_id or '',
                p.mensagem_enviada or '',
            ])

        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            if letter in ('O', 'U'):
                ws.column_dimensions[letter].width = 42
            elif letter in ('E', 'N', 'D'):
                ws.column_dimensions[letter].width = 28
            else:
                ws.column_dimensions[letter].width = 16
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
        origem_sufixo = f'_{origem}' if origem else ''
        nome = f'acionamentos_ajuda_gc{origem_sufixo}_{di.isoformat()}_{df.isoformat()}_{stamp}.xlsx'
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{nome}"'
        return resp


class EtapaErroAjudaGcListCreateView(generics.ListCreateAPIView):
    queryset = EtapaErroAjudaGc.objects.all().order_by('contexto', 'ordem', 'nome')
    serializer_class = EtapaErroAjudaGcSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        contexto = (self.request.query_params.get('contexto') or '').strip().lower()
        if contexto in (
            EtapaErroAjudaGc.CONTEXTO_AUDITORIA,
            EtapaErroAjudaGc.CONTEXTO_ESTEIRA,
        ):
            qs = qs.filter(contexto=contexto)
        ativo = self.request.query_params.get('ativo')
        if ativo is not None:
            ativo_norm = str(ativo).strip().lower()
            if ativo_norm in ('1', 'true', 'sim', 'yes'):
                qs = qs.filter(ativo=True)
            elif ativo_norm in ('0', 'false', 'nao', 'não', 'no'):
                qs = qs.filter(ativo=False)
        return qs


class EtapaErroAjudaGcDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EtapaErroAjudaGc.objects.all().order_by('contexto', 'ordem', 'nome')
    serializer_class = EtapaErroAjudaGcSerializer
    permission_classes = [permissions.IsAuthenticated]


class PedidoAjudaGcContextoView(APIView):
    """Prefill do modal de pedido de ajuda (dados da venda + etapas + GC)."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        origem = (request.query_params.get('origem') or '').strip().lower()
        if origem not in (PedidoAjudaGc.ORIGEM_AUDITORIA, PedidoAjudaGc.ORIGEM_ESTEIRA):
            return Response(
                {'detail': 'Informe origem=auditoria ou origem=esteira.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        venda_id = request.query_params.get('venda_id')
        venda = None
        if venda_id:
            try:
                venda = (
                    Venda.objects.select_related('cliente', 'vendedor')
                    .filter(id=int(venda_id))
                    .first()
                )
            except (TypeError, ValueError):
                venda = None
            if not venda:
                return Response({'detail': 'Venda não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        config = obter_config_gc()
        etapas = list(
            EtapaErroAjudaGc.objects.filter(contexto=origem, ativo=True)
            .order_by('ordem', 'nome')
            .values('id', 'nome', 'ordem')
        )
        return Response(
            {
                'origem': origem,
                'tipo': PedidoAjudaGc.TIPO_ABRIR_CHAMADO_TI,
                'pdv': PDV_PADRAO,
                'nome_gc': config.nome_gc or '',
                'email_gc': config.email_gc or '',
                'telefone_gc': config.telefone_gc or '',
                'contato': contato_do_usuario(request.user),
                'login_bo_sugerido': (getattr(request.user, 'matricula_pap', None) or '')
                or (getattr(request.user, 'username', None) or ''),
                'cpf_cnpj_cliente': cpf_cnpj_da_venda(venda),
                'numero_pedido': numero_pedido_da_venda(venda),
                'numero_registro': protocolo_registro_auditoria(venda)
                if origem == PedidoAjudaGc.ORIGEM_AUDITORIA
                else '',
                'cliente_nome': (
                    venda.cliente.nome_razao_social if venda and venda.cliente else ''
                ),
                'venda_id': venda.id if venda else None,
                'etapas': etapas,
            }
        )


class PedidoAjudaGcEnviarView(APIView):
    """POST multipart: envia pedido de ajuda (máscara TI) por e-mail e WhatsApp."""

    parser_classes = [MultiPartParser, FormParser, JSONParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        tipo = (request.data.get('tipo') or PedidoAjudaGc.TIPO_ABRIR_CHAMADO_TI).strip()
        if tipo != PedidoAjudaGc.TIPO_ABRIR_CHAMADO_TI:
            return Response(
                {'detail': 'Tipo de pedido ainda não suportado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        origem = (request.data.get('origem') or '').strip().lower()
        venda_id = request.data.get('venda_id')
        venda = None
        if venda_id not in (None, ''):
            try:
                venda = (
                    Venda.objects.select_related('cliente', 'vendedor')
                    .filter(id=int(venda_id))
                    .first()
                )
            except (TypeError, ValueError):
                venda = None
            if not venda:
                return Response({'detail': 'Venda não encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        evidencia = request.FILES.get('evidencia') or request.FILES.get('arquivo')
        pedido, erro = processar_pedido_abrir_chamado_ti(
            usuario=request.user,
            venda=venda,
            origem=origem,
            login_bo=request.data.get('login_bo') or '',
            login_vendedor=request.data.get('login_vendedor') or '',
            etapa_erro=request.data.get('etapa_erro') or '',
            detalhe_cenario=request.data.get('detalhe_cenario') or '',
            numero_registro=request.data.get('numero_registro') or '',
            evidencia_upload=evidencia,
            cpf_cnpj_override=request.data.get('cpf_cnpj_cliente') or '',
            numero_pedido_override=request.data.get('numero_pedido') or '',
            contato_override=request.data.get('contato') or '',
        )
        if erro and not pedido:
            return Response({'detail': erro}, status=status.HTTP_400_BAD_REQUEST)
        if erro and pedido:
            return Response(
                {
                    'detail': erro,
                    'id': pedido.id,
                    'enviado_email': pedido.enviado_email,
                    'enviado_whatsapp': pedido.enviado_whatsapp,
                    'erros': pedido.erros,
                },
                status=status.HTTP_502_BAD_GATEWAY,
            )

        return Response(
            {
                'id': pedido.id,
                'mensagem': pedido.mensagem_enviada,
                'enviado_email': pedido.enviado_email,
                'enviado_whatsapp': pedido.enviado_whatsapp,
                'erros': pedido.erros,
            },
            status=status.HTTP_201_CREATED,
        )
