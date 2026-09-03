# crm_app/historico_pap.py
"""Histórico PAP: venda / interesse / pré-venda, dedup pela coluna Pedido.

Não grava Venda. Só persiste protocolos já vistos e monta Excel para revisão.
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TIPOS_HISTORICO = ("VENDA", "INTERESSE", "PRE_VENDA")
TIPO_ALIASES = {
    "VENDA": "VENDA",
    "VENDAS": "VENDA",
    "INTERESSE": "INTERESSE",
    "INTERESSES": "INTERESSE",
    "PRE_VENDA": "PRE_VENDA",
    "PRE-VENDA": "PRE_VENDA",
    "PRE VENDA": "PRE_VENDA",
    "PREVENDAS": "PRE_VENDA",
    "PRE_VENDAS": "PRE_VENDA",
    "PREVENDA": "PRE_VENDA",
    "PRE-VENDAS": "PRE_VENDA",
}
TIPO_API_ALIASES = {
    "VENDA": ("VENDA",),
    "INTERESSE": ("INTERESSE",),
    "PRE_VENDA": ("PRE_VENDA", "PRE-VENDA", "PRE_VENDAS", "PREVENDAS", "PREVENDA"),
}

STATUS_LISTA_PADRAO = (
    "ANALISE_BO,CONTATO_AGENDADO,VENDA_NAO_CONFIRMADA,PEDIDO_GERADO,"
    "CLIENTE_SEM_CONTATO,VENDA_LIQUIDA,VENDA_GROSS"
)
SEGMENTO_PADRAO = "EMPRESARIAL,VAREJO"
PAP_API_VENDAS = "https://pap-api.niointernet.com.br/api/portal/vendas"
PAP_HISTORICO_URL = "https://pap.niointernet.com.br/administrativo/historico"
LIMIT_PAGINA = 15
MAX_DIAS_BUSCA = 93

PEDIDO_COLUNAS = (
    "pedido",
    "numero pedido",
    "numeropedido",
    "número do pedido",
    "protocolo",
    "pedido pap",
)

HEADERS = [
    ("tipo_venda", "Tipo"),
    ("pedido", "Pedido"),
    ("data_pedido", "Data do pedido"),
    ("status", "Status"),
    ("status_primario", "Status primário"),
    ("status_secundario", "Status secundário"),
    ("tipo_abertura_os", "Tipo abertura OS"),
    ("status_abertura_os", "Status abertura OS"),
    ("vendedor_matricula", "Matrícula vendedor"),
    ("vendedor_nome", "Vendedor"),
    ("vendedor_perfil", "Perfil vendedor"),
    ("operador_bo", "Operador BO"),
    ("cliente", "Cliente"),
    ("cpf", "CPF"),
    ("data_nascimento", "Data de nascimento"),
    ("nome_mae", "Nome da mãe"),
    ("genero", "Gênero"),
    ("email", "E-mail"),
    ("celular_principal", "Celular principal"),
    ("celular_2", "Celular 2"),
    ("protocolo_analise", "Protocolo de análise"),
    ("resultado_biometria", "Resultado biometria"),
    ("cep", "CEP"),
    ("logradouro", "Logradouro"),
    ("numero", "Número"),
    ("complemento", "Complemento"),
    ("bairro", "Bairro"),
    ("cidade", "Cidade"),
    ("uf", "UF"),
    ("ponto_referencia", "Ponto de referência"),
    ("plano", "Plano"),
    ("velocidade", "Velocidade"),
    ("valor_mensal", "Valor mensal"),
    ("valor_total_pedido", "Valor total pedido"),
    ("forma_pagamento", "Forma de pagamento"),
    ("preferencia_data", "Preferência instalação (data)"),
    ("preferencia_periodo", "Preferência instalação (período)"),
    ("os_instalacao", "OS instalação"),
    ("data_instalacao", "Data instalação"),
    ("periodo_instalacao", "Período instalação"),
    ("observacao_vendedor", "Observação vendedor"),
    ("observacao_operador", "Observação operador"),
]


def normalizar_pedido(valor: Any) -> str:
    digits = re.sub(r"\D", "", "" if valor is None else str(valor))
    if 12 <= len(digits) <= 24:
        return digits
    return ""


def normalizar_tipo(valor: Any) -> str:
    raw = unicodedata_fold("" if valor is None else str(valor))
    return TIPO_ALIASES.get(raw, "")


def unicodedata_fold(texto: str) -> str:
    import unicodedata

    nfd = unicodedata.normalize("NFD", texto or "")
    sem = "".join(ch for ch in nfd if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", sem).strip().upper().replace("É", "E")


def tipos_solicitados(valores: Iterable[Any] | None) -> list[str]:
    if not valores:
        return list(TIPOS_HISTORICO)
    out: list[str] = []
    for v in valores:
        tipo = normalizar_tipo(v)
        if tipo and tipo not in out:
            out.append(tipo)
    return out or list(TIPOS_HISTORICO)


def _get(obj: Any, *path, default=""):
    cur = obj
    for p in path:
        if cur is None:
            return default
        if isinstance(p, int):
            if not isinstance(cur, list) or p >= len(cur):
                return default
            cur = cur[p]
        else:
            if not isinstance(cur, dict):
                return default
            cur = cur.get(p)
        if cur is None:
            return default
    return cur


def _txt(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Sim" if val else "Não"
    s = str(val).strip()
    if s.lower() in {"none", "null", "undefined"}:
        return ""
    return s


def _only_digits(val: Any) -> str:
    return re.sub(r"\D", "", _txt(val))


def _fmt_cpf(val: Any) -> str:
    d = _only_digits(val)
    if len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return _txt(val)


def _fmt_cep(val: Any) -> str:
    d = _only_digits(val)
    if len(d) == 8:
        return f"{d[:5]}-{d[5:]}"
    return _txt(val)


def _fmt_phone(val: Any) -> str:
    d = _only_digits(val)
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    if len(d) == 10:
        return f"({d[:2]}) {d[2:6]}-{d[6:]}"
    return _txt(val)


def _fmt_dt(val: Any) -> str:
    if val is None:
        return ""
    s = _txt(val)
    if not s:
        return ""
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            raw = s.replace("Z", "+00:00") if s.endswith("Z") and "%z" in fmt else s
            dt = datetime.fromisoformat(raw) if "T" in s and fmt.startswith("%Y-%m-%dT") else datetime.strptime(s[:26], fmt)
            if "T" in s:
                return dt.strftime("%d/%m/%Y %H:%M")
            return dt.strftime("%d/%m/%Y")
        except Exception:
            continue
    return s


def _money(val: Any):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = _txt(val).replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return _txt(val)


def _end_inst(p: dict) -> dict:
    ends = p.get("enderecos") if isinstance(p.get("enderecos"), list) else []
    if isinstance(p.get("enderecoInstalacao"), dict):
        return p["enderecoInstalacao"]
    if ends:
        inst = next((e for e in ends if isinstance(e, dict) and re.search(r"instala", str(e.get("tipo") or ""), re.I)), None)
        return inst or (ends[0] if isinstance(ends[0], dict) else {})
    dados = p.get("dadosCliente") if isinstance(p.get("dadosCliente"), dict) else {}
    if isinstance(dados.get("endereco"), dict):
        return dados["endereco"]
    return {}


def _complemento(end: dict) -> str:
    comps = end.get("complementos") if isinstance(end.get("complementos"), list) else []
    parts = []
    for c in comps:
        if isinstance(c, dict):
            bit = " ".join(_txt(c.get(k)) for k in ("tipo", "numero", "descricao", "complemento") if _txt(c.get(k))).strip()
            if bit:
                parts.append(bit)
        elif _txt(c):
            parts.append(_txt(c))
    extra = _txt(end.get("complemento"))
    if extra:
        parts.append(extra)
    seen = []
    for p in parts:
        if p not in seen:
            seen.append(p)
    return " | ".join(seen)


def _celulares(dc: dict) -> tuple[str, str]:
    contatos = dc.get("contatos") if isinstance(dc, dict) else None
    if not isinstance(contatos, list):
        return "", ""
    ordered = sorted(
        [c for c in contatos if isinstance(c, dict)],
        key=lambda c: c.get("ordem") if isinstance(c.get("ordem"), int) else 99,
    )
    tels = [_fmt_phone(c.get("telefone")) for c in ordered if _txt(c.get("telefone"))]
    return (tels[0] if tels else "", tels[1] if len(tels) > 1 else "")


def map_pedido_api(p: dict, tipo_venda: str = "VENDA") -> dict:
    dc = p.get("dadosCliente") if isinstance(p.get("dadosCliente"), dict) else {}
    vend = p.get("vendedor") if isinstance(p.get("vendedor"), dict) else {}
    prod = p.get("produtos") if isinstance(p.get("produtos"), dict) else {}
    bl = prod.get("bandaLarga") if isinstance(prod.get("bandaLarga"), dict) else {}
    pag = p.get("pagamento") if isinstance(p.get("pagamento"), dict) else {}
    pref = p.get("preferenciaInstalacao") if isinstance(p.get("preferenciaInstalacao"), dict) else {}
    bio = dc.get("biometria") if isinstance(dc.get("biometria"), dict) else {}
    end = _end_inst(p)
    cel1, cel2 = _celulares(dc)
    cliente = _txt(p.get("cliente")) or _txt(dc.get("nome"))
    plano = (
        _txt(prod.get("comboContratado"))
        or _txt(prod.get("comboSolicitado"))
        or _txt(_get(prod, "combo", "ofertaPortal", "nome"))
    )
    tipo = normalizar_tipo(tipo_venda) or normalizar_tipo(p.get("tipoVenda")) or "VENDA"
    return {
        "tipo_venda": tipo,
        "pedido": normalizar_pedido(p.get("numeroPedido") or p.get("pedido")),
        "data_pedido": _fmt_dt(p.get("dataCriacao")),
        "status": _txt(p.get("status")),
        "status_primario": _txt(p.get("chaveStatusPrimario")),
        "status_secundario": _txt(p.get("subStatus")) or _txt(p.get("chaveStatusSecundario")),
        "tipo_abertura_os": _txt(p.get("tipoDeAbertura")),
        "status_abertura_os": _txt(p.get("statusAberturaAutoOS")),
        "vendedor_matricula": _txt(vend.get("matricula")),
        "vendedor_nome": _txt(vend.get("nome")),
        "vendedor_perfil": _txt(vend.get("perfil")),
        "operador_bo": _txt(p.get("operador")),
        "cliente": cliente,
        "cpf": _fmt_cpf(p.get("cpf") or dc.get("cpf")),
        "data_nascimento": _fmt_dt(dc.get("dataNascimento")),
        "nome_mae": _txt(dc.get("nomeMae")),
        "genero": _txt(dc.get("genero")),
        "email": _txt(dc.get("email")).lower(),
        "celular_principal": cel1,
        "celular_2": cel2,
        "protocolo_analise": _txt(bio.get("protocolo")),
        "resultado_biometria": _txt(bio.get("resultado")),
        "cep": _fmt_cep(end.get("cep")),
        "logradouro": _txt(end.get("logradouro")),
        "numero": _txt(end.get("numero")),
        "complemento": _complemento(end) if isinstance(end, dict) else "",
        "bairro": _txt(end.get("bairro")),
        "cidade": _txt(end.get("localidade") or end.get("cidade")),
        "uf": _txt(end.get("uf")),
        "ponto_referencia": _txt(end.get("pontoReferencia")),
        "plano": plano,
        "velocidade": _txt(bl.get("velocidade")),
        "valor_mensal": _money(prod.get("valorPlano")),
        "valor_total_pedido": _money(prod.get("valorTotalPedido") or prod.get("valorTotal")),
        "forma_pagamento": _txt(pag.get("formaPagamento")),
        "preferencia_data": _fmt_dt(pref.get("data")),
        "preferencia_periodo": _txt(pref.get("periodo")),
        "os_instalacao": _txt(bl.get("osInstalacao")),
        "data_instalacao": _fmt_dt(bl.get("dataInstalacao")),
        "periodo_instalacao": _txt(bl.get("periodoInstalacao")),
        "observacao_vendedor": _txt(p.get("observacaoVendedor")),
        "observacao_operador": _txt(p.get("observacaoOperador")),
        "pdv": _txt(p.get("pdv") or _get(p, "pontoVenda", "codigo") or _get(p, "pdvSap")),
    }


def extrair_lista_api(payload: Any) -> tuple[list[dict], int]:
    if payload is None:
        return [], 0
    if isinstance(payload, list):
        return [p for p in payload if isinstance(p, dict)], len(payload)
    if not isinstance(payload, dict):
        return [], 0
    lista = (
        payload.get("data")
        or payload.get("content")
        or payload.get("items")
        or payload.get("results")
        or payload.get("vendas")
        or payload.get("pedidos")
        or _get(payload, "lists", "pedidosValidos", default=None)
    )
    if isinstance(lista, dict):
        lista = (
            lista.get("pedidosValidos")
            or lista.get("content")
            or lista.get("items")
            or lista.get("data")
        )
    if not isinstance(lista, list):
        lista = []
    total = (
        payload.get("total")
        or payload.get("totalElements")
        or payload.get("count")
        or _get(payload, "fields", "total", default=None)
        or len(lista)
    )
    try:
        total_i = int(total)
    except (TypeError, ValueError):
        total_i = len(lista)
    return [p for p in lista if isinstance(p, dict)], total_i


def _norm_col(nome: Any) -> str:
    return unicodedata_fold(str(nome or "")).replace("_", " ")


def pedidos_de_dataframe(df: pd.DataFrame) -> list[tuple[str, str]]:
    if df is None or df.empty:
        return []
    col_pedido = None
    col_tipo = None
    for c in df.columns:
        n = _norm_col(c)
        if col_pedido is None and n in PEDIDO_COLUNAS:
            col_pedido = c
        if col_tipo is None and n in {"tipo", "tipo venda", "tipovenda"}:
            col_tipo = c
    if col_pedido is None:
        for c in df.columns:
            sample = df[c].astype(str).head(20).tolist()
            hits = sum(1 for v in sample if normalizar_pedido(v))
            if hits >= max(1, min(3, len(sample) // 2)):
                col_pedido = c
                break
    if col_pedido is None:
        return []
    out = []
    seen = set()
    for _, row in df.iterrows():
        ped = normalizar_pedido(row.get(col_pedido))
        if not ped or ped in seen:
            continue
        tipo = normalizar_tipo(row.get(col_tipo)) if col_tipo else "VENDA"
        seen.add(ped)
        out.append((ped, tipo or "VENDA"))
    return out


def pedidos_de_json(raw: Any) -> list[tuple[str, str, dict]]:
    lista, _ = extrair_lista_api(raw)
    if not lista and isinstance(raw, list):
        lista = [x for x in raw if isinstance(x, dict)]
    out = []
    seen = set()
    for p in lista:
        ped = normalizar_pedido(p.get("numeroPedido") or p.get("pedido"))
        if not ped or ped in seen:
            continue
        tipo = normalizar_tipo(p.get("tipoVenda")) or "VENDA"
        seen.add(ped)
        out.append((ped, tipo, p))
    return out


def montar_xlsx_historico(linhas: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    fill = PatternFill("solid", fgColor="1F6B4A")
    font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
    body = Font(name="Calibri", size=10, color="1F2A24")
    thin = Border(
        left=Side(style="thin", color="C5D5CC"),
        right=Side(style="thin", color="C5D5CC"),
        top=Side(style="thin", color="C5D5CC"),
        bottom=Side(style="thin", color="C5D5CC"),
    )
    for c, (_, lab) in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, lab)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    zebra = PatternFill("solid", fgColor="F3F8F5")
    for i, r in enumerate(linhas, 2):
        for c, (k, _) in enumerate(HEADERS, 1):
            cell = ws.cell(i, c, r.get(k) if r.get(k) is not None else "")
            cell.font = body
            cell.border = thin
            if i % 2 == 0:
                cell.fill = zebra
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val), 42))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), 42)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_arquivo_exportacao(nome: str, content: bytes) -> list[tuple[str, str, dict]]:
    nome_l = (nome or "").lower()
    if nome_l.endswith(".json"):
        data = json.loads(content.decode("utf-8"))
        return pedidos_de_json(data)
    from crm_app.legado_pap_osab import ler_excel_bytes, pick_sheet

    sheets = ler_excel_bytes(nome, content)
    df = pick_sheet(sheets, ("Pedidos", "PEDIDOS", "Pedido", "BASE", "Export"))
    pares = pedidos_de_dataframe(df)
    return [(ped, tipo, {"numeroPedido": ped, "tipoVenda": tipo}) for ped, tipo in pares]


def montar_url_vendas(
    *,
    data_inicio: str,
    data_fim: str,
    pdv: str,
    tipo_api: str,
    page: int,
    limit: int = LIMIT_PAGINA,
    status: str | None = STATUS_LISTA_PADRAO,
    segmento: str = SEGMENTO_PADRAO,
) -> str:
    from urllib.parse import urlencode

    params = {
        "dataInicio": data_inicio,
        "dataFim": data_fim,
        "segmento": segmento,
        "tipoVenda": tipo_api,
        "page": page,
        "limit": limit,
    }
    if (pdv or "").strip():
        params["pdv"] = (pdv or "").strip()
    if status:
        params["status"] = status
    return f"{PAP_API_VENDAS}?{urlencode(params)}"
