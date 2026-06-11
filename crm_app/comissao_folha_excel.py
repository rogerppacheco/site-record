"""Geração de XLSX da folha de comissão (extrato por venda)."""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUNAS_EXTRATO = [
    ('venda_id', 'VENDA'),
    ('grupo', 'GRUPO'),
    ('nome', 'NOME'),
    ('dacc', 'DACC'),
    ('cnpj', 'CNPJ'),
    ('classificacao_mei', 'MEI/NMEI'),
    ('plano', 'PLANO'),
    ('dt_pedido', 'DT PEDIDO'),
    ('dt_inst', 'DT INST'),
    ('os', 'OS'),
    ('situacao', 'SITUAÇÃO'),
    ('churn', 'CHURN'),
    ('adiantada', 'ADIANT.'),
    ('valor_comissao', 'COMISSÃO'),
    ('comissao_tipo', 'TIPO COMISSÃO'),
]

_LARGURAS = {
    'VENDA': 8,
    'GRUPO': 22,
    'NOME': 36,
    'DACC': 6,
    'CNPJ': 6,
    'MEI/NMEI': 10,
    'PLANO': 14,
    'DT PEDIDO': 11,
    'DT INST': 11,
    'OS': 12,
    'SITUAÇÃO': 24,
    'CHURN': 8,
    'ADIANT.': 9,
    'COMISSÃO': 12,
    'TIPO COMISSÃO': 28,
}


def _norm(txt) -> str:
    return (txt or '').strip().upper()


def _bloco_extrato(linha: Dict[str, Any]) -> Tuple[int, str]:
    situacao = _norm(linha.get('situacao'))
    is_churn = _norm(linha.get('churn')) == 'SIM'
    if situacao == 'INSTALADA' and not is_churn:
        return 0, 'INSTALADAS'
    if situacao == 'INSTALADA' and is_churn:
        return 1, 'INSTALADAS COM CHURN'
    if 'CANCELADA' in situacao:
        return 2, 'CANCELADAS'
    return 3, 'DEMAIS STATUS'


def _parse_br_date_key(dt: str) -> str:
    parts = (dt or '').strip().split('/')
    if len(parts) != 3:
        return '9999-99-99'
    dd, mm, yyyy = parts
    if not (dd and mm and yyyy):
        return '9999-99-99'
    return f'{yyyy.zfill(4)}-{mm.zfill(2)}-{dd.zfill(2)}'


def _tipo_comissao_label(tipo: str | None) -> str:
    """Legado: comissao_tipo já vem formatado do backend."""
    return (tipo or '').strip() or '—'


def _linhas_extrato_vendedor(vendedor_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    extrato = list(vendedor_data.get('extrato') or [])
    blocos: Dict[int, List[Dict[str, Any]]] = {0: [], 1: [], 2: [], 3: []}
    titulos = {
        0: 'INSTALADAS',
        1: 'INSTALADAS COM CHURN',
        2: 'CANCELADAS',
        3: 'DEMAIS STATUS',
    }
    for e in extrato:
        ordem, titulo = _bloco_extrato(e)
        blocos[ordem].append({**e, '_grupo': titulo})

    rows: List[Dict[str, Any]] = []
    for ordem in (0, 1, 2, 3):
        itens = sorted(
            blocos[ordem],
            key=lambda x: (_parse_br_date_key(x.get('dt_pedido')), (x.get('nome') or '').upper()),
        )
        if not itens:
            continue
        rows.append({'_is_grupo': True, 'grupo': f'{titulos[ordem]} — {len(itens)} venda(s)'})
        for e in itens:
            rows.append({
                'venda_id': e.get('venda_id'),
                'grupo': e.get('_grupo') or titulos[ordem],
                'nome': e.get('nome') or '',
                'dacc': e.get('dacc') or '',
                'cnpj': e.get('cnpj') or '',
                'classificacao_mei': e.get('classificacao_mei') or '-',
                'plano': e.get('plano') or '',
                'dt_pedido': e.get('dt_pedido') or '',
                'dt_inst': e.get('dt_inst') or '',
                'os': e.get('os') or '',
                'situacao': e.get('situacao') or '',
                'churn': e.get('churn') or '',
                'adiantada': e.get('adiantada') or '-',
                'valor_comissao': e.get('valor_comissao'),
                'comissao_tipo': _tipo_comissao_label(e.get('comissao_tipo')),
            })
    return rows


def _titulo_aba_seguro(nome: str, usados: set) -> str:
    base = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in (nome or 'Extrato'))[:28].strip() or 'Extrato'
    titulo = base
    n = 2
    while titulo in usados:
        suf = f'_{n}'
        titulo = (base[: 28 - len(suf)] + suf).strip()
        n += 1
    usados.add(titulo)
    return titulo


def _preencher_aba(ws, linhas: List[Dict[str, Any]]):
    headers = [h for _, h in COLUNAS_EXTRATO]
    col_keys = [k for k, _ in COLUNAS_EXTRATO]
    ws.append(headers)

    fill_hdr = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    font_hdr = Font(bold=True, color='FFFFFF')
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill_hdr
        c.font = font_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fill_grupo = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    idx_comissao = col_keys.index('valor_comissao') + 1
    dinheiro_fmt = 'R$ #,##0.00'

    for row in linhas:
        if row.get('_is_grupo'):
            ws.append([row.get('grupo')] + [''] * (len(headers) - 1))
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            cell = ws.cell(row=r, column=1)
            cell.fill = fill_grupo
            cell.font = Font(bold=True)
            continue
        ws.append([row.get(k, '') for k in col_keys])
        r = ws.max_row
        val = row.get('valor_comissao')
        if val is not None and val != '':
            com_cell = ws.cell(row=r, column=idx_comissao)
            com_cell.number_format = dinheiro_fmt
            com_cell.alignment = Alignment(horizontal='right')

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _LARGURAS.get(h, 12)
    ws.freeze_panes = 'A2'


def _estilo_cabecalho(ws, row: int = 1, ncols: int | None = None) -> None:
    fill_hdr = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    font_hdr = Font(bold=True, color='FFFFFF')
    ncols = ncols or ws.max_column
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill_hdr
        c.font = font_hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _fmt_money_cell(ws, row: int, col: int, valor: Any) -> None:
    cell = ws.cell(row=row, column=col, value=float(valor or 0))
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = Alignment(horizontal='right', vertical='center')


def _mapa_complemento_sabado(resumo: Dict[str, Any]) -> Dict[int, Dict[str, Any]]:
    out: Dict[int, Dict[str, Any]] = {}
    for item in resumo.get('detalhes_complemento_sabado') or []:
        vid = item.get('venda_id')
        if vid is not None:
            out[int(vid)] = item
    info = (resumo.get('info_comissao_adiantada') or {}).get('complemento_sabado') or {}
    for item in info.get('detalhes_vendas') or []:
        vid = item.get('venda_id')
        if vid is not None and int(vid) not in out:
            out[int(vid)] = item
    return out


def _aba_resumo(ws, vendedor_data: Dict[str, Any], periodo: str) -> None:
    resumo = vendedor_data.get('resumo') or {}
    nome = vendedor_data.get('vendedor_nome') or ''
    ws.append(['AUDITORIA DE COMISSÃO'])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=14)

    linhas_info = [
        ('Vendedor', nome),
        ('Período', periodo),
        ('Faixa aplicada', resumo.get('faixa_aplicada') or '-'),
        ('Total instalados (faixa)', resumo.get('qtd_instalada_faixa_complemento')),
        ('Vendas a pagar', resumo.get('total_qtd_instalada_a_pagar')),
        ('Vendas antecipadas', resumo.get('total_qtd_instalada_antecipada')),
        ('Total vendas no mês', resumo.get('total_qtd_vendas_folha')),
    ]
    row = 3
    for rotulo, valor in linhas_info:
        ws.cell(row=row, column=1, value=rotulo).font = Font(bold=True)
        ws.cell(row=row, column=2, value=valor)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='CÁLCULO DO LÍQUIDO').font = Font(bold=True, size=12)
    row += 1
    calculo = [
        ('Comissão (a pagar)', float(resumo.get('comissao_total_geral') or 0), False),
        ('Complemento sábado', float(resumo.get('complemento_sabado_total') or 0), False),
        ('Bônus / premiação', float(resumo.get('total_bonus') or 0), False),
        ('Total descontos', float(resumo.get('total_descontos') or 0), True),
        ('LÍQUIDO A PAGAR', float(resumo.get('liquido') or 0), False),
    ]
    ws.cell(row=row, column=1, value='Item').font = Font(bold=True)
    ws.cell(row=row, column=2, value='Valor').font = Font(bold=True)
    row += 1
    for rotulo, valor, negativo in calculo:
        ws.cell(row=row, column=1, value=rotulo)
        sinal = -valor if negativo else valor
        _fmt_money_cell(ws, row, 2, sinal)
        if rotulo == 'LÍQUIDO A PAGAR':
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

    row += 1
    comissao = float(resumo.get('comissao_total_geral') or 0)
    comp = float(resumo.get('complemento_sabado_total') or 0)
    bonus = float(resumo.get('total_bonus') or 0)
    desc = float(resumo.get('total_descontos') or 0)
    liq = float(resumo.get('liquido') or 0)
    ws.cell(row=row, column=1, value='Fórmula')
    ws.cell(
        row=row,
        column=2,
        value=f'{comissao:.2f} + {comp:.2f} + {bonus:.2f} - {desc:.2f} = {liq:.2f}',
    )

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22


def _aba_por_plano(ws, resumo: Dict[str, Any]) -> None:
    headers = [
        'PLANO', 'QTD A PAGAR', 'QTD ANTECIPADA', 'VALOR JÁ ADIANT.',
        'COMPLEMENTO SÁBADO', 'VALOR UNIT.', 'VALOR TOTAL', 'COMISSÃO',
    ]
    ws.append(headers)
    _estilo_cabecalho(ws)

    sum_q_ant = 0
    sum_adiant = 0.0
    sum_comp = 0.0
    sum_total = 0.0
    sum_com = 0.0
    total_q_pagar = int(resumo.get('total_qtd_instalada_a_pagar') or 0)

    for p in resumo.get('por_plano') or []:
        q_ant = int(p.get('qtd_antecipada') or 0)
        v_adiant = float(p.get('valor_total_antecipado') or 0)
        v_comp = float(p.get('valor_total_complemento_sabado') or 0)
        v_tot = float(p.get('valor_total_instalados') or 0)
        com = float(p.get('comissao_total') or 0)
        if not any([p.get('qtd_instalada_a_pagar'), q_ant, v_tot, v_adiant, v_comp, com]):
            continue
        sum_q_ant += q_ant
        sum_adiant += v_adiant
        sum_comp += v_comp
        sum_total += v_tot
        sum_com += com
        ws.append([
            p.get('plano'),
            int(p.get('qtd_instalada_a_pagar') or 0),
            q_ant,
            v_adiant,
            v_comp,
            p.get('valor_unitario_instalados'),
            v_tot,
            com,
        ])
        r = ws.max_row
        for col in (4, 5, 6, 7, 8):
            _fmt_money_cell(ws, r, col, ws.cell(row=r, column=col).value)

    ws.append(['TOTAL', total_q_pagar, sum_q_ant, sum_adiant, sum_comp, None, sum_total, sum_com])
    r = ws.max_row
    for col in range(1, 9):
        ws.cell(row=r, column=col).font = Font(bold=True)
    for col in (4, 5, 7, 8):
        _fmt_money_cell(ws, r, col, ws.cell(row=r, column=col).value)

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)
    ws.freeze_panes = 'A2'


def _aba_vendas(
    ws,
    vendedor_data: Dict[str, Any],
    formas_pagamento: Dict[int, str] | None = None,
) -> None:
    resumo = vendedor_data.get('resumo') or {}
    comp_map = _mapa_complemento_sabado(resumo)
    formas_pagamento = formas_pagamento or {}

    headers = [
        'VENDA_ID', 'NOME', 'CPF/CNPJ', 'FORMA PAGAMENTO', 'DACC', 'CNPJ', 'MEI/NMEI',
        'PLANO', 'DT PEDIDO', 'DT INST', 'OS', 'SITUAÇÃO', 'CHURN', 'ADIANT.',
        'COMISSÃO', 'TIPO COMISSÃO', 'PAGO SÁBADO', 'ALVO FAIXA', 'COMPLEMENTO SÁBADO',
        'ENTRA COMISSÃO A PAGAR',
    ]
    ws.append(headers)
    _estilo_cabecalho(ws, ncols=len(headers))

    extrato = sorted(
        vendedor_data.get('extrato') or [],
        key=lambda x: (_parse_br_date_key(x.get('dt_pedido')), (x.get('nome') or '').upper()),
    )
    idx_com = headers.index('COMISSÃO') + 1
    idx_pago = headers.index('PAGO SÁBADO') + 1
    idx_alvo = headers.index('ALVO FAIXA') + 1
    idx_comp = headers.index('COMPLEMENTO SÁBADO') + 1
    dinheiro_fmt = 'R$ #,##0.00'

    for e in extrato:
        vid = e.get('venda_id')
        comp = comp_map.get(int(vid)) if vid is not None else None
        tipo = (e.get('comissao_tipo_codigo') or e.get('comissao_tipo') or '').lower()
        adiant = (e.get('adiantada') or '').upper() == 'SIM'
        situacao = _norm(e.get('situacao'))
        entra_a_pagar = (
            situacao == 'INSTALADA'
            and (e.get('churn') or '').upper() != 'SIM'
            and not adiant
        )
        ws.append([
            vid,
            e.get('nome') or '',
            e.get('cpf_cnpj') or '',
            formas_pagamento.get(int(vid), '') if vid is not None else '',
            e.get('dacc') or '',
            e.get('cnpj') or '',
            e.get('classificacao_mei') or '-',
            e.get('plano') or '',
            e.get('dt_pedido') or '',
            e.get('dt_inst') or '',
            e.get('os') or '',
            e.get('situacao') or '',
            e.get('churn') or '',
            e.get('adiantada') or '-',
            e.get('valor_comissao'),
            e.get('comissao_tipo') or '',
            comp.get('pago') if comp else None,
            comp.get('alvo') if comp else None,
            comp.get('complemento') if comp else None,
            'SIM' if entra_a_pagar else 'NÃO',
        ])
        r = ws.max_row
        for col in (idx_com, idx_pago, idx_alvo, idx_comp):
            val = ws.cell(row=r, column=col).value
            if val is not None and val != '':
                ws.cell(row=r, column=col).number_format = dinheiro_fmt
                ws.cell(row=r, column=col).alignment = Alignment(horizontal='right')

    widths = {
        'NOME': 34, 'PLANO': 14, 'SITUAÇÃO': 22, 'TIPO COMISSÃO': 30, 'FORMA PAGAMENTO': 18,
    }
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 12)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'


def _aba_lista_simples(ws, titulo: str, headers: List[str], linhas: List[List[Any]], money_cols: List[int] | None = None) -> None:
    ws.append([titulo])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=12)
    ws.append(headers)
    _estilo_cabecalho(ws, row=2, ncols=len(headers))
    money_cols = money_cols or []
    for linha in linhas:
        ws.append(linha)
        r = ws.max_row
        for col in money_cols:
            val = ws.cell(row=r, column=col).value
            if val is not None and val != '':
                _fmt_money_cell(ws, r, col, val)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 2)
    ws.freeze_panes = 'A3'


def gerar_xlsx_auditoria_comissao(
    vendedor_data: Dict[str, Any],
    periodo: str,
    formas_pagamento: Dict[int, str] | None = None,
) -> BytesIO:
    """
    Planilha detalhada de auditoria: resumo, por plano, vendas, complementos, descontos e bônus.
    """
    resumo = vendedor_data.get('resumo') or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_resumo = wb.create_sheet('Resumo')
    _aba_resumo(ws_resumo, vendedor_data, periodo)

    ws_plano = wb.create_sheet('Por Plano')
    _aba_por_plano(ws_plano, resumo)

    ws_vendas = wb.create_sheet('Vendas')
    _aba_vendas(ws_vendas, vendedor_data, formas_pagamento)

    comp_linhas = []
    for item in resumo.get('detalhes_complemento_sabado') or []:
        comp_linhas.append([
            item.get('venda_id'),
            item.get('os'),
            item.get('plano'),
            item.get('faixa_nome'),
            item.get('pago'),
            item.get('alvo'),
            item.get('complemento'),
        ])
    ws_comp = wb.create_sheet('Complemento Sabado')
    _aba_lista_simples(
        ws_comp,
        'Complemento de adiantamento sábado (faixa alcançada)',
        ['VENDA_ID', 'OS', 'PLANO', 'FAIXA', 'PAGO SÁBADO', 'ALVO FAIXA', 'COMPLEMENTO'],
        comp_linhas,
        money_cols=[5, 6, 7],
    )
    if comp_linhas:
        r = ws_comp.max_row + 1
        total = sum(float(x[6] or 0) for x in comp_linhas)
        ws_comp.cell(row=r, column=6, value='TOTAL').font = Font(bold=True)
        _fmt_money_cell(ws_comp, r, 7, total)

    desc_linhas = []
    for d in resumo.get('detalhes_descontos') or []:
        desc_linhas.append([
            d.get('motivo'),
            d.get('tipo_exibicao'),
            d.get('quantidade'),
            d.get('valor'),
        ])
    ws_desc = wb.create_sheet('Descontos')
    _aba_lista_simples(
        ws_desc,
        'Lançamentos descontados na folha',
        ['MOTIVO', 'TIPO', 'QUANTIDADE', 'VALOR'],
        desc_linhas,
        money_cols=[4],
    )
    if desc_linhas:
        r = ws_desc.max_row + 1
        total = sum(float(x[3] or 0) for x in desc_linhas)
        ws_desc.cell(row=r, column=3, value='TOTAL').font = Font(bold=True)
        _fmt_money_cell(ws_desc, r, 4, total)

    bonus_linhas = [[b.get('motivo'), b.get('valor')] for b in (resumo.get('detalhes_bonus') or [])]
    ws_bonus = wb.create_sheet('Bonus')
    _aba_lista_simples(
        ws_bonus,
        'Bônus e premiação',
        ['MOTIVO', 'VALOR'],
        bonus_linhas or [['Nenhum bônus no período', 0]],
        money_cols=[2],
    )

    info_ad = resumo.get('info_comissao_adiantada') or {}
    if int(info_ad.get('quantidade_total') or 0) > 0:
        ref_linhas = []
        for origem, dados in (info_ad.get('por_origem') or {}).items():
            ref_linhas.append([origem, dados.get('quantidade'), dados.get('valor_total')])
        ws_ref = wb.create_sheet('Adiantamentos Ref')
        _aba_lista_simples(
            ws_ref,
            'Referência — adiantamentos (não entram em descontos automáticos da mesma forma)',
            ['ORIGEM', 'QUANTIDADE', 'VALOR TOTAL'],
            ref_linhas,
            money_cols=[3],
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_xlsx_extrato_comissao(vendedores_data: List[Dict[str, Any]]) -> BytesIO:
    """Gera XLSX com extrato igual à tela (uma aba por vendedor)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    usados: set = set()

    if not vendedores_data:
        ws = wb.create_sheet(title='Extrato')
        ws.append(['Sem dados de extrato para o período'])
    else:
        for vd in vendedores_data:
            nome = vd.get('vendedor_nome') or f"ID {vd.get('vendedor_id')}"
            ws = wb.create_sheet(title=_titulo_aba_seguro(nome, usados))
            _preencher_aba(ws, _linhas_extrato_vendedor(vd))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
