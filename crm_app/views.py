from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes
from rest_framework import permissions
from rest_framework.response import Response

# Endpoint de validação de e-mail
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def api_verificar_email(request, email=None):
    """
    Valida se um e-mail tem formato correto.
    Retorna {"valido": true/false, "mensagem": "..."}
    """
    import re
    
    if not email or not email.strip():
        return Response({"valido": False, "mensagem": "E-mail não informado"})
    
    email = email.strip().lower()
    
    # Regex para validação de e-mail (padrão RFC 5322 simplificado)
    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    if not re.match(email_regex, email):
        return Response({"valido": False, "mensagem": "Formato de e-mail inválido"})
    
    # Lista de domínios comuns conhecidos (opcional: pode adicionar mais validações)
    dominios_suspeitos = ['teste.com', 'test.com', 'exemplo.com', 'example.com']
    dominio = email.split('@')[-1]
    
    if dominio in dominios_suspeitos:
        return Response({"valido": False, "mensagem": f"Domínio '{dominio}' não é aceito"})
    
    return Response({"valido": True, "mensagem": "E-mail válido", "email": email})

# Endpoint de validação WhatsApp
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def api_verificar_whatsapp(request, telefone=None):
    """
    Valida se um número possui WhatsApp ativo via Z-API.
    Aceita telefone como parâmetro de URL ou query string.
    """
    # Pega telefone da URL ou query string
    if not telefone:
        telefone = request.GET.get('numero') or request.GET.get('telefone')
    
    if not telefone:
        return Response({
            "exists": False,
            "whatsapp_valido": False,
            "possui_whatsapp": False,
            "erro": "Número não informado"
        }, status=400)
    
    # Limpa o telefone (remove caracteres não numéricos)
    import re
    telefone_limpo = re.sub(r'\D', '', str(telefone))
    
    if len(telefone_limpo) < 10:
        return Response({
            "exists": False,
            "whatsapp_valido": False,
            "possui_whatsapp": False,
            "erro": "Número inválido (mínimo 10 dígitos)"
        }, status=400)
    
    try:
        from crm_app.whatsapp_service import WhatsAppService
        
        service = WhatsAppService()
        
        # Verifica se a API está configurada
        if not service.instance_id or not service.token:
            return Response({
                "exists": True,  # Retorna True para não bloquear o cadastro
                "whatsapp_valido": True,
                "possui_whatsapp": True,
                "aviso": "API WhatsApp não configurada. Validação ignorada."
            }, status=200)
        
        # Formata telefone para a API (adiciona 55 se necessário)
        if len(telefone_limpo) <= 11:
            telefone_api = f"55{telefone_limpo}"
        else:
            telefone_api = telefone_limpo
        
        # Verifica se o número existe no WhatsApp
        try:
            existe = service.verificar_numero_existe(telefone_api)
            
            # Se a API retornou None (erro), trata como se não existisse mas permite salvar
            if existe is None:
                return Response({
                    "exists": True,  # Permite salvar mesmo se não conseguir verificar
                    "whatsapp_valido": True,
                    "possui_whatsapp": True,
                    "aviso": "Não foi possível verificar WhatsApp (API offline). Você pode salvar mesmo assim."
                }, status=200)
            
            return Response({
                "exists": existe,
                "whatsapp_valido": existe,
                "possui_whatsapp": existe,
                "telefone": telefone_limpo
            }, status=200)
        except Exception as e_inner:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception(f"Erro ao chamar verificar_numero_existe: {e_inner}")
            # Em caso de erro na verificação, permite salvar mas avisa
            return Response({
                "exists": True,
                "whatsapp_valido": True,
                "possui_whatsapp": True,
                "aviso": f"Erro ao verificar: {str(e_inner)}. Você pode salvar mesmo assim."
            }, status=200)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception(f"Erro ao verificar WhatsApp: {e}")
        
        # Retorna True com aviso para não bloquear o cadastro em caso de erro
        return Response({
            "exists": True,
            "whatsapp_valido": True,
            "possui_whatsapp": True,
            "aviso": f"Erro ao verificar WhatsApp: {str(e)}. Validação ignorada."
        }, status=200)
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from rest_framework import status

# --- RESTAURAÇÃO: WebhookWhatsAppView ---
# URL de produção: https://www.recordpap.com.br/api/crm/webhook-whatsapp/
# URL alternativa: https://site-record-production.up.railway.app/api/crm/webhook-whatsapp/
# Configurar no Z-API: Método POST, Eventos: Mensagens recebidas
class WebhookWhatsAppView(APIView):
    permission_classes = [AllowAny]  # Permite acesso sem autenticação para webhooks

    def post(self, request, *args, **kwargs):
        """
        Endpoint para receber eventos do WhatsApp e processar fluxos.
        Sempre retorna uma resposta HTTP para evitar 502 (ngrok/Z-API).
        """
        import logging
        logger_webhook = logging.getLogger(__name__)
        print("[WEBHOOK] >>> POST /api/crm/webhook-whatsapp/ recebido <<<", flush=True)
        logger_webhook.info("[WebhookWhatsAppView] POST recebido (início)")
        
        try:
            data = request.data
            if isinstance(request.data, dict):
                data = request.data
            else:
                try:
                    import json
                    data = json.loads(request.body) if request.body else {}
                except Exception:
                    data = {}
            logger_webhook.info(f"[WebhookWhatsAppView] Dados recebidos (keys): {list(data.keys()) if isinstance(data, dict) else type(data)}")
        except Exception as e:
            logger_webhook.exception(f"[WebhookWhatsAppView] Erro ao ler request.data: {e}")
            return Response({'status': 'ok', 'mensagem': 'Payload inválido'}, status=200)
        
        try:
            from crm_app.whatsapp_webhook_handler import processar_webhook_whatsapp
            resultado = processar_webhook_whatsapp(data, request=request)
            logger_webhook.info(f"[WebhookWhatsAppView] Resultado: {resultado.get('status', '?')}")
            return Response(resultado, status=200 if resultado.get('status') == 'ok' else 500)
        except Exception as e:
            logger_webhook.exception(f"[WebhookWhatsAppView] Erro no processamento: {e}")
            return Response({'status': 'erro', 'mensagem': str(e)}, status=500)


def serve_pdf_view(request, token):
    """
    Serve um PDF da pasta downloads/ com token assinado (para Z-API buscar o arquivo por URL).
    Uso: /api/crm/serve-pdf/<token>/ → retorna o PDF.
    Token = base64url(filename) + "." + hmac(SECRET_KEY, filename).hexdigest()[:32]
    """
    import hmac
    import base64
    from django.http import FileResponse, HttpResponseNotFound

    if not token or "." not in token:
        return HttpResponseNotFound("Token inválido")
    parts = token.split(".", 1)
    try:
        payload_b64 = parts[0]
        payload_b64 += "=" * (4 - len(payload_b64) % 4)  # padding
        filename = base64.urlsafe_b64decode(payload_b64).decode("utf-8")
    except Exception:
        return HttpResponseNotFound("Token inválido")
    if not re.match(r"^[a-zA-Z0-9_.-]+$", filename):
        return HttpResponseNotFound("Nome de arquivo inválido")
    secret = (getattr(settings, "SECRET_KEY", "") or "").encode("utf-8")
    expected_sig = hmac.new(secret, filename.encode("utf-8"), "sha256").hexdigest()[:32]
    if parts[1] != expected_sig:
        return HttpResponseNotFound("Token inválido")
    downloads_dir = os.path.join(settings.BASE_DIR, "downloads")
    filepath = os.path.join(downloads_dir, filename)
    if not os.path.abspath(filepath).startswith(os.path.abspath(downloads_dir)) or not os.path.isfile(filepath):
        return HttpResponseNotFound("Arquivo não encontrado")
    response = FileResponse(open(filepath, "rb"), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# Endpoint para duplicar venda (Reemissão)
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Venda

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def duplicar_venda(request):
    """
    Duplica uma venda (reemissão). Valida entrada e delega ao ReemissaoVendaService.
    """
    id_venda = request.data.get("id_venda")
    nova_os = request.data.get("nova_os") or request.data.get("ordem_servico")
    nova_data = request.data.get("nova_data_agendamento") or request.data.get("data_agendamento")
    novo_turno = request.data.get("novo_turno") or request.data.get("periodo_agendamento")

    if not (id_venda and nova_os and nova_data and novo_turno):
        return Response(
            {
                "detail": "Dados obrigatórios faltando: id_venda, ordem_servico, data_agendamento, periodo_agendamento."
            },
            status=400,
        )

    try:
        from crm_app.services.reemissao_venda_service import (
            ReemissaoVendaError,
            duplicar as reemissao_duplicar,
        )

        venda_nova = reemissao_duplicar(
            id_venda=int(id_venda),
            nova_os=str(nova_os).strip(),
            nova_data=nova_data,
            novo_turno=str(novo_turno).strip(),
            enviar_whatsapp=True,
        )
        return Response(
            {
                "success": True,
                "nova_venda_id": venda_nova.id,
                "message": "Reemissão criada com sucesso!",
            }
        )
    except Venda.DoesNotExist:
        return Response({"detail": "Venda não encontrada."}, status=404)
    except ReemissaoVendaError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception("Erro ao duplicar venda: %s", e)
        return Response({"detail": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def liberar_pap_bo_view(request):
    """
    Libera todos os logins PAP (remove locks PapBoEmUso).
    Use quando os logins ficaram travados sem uso (ex.: sessão caiu).
    Requer usuário autenticado.
    """
    from crm_app.pool_bo_pap import liberar_todos_bos
    qtd, msg = liberar_todos_bos()
    return Response({"success": True, "liberados": qtd, "message": msg})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def historico_consultas_pap_bo_view(request):
    """
    Lista histórico de consultas das automações ao pool de login PAP.
    """
    from crm_app.models import HistoricoConsultaAutomacaoPAP

    limite_raw = request.GET.get("limite", "200")
    try:
        limite = max(1, min(int(limite_raw), 1000))
    except (TypeError, ValueError):
        limite = 200

    queryset = (
        HistoricoConsultaAutomacaoPAP.objects.select_related(
            "solicitado_por",
            "login_pap_utilizado",
        )
        .order_by("-criado_em")[:limite]
    )

    data = []
    for item in queryset:
        data.append(
            {
                "id": item.id,
                "data_hora": item.criado_em,
                "automacao": item.get_tipo_automacao_display() if item.tipo_automacao else "Não informado",
                "automacao_cod": item.tipo_automacao or "",
                "usuario_solicitante": (
                    item.solicitado_por.username if item.solicitado_por else None
                ),
                "telefone_solicitante": item.telefone_solicitante or "",
                "login_pap_usuario": (
                    item.login_pap_utilizado.username if item.login_pap_utilizado else None
                ),
                "matricula_pap": item.matricula_pap_utilizada or "",
                "status_execucao": item.status_execucao or "",
                "status_execucao_label": item.get_status_execucao_display() if item.status_execucao else "",
                "mensagem_resultado": item.mensagem_resultado or "",
            }
        )

    return Response({"count": len(data), "results": data})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def consultar_biometria_brpronto_view(request):
    """
    Consulta biometria no Br Pronto PDV para um CPF.
    Usa as credenciais Br Pronto do usuário logado.
    Body: { "cpf": "12345678900" } ou { "venda_id": 123 }.
    """
    import re
    from crm_app.services_brpronto import consultar_biometria_brpronto
    from crm_app.models import Venda

    user = request.user
    login = getattr(user, "brpronto_login", None) or ""
    senha = getattr(user, "brpronto_senha", None) or ""
    dominio = getattr(user, "brpronto_dominio", None) or ""

    cpf = request.data.get("cpf")
    venda_id = request.data.get("venda_id")

    if venda_id is not None:
        try:
            venda = Venda.objects.select_related("cliente").get(pk=venda_id)
        except Venda.DoesNotExist:
            return Response({"ok": False, "error": "Venda não encontrada."}, status=404)
        cliente_cpf_cnpj = (venda.cliente.cpf_cnpj or "").replace(".", "").replace("-", "").replace("/", "").strip()
        if len(cliente_cpf_cnpj) == 11:
            cpf = cliente_cpf_cnpj
        else:
            cpf = (venda.cpf_representante_legal or "").replace(".", "").replace("-", "").strip() or cliente_cpf_cnpj
    if not cpf:
        return Response({"ok": False, "error": "Informe cpf ou venda_id."}, status=400)

    cpf_limpo = re.sub(r"\D", "", str(cpf))
    if len(cpf_limpo) != 11:
        return Response({"ok": False, "error": "CPF deve ter 11 dígitos."}, status=400)

    sucesso, msg_erro, resultado = consultar_biometria_brpronto(
        login=login,
        senha=senha,
        cpf=cpf_limpo,
        dominio=dominio or None,
        headless=True,
    )
    if not sucesso:
        return Response({"ok": False, "error": msg_erro or "Erro ao consultar Br Pronto."}, status=400)
    return Response({
        "ok": True,
        "aprovada": resultado.get("aprovada", False),
        "data_mais_recente_apta": resultado.get("data_mais_recente_apta"),
        "registros": resultado.get("registros", []),
    })


from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
# --- NOVO ENDPOINT: Buscar Fatura NIO para Bonus M-10 ---
# Plano A = mesma consulta do WhatsApp: consultar_dividas_nio (API /params + REST). Sem Playwright.
@api_view(["POST"])
@permission_classes([AllowAny])
def buscar_fatura_nio_bonus_m10(request):
    import logging
    import re
    import requests
    from datetime import datetime

    logger = logging.getLogger(__name__)
    cpf = request.data.get("cpf")

    if not cpf:
        return Response({"error": "CPF não informado."}, status=400)

    cpf_limpo = re.sub(r"\D", "", str(cpf))
    if not cpf_limpo or len(cpf_limpo) < 11:
        return Response({"error": "CPF inválido."}, status=400)

    def _fmt_date(val):
        if val is None:
            return None
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y%m%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
        return None

    try:
        from crm_app.nio_api import consultar_dividas_nio, get_invoice_pdf_url

        api_result = consultar_dividas_nio(cpf_limpo, offset=0, limit=50, headless=True)
        invoices = api_result.get("invoices") or []

        if not invoices:
            return Response({
                "success": False,
                "message": "Nenhuma fatura encontrada para este CPF.",
            }, status=404)

        inv = invoices[0]
        valor = inv.get("amount")
        codigo_pix = inv.get("pix") or inv.get("codigo_pix")
        codigo_barras = inv.get("barcode") or inv.get("codigo_barras")
        due = inv.get("due_date_raw") or inv.get("data_vencimento")
        data_vencimento = None
        if due:
            if hasattr(due, "strftime"):
                data_vencimento = due
            elif isinstance(due, str) and len(due) >= 8:
                try:
                    s = due[:10].replace("/", "-")
                    if "-" in s:
                        data_vencimento = datetime.strptime(s, "%Y-%m-%d").date()
                    elif due[:8].isdigit():
                        data_vencimento = datetime.strptime(due[:8], "%Y%m%d").date()
                except Exception:
                    pass

        pdf_url = None
        if api_result.get("token") and api_result.get("api_base") and api_result.get("session_id"):
            sess = requests.Session()
            pdf_url = get_invoice_pdf_url(
                api_result["api_base"],
                api_result["token"],
                api_result["session_id"],
                inv.get("debt_id", ""),
                str(inv.get("invoice_id", "")),
                cpf_limpo,
                inv.get("reference_month", "") or "",
                sess,
            )

        dv = data_vencimento
        return Response({
            "success": True,
            "valor": valor,
            "codigo_pix": codigo_pix,
            "codigo_barras": codigo_barras,
            "data_vencimento": dv.strftime("%Y-%m-%d") if hasattr(dv, "strftime") else _fmt_date(dv),
            "pdf_url": pdf_url,
            "metodo_usado": "api_nio",
        })
    except Exception as e:
        logger.exception(f"[Bonus M-10] Erro ao buscar fatura (Plano A / API Nio): {e}")
        return Response({
            "success": False,
            "message": "Não foi possível buscar a fatura. Verifique o CPF ou tente novamente.",
        }, status=500)
import logging
import pandas as pd
import numpy as np
from decimal import Decimal
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from operator import itemgetter
import base64
from io import BytesIO
import json
import os
import calendar
import re
import xml.etree.ElementTree as ET
import unicodedata
import threading
import requests
from email_validator import validate_email, EmailNotValidError


# --- CORREÇÃO CRÍTICA: Importar transaction e IntegrityError ---
from django.db import transaction, IntegrityError
from django.db.models import Count, Q, Sum, F
from django.utils import timezone
from django.utils.timezone import now
from django.contrib.auth import get_user_model, authenticate
from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.shortcuts import render  # <--- Adicionado para renderizar HTML
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from django.contrib.staticfiles import finders
from django.core.exceptions import ValidationError
from usuarios.models import Usuario
from .models import CdoiSolicitacao, CdoiBloco, PreVenda, LinkPublicoPreVenda
from .onedrive_service import OneDriveUploader

# Importar mapeamento de status FPD
from .fpd_status_mapping import normalizar_status_fpd

from rest_framework import generics, viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.exceptions import PermissionDenied, NotFound
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.pagination import PageNumberPagination
from openpyxl.utils import get_column_letter


class VendaPagination(PageNumberPagination):
    """Paginação que aceita page_size na query (até 1000), para esteira e listagens grandes."""
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 1000


# --- IMPORTS EXTRAS DO PROJETO ---
from core.models import DiaFiscal, RegraAutomacao
from core.validators import validar_cpf, validar_cnpj, validar_cpf_ou_cnpj
from .whatsapp_service import WhatsAppService
from xhtml2pdf import pisa
from usuarios.permissions import CheckAPIPermission, VendaPermission
import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import GrupoDisparo
from .serializers import GrupoDisparoSerializer
from .models import LancamentoFinanceiro
from .serializers import LancamentoFinanceiroSerializer
from .nio_api import consultar_dividas_nio


# Funções de Mapa (Geometria e Busca)
from .utils import (
    is_member,
    verificar_viabilidade_por_cep,
    verificar_viabilidade_por_coordenadas,
    verificar_viabilidade_exata,
    montar_resumo_plano_para_whatsapp,
    validar_venda_para_resumo_auditoria,
    vendedor_ou_supervisor_restrito_mes,
    mes_completo_vendedor_supervisor_valido,
    q_venda_acesso_retrieve_vendedor_supervisor,
)

# Modelos do App
from .models import (
    Operadora, Plano, FormaPagamento, StatusCRM, MotivoPendencia,
    RegraComissao, Cliente, Venda, ImportacaoOsab, ImportacaoChurn,
    CicloPagamento, HistoricoAlteracaoVenda, PagamentoComissao, PagamentoComissaoItem,
    Campanha, ComissaoOperadora, Comunicado, AreaVenda,
    SessaoWhatsapp, DFV, GrupoDisparo, LancamentoFinanceiro,
    AgendamentoDisparo, LogEnvioPerformance, ImportacaoAgendamento, ImportacaoRecompra,
    LogImportacaoAgendamento, LogImportacaoLegado, LogImportacaoRecompra, EstatisticaBotWhatsApp,
    RegraComissaoFaixa, ConfigComissaoVendedor,
    AnteciparInstalacaoConfig, AnteciparInstalacaoSolicitacao,
    PapConfirmacaoCliente, LembreteInstalacaoEnviado,
    ControleTTDiaTratado,
    BoasVindasEnviado, MensagemClienteBoasVindas, StatusBoasVindas, FilaEnvioBoasVindas,
)
from .antecipar_instalacao_utils import mensagem_resposta_gc_para_vendedor

# Serializers do App
from .serializers import (
    OperadoraSerializer, PlanoSerializer, FormaPagamentoSerializer,
    StatusCRMSerializer, MotivoPendenciaSerializer, RegraComissaoSerializer,
    RegraComissaoFaixaSerializer, ConfigComissaoVendedorSerializer,
    VendaSerializer, VendaCreateSerializer, ClienteSerializer,
    VendaUpdateSerializer, ImportacaoOsabSerializer, ImportacaoChurnSerializer,
    CicloPagamentoSerializer, VendaDetailSerializer,
    CampanhaSerializer, ComissaoOperadoraSerializer, ComunicadoSerializer,
    FaturaM10Serializer
)

logger = logging.getLogger(__name__)

def get_osab_bot_user():
    User = get_user_model()
    bot, created = User.objects.get_or_create(
        username='OSAB_IMPORT',
        defaults={
            'first_name': 'OSAB',
            'last_name': 'Automático',
            'email': 'osab_import@sistema.local',
            'is_active': True,
            'password': 'senha_temporaria_validacao_sistema'
        }
    )
    if created:
        bot.set_unusable_password()
        bot.save()
    return bot

class OperadoraListCreateView(generics.ListCreateAPIView):
    queryset = Operadora.objects.filter(ativo=True)
    serializer_class = OperadoraSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'operadora'

class OperadoraDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Operadora.objects.all()
    serializer_class = OperadoraSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'operadora'

class PlanoListCreateView(generics.ListCreateAPIView):
    serializer_class = PlanoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Plano.objects.filter(ativo=True)
        operadora_id = self.request.query_params.get('operadora')
        if operadora_id:
            queryset = queryset.filter(operadora_id=operadora_id)
        return queryset

class PlanoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Plano.objects.all()
    serializer_class = PlanoSerializer
    permission_classes = [permissions.IsAuthenticated]

class FormaPagamentoListCreateView(generics.ListCreateAPIView):
    queryset = FormaPagamento.objects.filter(ativo=True)
    serializer_class = FormaPagamentoSerializer
    permission_classes = [permissions.IsAuthenticated]

class FormaPagamentoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = FormaPagamento.objects.all()
    serializer_class = FormaPagamentoSerializer
    permission_classes = [permissions.IsAuthenticated]

class CampanhaListCreateView(generics.ListCreateAPIView):
    queryset = Campanha.objects.filter(ativo=True)
    serializer_class = CampanhaSerializer
    permission_classes = [permissions.IsAuthenticated]

class CampanhaDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Campanha.objects.all()
    serializer_class = CampanhaSerializer
    permission_classes = [permissions.IsAuthenticated]

class ComissaoOperadoraViewSet(viewsets.ModelViewSet):
    queryset = ComissaoOperadora.objects.select_related('plano').all()
    serializer_class = ComissaoOperadoraSerializer
    permission_classes = [permissions.IsAuthenticated]

class StatusCRMListCreateView(generics.ListCreateAPIView):
    serializer_class = StatusCRMSerializer
    permission_classes = [permissions.IsAuthenticated]
    resource_name = 'statuscrm'
    pagination_class = None  # Retorna todos os status (lista pequena); evita status “sumidos” na aba e no modal Reprovar

    def get_queryset(self):
        user = self.request.user
        # Listar status: quem gerencia cadastros (Diretoria, BackOffice, Admin, Supervisor) ou quem usa na auditoria (Auditoria, Qualidade)
        if is_member(user, ['Diretoria', 'BackOffice', 'Admin', 'Supervisor', 'Auditoria', 'Qualidade']):
            queryset = StatusCRM.objects.all()
            tipo = self.request.query_params.get('tipo', None)
            if tipo:
                queryset = queryset.filter(tipo__iexact=tipo)
            return queryset.order_by('nome')
        return StatusCRM.objects.none()

    def perform_create(self, serializer):
        """Evita 500 por IntegrityError (nome+tipo duplicado); retorna 400 com mensagem."""
        from rest_framework.exceptions import ValidationError as DRFValidationError
        try:
            serializer.save()
        except IntegrityError as e:
            if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
                raise DRFValidationError(
                    {'nome': 'Já existe um status com este nome e tipo.'}
                )
            raise

class StatusCRMDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = StatusCRM.objects.all()
    serializer_class = StatusCRMSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'statuscrm'

    def perform_update(self, serializer):
        """Evita 500 por IntegrityError (nome+tipo duplicado) no PATCH."""
        from rest_framework.exceptions import ValidationError as DRFValidationError
        try:
            serializer.save()
        except IntegrityError as e:
            if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
                raise DRFValidationError(
                    {'nome': 'Já existe um status com este nome e tipo.'}
                )
            raise

class MotivoPendenciaListCreateView(generics.ListCreateAPIView):
    queryset = MotivoPendencia.objects.all().order_by('nome')
    serializer_class = MotivoPendenciaSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None  # Desabilita paginação - retorna todos os resultados
    
    def perform_create(self, serializer):
        """
        Sobrescreve perform_create para tratar erro de sequência dessincronizada.
        Se ocorrer erro de chave única, tenta corrigir a sequência e recriar.
        """
        from django.db import connection
        try:
            serializer.save()
        except IntegrityError as e:
            # Se for erro de sequência (chave primária duplicada)
            if 'pkey' in str(e) or 'unique constraint' in str(e).lower():
                # Tenta corrigir a sequência
                table_name = MotivoPendencia._meta.db_table
                with connection.cursor() as cursor:
                    cursor.execute(f"SELECT COALESCE(MAX(id), 0) FROM {table_name};")
                    max_id = cursor.fetchone()[0]
                    cursor.execute(f"SELECT pg_get_serial_sequence('{table_name}', 'id');")
                    seq_name = cursor.fetchone()[0]
                    if seq_name:
                        cursor.execute(f"SELECT setval(%s, %s, true);", [seq_name, max_id])
                # Tenta salvar novamente
                serializer.save()
            else:
                # Se for outro tipo de IntegrityError, re-lança
                raise

class MotivoPendenciaDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = MotivoPendencia.objects.all().order_by('nome')
    serializer_class = MotivoPendenciaSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'motivopendencia'

class RegraComissaoListCreateView(generics.ListCreateAPIView):
    queryset = RegraComissao.objects.all()
    serializer_class = RegraComissaoSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'regracomissao'

class RegraComissaoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = RegraComissao.objects.all()
    serializer_class = RegraComissaoSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'regracomissao'


class RegraComissaoFaixaListCreateView(generics.ListCreateAPIView):
    """Regras por faixa (REGRAS_FAIXAS). Filtros: perfil, vendedor_id."""
    queryset = RegraComissaoFaixa.objects.select_related('vendedor').all().order_by('perfil', 'vendedor', 'min_vendas')
    serializer_class = RegraComissaoFaixaSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        perfil = self.request.query_params.get('perfil')
        vendedor_id = self.request.query_params.get('vendedor_id')
        if perfil:
            qs = qs.filter(perfil=perfil)
        if vendedor_id:
            qs = qs.filter(vendedor_id=vendedor_id)
        return qs


class RegraComissaoFaixaDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = RegraComissaoFaixa.objects.all()
    serializer_class = RegraComissaoFaixaSerializer
    permission_classes = [permissions.IsAuthenticated]


def _config_vendedor_padrao(u):
    """Item padrão quando não existe config para o usuário."""
    return {
        'usuario_id': u.id,
        'username': u.username,
        'perfil_comissao': 'Vendedor',
        'usar_valor_manual': False,
        'valor_500mb_pap_manual': None,
        'valor_700mb_pap_manual': None,
        'valor_1gb_pap_manual': None,
        'valor_500mb_cnpj_manual': None,
        'valor_700mb_cnpj_manual': None,
        'valor_1gb_cnpj_manual': None,
        'desconta_dacc_pap': False,
        'desconto_boleto': None,
        'desconto_inclusao': None,
        'desconto_instalacao': None,
        'adiantar_cnpj': None,
        'inss_valor': None,
        'adiantamento': None,
        'premiação': None,
        'bonus_cartao_credito': None,
        'cartao_trafego': None,
        'gestor_trafego': None,
    }


class ConfigComissaoVendedorListView(APIView):
    """Lista todos os usuários ativos; retorna config de comissão (por mês se ano/mes informados). GET ?ano=&mes= para regras daquele mês."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        usuarios = User.objects.filter(is_active=True).order_by('username')
        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        try:
            ano = int(ano) if ano else None
            mes = int(mes) if mes else None
        except (TypeError, ValueError):
            ano, mes = None, None

        if ano is not None and mes is not None:
            # Regras do mês: (usuario, ano, mes) ou fallback (usuario, null, null)
            configs_mes = {c.usuario_id: c for c in ConfigComissaoVendedor.objects.filter(ano=ano, mes=mes).select_related('usuario')}
            configs_template = {c.usuario_id: c for c in ConfigComissaoVendedor.objects.filter(ano__isnull=True, mes__isnull=True).select_related('usuario')}
        else:
            configs_mes = {}
            configs_template = {c.usuario_id: c for c in ConfigComissaoVendedor.objects.filter(ano__isnull=True, mes__isnull=True).select_related('usuario')}

        lista = []
        for u in usuarios:
            c = configs_mes.get(u.id) or configs_template.get(u.id)
            if c:
                item = ConfigComissaoVendedorSerializer(c).data
                item['usuario_id'] = item.get('usuario') or u.id
                item['username'] = item.get('username') or u.username
            else:
                item = _config_vendedor_padrao(u)
            lista.append(item)
        return Response(lista)


class ConfigComissaoVendedorDetailView(APIView):
    """GET retorna config por user_id (e opcionalmente ano/mes). PUT faz upsert (aceita ano/mes no body para regra do mês)."""
    permission_classes = [permissions.IsAuthenticated]

    def _get_config(self, user_id, ano=None, mes=None):
        if ano is not None and mes is not None:
            c = ConfigComissaoVendedor.objects.filter(usuario_id=user_id, ano=ano, mes=mes).select_related('usuario').first()
            if c:
                return c
        return ConfigComissaoVendedor.objects.filter(usuario_id=user_id, ano__isnull=True, mes__isnull=True).select_related('usuario').first()

    def get(self, request, user_id):
        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        try:
            ano = int(ano) if ano else None
            mes = int(mes) if mes else None
        except (TypeError, ValueError):
            ano, mes = None, None
        config = self._get_config(user_id, ano, mes)
        if not config:
            return Response({'detail': 'Configuração não encontrada.'}, status=404)
        serializer = ConfigComissaoVendedorSerializer(config)
        return Response(serializer.data)

    def put(self, request, user_id):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        if not User.objects.filter(id=user_id).exists():
            return Response({'error': 'Usuário não encontrado'}, status=404)
        data = dict(request.data)
        ano = data.pop('ano', None)
        mes = data.pop('mes', None)
        try:
            ano = int(ano) if ano is not None else None
            mes = int(mes) if mes is not None else None
        except (TypeError, ValueError):
            ano, mes = None, None
        if ano is not None and mes is not None:
            config, _ = ConfigComissaoVendedor.objects.get_or_create(
                usuario_id=user_id, ano=ano, mes=mes, defaults={}
            )
        else:
            config, _ = ConfigComissaoVendedor.objects.get_or_create(
                usuario_id=user_id, ano=None, mes=None, defaults={'usuario_id': user_id}
            )
        serializer = ConfigComissaoVendedorSerializer(config, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    def patch(self, request, user_id):
        ano = request.data.get('ano')
        mes = request.data.get('mes')
        try:
            ano = int(ano) if ano is not None else None
            mes = int(mes) if mes is not None else None
        except (TypeError, ValueError):
            ano, mes = None, None
        config = self._get_config(user_id, ano, mes)
        if not config:
            return Response({'detail': 'Configuração não encontrada.'}, status=404)
        serializer = ConfigComissaoVendedorSerializer(config, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)


class ConfigComissaoVendedorSalvarMesView(APIView):
    """POST config-comissao-vendedor/salvar-mes/ - Body: { ano, mes, configs: [ { usuario_id, perfil_comissao, usar_valor_manual, ... }, ... ] }. Salva todas as regras daquele mês."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ano = request.data.get('ano')
        mes = request.data.get('mes')
        configs = request.data.get('configs') or []
        try:
            ano = int(ano)
            mes = int(mes)
        except (TypeError, ValueError):
            return Response({'error': 'ano e mes são obrigatórios e devem ser números.'}, status=400)
        if not 1 <= mes <= 12:
            return Response({'error': 'mes deve ser entre 1 e 12.'}, status=400)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        salvos = 0
        erros = []
        for c in configs:
            uid = c.get('usuario_id') or c.get('usuario')
            if not uid:
                continue
            try:
                if not User.objects.filter(id=uid).exists():
                    erros.append(f"Usuário {uid} não encontrado")
                    continue
                data = dict(c)
                data.pop('usuario_id', None)
                data.pop('usuario', None)
                data.pop('username', None)
                data.pop('ano', None)
                data.pop('mes', None)
                config, created = ConfigComissaoVendedor.objects.get_or_create(
                    usuario_id=uid, ano=ano, mes=mes, defaults={}
                )
                serializer = ConfigComissaoVendedorSerializer(config, data=data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    salvos += 1
                else:
                    erros.append(f"User {uid}: " + str(serializer.errors))
            except Exception as e:
                erros.append(f"User {uid}: {e}")
        return Response({'salvos': salvos, 'erros': erros})


def _decimal_ou_none(val):
    from decimal import Decimal
    if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
        return None
    try:
        return Decimal(str(val).replace(',', '.'))
    except Exception:
        return None


def _int_ou_none(val):
    if val is None or val == '' or (isinstance(val, str) and val.strip() == ''):
        return None
    try:
        return int(float(val))
    except Exception:
        return None


class RegraComissaoFaixaExportarView(APIView):
    """GET regras-comissao-faixa/exportar/?formato=xlsx|csv - Download Excel ou CSV."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        formato = (request.query_params.get('formato') or 'xlsx').lower()
        qs = RegraComissaoFaixa.objects.select_related('vendedor').all().order_by('perfil', 'vendedor', 'min_vendas')
        rows = [[
            'PERFIL', 'FINALIDADE', 'FAIXA_NOME', 'MIN_VENDAS', 'MAX_VENDAS',
            'VALOR_500MB_PAP', 'VALOR_700MB_PAP', 'VALOR_1GB_PAP',
            'VALOR_500MB_CNPJ', 'VALOR_700MB_CNPJ', 'VALOR_1GB_CNPJ',
        ]]
        for r in qs:
            perfil = r.vendedor.username if r.vendedor_id else (r.perfil or '')
            fin = getattr(r, 'finalidade', None) or 'COMISSAO'
            rows.append([
                perfil, fin, r.faixa_nome, r.min_vendas, r.max_vendas,
                r.valor_500mb_pap, r.valor_700mb_pap, r.valor_1gb_pap,
                r.valor_500mb_cnpj, r.valor_700mb_cnpj, r.valor_1gb_cnpj,
            ])
        if formato == 'csv':
            from django.http import HttpResponse
            import csv
            from io import StringIO
            buf = StringIO()
            w = csv.writer(buf, delimiter=';')
            w.writerows(rows)
            resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = 'attachment; filename="regras_faixa.csv"'
            return resp
        from openpyxl import Workbook
        from django.http import HttpResponse
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        ws.title = 'REGRAS_FAIXAS'
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="regras_faixa.xlsx"'
        return resp


class RegraComissaoFaixaImportarView(APIView):
    """POST regras-comissao-faixa/importar/ - Body: arquivo (xlsx ou csv)."""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        arquivo = request.FILES.get('arquivo') or request.FILES.get('file')
        if not arquivo:
            return Response({'error': 'Envie o arquivo (arquivo ou file).'}, status=400)
        erros = []
        importados = 0
        nome = (arquivo.name or '').lower()
        try:
            if nome.endswith('.xlsx') or nome.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            elif nome.endswith('.csv'):
                import csv
                from io import TextIOWrapper
                rows = list(csv.reader(TextIOWrapper(arquivo, encoding='utf-8-sig'), delimiter=';'))
                if not rows and arquivo.tell() == 0:
                    arquivo.seek(0)
                    rows = list(csv.reader(TextIOWrapper(arquivo, encoding='latin-1'), delimiter=';'))
            else:
                return Response({'error': 'Use arquivo .xlsx ou .csv'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
        if not rows or len(rows) < 2:
            return Response({'error': 'Arquivo vazio ou sem dados.'}, status=400)
        header = [str(c).strip().upper() if c is not None else '' for c in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        for i, row in enumerate(rows[1:], start=2):
            row = list(row) if row else []
            while len(row) < len(header):
                row.append('')
            perfil_str = (row[idx.get('PERFIL', 0)] or '').strip()
            finalidade_str = (row[idx.get('FINALIDADE', 1)] or '').strip().upper()
            if finalidade_str not in ('COMISSAO', 'ADIANTAMENTO'):
                finalidade_str = 'COMISSAO'
            faixa_nome = (row[idx.get('FAIXA_NOME', 1)] or '').strip() or f'Faixa {i}'
            min_v = _int_ou_none(row[idx.get('MIN_VENDAS', 2)] if len(row) > 2 else None)
            max_v = _int_ou_none(row[idx.get('MAX_VENDAS', 3)] if len(row) > 3 else None)
            min_v = min_v if min_v is not None else 0
            max_v = max_v if max_v is not None else 99999
            vendedor = None
            perfil = None
            if perfil_str and perfil_str.upper() in ('SUPERVISOR', 'VENDEDOR'):
                perfil = perfil_str.capitalize()
            else:
                user = User.objects.filter(username__iexact=perfil_str).first() if perfil_str else None
                if user:
                    vendedor = user
            try:
                RegraComissaoFaixa.objects.create(
                    perfil=perfil,
                    vendedor=vendedor,
                    finalidade=finalidade_str,
                    faixa_nome=faixa_nome,
                    min_vendas=min_v,
                    max_vendas=max_v,
                    valor_500mb_pap=_decimal_ou_none(row[idx.get('VALOR_500MB_PAP', 4)]),
                    valor_700mb_pap=_decimal_ou_none(row[idx.get('VALOR_700MB_PAP', 5)]),
                    valor_1gb_pap=_decimal_ou_none(row[idx.get('VALOR_1GB_PAP', 6)]),
                    valor_500mb_cnpj=_decimal_ou_none(row[idx.get('VALOR_500MB_CNPJ', 7)]),
                    valor_700mb_cnpj=_decimal_ou_none(row[idx.get('VALOR_700MB_CNPJ', 8)]),
                    valor_1gb_cnpj=_decimal_ou_none(row[idx.get('VALOR_1GB_CNPJ', 9)]),
                )
                importados += 1
            except Exception as e:
                erros.append(f'Linha {i}: {e}')
        return Response({'importados': importados, 'erros': erros})


class ConfigComissaoVendedorExportarView(APIView):
    """GET config-comissao-vendedor/exportar/?formato=xlsx|csv - Download Excel ou CSV."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        formato = (request.query_params.get('formato') or 'xlsx').lower()
        qs = ConfigComissaoVendedor.objects.select_related('usuario').all().order_by('usuario__username')
        rows = [[
            'VENDEDOR', 'PERFIL_COMISSAO', 'USAR_VALOR_MANUAL?',
            '500MB_PAP_MANUAL', '700MB_PAP_MANUAL', '1GB_PAP_MANUAL',
            '500MB_CNPJ_MANUAL', '700MB_CNPJ_MANUAL', '1GB_CNPJ_MANUAL',
            'DESCONTA DACC PAP?', 'VENDAS BOLETO', 'INCLUSÃO', 'INSTALAÇÃO', 'ADIANTAR CNPJ',
            'INSS_VALOR', 'ADIANTAMENTO', 'PREMIAÇÃO', 'CARTÃO TRAFEGO', 'GESTOR TRAFEGO',
        ]]
        for c in qs:
            rows.append([
                c.usuario.username,
                c.perfil_comissao or 'Vendedor',
                'SIM' if c.usar_valor_manual else 'NÃO',
                c.valor_500mb_pap_manual, c.valor_700mb_pap_manual, c.valor_1gb_pap_manual,
                c.valor_500mb_cnpj_manual, c.valor_700mb_cnpj_manual, c.valor_1gb_cnpj_manual,
                'SIM' if c.desconta_dacc_pap else 'NÃO',
                c.desconto_boleto, c.desconto_inclusao, c.desconto_instalacao, c.adiantar_cnpj,
                c.inss_valor, c.adiantamento, getattr(c, 'premiação', None), c.cartao_trafego, c.gestor_trafego,
            ])
        if formato == 'csv':
            from django.http import HttpResponse
            import csv
            from io import StringIO
            buf = StringIO()
            w = csv.writer(buf, delimiter=';')
            w.writerows(rows)
            resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = 'attachment; filename="regras_vendedores.csv"'
            return resp
        from openpyxl import Workbook
        from django.http import HttpResponse
        from io import BytesIO
        wb = Workbook()
        ws = wb.active
        ws.title = 'REGRAS_VENDEDORES'
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="regras_vendedores.xlsx"'
        return resp


class ConfigComissaoVendedorImportarView(APIView):
    """POST config-comissao-vendedor/importar/ - Body: arquivo (xlsx ou csv)."""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        arquivo = request.FILES.get('arquivo') or request.FILES.get('file')
        if not arquivo:
            return Response({'error': 'Envie o arquivo (arquivo ou file).'}, status=400)
        erros = []
        importados = 0
        nome = (arquivo.name or '').lower()
        try:
            if nome.endswith('.xlsx') or nome.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            elif nome.endswith('.csv'):
                import csv
                from io import TextIOWrapper
                rows = list(csv.reader(TextIOWrapper(arquivo, encoding='utf-8-sig'), delimiter=';'))
                if not rows and arquivo.tell() == 0:
                    arquivo.seek(0)
                    rows = list(csv.reader(TextIOWrapper(arquivo, encoding='latin-1'), delimiter=';'))
            else:
                return Response({'error': 'Use arquivo .xlsx ou .csv'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=400)
        if not rows or len(rows) < 2:
            return Response({'error': 'Arquivo vazio ou sem dados.'}, status=400)
        header = [str(c).strip().upper().replace('?', '').replace(' ', '_') if c is not None else '' for c in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        def col(k, default_idx):
            for alias in (k, k.replace('_', ' '), k.replace(' ', '_')):
                if alias in idx:
                    return idx[alias]
            return default_idx
        for i, row in enumerate(rows[1:], start=2):
            row = list(row) if row else []
            while len(row) < max(idx.values()) + 1:
                row.append('')
            username = (row[col('VENDEDOR', 0)] or '').strip()
            if not username:
                continue
            user = User.objects.filter(username__iexact=username).first()
            if not user:
                erros.append(f'Linha {i}: Vendedor "{username}" não encontrado.')
                continue
            try:
                config, _ = ConfigComissaoVendedor.objects.get_or_create(usuario=user)
                perfil = (row[col('PERFIL_COMISSAO', 1)] or '').strip()
                if perfil and perfil.upper() in ('SUPERVISOR', 'VENDEDOR'):
                    config.perfil_comissao = perfil.capitalize()
                manual = (row[col('USAR_VALOR_MANUAL', 2)] or '').strip().upper()
                config.usar_valor_manual = manual in ('SIM', 'S', '1', 'TRUE')
                config.valor_500mb_pap_manual = _decimal_ou_none(row[col('500MB_PAP_MANUAL', 3)])
                config.valor_700mb_pap_manual = _decimal_ou_none(row[col('700MB_PAP_MANUAL', 4)])
                config.valor_1gb_pap_manual = _decimal_ou_none(row[col('1GB_PAP_MANUAL', 5)])
                config.valor_500mb_cnpj_manual = _decimal_ou_none(row[col('500MB_CNPJ_MANUAL', 6)])
                config.valor_700mb_cnpj_manual = _decimal_ou_none(row[col('700MB_CNPJ_MANUAL', 7)])
                config.valor_1gb_cnpj_manual = _decimal_ou_none(row[col('1GB_CNPJ_MANUAL', 8)])
                dacc = (row[col('DESCONTA_DACC_PAP', 9)] or '').strip().upper()
                config.desconta_dacc_pap = dacc in ('SIM', 'S', '1')
                config.desconto_boleto = _decimal_ou_none(row[col('VENDAS_BOLETO', 10)]) or 0
                config.desconto_inclusao = _decimal_ou_none(row[col('INCLUSÃO', 11)]) or 0
                config.desconto_instalacao = _decimal_ou_none(row[col('INSTALAÇÃO', 12)]) or 0
                config.adiantar_cnpj = _decimal_ou_none(row[col('ADIANTAR_CNPJ', 13)]) or 0
                config.inss_valor = _decimal_ou_none(row[col('INSS_VALOR', 14)])
                config.adiantamento = _decimal_ou_none(row[col('ADIANTAMENTO', 15)])
                setattr(config, 'premiação', _decimal_ou_none(row[col('PREMIAÇÃO', 16)]))
                config.cartao_trafego = _decimal_ou_none(row[col('CARTÃO_TRAFEGO', 17)])
                config.gestor_trafego = _decimal_ou_none(row[col('GESTOR_TRAFEGO', 18)])
                config.save()
                importados += 1
            except Exception as e:
                erros.append(f'Linha {i}: {e}')
        return Response({'importados': importados, 'erros': erros})


class ComunicadoViewSet(viewsets.ModelViewSet):
    queryset = Comunicado.objects.all().order_by('-id')
    serializer_class = ComunicadoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def _processar_envio_comunicado(self, comunicado):
        """
        Processa e envia um comunicado via WhatsApp.
        Mapeia perfis para grupos do WhatsApp.
        """
        from django.utils import timezone
        from datetime import datetime, date, time
        
        try:
            whatsapp_service = WhatsAppService()
            
            # Verificar se é envio imediato (data/hora atual ou passada)
            agora = timezone.now()
            data_envio = comunicado.data_programada
            hora_envio = comunicado.hora_programada
            
            # Combinar data e hora para comparação
            if isinstance(hora_envio, time):
                datetime_envio = datetime.combine(data_envio, hora_envio)
                if timezone.is_naive(datetime_envio):
                    from django.utils import timezone as tz
                    datetime_envio = tz.make_aware(datetime_envio)
            else:
                datetime_envio = agora  # Se não tiver hora, considera agora
            
            # Se a data/hora programada já passou ou é agora, processa imediatamente
            enviar_agora = datetime_envio <= agora
            
            if not enviar_agora:
                # Agendado para o futuro, não processa ainda
                return False
            
            # Mapear perfil para grupos do WhatsApp
            from django.db.models import Q
            
            # Buscar grupos ativos no banco
            if comunicado.perfil_destino == 'TODOS':
                # Para TODOS, buscar todos os grupos ativos
                grupos_ativos = GrupoDisparo.objects.filter(ativo=True)
            else:
                # Para perfis específicos, buscar grupos que contenham o nome do perfil no nome
                # Mapeamento: perfil_destino -> termos de busca
                perfil_termos = {
                    'DIRETORIA': ['Diretoria', 'diretor'],
                    'BACKOFFICE': ['BackOffice', 'Back Office', 'backoffice'],
                    'SUPERVISOR': ['Supervisor', 'supervisor'],
                    'VENDEDOR': ['Vendedor', 'vendedor'],
                }
                
                termos = perfil_termos.get(comunicado.perfil_destino, [comunicado.perfil_destino])
                
                # Criar filtro: ativo=True AND (nome contém termo1 OU nome contém termo2 OU ...)
                filtro_nome = Q()
                for termo in termos:
                    filtro_nome |= Q(nome__icontains=termo)
                
                grupos_ativos = GrupoDisparo.objects.filter(Q(ativo=True) & filtro_nome)
            
            grupos_ids = list(grupos_ativos.values_list('chat_id', flat=True))
            
            if not grupos_ids:
                logger.warning(f"Nenhum grupo encontrado para perfil {comunicado.perfil_destino}")
                comunicado.status = 'ERRO'
                comunicado.save()
                return False
            
            # Enviar para cada grupo
            sucesso_total = True
            for grupo_id in grupos_ids:
                try:
                    resultado, resposta = whatsapp_service.enviar_mensagem_texto(grupo_id, comunicado.mensagem)
                    if not resultado:
                        sucesso_total = False
                        logger.error(f"Erro ao enviar comunicado {comunicado.id} para grupo {grupo_id}")
                except Exception as e:
                    sucesso_total = False
                    logger.error(f"Exceção ao enviar comunicado {comunicado.id} para grupo {grupo_id}: {e}")
            
            # Atualizar status
            if sucesso_total:
                comunicado.status = 'ENVIADO'
                comunicado.save()
                return True
            else:
                comunicado.status = 'ERRO'
                comunicado.save()
                return False
                
        except Exception as e:
            logger.error(f"Erro ao processar comunicado {comunicado.id}: {e}")
            import traceback
            traceback.print_exc()
            comunicado.status = 'ERRO'
            comunicado.save()
            return False

    def perform_create(self, serializer):
        from django.utils import timezone
        from datetime import datetime, date, time
        
        comunicado = serializer.save(criado_por=self.request.user)
        
        # Se for envio imediato (data/hora atual ou passada), processa agora
        agora = timezone.now()
        data_envio = comunicado.data_programada
        hora_envio = comunicado.hora_programada
        
        if isinstance(hora_envio, time):
            datetime_envio = datetime.combine(data_envio, hora_envio)
            if timezone.is_naive(datetime_envio):
                from django.utils import timezone as tz
                datetime_envio = tz.make_aware(datetime_envio)
        else:
            datetime_envio = agora
        
        # Se a data/hora programada já passou ou é agora, processa imediatamente
        if datetime_envio <= agora:
            # Processar em thread para não bloquear a resposta
            import threading
            thread = threading.Thread(target=self._processar_envio_comunicado, args=(comunicado,))
            thread.daemon = True
            thread.start()

    @action(detail=True, methods=['post'], url_path='enviar-agora', permission_classes=[permissions.IsAuthenticated])
    def enviar_agora(self, request, pk=None):
        """
        Action para processar e enviar um comunicado pendente imediatamente
        """
        comunicado = self.get_object()
        
        if comunicado.status == 'ENVIADO':
            return Response({"detail": "Comunicado já foi enviado."}, status=status.HTTP_400_BAD_REQUEST)
        
        if comunicado.status == 'CANCELADO':
            return Response({"detail": "Comunicado foi cancelado e não pode ser enviado."}, status=status.HTTP_400_BAD_REQUEST)
        
        sucesso = self._processar_envio_comunicado(comunicado)
        
        if sucesso:
            return Response({"detail": "Comunicado enviado com sucesso!", "status": comunicado.status})
        else:
            return Response({"detail": "Erro ao enviar comunicado. Verifique os logs.", "status": comunicado.status}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class EstatisticasBotWhatsAppView(APIView):
    """
    API para buscar estatísticas do bot WhatsApp
    Retorna contagem por comando e por vendedor
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        try:
            from django.db.models import Count, Q
            from django.utils import timezone
            from datetime import timedelta
            
            # Parâmetros opcionais de filtro
            dias = request.query_params.get('dias', 30)  # Padrão: últimos 30 dias
            try:
                dias = int(dias)
            except:
                dias = 30
            
            data_inicio = timezone.now() - timedelta(days=dias)
            
            # Filtrar estatísticas do período
            estatisticas = EstatisticaBotWhatsApp.objects.filter(
                data_envio__gte=data_inicio
            ).select_related('vendedor')
            
            # 1. Contagem por comando
            por_comando = estatisticas.values('comando').annotate(
                total=Count('id')
            ).order_by('comando')
            
            comando_dict = {item['comando']: item['total'] for item in por_comando}

            # Comandos exibidos nos 4 primeiros cartões da UI
            principais = {'FACHADA', 'VIABILIDADE', 'FATURA', 'STATUS'}
            outros_tipos = sum(
                v for k, v in comando_dict.items()
                if k not in principais
            )
            
            # 2. Contagem por vendedor (com coluna "outros": VENDER, PEDIDO, ANDAMENTO, CREDITO, etc.)
            def _annotate_por_grupo(qs):
                return qs.values(
                    'vendedor__id',
                    'vendedor__username',
                    'vendedor__first_name',
                    'vendedor__last_name',
                ).annotate(
                    total=Count('id'),
                    fachada=Count('id', filter=Q(comando='FACHADA')),
                    viabilidade=Count('id', filter=Q(comando='VIABILIDADE')),
                    fatura=Count('id', filter=Q(comando='FATURA')),
                    status=Count('id', filter=Q(comando='STATUS')),
                    outros=Count('id', filter=~Q(comando__in=list(principais))),
                ).order_by('-total')

            por_vendedor = _annotate_por_grupo(estatisticas.filter(vendedor__isnull=False))
            por_sem_vendedor = _annotate_por_grupo(estatisticas.filter(vendedor__isnull=True))
            
            vendedores_data = []
            for item in por_vendedor:
                nome_completo = item.get('vendedor__first_name', '') or ''
                sobrenome = item.get('vendedor__last_name', '') or ''
                if sobrenome:
                    nome_completo = f"{nome_completo} {sobrenome}".strip()
                if not nome_completo:
                    nome_completo = item.get('vendedor__username', 'N/A')
                
                vendedores_data.append({
                    'vendedor_id': item['vendedor__id'],
                    'vendedor_nome': nome_completo,
                    'vendedor_username': item.get('vendedor__username', 'N/A'),
                    'total': item['total'],
                    'fachada': item['fachada'],
                    'viabilidade': item['viabilidade'],
                    'fatura': item['fatura'],
                    'status': item['status'],
                    'outros': item['outros'],
                    'sem_cadastro': False,
                })

            # Quem usa o bot mas o número não bate com tel_whatsapp 1–3 no CRM
            for item in por_sem_vendedor:
                vendedores_data.append({
                    'vendedor_id': None,
                    'vendedor_nome': 'Sem vínculo no cadastro (só telefone)',
                    'vendedor_username': '—',
                    'total': item['total'],
                    'fachada': item['fachada'],
                    'viabilidade': item['viabilidade'],
                    'fatura': item['fatura'],
                    'status': item['status'],
                    'outros': item['outros'],
                    'sem_cadastro': True,
                })

            vendedores_data.sort(key=lambda x: x['total'], reverse=True)
            
            # 3. Totais gerais
            total_geral = estatisticas.count()
            total_sem_vendedor = estatisticas.filter(vendedor__isnull=True).count()
            
            return Response({
                'periodo_dias': dias,
                'data_inicio': data_inicio.isoformat(),
                'totais': {
                    'geral': total_geral,
                    'sem_vendedor': total_sem_vendedor,
                    'com_vendedor': total_geral - total_sem_vendedor
                },
                'por_comando': {
                    'FACHADA': comando_dict.get('FACHADA', 0),
                    'VIABILIDADE': comando_dict.get('VIABILIDADE', 0),
                    'FATURA': comando_dict.get('FATURA', 0),
                    'STATUS': comando_dict.get('STATUS', 0),
                    'OUTROS': outros_tipos,
                },
                'por_comando_detalhe': comando_dict,
                'por_vendedor': vendedores_data
            })
        except Exception as e:
            logger.error(f"Erro ao buscar estatísticas do bot WhatsApp: {e}")
            import traceback
            traceback.print_exc()
            return Response({
                'periodo_dias': 30,
                'data_inicio': None,
                'totais': {'geral': 0, 'sem_vendedor': 0, 'com_vendedor': 0},
                'por_comando': {'FACHADA': 0, 'VIABILIDADE': 0, 'FATURA': 0, 'STATUS': 0, 'OUTROS': 0},
                'por_comando_detalhe': {},
                'por_vendedor': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _pode_gerir_adiantamento_sabado(user):
    return bool(user and user.is_superuser) or is_member(user, ['Diretoria', 'Admin', 'BackOffice'])


def _quitar_adiantamento_sabado_na_instalacao(venda, status_esteira_antes):
    """Ao instalar venda que tinha adiantamento sábado: marca antecipação e quitação sem novo lançamento."""
    nome_new = (venda.status_esteira.nome if venda.status_esteira else '') or ''
    nome_old = (status_esteira_antes.nome if status_esteira_antes else '') or ''
    if 'INSTALADA' not in nome_new.upper():
        return
    if 'INSTALADA' in nome_old.upper():
        return
    if not getattr(venda, 'adiantamento_sabado_marcado', False):
        return
    if venda.adiantamento_sabado_quitado_em:
        return
    venda.antecipacao_comissao = True
    venda.adiantamento_sabado_quitado_em = timezone.now()
    venda.save(update_fields=['antecipacao_comissao', 'adiantamento_sabado_quitado_em'])


def _marcar_adiantamento_sabado_exec(venda, user, manual=False, obs=''):
    """
    Marca adiantamento sábado para uma venda (mesmas regras do endpoint unitário).
    Levanta ValueError com mensagem amigável em caso de falha.
    Retorna dict com adiantamento_sabado_valor (float) e data_lancamento (iso str).
    """
    if venda.adiantamento_sabado_marcado:
        raise ValueError('Venda já marcada como adiantamento sábado.')
    st = (venda.status_esteira.nome if venda.status_esteira else '') or ''
    if 'AGENDADO' not in st.upper():
        raise ValueError('Somente vendas com status AGENDADO na esteira.')
    if not venda.vendedor_id:
        raise ValueError('Venda sem vendedor.')
    recebe = getattr(venda.vendedor, 'recebe_adiantamento_sabado', False)

    if not venda.data_abertura:
        raise ValueError('Informe a data/hora de abertura da O.S.')
    dt_ab_local = timezone.localtime(venda.data_abertura).date()
    eh_sabado = dt_ab_local.weekday() == 5

    if not manual:
        if not recebe:
            raise ValueError(
                'Vendedor sem permissão no cadastro (recebe adiantamento sábado). '
                'Use marcação manual com observação.'
            )
        if not eh_sabado:
            raise ValueError('A data de abertura da O.S. deve ser um sábado.')

    faixa_adiantamento = RegraComissaoFaixa.objects.filter(finalidade='ADIANTAMENTO').first()
    if not faixa_adiantamento:
        faixa_adiantamento = RegraComissaoFaixa.objects.filter(faixa_nome__iexact='Adiantamento').first()
    doc = re.sub(r'\D', '', (venda.cliente.cpf_cnpj if venda.cliente else '') or '')
    is_cnpj_venda = len(doc) == 14
    valor_unit = Decimal(
        str(LancamentoFinanceiroViewSet._valor_comissao_estimado_venda(venda, faixa_adiantamento, is_cnpj_venda))
    )
    if valor_unit <= 0:
        raise ValueError('Valor zero nas Regras por Faixa (Finalidade = Adiantamento).')

    data_lanc = dt_ab_local
    descricao = 'Adiantamento sábado (Agendados)'
    with transaction.atomic():
        venda.adiantamento_sabado_marcado = True
        venda.adiantamento_sabado_valor = valor_unit
        venda.adiantamento_sabado_marcado_em = timezone.now()
        venda.adiantamento_sabado_marcado_por = user
        venda.adiantamento_sabado_manual = manual
        venda.adiantamento_sabado_obs_manual = obs if manual else ''
        venda.save(
            update_fields=[
                'adiantamento_sabado_marcado',
                'adiantamento_sabado_valor',
                'adiantamento_sabado_marcado_em',
                'adiantamento_sabado_marcado_por',
                'adiantamento_sabado_manual',
                'adiantamento_sabado_obs_manual',
            ]
        )
        lanc = LancamentoFinanceiro.objects.filter(
            usuario_id=venda.vendedor_id,
            tipo='ADIANTAMENTO_COMISSAO',
            data=data_lanc,
            descricao=descricao,
        ).first()
        venda_ids = []
        if lanc and isinstance(lanc.metadados, dict):
            venda_ids = list(lanc.metadados.get('venda_ids') or [])
        if venda.id not in venda_ids:
            venda_ids.append(venda.id)
        meta_old = lanc.metadados if lanc and isinstance(lanc.metadados, dict) else {}
        valores = dict(meta_old.get('valores_por_venda_id') or {})
        valores[str(venda.id)] = str(valor_unit)
        meta = {
            'origem': 'esteira_sabado_agendados',
            'venda_ids': venda_ids,
            'valores_por_venda_id': valores,
        }
        if lanc:
            lanc.valor = Decimal(str(lanc.valor or 0)) + valor_unit
            lanc.quantidade_vendas = max(int(lanc.quantidade_vendas or 0), 0) + 1
            lanc.metadados = meta
            lanc.save(update_fields=['valor', 'quantidade_vendas', 'metadados'])
        else:
            LancamentoFinanceiro.objects.create(
                usuario_id=venda.vendedor_id,
                tipo='ADIANTAMENTO_COMISSAO',
                data=data_lanc,
                valor=valor_unit,
                quantidade_vendas=1,
                descricao=descricao,
                metadados=meta,
                criado_por=user,
            )

    return {
        'adiantamento_sabado_valor': float(valor_unit),
        'data_lancamento': data_lanc.isoformat(),
    }


class VendaViewSet(viewsets.ModelViewSet):
    permission_classes = [VendaPermission]
    resource_name = 'venda'
    queryset = Venda.objects.filter(ativo=True).order_by('-data_criacao')
    pagination_class = VendaPagination

    def get_serializer_context(self):
        context = super(VendaViewSet, self).get_serializer_context()
        context.update({"request": self.request})
        return context

    def get_serializer_class(self):
        if self.action == 'retrieve': return VendaDetailSerializer
        if self.action == 'create': return VendaCreateSerializer
        if self.action in ['update', 'partial_update']: return VendaUpdateSerializer
        return VendaSerializer

    def get_queryset(self):
        queryset = Venda.objects.filter(ativo=True).select_related(
            'vendedor', 'cliente', 'plano', 'forma_pagamento',
            'status_tratamento', 'status_esteira', 'status_comissionamento',
            'motivo_pendencia', 'auditor_atual', 'editado_por'
        )
        
        # ✅ OTIMIZAÇÃO: Carrega histórico APENAS quando recuperando detalhes (retrieve)
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related('historico_alteracoes__usuario')
        
        queryset = queryset.order_by('-data_criacao')
        
        user = self.request.user
        view_type = self.request.query_params.get('view')
        flow = self.request.query_params.get('flow')
        search = self.request.query_params.get('search')
        data_inicio_str = self.request.query_params.get('data_inicio')
        data_fim_str = self.request.query_params.get('data_fim')
        data_tipo = (self.request.query_params.get('data_tipo') or 'criacao').strip().lower()
        status_filter_raw = (self.request.query_params.get('status') or '').strip().upper()
        
        # --- REGRA DE DATA OBRIGATÓRIA (MÊS ATUAL) ---
        grupos_livres = ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']
        eh_gestao_total = is_member(user, grupos_livres)
        eh_vs_mes = vendedor_ou_supervisor_restrito_mes(user)
        hoje_local = timezone.localtime(timezone.now())
        hoje_d = hoje_local.date()

        if not eh_gestao_total and not search:
            hoje = hoje_local
            inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            aplica_padrao_mes_atual = True
            if eh_vs_mes and data_inicio_str and data_fim_str:
                try:
                    di = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
                    df = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                    if mes_completo_vendedor_supervisor_valido(di, df, hoje_d):
                        aplica_padrao_mes_atual = False
                except ValueError:
                    pass
            if aplica_padrao_mes_atual:
                if status_filter_raw == 'INSTALADA' or flow == 'esteira_instaladas':
                    queryset = queryset.filter(
                        _filtro_data_efetiva_instalacao_intervalo_venda(inicio_mes.date(), hoje_d)
                    )
                else:
                    queryset = queryset.filter(Q(data_criacao__gte=inicio_mes) | Q(data_instalacao__gte=inicio_mes))

        # --- FILTRO DE STATUS ---
        status_filter = self.request.query_params.get('status')
        status_instalada_exata = False
        if status_filter:
            status_upper = status_filter.upper()
            if status_upper == 'CANCELADO':
                queryset = queryset.filter(status_esteira__nome__icontains='CANCELAD')
            elif 'PENDEN' in status_upper:
                queryset = queryset.filter(status_esteira__nome__icontains='PENDEN')
            else:
                queryset = queryset.filter(status_esteira__nome__iexact=status_filter)
            status_instalada_exata = status_upper == 'INSTALADA'

        # --- FILTRO DE BUSCA GLOBAL ---
        if search:
            search_clean = re.sub(r'\D', '', search)
            search_strip = (search or '').strip()
            filters = Q(ordem_servico__icontains=search_strip) | \
                      Q(cliente__nome_razao_social__icontains=search_strip) | \
                      Q(cliente__cpf_cnpj__icontains=search_strip)
            if search_clean:
                filters |= Q(cliente__cpf_cnpj__icontains=search_clean)
                filters |= Q(ordem_servico__icontains=search_clean)
            # Incluir busca por consultor/vendedor (nome ou username)
            if search_strip:
                filters |= Q(vendedor__username__icontains=search_strip) | \
                           Q(vendedor__first_name__icontains=search_strip) | \
                           Q(vendedor__last_name__icontains=search_strip)
            queryset = queryset.filter(filters)

        # --- PERMISSÕES DE VISUALIZAÇÃO ---
        acoes_gestao = [
            'retrieve', 'update', 'partial_update', 'destroy',
            'alocar_auditoria', 'liberar_auditoria', 'finalizar_auditoria',
            'pendentes_auditoria', 'resumo_auditoria',
            'reenviar_whatsapp_aprovacao', 'enviar_resumo_plano_whatsapp',
            'toggle_adiantamento_comissao',
            'toggle_adiantamento_cnpj',
            'marcar_adiantamento_sabado',
            'desmarcar_adiantamento_sabado',
            'marcar_adiantamento_sabado_lote',
        ]

        if self.action in acoes_gestao:
            grupos_gestao_acao = ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']
            if user.is_superuser or is_member(user, grupos_gestao_acao):
                return queryset
            q_vs_retrieve = q_venda_acesso_retrieve_vendedor_supervisor()
            if is_member(user, ['Supervisor']):
                liderados_ids = list(user.liderados.values_list('id', flat=True))
                liderados_ids.append(user.id)
                qs_sup = queryset.filter(vendedor_id__in=liderados_ids)
                if self.action == 'retrieve':
                    qs_sup = qs_sup.filter(q_vs_retrieve)
                return qs_sup
            qs_v = queryset.filter(vendedor=user)
            if self.action == 'retrieve':
                qs_v = qs_v.filter(q_vs_retrieve)
            return qs_v

        if not view_type:
            if flow and is_member(user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
                view_type = 'geral'
            else:
                view_type = 'minhas_vendas'

        if view_type == 'minhas_vendas':
            queryset = queryset.filter(vendedor=user)
        
        elif view_type == 'visao_equipe' or view_type == 'geral':
            if is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
                pass 
            elif is_member(user, ['Auditoria', 'Qualidade']):
                pass # Já filtrado no inicio
            elif is_member(user, ['Supervisor']):
                liderados_ids = list(user.liderados.values_list('id', flat=True))
                liderados_ids.append(user.id)
                queryset = queryset.filter(vendedor_id__in=liderados_ids)
            else:
                return queryset.none()

        # Filtros Avançados (Data, OS, Consultor)
        ordem_servico = self.request.query_params.get('ordem_servico')
        consultor_id = self.request.query_params.get('consultor_id')
        
        if consultor_id:
            if view_type != 'minhas_vendas':
                queryset = queryset.filter(vendedor_id=consultor_id)

        if flow == 'auditoria':
            queryset = queryset.filter(status_tratamento__isnull=False, status_esteira__isnull=True)
            # Filtro opcional por status do tratamento (auditoria)
            status_tratamento_id = self.request.query_params.get('status_tratamento_id')
            if status_tratamento_id and str(status_tratamento_id).isdigit():
                queryset = queryset.filter(status_tratamento_id=int(status_tratamento_id))
        elif flow == 'esteira':
            queryset = queryset.filter(status_esteira__isnull=False, status_esteira__estado__iexact='ABERTO')
        elif flow == 'esteira_todas':
            if not is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
                return queryset.none()
            queryset = queryset.filter(status_esteira__isnull=False)
        elif flow == 'esteira_instaladas':
            # Vendas instaladas com filtros dedicados (vendedor e período de instalação).
            queryset = queryset.filter(
                status_esteira__isnull=False,
                status_esteira__nome__iexact='Instalada'
            ).order_by('-data_instalacao', '-id')
            vendedor_id = self.request.query_params.get('vendedor_id')
            if vendedor_id and str(vendedor_id).isdigit():
                queryset = queryset.filter(vendedor_id=int(vendedor_id))
            data_inicio_inst = (self.request.query_params.get('data_inicio_inst') or '').strip()
            data_fim_inst = (self.request.query_params.get('data_fim_inst') or '').strip()
            if data_inicio_inst or data_fim_inst:
                if not (data_inicio_inst and data_fim_inst):
                    return queryset.none()
                try:
                    dt_ini_inst = datetime.strptime(data_inicio_inst, '%Y-%m-%d').date()
                    dt_fim_inst = datetime.strptime(data_fim_inst, '%Y-%m-%d').date()
                    if dt_ini_inst > dt_fim_inst:
                        return queryset.none()
                    queryset = queryset.filter(
                        _filtro_data_efetiva_instalacao_intervalo_venda(dt_ini_inst, dt_fim_inst)
                    )
                except ValueError:
                    return queryset.none()
        elif flow == 'comissionamento':
            queryset = queryset.filter(status_esteira__nome__iexact='Instalada').exclude(status_comissionamento__nome__iexact='Pago')

        if ordem_servico:
            queryset = queryset.filter(ordem_servico__icontains=ordem_servico)

        # Se houver datas especificas vindas do front (gestão livre ou vendedor/supervisor em mês permitido)
        pode_filtrar_datas = False
        dt_ini = dt_fim = None
        if data_inicio_str and data_fim_str:
            try:
                dt_ini = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
                dt_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            except ValueError:
                dt_ini = dt_fim = None
        if dt_ini and dt_fim:
            if eh_gestao_total:
                pode_filtrar_datas = True
            elif eh_vs_mes and mes_completo_vendedor_supervisor_valido(dt_ini, dt_fim, hoje_d):
                pode_filtrar_datas = True
        if pode_filtrar_datas and dt_ini and dt_fim:
            try:
                if flow == 'auditoria':
                    queryset = queryset.filter(
                        data_criacao__date__gte=dt_ini,
                        data_criacao__date__lte=dt_fim
                    )
                else:
                    dt_fim_plus = dt_fim + timedelta(days=1)
                    if data_tipo == 'criacao':
                        queryset = queryset.filter(
                            data_criacao__date__gte=dt_ini,
                            data_criacao__date__lte=dt_fim
                        )
                    elif data_tipo == 'agendamento':
                        queryset = queryset.filter(
                            data_agendamento__gte=dt_ini,
                            data_agendamento__lte=dt_fim
                        )
                    elif data_tipo == 'instalacao':
                        queryset = queryset.filter(
                            _filtro_data_efetiva_instalacao_intervalo_venda(dt_ini, dt_fim)
                        )
                    elif status_instalada_exata or flow == 'esteira_instaladas':
                        # Para aba/filtro de instaladas, usar data efetiva de instalação
                        # (física quando existir; senão OSAB), alinhando com Performance.
                        queryset = queryset.filter(
                            _filtro_data_efetiva_instalacao_intervalo_venda(dt_ini, dt_fim)
                        )
                    else:
                        queryset = queryset.filter(
                            Q(data_criacao__range=(dt_ini, dt_fim_plus))
                            | Q(data_instalacao__range=(dt_ini, dt_fim_plus))
                        )
            except Exception:
                pass

        return queryset
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Verificar se a venda está travada para outro auditor
        if instance.auditor_atual and instance.auditor_atual != request.user:
            # Permitir edição apenas para supervisores/Diretoria/Admin/BackOffice
            grupos_permitidos = ['Diretoria', 'Admin', 'BackOffice', 'Supervisor']
            if not request.user.is_superuser and not is_member(request.user, grupos_permitidos):
                return Response(
                    {"detail": f"Esta venda está sendo auditada por {instance.auditor_atual.get_full_name() or instance.auditor_atual.username}. Você não tem permissão para editá-la."},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Validar CPF/CNPJ antes de processar
        if 'cliente_cpf_cnpj' in request.data:
            try:
                validar_cpf_ou_cnpj(request.data['cliente_cpf_cnpj'])
            except ValidationError as e:
                return Response({"cliente_cpf_cnpj": [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar CPF do Representante Legal se fornecido
        if 'cpf_representante_legal' in request.data and request.data['cpf_representante_legal']:
            try:
                validar_cpf(request.data['cpf_representante_legal'])
            except ValidationError as e:
                return Response({"cpf_representante_legal": [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        
        if not serializer.is_valid():
            # --- LOG DE ERRO NO TERMINAL ---
            import json
            print("\n" + "="*60)
            print(f"!!! ERRO 400 AO SALVAR VENDA #{instance.id} !!!")
            print(f"DADOS ENVIADOS: {json.dumps(request.data, indent=2, default=str)}")
            print("-" * 10)
            print(f"MOTIVO DO ERRO: {json.dumps(serializer.errors, indent=2, default=str)}")
            print("="*60 + "\n")
            # -------------------------------
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            self.perform_update(serializer)
        except Exception as e:
            erro_texto = str(e).lower()
            if hasattr(e, 'detail'):
                return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)

            if isinstance(e, IntegrityError):
                if 'cliente_cpf_cnpj_key' in erro_texto or ('cpf_cnpj' in erro_texto and ('duplicate' in erro_texto or 'unique' in erro_texto)):
                    return Response(
                        {"cliente_cpf_cnpj": ["Este CPF/CNPJ já está cadastrado em outro cliente."]},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                logger.error(f"Erro de integridade ao salvar venda #{instance.id}: {e}", exc_info=True)
                return Response(
                    {"detail": "Não foi possível salvar por conflito de dados. Revise os campos e tente novamente."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            logger.error(f"Erro inesperado ao salvar venda #{instance.id}: {e}", exc_info=True)
            return Response(
                {"detail": "Não foi possível salvar a venda. Tente novamente em instantes."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        if getattr(instance, '_prefetched_objects_cache', None):
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        print("[CRM DEBUG] Dados recebidos no POST /crm/vendas/:", dict(request.data))
        # Validar CPF/CNPJ antes de processar
        if 'cliente_cpf_cnpj' in request.data:
            try:
                validar_cpf_ou_cnpj(request.data['cliente_cpf_cnpj'])
            except ValidationError as e:
                print("[CRM DEBUG] Erro de validação CPF/CNPJ:", e)
                return Response({"cliente_cpf_cnpj": [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        # Validar CPF do Representante Legal se fornecido
        if 'cpf_representante_legal' in request.data and request.data['cpf_representante_legal']:
            try:
                validar_cpf(request.data['cpf_representante_legal'])
            except ValidationError as e:
                print("[CRM DEBUG] Erro de validação CPF Rep. Legal:", e)
                return Response({"cpf_representante_legal": [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        # Validar e-mail por fluxo:
        # - VIA APP: e-mail opcional nesta etapa
        # - SEM APP: e-mail obrigatório e válido
        email_raw = request.data.get('cliente_email')
        forma_entrada = str(request.data.get('forma_entrada') or '').strip().upper()
        email = str(email_raw or '').strip()
        if forma_entrada == 'SEM_APP':
            if not email:
                return Response({"cliente_email": ["No fluxo Sem App, o e-mail é obrigatório."]}, status=status.HTTP_400_BAD_REQUEST)
            try:
                validate_email(email, check_deliverability=False)
            except EmailNotValidError as e:
                return Response({"cliente_email": [f"E-mail inválido: {str(e)}"]}, status=status.HTTP_400_BAD_REQUEST)
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("[CRM DEBUG] Erros de validação do serializer:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        try:
            self.perform_create(serializer)
        except IntegrityError as e:
            erro_texto = str(e).lower()
            if 'cliente_pkey' in erro_texto or ('crm_cliente' in erro_texto and 'duplicate' in erro_texto):
                return Response(
                    {"detail": "Conflito técnico ao criar cliente (sequência de ID). Tente novamente."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if 'cliente_cpf_cnpj_key' in erro_texto or ('cpf_cnpj' in erro_texto and ('duplicate' in erro_texto or 'unique' in erro_texto)):
                return Response(
                    {"cliente_cpf_cnpj": ["Este CPF/CNPJ já está cadastrado em outro cliente."]},
                    status=status.HTTP_400_BAD_REQUEST
                )
            logger.error("Erro de integridade ao criar venda: %s", e, exc_info=True)
            return Response(
                {"detail": "Não foi possível salvar por conflito de dados. Revise e tente novamente."},
                status=status.HTTP_400_BAD_REQUEST
            )
        venda = serializer.instance
        # Notificar o vendedor responsável (cadastro pelo vendedor ou pelo backoffice)
        if venda.vendedor and getattr(venda.vendedor, 'tel_whatsapp', None):
            try:
                from .whatsapp_service import WhatsAppService
                nome_cliente = venda.cliente.nome_razao_social if venda.cliente else 'N/A'
                msg = (
                    f"Sua venda, do cliente {nome_cliente} foi recebida pelo backOffice, está na etapa da auditoria, "
                    f"aguarde o tratamento e acompanhe o status pelo bot, enviando a palavra \"Status\".\n"
                    f"Protocolo: {venda.id}"
                )
                svc = WhatsAppService()
                svc.enviar_mensagem_texto(venda.vendedor.tel_whatsapp, msg)
                logger.info(f"Notificação de venda recebida enviada para vendedor {venda.vendedor.username} (venda #{venda.id})")
            except Exception as e:
                logger.warning(f"Falha ao enviar WhatsApp de venda recebida para {venda.vendedor.username}: {e}")
        headers = self.get_success_headers(serializer.data)
        print("[CRM DEBUG] Venda criada com sucesso! Dados:", serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=True, methods=['post'], url_path='reenviar-whatsapp-aprovacao', permission_classes=[permissions.IsAuthenticated])
    def reenviar_whatsapp_aprovacao(self, request, pk=None):
        venda = self.get_object()
        if not venda.vendedor or not venda.vendedor.tel_whatsapp:
            return Response({"detail": "Venda sem vendedor ou vendedor sem WhatsApp cadastrado."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            svc = WhatsAppService()
            svc.enviar_mensagem_cadastrada(venda, telefone_destino=venda.vendedor.tel_whatsapp)
            return Response({"detail": "Mensagem reenviada com sucesso!"})
        except Exception as e:
            logger.error(f"Erro reenvio zap: {e}")
            return Response({"detail": "Erro ao tentar enviar mensagem."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='enviar-resumo-plano-whatsapp', permission_classes=[permissions.IsAuthenticated])
    def enviar_resumo_plano_whatsapp(self, request, pk=None):
        """Envia o resumo do plano (mesmo formato do fluxo VENDER) para o celular 1 do cadastro da venda (cliente)."""
        venda = self.get_object()
        telefone = venda.telefone1
        if not telefone or not str(telefone).strip():
            return Response(
                {"detail": "Cadastro da venda não possui Celular 1 preenchido. Preencha o telefone na aba Contato."},
                status=status.HTTP_400_BAD_REQUEST
            )
        ok, msg_erro = validar_venda_para_resumo_auditoria(venda)
        if not ok:
            return Response({"detail": msg_erro}, status=status.HTTP_400_BAD_REQUEST)
        try:
            resumo = montar_resumo_plano_para_whatsapp(venda)
            if not resumo:
                return Response({"detail": "Não foi possível montar o resumo da venda."}, status=status.HTTP_400_BAD_REQUEST)
            svc = WhatsAppService()
            svc.enviar_mensagem_texto(telefone, resumo)
            # Registrar pendência de confirmação do cliente (resumo enviado da auditoria)
            # para o webhook não responder "usuário não ativo" e gerar protocolo ao confirmar
            celular_limpo = "".join(filter(str.isdigit, str(telefone)))
            if celular_limpo.startswith("55") and len(celular_limpo) > 12:
                celular_limpo = celular_limpo[2:]
            if celular_limpo:
                PapConfirmacaoCliente.objects.create(
                    celular_cliente=celular_limpo,
                    confirmado=False,
                    sessao=None,
                    venda=venda,
                    enviado_por=request.user,
                )
            return Response({"detail": "Resumo enviado para o Cliente com sucesso!"})
        except Exception as e:
            logger.error(f"Erro ao enviar resumo plano WhatsApp (venda {venda.id}): {e}", exc_info=True)
            return Response(
                {"detail": "Erro ao enviar mensagem. Verifique o número e tente novamente."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def pendentes_auditoria(self, request):
        request.GET._mutable = True
        request.GET['flow'] = 'auditoria'
        request.GET['view'] = 'geral'
        request.GET._mutable = False
        qs = self.filter_queryset(self.get_queryset())
        qs = qs.exclude(status_tratamento__estado__iexact='FECHADO').order_by('-id')

        # Aplicar filtros de data e status aqui (garantia: não depender só do get_queryset)
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        if data_inicio_str and data_fim_str:
            try:
                dt_ini = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
                dt_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
                qs = qs.filter(
                    data_criacao__date__gte=dt_ini,
                    data_criacao__date__lte=dt_fim
                )
            except (ValueError, TypeError):
                pass
        status_tratamento_id = request.query_params.get('status_tratamento_id')
        if status_tratamento_id and str(status_tratamento_id).isdigit():
            qs = qs.filter(status_tratamento_id=int(status_tratamento_id))

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='resumo_auditoria')
    def resumo_auditoria(self, request):
        """Resumo do mês atual: total vendas, envios de resumo pelo BO e confirmações pelo cliente."""
        request.GET._mutable = True
        request.GET['flow'] = 'auditoria'
        request.GET['view'] = 'geral'
        request.GET._mutable = False
        qs = self.filter_queryset(self.get_queryset())
        hoje = timezone.localdate()
        primeiro_dia = hoje.replace(day=1)
        ultimo_dia = (primeiro_dia + relativedelta(months=1)) - timedelta(days=1)

        vendas_mes = qs.filter(
            data_criacao__date__gte=primeiro_dia,
            data_criacao__date__lte=ultimo_dia
        ).order_by('-data_criacao')

        total_vendas_mes = vendas_mes.count()

        venda_ids_permitidos = list(qs.values_list('id', flat=True))
        envio_venda_ids = list(
            PapConfirmacaoCliente.objects.filter(
                venda_id__in=venda_ids_permitidos,
                criado_em__date__gte=primeiro_dia,
                criado_em__date__lte=ultimo_dia
            ).values_list('venda_id', flat=True).distinct()
        )
        total_envios_resumo = len(envio_venda_ids)
        lista_envios = vendas_mes.filter(id__in=envio_venda_ids) if envio_venda_ids else vendas_mes.none()

        lista_confirmacoes = vendas_mes.filter(
            cliente_confirmou_auditoria=True,
            data_confirmacao_auditoria__date__gte=primeiro_dia,
            data_confirmacao_auditoria__date__lte=ultimo_dia
        ).order_by('-data_confirmacao_auditoria')
        total_confirmacoes = lista_confirmacoes.count()

        # Lista por BO: username, tratou, enviou, confirmaram (mês atual)
        vendas_mes_ids = list(vendas_mes.values_list('id', flat=True))
        User = get_user_model()
        user_ids_hist = HistoricoAlteracaoVenda.objects.filter(
            venda_id__in=vendas_mes_ids,
            data_alteracao__date__gte=primeiro_dia,
            data_alteracao__date__lte=ultimo_dia,
            usuario__isnull=False
        ).values_list('usuario_id', flat=True).distinct()
        user_ids_envio = PapConfirmacaoCliente.objects.filter(
            venda_id__in=venda_ids_permitidos,
            criado_em__date__gte=primeiro_dia,
            criado_em__date__lte=ultimo_dia,
            enviado_por_id__isnull=False
        ).values_list('enviado_por_id', flat=True).distinct()
        all_user_ids = list(set(user_ids_hist) | set(user_ids_envio))
        lista_por_bo = []
        for uid in all_user_ids:
            user = User.objects.filter(pk=uid).first()
            if not user:
                continue
            tratou = HistoricoAlteracaoVenda.objects.filter(
                usuario_id=uid,
                venda_id__in=vendas_mes_ids,
                data_alteracao__date__gte=primeiro_dia,
                data_alteracao__date__lte=ultimo_dia
            ).values_list('venda_id', flat=True).distinct().count()
            enviou = PapConfirmacaoCliente.objects.filter(
                enviado_por_id=uid,
                venda_id__in=venda_ids_permitidos,
                criado_em__date__gte=primeiro_dia,
                criado_em__date__lte=ultimo_dia
            ).values_list('venda_id', flat=True).distinct().count()
            venda_ids_enviados_user = list(PapConfirmacaoCliente.objects.filter(
                enviado_por_id=uid,
                venda_id__in=venda_ids_permitidos,
                criado_em__date__gte=primeiro_dia,
                criado_em__date__lte=ultimo_dia
            ).values_list('venda_id', flat=True).distinct())
            confirmaram = Venda.objects.filter(
                id__in=venda_ids_enviados_user,
                cliente_confirmou_auditoria=True,
                data_confirmacao_auditoria__date__gte=primeiro_dia,
                data_confirmacao_auditoria__date__lte=ultimo_dia
            ).count() if venda_ids_enviados_user else 0
            lista_por_bo.append({
                'username': user.username,
                'tratou': tratou,
                'enviou': enviou,
                'confirmaram': confirmaram,
            })
        lista_por_bo.sort(key=lambda x: (-x['tratou'], -x['enviou'], x['username']))

        serializer = self.get_serializer(vendas_mes, many=True)
        serializer_envios = self.get_serializer(lista_envios, many=True)
        serializer_confirmacoes = self.get_serializer(lista_confirmacoes, many=True)

        return Response({
            'periodo': {
                'primeiro_dia': primeiro_dia.isoformat(),
                'ultimo_dia': ultimo_dia.isoformat(),
                'mes_ano': f"{primeiro_dia.month:02d}/{primeiro_dia.year}",
            },
            'total_vendas_mes': total_vendas_mes,
            'total_envios_resumo': total_envios_resumo,
            'total_confirmacoes': total_confirmacoes,
            'lista_vendas': serializer.data,
            'lista_envios_resumo': serializer_envios.data,
            'lista_confirmacoes': serializer_confirmacoes.data,
            'lista_por_bo': lista_por_bo,
        })

    @action(detail=True, methods=['post'], url_path='alocar-auditoria', permission_classes=[permissions.IsAuthenticated])
    def alocar_auditoria(self, request, pk=None):
        venda = self.get_object()
        usuario = request.user
        grupos_permitidos = ['Diretoria', 'Admin', 'BackOffice', 'Supervisor', 'Auditoria', 'Qualidade']
        if not is_member(usuario, grupos_permitidos):
             return Response({"detail": "Permissão negada. Apenas auditores podem alocar vendas."}, status=status.HTTP_403_FORBIDDEN)
        if venda.auditor_atual and venda.auditor_atual != usuario:
            return Response({"detail": f"Esta venda já está sendo auditada por {venda.auditor_atual}."}, status=status.HTTP_409_CONFLICT)
        venda.auditor_atual = usuario
        venda.save()
        serializer = self.get_serializer(venda)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='liberar-auditoria', permission_classes=[permissions.IsAuthenticated])
    def liberar_auditoria(self, request, pk=None):
        venda = self.get_object()
        usuario = request.user
        is_supervisor = is_member(usuario, ['Diretoria', 'Admin', 'Supervisor'])
        if venda.auditor_atual and venda.auditor_atual != usuario and not is_supervisor:
             return Response({"detail": "Você não tem permissão para liberar uma venda travada por outro auditor."}, status=status.HTTP_403_FORBIDDEN)
        venda.auditor_atual = None
        venda.save()
        return Response({"detail": "Venda liberada com sucesso."})

    @action(detail=True, methods=['post'], url_path='finalizar_auditoria', permission_classes=[permissions.IsAuthenticated])
    def finalizar_auditoria(self, request, pk=None):
        venda = self.get_object()
        grupos_permitidos = ['Diretoria', 'Admin', 'BackOffice', 'Supervisor', 'Auditoria', 'Qualidade']
        if not is_member(request.user, grupos_permitidos):
             return Response({"detail": "Permissão negada."}, status=status.HTTP_403_FORBIDDEN)

        status_novo_id = request.data.get('status')
        observacoes = request.data.get('observacoes', '')
        dados_edicao = request.data.get('dados_atualizados', {})

        if not status_novo_id:
             return Response({"detail": "Status inválido."}, status=status.HTTP_400_BAD_REQUEST)

        status_obj = None
        if str(status_novo_id).isdigit():
             status_obj = StatusCRM.objects.filter(id=int(status_novo_id)).first()
        else:
             termo = str(status_novo_id).upper()
             if 'APROVAD' in termo: termo = 'AUDITADA'
             status_obj = StatusCRM.objects.filter(nome__iexact=termo, tipo='Tratamento').first()
             if not status_obj:
                 status_obj = StatusCRM.objects.filter(nome__icontains=termo, tipo='Tratamento').first()

        if not status_obj:
             return Response({"detail": f"Status '{status_novo_id}' não encontrado no banco."}, status=status.HTTP_400_BAD_REQUEST)

        def os_valida(valor):
            if not valor:
                return False
            return bool(re.fullmatch(r'(\d{8}|\d-\d{12})', str(valor).strip()))

        os_informada = (dados_edicao.get('ordem_servico', venda.ordem_servico) or '').strip()
        if os_informada and not os_valida(os_informada):
            return Response(
                {"detail": "Formato de O.S inválido. Use 8 dígitos (08907507) ou X-12DÍGITOS (4-212051254235)."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if status_obj.nome.upper() == 'CADASTRADA' and not os_valida(os_informada):
            return Response(
                {"detail": "Para mudar o status para CADASTRADA é obrigatório informar O.S válida."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                if dados_edicao:
                    email_editado = dados_edicao.get('cliente_email')
                    if email_editado is not None:
                        email_editado = str(email_editado).strip()
                        if email_editado:
                            try:
                                validate_email(email_editado, check_deliverability=False)
                            except EmailNotValidError:
                                return Response(
                                    {"detail": "O campo de e-mail está fora do padrão esperado (ex: nome@dominio.com)."},
                                    status=status.HTTP_400_BAD_REQUEST
                                )
                        dados_edicao['cliente_email'] = email_editado

                    novo_cpf = None
                    if 'cliente_cpf' in dados_edicao:
                        novo_cpf = re.sub(r'\D', '', dados_edicao['cliente_cpf'])
                    
                    # Se o CPF foi alterado, verifica se já existe outro cliente com esse CPF
                    if novo_cpf and novo_cpf != venda.cliente.cpf_cnpj:
                        cliente_existente = Cliente.objects.filter(cpf_cnpj=novo_cpf).first()
                        if cliente_existente:
                            # Se existe, vincula a venda a esse cliente e atualiza os dados
                            venda.cliente = cliente_existente
                            if 'cliente_nome' in dados_edicao: cliente_existente.nome_razao_social = dados_edicao['cliente_nome'].upper()
                            if 'cliente_email' in dados_edicao: cliente_existente.email = dados_edicao['cliente_email']
                            cliente_existente.save()
                        else:
                            # Se não existe, atualiza o cliente atual normalmente
                            venda.cliente.cpf_cnpj = novo_cpf
                            if 'cliente_nome' in dados_edicao: venda.cliente.nome_razao_social = dados_edicao['cliente_nome'].upper()
                            if 'cliente_email' in dados_edicao: venda.cliente.email = dados_edicao['cliente_email']
                            venda.cliente.save()
                    else:
                        # CPF não mudou ou não foi enviado, apenas atualiza nome/email do cliente atual
                        if 'cliente_nome' in dados_edicao: venda.cliente.nome_razao_social = dados_edicao['cliente_nome'].upper()
                        if 'cliente_email' in dados_edicao: venda.cliente.email = dados_edicao['cliente_email']
                        venda.cliente.save()

                    if 'nome_mae' in dados_edicao:
                        nm = dados_edicao['nome_mae']
                        venda.nome_mae = (nm or '').strip() or None
                    if 'data_nascimento' in dados_edicao: 
                        dt = dados_edicao['data_nascimento']
                        venda.data_nascimento = None if dt == "" else dt
                    if 'mes_nascimento_pap' in dados_edicao:
                        mnp = dados_edicao['mes_nascimento_pap']
                        if mnp is None or mnp == '':
                            venda.mes_nascimento_pap = None
                        else:
                            try:
                                mi = int(mnp)
                                venda.mes_nascimento_pap = mi if 1 <= mi <= 12 else None
                            except (TypeError, ValueError):
                                venda.mes_nascimento_pap = None
                    if 'telefone1' in dados_edicao: venda.telefone1 = dados_edicao['telefone1']
                    if 'telefone2' in dados_edicao: venda.telefone2 = dados_edicao['telefone2']
                    if 'cep' in dados_edicao: venda.cep = str(dados_edicao['cep'])[:9]
                    if 'logradouro' in dados_edicao: venda.logradouro = (dados_edicao['logradouro'] or '').upper()
                    if 'numero' in dados_edicao: venda.numero_residencia = dados_edicao['numero']
                    if 'complemento' in dados_edicao: venda.complemento = (dados_edicao['complemento'] or '').upper()
                    if 'bairro' in dados_edicao: venda.bairro = (dados_edicao['bairro'] or '').upper()
                    if 'cidade' in dados_edicao: venda.cidade = (dados_edicao['cidade'] or '').upper()
                    if 'estado' in dados_edicao: venda.estado = str(dados_edicao['estado'] or '').upper()[:2]
                    if 'referencia' in dados_edicao: venda.ponto_referencia = (dados_edicao['referencia'] or '').upper()
                    if 'plano' in dados_edicao and dados_edicao['plano']: venda.plano_id = dados_edicao['plano']
                    if 'forma_pagamento' in dados_edicao and dados_edicao['forma_pagamento']: venda.forma_pagamento_id = dados_edicao['forma_pagamento']
                    if 'data_agendamento' in dados_edicao: 
                        dt_ag = dados_edicao['data_agendamento']
                        venda.data_agendamento = None if dt_ag == "" else dt_ag
                    if 'periodo_agendamento' in dados_edicao: venda.periodo_agendamento = dados_edicao['periodo_agendamento']
                    if 'ordem_servico' in dados_edicao: venda.ordem_servico = (dados_edicao['ordem_servico'] or '').strip()

                venda.status_tratamento = status_obj
                if observacoes: venda.observacoes = observacoes
                venda.auditor_atual = None
                
                # --- BLOCO CORRIGIDO ---
                if status_obj.nome.upper() == 'CADASTRADA':
                    # Esta linha e as abaixo devem estar alinhadas (identadas)
                    venda.data_abertura = timezone.now()
                    
                    if not venda.status_esteira:
                        st_agendado = StatusCRM.objects.filter(nome__iexact='AGENDADO', tipo='Esteira').first()
                        if st_agendado: 
                            venda.status_esteira = st_agendado

                venda.save()
                
                HistoricoAlteracaoVenda.objects.create(
                    venda=venda,
                    usuario=request.user,
                    alteracoes={'status_tratamento': f"Auditoria finalizada: {status_obj.nome}", 'dados_editados': bool(dados_edicao)}
                )

                nm_st = status_obj.nome.upper()
                STATUS_SUCESSO = ['AUDITADA', 'CADASTRADA', 'APROVADA', 'INSTALADA', 'AGENDADO', 'CONCLUIDA', 'CONCLUÍDA']
                eh_repro = not any(s in nm_st for s in STATUS_SUCESSO)
                
                if eh_repro and venda.vendedor and venda.vendedor.tel_whatsapp:
                    try:
                        svc = WhatsAppService()
                        end_parts = []
                        if venda.logradouro: end_parts.append(venda.logradouro)
                        if venda.numero_residencia: end_parts.append(venda.numero_residencia)
                        if venda.bairro: end_parts.append(venda.bairro)
                        end_str = ", ".join(end_parts) if end_parts else "Endereço não informado"

                        msg_text = (
                            f"*Nome Completo:* {venda.cliente.nome_razao_social}\n"
                            f"*CPF/CNPJ:* {venda.cliente.cpf_cnpj}\n"
                            f"*Endereço de Instalação:* {end_str}\n"
                            f"*Status de Auditoria:* {status_obj.nome}\n"
                            f"*Observações:* {observacoes or 'Verificar no sistema.'}"
                        )
                        svc.enviar_mensagem_texto(venda.vendedor.tel_whatsapp, msg_text)
                        logger.info(f"Zap de reprovação enviado para {venda.vendedor.username}")
                    except Exception as e_zap:
                        logger.error(f"Erro ao enviar Zap Reprovação: {e_zap}")

            return Response({"status": "Auditoria finalizada com sucesso."})

        except Exception as e:
            logger.error(f"Erro crítico auditoria: {str(e)}", exc_info=True)
            return Response({"detail": f"Erro ao salvar dados: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def _sincronizar_sequencia_cliente(self):
        from django.db import connection
        table_name = Cliente._meta.db_table
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT COALESCE(MAX(id), 0) FROM {table_name};")
            max_id = cursor.fetchone()[0]
            cursor.execute("SELECT pg_get_serial_sequence(%s, 'id');", [table_name])
            seq_name = cursor.fetchone()[0]
            if seq_name:
                cursor.execute("SELECT setval(%s, %s, true);", [seq_name, max_id])

    def perform_create(self, serializer):
        raw_cpf = serializer.validated_data.pop('cliente_cpf_cnpj')
        cpf_limpo = re.sub(r'\D', '', raw_cpf)
        
        nome = serializer.validated_data.pop('cliente_nome_razao_social')
        email = serializer.validated_data.pop('cliente_email', None)
        
        try:
            cliente, created = Cliente.objects.get_or_create(
                cpf_cnpj=cpf_limpo,
                defaults={'nome_razao_social': nome, 'email': email}
            )
        except IntegrityError as e:
            erro_texto = str(e).lower()
            if 'cliente_pkey' in erro_texto or ('crm_cliente' in erro_texto and 'duplicate' in erro_texto):
                self._sincronizar_sequencia_cliente()
                cliente, created = Cliente.objects.get_or_create(
                    cpf_cnpj=cpf_limpo,
                    defaults={'nome_razao_social': nome, 'email': email}
                )
            else:
                raise
        if not created:
            cliente.nome_razao_social = nome
            if email: cliente.email = email
            cliente.save()
            
        status_inicial = StatusCRM.objects.filter(nome__iexact="SEM TRATAMENTO", tipo__iexact="Tratamento").first()
        if not status_inicial:
            status_inicial = StatusCRM.objects.filter(tipo__iexact="Tratamento").first()

        vendedor = self.request.user
        vendedor_data = serializer.validated_data.pop('vendedor', None)
        if vendedor_data and is_member(self.request.user, ['Diretoria', 'Admin', 'BackOffice', 'Supervisor']):
            vendedor = vendedor_data
        gerada_os = serializer.validated_data.pop('gerada_os_automatica', False)

        serializer.save(vendedor=vendedor, cliente=cliente, status_tratamento=status_inicial, gerada_os_automatica=gerada_os)

    def perform_update(self, serializer):
        venda_antes = self.get_object()
        status_esteira_antes = venda_antes.status_esteira
        status_tratamento_antes = venda_antes.status_tratamento
        
        # Capturar valores antes da atualização para detectar mudanças nos dados
        data_agendamento_antes = venda_antes.data_agendamento
        periodo_agendamento_antes = venda_antes.periodo_agendamento
        data_instalacao_antes = venda_antes.data_instalacao
        motivo_pendencia_antes = venda_antes.motivo_pendencia
        
        novo_status = serializer.validated_data.get('status_esteira')
        extra_updates = {}

        if novo_status:
            nome_status = novo_status.nome.upper()
            if 'PENDEN' not in nome_status and 'PENDÊN' not in nome_status:
                extra_updates['motivo_pendencia'] = None
            if 'AGENDADO' not in nome_status and 'INSTALADA' not in nome_status:
                extra_updates['data_agendamento'] = None
                extra_updates['periodo_agendamento'] = None

        cliente_data = serializer.validated_data.get('cliente')
        if cliente_data and 'cpf_cnpj' in cliente_data:
             cliente_data['cpf_cnpj'] = re.sub(r'\D', '', cliente_data['cpf_cnpj'])

        venda_atualizada = serializer.save(**extra_updates)

        # Quitação do adiantamento sábado ao passar para INSTALADA (evita segundo pagamento na aba Instaladas)
        try:
            _quitar_adiantamento_sabado_na_instalacao(venda_atualizada, status_esteira_antes)
        except Exception:
            logger.exception('Erro ao quitar adiantamento sábado na instalação')

        alteracoes = {}
        if venda_antes.status_esteira != venda_atualizada.status_esteira:
            alteracoes['status_esteira'] = {
                'de': str(status_esteira_antes), 
                'para': str(venda_atualizada.status_esteira)
            }
        if venda_antes.status_tratamento != venda_atualizada.status_tratamento:
            alteracoes['status_tratamento'] = {
                'de': str(status_tratamento_antes), 
                'para': str(venda_atualizada.status_tratamento)
            }

        if alteracoes:
            try:
                HistoricoAlteracaoVenda.objects.create(
                    venda=venda_atualizada,
                    usuario=self.request.user,
                    alteracoes=alteracoes
                )
            except Exception as e:
                logger.error(f"Erro ao salvar histórico: {e}")

        # --- 2. NOVA LÓGICA DE NOTIFICAÇÃO WHATSAPP (ESTEIRA) ---
        # Verifica se status mudou OU se dados importantes foram alterados
        status_mudou = venda_atualizada.status_esteira and venda_atualizada.status_esteira != status_esteira_antes
        dados_mudaram = (
            (data_agendamento_antes != venda_atualizada.data_agendamento) or
            (periodo_agendamento_antes != venda_atualizada.periodo_agendamento) or
            (data_instalacao_antes != venda_atualizada.data_instalacao) or
            (motivo_pendencia_antes != venda_atualizada.motivo_pendencia)
        )

        enviar_whatsapp = self.request.data.get('enviar_whatsapp', True)
        if isinstance(enviar_whatsapp, str):
            enviar_whatsapp = enviar_whatsapp.lower() not in ['false', '0', 'no', 'nao', 'off']
        
        if enviar_whatsapp and venda_atualizada.status_esteira and (status_mudou or dados_mudaram):
            novo_status_nome = venda_atualizada.status_esteira.nome.upper()
            
            if ('PENDEN' in novo_status_nome or 'AGENDADO' in novo_status_nome or 'INSTALADA' in novo_status_nome) and 'CANCEL' not in novo_status_nome:
                
                if venda_atualizada.vendedor and venda_atualizada.vendedor.tel_whatsapp:
                    try:
                        svc = WhatsAppService()
                        msg = ""
                        
                        # Determina se é alteração de dados ou mudança de status
                        prefixo = "🔄 *ATUALIZAÇÃO - " if (dados_mudaram and not status_mudou) else ""
                        
                        cliente_nome = venda_atualizada.cliente.nome_razao_social
                        os_num = venda_atualizada.ordem_servico or "Não informada"
                        status_label = venda_atualizada.status_esteira.nome
                        obs = venda_atualizada.observacoes or "Sem observações"

                        if 'PENDEN' in novo_status_nome:
                            motivo = venda_atualizada.motivo_pendencia.nome if venda_atualizada.motivo_pendencia else "Não informado"
                            titulo = f"{prefixo}VENDA PENDENCIADA*" if prefixo else "⚠️ *VENDA PENDENCIADA*"
                            msg = (
                                f"{titulo}\n\n"
                                f"*Nome do cliente:* {cliente_nome}\n"
                                f"*O.S:* {os_num}\n"
                                f"*Status:* {status_label}\n"
                                f"*Motivo pendência:* {motivo}\n"
                                f"*Observação:* {obs}"
                            )

                        elif 'AGENDADO' in novo_status_nome:
                            data_ag = venda_atualizada.data_agendamento.strftime('%d/%m/%Y') if venda_atualizada.data_agendamento else "Não informada"
                            turno = venda_atualizada.get_periodo_agendamento_display() if venda_atualizada.periodo_agendamento else "Não informado"
                            titulo = f"{prefixo}VENDA AGENDADA*" if prefixo else "📅 *VENDA AGENDADA*"
                            msg = (
                                f"{titulo}\n\n"
                                f"*Nome do cliente:* {cliente_nome}\n"
                                f"*O.S:* {os_num}\n"
                                f"*Status:* {status_label}\n"
                                f"*Data e turno agendado:* {data_ag} - {turno}\n"
                                f"*Observação:* {obs}\n\n"
                                f"Lembrete: Peça ao seu cliente que salve o número 21 4040-1810, o técnico toda vez que coloca uma pendência o CO Digital faz uma ligação automática ao cliente para confirmar. Salvando o número ele atende e evita pendências indevidas!"
                            )

                        elif 'INSTALADA' in novo_status_nome:
                            data_inst = venda_atualizada.data_instalacao.strftime('%d/%m/%Y') if venda_atualizada.data_instalacao else date.today().strftime('%d/%m/%Y')
                            titulo = f"{prefixo}VENDA INSTALADA*" if prefixo else "✅ *VENDA INSTALADA*"
                            msg = (
                                f"{titulo}\n\n"
                                f"*Nome do cliente:* {cliente_nome}\n"
                                f"*O.S:* {os_num}\n"
                                f"*Status:* {status_label}\n"
                                f"*Data instalada:* {data_inst}\n"
                                f"*Observação:* {obs}"
                            )
                        
                        if msg:
                            svc.enviar_mensagem_texto(venda_atualizada.vendedor.tel_whatsapp, msg)
                            logger.info(f"Zap Esteira ({novo_status_nome}) enviado para {venda_atualizada.vendedor.username}")

                    except Exception as e:
                        logger.error(f"Erro ao enviar Zap Esteira: {e}")

        return venda_atualizada

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        try:
            instance.ativo = False
            instance.save()
            HistoricoAlteracaoVenda.objects.create(
                venda=instance,
                usuario=request.user,
                alteracoes={"acao": "exclusao_logica", "detalhe": f"Venda inativada por {request.user.username}."}
            )
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            logger.error(f"Erro ao excluir venda {kwargs.get('pk')}: {e}", exc_info=True)
            return Response({"detail": "Erro ao excluir venda."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # --- NOVA AÇÃO: EXPORTAR EXCEL ---
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        import time
        import pandas as pd
        from django.http import HttpResponse
        from io import BytesIO
        start_time = time.time()
        print(f"[EXPORTAR_EXCEL] Início: {start_time}")

        user = request.user
        if not is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({"detail": "Acesso negado."}, status=status.HTTP_403_FORBIDDEN)

        base_completa = request.query_params.get('base_completa', '').strip().lower() in ('1', 'true', 'sim', 's')
        if base_completa:
            vendas = Venda.objects.filter(ativo=True).select_related(
                'vendedor', 'vendedor__supervisor', 'cliente', 'plano', 'forma_pagamento',
                'status_tratamento', 'status_esteira', 'status_comissionamento',
                'motivo_pendencia',
            ).order_by('-data_criacao')
        else:
            vendas = self.filter_queryset(self.get_queryset())

        headers = [
            'ID', 'Reemissão', 'Data Criação', 'Data Abertura (OS)', 'Vendedor', 'Supervisor', 'Canal',
            'Cliente', 'CPF/CNPJ', 'Telefone 1', 'Telefone 2', 'Email',
            'Plano', 'Valor', 'Forma Pagamento',
            'Status Esteira', 'Status Tratamento', 'Status Comissionamento',
            'OS', 'Data Agendamento', 'Turno', 'Data Instalação', 'Data Física (no cliente)',
            'Adiant. CNPJ', 'Adiantamento de Comissão',
            'Motivo Pendência', 'Observações',
            'CEP', 'Logradouro', 'Número', 'Complemento', 'Bairro', 'Cidade', 'UF', 'Ponto Ref.'
        ]

        data = []
        for v in vendas.iterator():
            sup_nome = v.vendedor.supervisor.username if v.vendedor and v.vendedor.supervisor else '-'
            canal_venda = getattr(v.vendedor, 'canal', '-') if v.vendedor else '-'
            from django.utils import timezone
            dt_criacao = timezone.localtime(v.data_criacao).strftime('%d/%m/%Y %H:%M') if v.data_criacao else '-'
            dt_abertura = timezone.localtime(v.data_abertura).strftime('%d/%m/%Y %H:%M') if v.data_abertura else '-'
            dt_agendamento = v.data_agendamento.strftime('%d/%m/%Y') if v.data_agendamento else '-'
            dt_instalacao = v.data_instalacao.strftime('%d/%m/%Y') if v.data_instalacao else '-'
            dt_instalacao_fisica = v.data_instalacao_fisica.strftime('%d/%m/%Y') if v.data_instalacao_fisica else '-'
            data.append([
                v.id,
                'Sim' if getattr(v, 'reemissao', False) else 'Não',
                dt_criacao,
                dt_abertura,
                v.vendedor.username if v.vendedor else '-',
                sup_nome,
                canal_venda,
                v.cliente.nome_razao_social if v.cliente else '-',
                v.cliente.cpf_cnpj if v.cliente else '-',
                v.telefone1 or '-',
                v.telefone2 or '-',
                v.cliente.email if v.cliente else '-',
                v.plano.nome if v.plano else '-',
                v.plano.valor if v.plano else 0.00,
                v.forma_pagamento.nome if v.forma_pagamento else '-',
                v.status_esteira.nome if v.status_esteira else '-',
                v.status_tratamento.nome if v.status_tratamento else '-',
                v.status_comissionamento.nome if v.status_comissionamento else '-',
                v.ordem_servico or '-',
                dt_agendamento,
                v.get_periodo_agendamento_display() or '-',
                dt_instalacao,
                dt_instalacao_fisica,
                'Sim' if v.flag_adiant_cnpj else 'Não',
                'Sim' if v.antecipacao_comissao else 'Não',
                v.motivo_pendencia.nome if v.motivo_pendencia else '-',
                v.observacoes or '-',
                v.cep or '-',
                v.logradouro or '-',
                v.numero_residencia or '-',
                v.complemento or '-',
                v.bairro or '-',
                v.cidade or '-',
                v.estado or '-',
                v.ponto_referencia or '-'
            ])

        df = pd.DataFrame(data, columns=headers)
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        sufixo = 'Completa' if base_completa else 'Filtrada'
        filename = f"Base_Vendas_{sufixo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        end_time = time.time()
        print(f"[EXPORTAR_EXCEL] Fim: {end_time} | Duração: {end_time - start_time:.2f} segundos")
        return response

    @action(detail=False, methods=['post'], url_path='marcar-adiantamento-cnpj-semana')
    def marcar_adiantamento_cnpj_semana(self, request):
        if not request.user.is_superuser and not is_member(request.user, ['Diretoria', 'Admin']):
            return Response({'detail': 'Acesso negado. Apenas Diretoria/Admin podem executar esta ação.'}, status=status.HTTP_403_FORBIDDEN)
        vendedor_id = request.data.get('vendedor_id')
        data_inicio_venda = (request.data.get('data_inicio_inst') or '').strip()
        data_fim_venda = (request.data.get('data_fim_inst') or '').strip()
        if (data_inicio_venda or data_fim_venda) and not (data_inicio_venda and data_fim_venda):
            return Response({'detail': 'Informe Data início e Data fim.'}, status=status.HTTP_400_BAD_REQUEST)
        if data_inicio_venda and data_fim_venda:
            try:
                dt_ini_venda = datetime.strptime(data_inicio_venda, '%Y-%m-%d').date()
                dt_fim_venda = datetime.strptime(data_fim_venda, '%Y-%m-%d').date()
                if dt_ini_venda > dt_fim_venda:
                    return Response({'detail': 'Data início não pode ser maior que data fim.'}, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                return Response({'detail': 'Período inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = str(request.data.get('dry_run', '')).lower() in ('1', 'true', 'sim', 's')
        agora_sp = timezone.localtime(timezone.now())
        semana_ini_atual = (agora_sp - timedelta(days=agora_sp.weekday())).date()
        # Regra acordada: segunda 00:00 até segunda 23:59:59 (8 dias corridos)
        semana_fim_atual = semana_ini_atual + timedelta(days=7)
        # Período filtrado representa a janela de VENDA. Sem filtro, usa semana atual (seg->segunda seguinte).
        ref_ini = dt_ini_venda if (data_inicio_venda and data_fim_venda) else semana_ini_atual
        ref_fim = dt_fim_venda if (data_inicio_venda and data_fim_venda) else semana_fim_atual

        vendas = Venda.objects.filter(
            ativo=True,
            status_esteira__nome__iexact='INSTALADA',
        ).select_related('cliente', 'vendedor')
        if vendedor_id and str(vendedor_id).isdigit():
            vendas = vendas.filter(vendedor_id=int(vendedor_id))
        if data_inicio_venda and data_fim_venda:
            vendas = vendas.filter(data_criacao__date__gte=dt_ini_venda, data_criacao__date__lte=dt_fim_venda)

        elegiveis_ids = []
        ignoradas = 0
        total_avaliadas = 0
        for v in vendas:
            total_avaliadas += 1
            doc = re.sub(r'\D', '', (v.cliente.cpf_cnpj if v.cliente else '') or '')
            if len(doc) != 14:
                ignoradas += 1
                continue
            if not v.data_criacao:
                ignoradas += 1
                continue
            sale_date = v.data_criacao.date()
            if not (ref_ini <= sale_date <= ref_fim):
                ignoradas += 1
                continue
            week_start = sale_date - timedelta(days=sale_date.weekday())
            week_end_sale = week_start + timedelta(days=6)
            install_deadline = week_start + timedelta(days=7)
            if not (week_start <= sale_date <= week_end_sale):
                ignoradas += 1
                continue
            if not (v.data_instalacao and week_start <= v.data_instalacao <= install_deadline):
                ignoradas += 1
                continue
            if not v.vendedor or getattr(v.vendedor, 'recebe_adiantamento_cnpj', True) is False:
                ignoradas += 1
                continue
            if v.flag_adiant_cnpj:
                ignoradas += 1
                continue
            elegiveis_ids.append(v.id)

        if elegiveis_ids and not dry_run:
            vendas_para_lancar = list(
                Venda.objects.filter(id__in=elegiveis_ids).select_related('vendedor', 'cliente')
            )
            hoje = timezone.localdate()
            por_vendedor = defaultdict(list)
            for v in vendas_para_lancar:
                if v.vendedor_id:
                    por_vendedor[v.vendedor_id].append(v)
            with transaction.atomic():
                Venda.objects.filter(id__in=elegiveis_ids).update(
                    flag_adiant_cnpj=True,
                    adiantamento_cnpj_realizado_em=timezone.now(),
                    adiantamento_cnpj_realizado_por=request.user,
                )
                for vid, lista_v in por_vendedor.items():
                    lanc = LancamentoFinanceiro.objects.filter(
                        usuario_id=vid,
                        tipo='ADIANTAMENTO_CNPJ',
                        data=hoje,
                        descricao='Adiantamento CNPJ (Instaladas)',
                    ).first()
                    meta_old = {}
                    if lanc is not None and isinstance(
                        getattr(lanc, 'metadados', None), dict
                    ):
                        meta_old = lanc.metadados
                    venda_ids = list(meta_old.get('venda_ids') or [])
                    valor_acum = Decimal(str(lanc.valor or 0)) if lanc else Decimal('0')
                    for v in lista_v:
                        cons = v.vendedor
                        if not cons:
                            continue
                        unit = Decimal(
                            str(getattr(cons, 'adiantamento_cnpj', None) or 0)
                        )
                        if unit <= 0:
                            continue
                        if v.id in venda_ids:
                            continue
                        venda_ids.append(v.id)
                        valor_acum += unit
                    if not venda_ids:
                        continue
                    qtd = len(venda_ids)
                    meta_new = {'origem': 'instaladas_cnpj', 'venda_ids': venda_ids}
                    if lanc:
                        lanc.valor = valor_acum
                        lanc.quantidade_vendas = qtd
                        lanc.metadados = meta_new
                        lanc.save(
                            update_fields=['valor', 'quantidade_vendas', 'metadados']
                        )
                    else:
                        LancamentoFinanceiro.objects.create(
                            usuario_id=vid,
                            tipo='ADIANTAMENTO_CNPJ',
                            data=hoje,
                            valor=valor_acum,
                            quantidade_vendas=qtd,
                            descricao='Adiantamento CNPJ (Instaladas)',
                            metadados=meta_new,
                            criado_por=request.user,
                        )
        return Response({
            'elegiveis_marcadas': len(elegiveis_ids),
            'ignoradas': ignoradas,
            'total_avaliadas': total_avaliadas,
            'semana_inicio': ref_ini.isoformat(),
            'semana_fim': ref_fim.isoformat(),
            'dry_run': dry_run,
        })

    @action(detail=True, methods=['post'], url_path='toggle-adiantamento-comissao')
    def toggle_adiantamento_comissao(self, request, pk=None):
        if not request.user.is_superuser and not is_member(request.user, ['Diretoria', 'Admin']):
            return Response({'detail': 'Acesso negado. Apenas Diretoria/Admin podem editar Adiantamento de Comissão.'}, status=status.HTTP_403_FORBIDDEN)
        venda = self.get_object()
        marcar = str(request.data.get('marcar', '')).lower() in ('1', 'true', 'sim', 's')
        if not venda.vendedor_id:
            return Response({'detail': 'Venda sem vendedor.'}, status=status.HTTP_400_BAD_REQUEST)
        if not venda.status_esteira or (venda.status_esteira.nome or '').strip().upper() != 'INSTALADA':
            return Response({'detail': 'Apenas vendas instaladas podem receber adiantamento de comissão.'}, status=status.HTTP_400_BAD_REQUEST)

        faixa_adiantamento = RegraComissaoFaixa.objects.filter(finalidade='ADIANTAMENTO').first()
        if not faixa_adiantamento:
            faixa_adiantamento = RegraComissaoFaixa.objects.filter(faixa_nome__iexact='Adiantamento').first()
        doc = re.sub(r'\D', '', (venda.cliente.cpf_cnpj if venda.cliente else '') or '')
        is_cnpj_venda = len(doc) == 14
        valor_unit = Decimal(str(LancamentoFinanceiroViewSet._valor_comissao_estimado_venda(venda, faixa_adiantamento, is_cnpj_venda)))
        if valor_unit <= 0:
            return Response({'detail': 'Valor de adiantamento da venda é zero. Revise Regras por Faixa (Finalidade = Adiantamento).'}, status=status.HTTP_400_BAD_REQUEST)

        hoje = timezone.localdate()
        with transaction.atomic():
            if marcar:
                # Já pago na esteira (Agendados) — não duplica lançamento do dia em Instaladas
                if getattr(venda, 'adiantamento_sabado_marcado', False):
                    if venda.antecipacao_comissao:
                        return Response({
                            'ok': True,
                            'antecipacao_comissao': True,
                            'adiantamento_sabado_sem_novo_lancamento': True,
                        })
                    venda.antecipacao_comissao = True
                    if not venda.adiantamento_sabado_quitado_em:
                        venda.adiantamento_sabado_quitado_em = timezone.now()
                    venda.save(update_fields=['antecipacao_comissao', 'adiantamento_sabado_quitado_em'])
                    return Response({
                        'ok': True,
                        'antecipacao_comissao': True,
                        'adiantamento_sabado_sem_novo_lancamento': True,
                    })
                if venda.antecipacao_comissao:
                    return Response({'ok': True, 'antecipacao_comissao': True})
                venda.antecipacao_comissao = True
                venda.save(update_fields=['antecipacao_comissao'])
                lanc = LancamentoFinanceiro.objects.filter(
                    usuario_id=venda.vendedor_id,
                    tipo='ADIANTAMENTO_COMISSAO',
                    data=hoje,
                    descricao='Adiantamento comissão (Instaladas)',
                ).first()
                venda_ids = []
                if lanc and isinstance(lanc.metadados, dict):
                    venda_ids = list(lanc.metadados.get('venda_ids') or [])
                if venda.id not in venda_ids:
                    venda_ids.append(venda.id)
                if lanc:
                    lanc.valor = Decimal(str(lanc.valor or 0)) + valor_unit
                    lanc.quantidade_vendas = max(int(lanc.quantidade_vendas or 0), 0) + 1
                    lanc.metadados = {'origem': 'instaladas_comissao', 'venda_ids': venda_ids}
                    lanc.save(update_fields=['valor', 'quantidade_vendas', 'metadados'])
                else:
                    LancamentoFinanceiro.objects.create(
                        usuario_id=venda.vendedor_id,
                        tipo='ADIANTAMENTO_COMISSAO',
                        data=hoje,
                        valor=valor_unit,
                        quantidade_vendas=1,
                        descricao='Adiantamento comissão (Instaladas)',
                        metadados={'origem': 'instaladas_comissao', 'venda_ids': [venda.id]},
                        criado_por=request.user,
                    )
            else:
                if not venda.antecipacao_comissao:
                    return Response({'ok': True, 'antecipacao_comissao': False})
                venda.antecipacao_comissao = False
                venda.save(update_fields=['antecipacao_comissao'])
                lancs = LancamentoFinanceiro.objects.filter(
                    usuario_id=venda.vendedor_id,
                    tipo='ADIANTAMENTO_COMISSAO',
                    descricao='Adiantamento comissão (Instaladas)',
                    data=hoje,
                )
                for lanc in lancs:
                    meta = lanc.metadados if isinstance(lanc.metadados, dict) else {}
                    venda_ids = list(meta.get('venda_ids') or [])
                    if venda.id not in venda_ids:
                        continue
                    venda_ids = [vid for vid in venda_ids if int(vid) != int(venda.id)]
                    novo_valor = Decimal(str(lanc.valor or 0)) - valor_unit
                    nova_qtd = max(int(lanc.quantidade_vendas or 0) - 1, 0)
                    if nova_qtd <= 0 or novo_valor <= 0:
                        lanc.delete()
                    else:
                        lanc.valor = novo_valor
                        lanc.quantidade_vendas = nova_qtd
                        lanc.metadados = {'origem': 'instaladas_comissao', 'venda_ids': venda_ids}
                        lanc.save(update_fields=['valor', 'quantidade_vendas', 'metadados'])
                    break
        return Response({'ok': True, 'antecipacao_comissao': venda.antecipacao_comissao})

    @action(detail=True, methods=['post'], url_path='toggle-adiantamento-cnpj')
    def toggle_adiantamento_cnpj(self, request, pk=None):
        if not request.user.is_superuser and not is_member(
            request.user, ['Diretoria', 'Admin']
        ):
            return Response(
                {
                    'detail': 'Acesso negado. Apenas Diretoria/Admin podem editar Adiantamento CNPJ.'
                },
                status=status.HTTP_403_FORBIDDEN,
            )
        venda = self.get_object()
        marcar = str(request.data.get('marcar', '')).lower() in ('1', 'true', 'sim', 's')

        if marcar:
            if not venda.vendedor_id:
                return Response(
                    {'detail': 'Venda sem vendedor.'}, status=status.HTTP_400_BAD_REQUEST
                )
            if not venda.status_esteira or (venda.status_esteira.nome or '').strip().upper() != 'INSTALADA':
                return Response(
                    {'detail': 'Apenas vendas instaladas.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            doc = re.sub(
                r'\D', '', (venda.cliente.cpf_cnpj if venda.cliente else '') or ''
            )
            if len(doc) != 14:
                return Response(
                    {
                        'detail': 'Adiantamento CNPJ só se aplica a cliente CNPJ (14 dígitos).'
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if getattr(venda.vendedor, 'recebe_adiantamento_cnpj', True) is False:
                return Response(
                    {
                        'detail': 'Este vendedor não está habilitado para adiantamento CNPJ (cadastro).'
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            unit = Decimal(
                str(getattr(venda.vendedor, 'adiantamento_cnpj', None) or 0)
            )
            if unit <= 0:
                return Response(
                    {
                        'detail': 'Valor de adiantamento CNPJ do vendedor é zero. Ajuste na configuração do consultor.'
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            unit = (
                Decimal(
                    str(
                        getattr(venda.vendedor, 'adiantamento_cnpj', None) or 0
                    )
                )
                if venda.vendedor_id
                else Decimal('0')
            )

        hoje = timezone.localdate()
        with transaction.atomic():
            if marcar:
                if venda.flag_adiant_cnpj:
                    return Response({'ok': True, 'flag_adiant_cnpj': True})
                venda.flag_adiant_cnpj = True
                venda.adiantamento_cnpj_realizado_em = timezone.now()
                venda.adiantamento_cnpj_realizado_por = request.user
                venda.save(
                    update_fields=[
                        'flag_adiant_cnpj',
                        'adiantamento_cnpj_realizado_em',
                        'adiantamento_cnpj_realizado_por',
                    ]
                )
                lanc = LancamentoFinanceiro.objects.filter(
                    usuario_id=venda.vendedor_id,
                    tipo='ADIANTAMENTO_CNPJ',
                    data=hoje,
                    descricao='Adiantamento CNPJ (Instaladas)',
                ).first()
                venda_ids = []
                if lanc is not None and isinstance(
                    getattr(lanc, 'metadados', None), dict
                ):
                    venda_ids = list(lanc.metadados.get('venda_ids') or [])
                if venda.id not in venda_ids:
                    venda_ids.append(venda.id)
                if lanc:
                    lanc.valor = Decimal(str(lanc.valor or 0)) + unit
                    lanc.quantidade_vendas = len(venda_ids)
                    lanc.metadados = {'origem': 'instaladas_cnpj', 'venda_ids': venda_ids}
                    lanc.save(update_fields=['valor', 'quantidade_vendas', 'metadados'])
                else:
                    LancamentoFinanceiro.objects.create(
                        usuario_id=venda.vendedor_id,
                        tipo='ADIANTAMENTO_CNPJ',
                        data=hoje,
                        valor=unit,
                        quantidade_vendas=1,
                        descricao='Adiantamento CNPJ (Instaladas)',
                        metadados={'origem': 'instaladas_cnpj', 'venda_ids': [venda.id]},
                        criado_por=request.user,
                    )
            else:
                if not venda.flag_adiant_cnpj:
                    return Response({'ok': True, 'flag_adiant_cnpj': False})
                venda.flag_adiant_cnpj = False
                venda.adiantamento_cnpj_realizado_em = None
                venda.adiantamento_cnpj_realizado_por = None
                venda.save(
                    update_fields=[
                        'flag_adiant_cnpj',
                        'adiantamento_cnpj_realizado_em',
                        'adiantamento_cnpj_realizado_por',
                    ]
                )
                if not venda.vendedor_id:
                    return Response({'ok': True, 'flag_adiant_cnpj': False})
                lancs = LancamentoFinanceiro.objects.filter(
                    usuario_id=venda.vendedor_id,
                    tipo='ADIANTAMENTO_CNPJ',
                    descricao='Adiantamento CNPJ (Instaladas)',
                    data=hoje,
                )
                for lanc in lancs:
                    meta = (
                        lanc.metadados if isinstance(lanc.metadados, dict) else {}
                    )
                    old_ids = list(meta.get('venda_ids') or [])
                    if venda.id not in old_ids:
                        continue
                    n_before = len(old_ids)
                    valor_lanc = Decimal(str(lanc.valor or 0))
                    if unit > 0:
                        valor_sub = unit
                    elif n_before > 0:
                        valor_sub = valor_lanc / Decimal(n_before)
                    else:
                        valor_sub = Decimal('0')
                    venda_ids = [
                        vid for vid in old_ids if int(vid) != int(venda.id)
                    ]
                    novo_valor = valor_lanc - valor_sub
                    if novo_valor < 0:
                        novo_valor = Decimal('0')
                    nova_qtd = len(venda_ids)
                    if nova_qtd <= 0 or novo_valor <= 0:
                        lanc.delete()
                    else:
                        lanc.valor = novo_valor
                        lanc.quantidade_vendas = nova_qtd
                        lanc.metadados = {
                            'origem': 'instaladas_cnpj',
                            'venda_ids': venda_ids,
                        }
                        lanc.save(
                            update_fields=['valor', 'quantidade_vendas', 'metadados']
                        )
                    break
        return Response({'ok': True, 'flag_adiant_cnpj': venda.flag_adiant_cnpj})

    @action(detail=True, methods=['post'], url_path='marcar-adiantamento-sabado')
    def marcar_adiantamento_sabado(self, request, pk=None):
        if not _pode_gerir_adiantamento_sabado(request.user):
            return Response(
                {'detail': 'Acesso negado. Apenas Diretoria, Admin ou BackOffice.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        venda = self.get_object()
        manual = str(request.data.get('manual', '')).lower() in ('1', 'true', 'sim', 's')
        obs = (request.data.get('observacao') or '').strip()
        if manual and len(obs) < 3:
            return Response(
                {'detail': 'Informe a observação para marcação manual (mín. 3 caracteres).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            out = _marcar_adiantamento_sabado_exec(venda, request.user, manual=manual, obs=obs)
        except ValueError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'ok': True, **out})

    @action(detail=False, methods=['post'], url_path='marcar-adiantamento-sabado-lote')
    def marcar_adiantamento_sabado_lote(self, request):
        """
        Marca todas as vendas elegíveis cuja data de abertura da O.S. (timezone local)
        cai no sábado informado. Body: { "data": "YYYY-MM-DD", "dry_run": false }.
        Só vendedores com recebe_adiantamento_sabado=True; status AGENDADO; não marcadas.
        """
        if not _pode_gerir_adiantamento_sabado(request.user):
            return Response(
                {'detail': 'Acesso negado. Apenas Diretoria, Admin ou BackOffice.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        data_str = (request.data.get('data') or '').strip()
        dry_run = str(request.data.get('dry_run', '')).lower() in ('1', 'true', 'sim', 's')
        if not data_str:
            return Response(
                {'detail': 'Informe data (YYYY-MM-DD): sábado correspondente ao dia de abertura da O.S.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            ref_date = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        if ref_date.weekday() != 5:
            return Response(
                {'detail': 'A data deve ser um sábado (dia da abertura da O.S. no calendário).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from datetime import time as dt_time

        tz = timezone.get_current_timezone()
        start_dt = timezone.make_aware(datetime.combine(ref_date, dt_time.min), tz)
        end_dt = start_dt + timedelta(days=1)

        qs = (
            Venda.objects.filter(
                ativo=True,
                adiantamento_sabado_marcado=False,
                data_abertura__gte=start_dt,
                data_abertura__lt=end_dt,
                vendedor__recebe_adiantamento_sabado=True,
                status_esteira__nome__icontains='AGENDADO',
            )
            .select_related('vendedor', 'status_esteira', 'cliente', 'plano')
            .order_by('id')
        )

        candidatas = list(qs)
        if dry_run:
            return Response(
                {
                    'ok': True,
                    'dry_run': True,
                    'data': data_str,
                    'quantidade_elegiveis': len(candidatas),
                    'venda_ids': [v.id for v in candidatas],
                }
            )

        marcadas = []
        ignoradas = []
        for v in candidatas:
            try:
                out = _marcar_adiantamento_sabado_exec(v, request.user, manual=False, obs='')
                marcadas.append({'venda_id': v.id, **out})
            except ValueError as e:
                ignoradas.append({'venda_id': v.id, 'motivo': str(e)})

        return Response(
            {
                'ok': True,
                'data': data_str,
                'total_candidatas': len(candidatas),
                'marcadas': len(marcadas),
                'marcadas_detalhe': marcadas,
                'ignoradas': ignoradas,
            }
        )

    @action(detail=True, methods=['post'], url_path='desmarcar-adiantamento-sabado')
    def desmarcar_adiantamento_sabado(self, request, pk=None):
        if not _pode_gerir_adiantamento_sabado(request.user):
            return Response(
                {'detail': 'Acesso negado. Apenas Diretoria, Admin ou BackOffice.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        venda = self.get_object()
        st = (venda.status_esteira.nome if venda.status_esteira else '') or ''
        if 'AGENDADO' not in st.upper():
            return Response(
                {'detail': 'Só é possível desmarcar enquanto a venda estiver AGENDADA.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not venda.adiantamento_sabado_marcado:
            return Response(
                {'detail': 'Venda não está marcada como adiantamento sábado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data_lanc = (
            timezone.localtime(venda.data_abertura).date()
            if venda.data_abertura
            else timezone.localdate()
        )
        descricao = 'Adiantamento sábado (Agendados)'
        valor_unit = Decimal(str(venda.adiantamento_sabado_valor or 0))

        with transaction.atomic():
            for lanc in LancamentoFinanceiro.objects.filter(
                usuario_id=venda.vendedor_id,
                tipo='ADIANTAMENTO_COMISSAO',
                data=data_lanc,
                descricao=descricao,
            ):
                meta = lanc.metadados if isinstance(lanc.metadados, dict) else {}
                if meta.get('origem') != 'esteira_sabado_agendados':
                    continue
                vids = [int(x) for x in (meta.get('venda_ids') or []) if x is not None]
                if venda.id not in vids:
                    continue
                vids_new = [x for x in vids if int(x) != int(venda.id)]
                valores = dict(meta.get('valores_por_venda_id') or {})
                sub = valor_unit
                if sub <= 0:
                    sub = Decimal(str(valores.get(str(venda.id)) or 0))
                valores.pop(str(venda.id), None)
                novo_valor = Decimal(str(lanc.valor or 0)) - sub
                if novo_valor <= 0 or len(vids_new) == 0:
                    lanc.delete()
                else:
                    lanc.valor = novo_valor
                    lanc.quantidade_vendas = len(vids_new)
                    lanc.metadados = {
                        'origem': 'esteira_sabado_agendados',
                        'venda_ids': vids_new,
                        'valores_por_venda_id': valores,
                    }
                    lanc.save(update_fields=['valor', 'quantidade_vendas', 'metadados'])
                break

            venda.adiantamento_sabado_marcado = False
            venda.adiantamento_sabado_valor = None
            venda.adiantamento_sabado_marcado_em = None
            venda.adiantamento_sabado_marcado_por = None
            venda.adiantamento_sabado_manual = False
            venda.adiantamento_sabado_obs_manual = ''
            venda.save(
                update_fields=[
                    'adiantamento_sabado_marcado',
                    'adiantamento_sabado_valor',
                    'adiantamento_sabado_marcado_em',
                    'adiantamento_sabado_marcado_por',
                    'adiantamento_sabado_manual',
                    'adiantamento_sabado_obs_manual',
                ]
            )

        return Response({'ok': True})

class VendasStatusCountView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        venda_viewset = VendaViewSet()
        venda_viewset.request = request
        venda_viewset.action = 'list'
        queryset = venda_viewset.get_queryset()

        status_counts = list(queryset.values('status_esteira__nome').annotate(count=Count('id')).order_by('-count'))
        order_priority = {"INSTALADA": 0, "AGENDADO": 1}
        filtered_counts = [item for item in status_counts if item.get('status_esteira__nome')]

        def sort_key(item):
            nome = item['status_esteira__nome'].upper()
            return (order_priority.get(nome, 99), -item['count'])

        sorted_counts = sorted(filtered_counts, key=sort_key)
        return Response({item['status_esteira__nome']: item['count'] for item in sorted_counts})

class DashboardResumoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        User = get_user_model()
        user = request.user
        
        data_inicio_str = request.query_params.get('data_inicio')
        data_fim_str = request.query_params.get('data_fim')
        
        hoje = now()
        agora_local = timezone.localtime(hoje)
        hoje_date = agora_local.date()
        
        if data_inicio_str and data_fim_str:
            try:
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
                data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d')
                data_fim_ajustada = data_fim + timedelta(days=1)
                data_fim_date = data_fim.date()
                if vendedor_ou_supervisor_restrito_mes(user) and not mes_completo_vendedor_supervisor_valido(
                    data_inicio.date(), data_fim.date(), hoje_date
                ):
                    raise ValueError('periodo nao permitido')
            except ValueError:
                data_inicio = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                data_fim_ajustada = (data_inicio + timedelta(days=32)).replace(day=1)
                data_fim_date = (data_fim_ajustada - timedelta(days=1)).date()
        else:
            data_inicio = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            data_fim_ajustada = (data_inicio + timedelta(days=32)).replace(day=1)
            data_fim_date = (data_fim_ajustada - timedelta(days=1)).date()

        last_day_of_month = calendar.monthrange(data_inicio.year, data_inicio.month)[1]
        data_fim_mes_projetado = data_inicio.replace(day=last_day_of_month).date()

        dias_fiscais_db = DiaFiscal.objects.filter(
            data__range=(data_inicio.date(), data_fim_mes_projetado)
        )
        mapa_fiscais = {d.data: d for d in dias_fiscais_db}

        peso_total_mes_venda = 0.0
        peso_realizado_venda = 0.0
        peso_total_mes_inst = 0.0
        peso_realizado_inst = 0.0
        
        limite_realizado = min(hoje_date, data_fim_date)

        data_iter = data_inicio.date()
        while data_iter <= data_fim_mes_projetado:
            if data_iter in mapa_fiscais:
                p_venda = float(mapa_fiscais[data_iter].peso_venda)
                p_inst = float(mapa_fiscais[data_iter].peso_instalacao)
            else:
                weekday = data_iter.weekday()
                if weekday == 6: p_venda = 0.0; p_inst = 0.0
                elif weekday == 5: p_venda = 0.5; p_inst = 0.5
                else: p_venda = 1.0; p_inst = 1.0
            
            peso_total_mes_venda += p_venda
            peso_total_mes_inst += p_inst
            
            if data_iter <= limite_realizado:
                # Hoje ainda em andamento: não contar o peso fiscal do dia inteiro como "realizado",
                # senão peso_realizado == peso_total e a projeção fica igual ao acumulado.
                if data_iter == hoje_date:
                    meia_noite = agora_local.replace(hour=0, minute=0, second=0, microsecond=0)
                    decorrido_s = (agora_local - meia_noite).total_seconds()
                    fracao_dia = min(max(decorrido_s / 86400.0, 1.0 / 24.0), 1.0)
                    peso_realizado_venda += p_venda * fracao_dia
                    peso_realizado_inst += p_inst * fracao_dia
                else:
                    peso_realizado_venda += p_venda
                    peso_realizado_inst += p_inst
            
            data_iter += timedelta(days=1)

        def calcular_projecao(valor_atual, peso_realizado, peso_total):
            if not peso_realizado or float(peso_realizado) == 0:
                return 0
            return (float(valor_atual) / float(peso_realizado)) * float(peso_total)

        consultor_filtro_id = request.query_params.get('consultor_id')
        if consultor_filtro_id:
            usuarios_para_calcular = User.objects.filter(id=consultor_filtro_id)
        elif is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
            base_dash = User.objects.exclude(username__in=['OSAB_IMPORT', 'admin', 'root'])
            usuarios_para_calcular = _aplicar_filtro_vendedor_ativo_perf(
                base_dash, user, request.query_params.get('vendedor_ativo')
            )
        elif is_member(user, ['Supervisor']):
            usuarios_para_calcular = [user]
        else:
            usuarios_para_calcular = [user]

        if is_member(user, ['Diretoria', 'Admin']): exibir_comissao = True
        elif is_member(user, ['BackOffice']): exibir_comissao = False
        else: exibir_comissao = True
        
        is_diretoria = is_member(user, ['Diretoria'])
        mapa_comissao_operadora = {}
        if is_diretoria:
            configs = ComissaoOperadora.objects.all()
            for c in configs:
                mapa_comissao_operadora[c.plano_id] = {
                    'base': float(c.valor_base),
                    'bonus': float(c.bonus_transicao),
                    'fim_bonus': c.data_fim_bonus
                }

        total_registradas_geral = 0
        total_instaladas_geral = 0
        comissao_total_geral = 0.0
        status_counts_geral = defaultdict(int)
        meta_display = 0
        detalhes_listas = defaultdict(list)

        faturamento_operadora_real = 0.0
        mix_velocidade = defaultdict(int)
        mix_pagamento = defaultdict(int)

        for vendedor in usuarios_para_calcular:
            meta_individual = getattr(vendedor, 'meta_comissao', 0) or 0
            meta_display += float(meta_individual)

            vendas_registro = Venda.objects.filter(
                vendedor=vendedor, ativo=True,
                data_criacao__gte=data_inicio, data_criacao__lt=data_fim_ajustada
            ).select_related('cliente', 'status_esteira')
            
            qtd_registradas = vendas_registro.count()
            bateu_meta = qtd_registradas >= float(meta_individual)

            for v in vendas_registro:
                nome_status = v.status_esteira.nome.upper() if v.status_esteira else 'AGUARDANDO'
                if nome_status != 'INSTALADA':
                    status_counts_geral[nome_status] += 1
                
                obj_venda = {
                    'id': v.id, 'cliente': v.cliente.nome_razao_social if v.cliente else 'S/C',
                    'status': nome_status, 'data_iso': v.data_criacao.isoformat(), 'vendedor': vendedor.username
                }
                detalhes_listas['TOTAL_REGISTRADAS'].append(obj_venda)
                if nome_status != 'INSTALADA': detalhes_listas[nome_status].append(obj_venda)

            vendas_instaladas = Venda.objects.filter(
                vendedor=vendedor, ativo=True,
                status_esteira__nome__iexact='INSTALADA',
                ordem_servico__isnull=False,
            ).filter(
                ~Q(ordem_servico=''),
                _filtro_data_efetiva_instalacao_intervalo_venda(
                    data_inicio.date(),
                    data_fim_date
                )
            ).select_related('plano', 'cliente', 'forma_pagamento')

            qtd_instaladas = vendas_instaladas.count()
            status_counts_geral['INSTALADA'] += qtd_instaladas

            for vi in vendas_instaladas:
                obj_inst = {
                    'id': vi.id, 'cliente': vi.cliente.nome_razao_social if vi.cliente else 'S/C',
                    'status': 'INSTALADA',
                    'data_iso': (
                        (vi.data_instalacao_fisica or vi.data_instalacao).isoformat()
                        if (vi.data_instalacao_fisica or vi.data_instalacao) else None
                    ),
                    'vendedor': vendedor.username
                }
                detalhes_listas['TOTAL_INSTALADAS'].append(obj_inst)
                detalhes_listas['INSTALADA'].append(obj_inst)

                if is_diretoria:
                    nome_plano = vi.plano.nome if vi.plano else "Outros"
                    mix_velocidade[nome_plano] += 1
                    nome_pgto = vi.forma_pagamento.nome if vi.forma_pagamento else "Não Informado"
                    mix_pagamento[nome_pgto] += 1

                    if vi.plano_id in mapa_comissao_operadora:
                        cfg = mapa_comissao_operadora[vi.plano_id]
                        valor_venda = cfg['base']
                        if cfg['bonus'] > 0:
                            if not cfg['fim_bonus'] or (vi.data_instalacao and vi.data_instalacao <= cfg['fim_bonus']):
                                valor_venda += cfg['bonus']
                        faturamento_operadora_real += valor_venda

            if exibir_comissao:
                regras = RegraComissao.objects.filter(consultor=vendedor).select_related('plano')
                comissao_vendedor = 0.0
                for v in vendas_instaladas:
                    valor_item = 0.0 
                    doc = v.cliente.cpf_cnpj if v.cliente else ''
                    doc_limpo = ''.join(filter(str.isdigit, doc))
                    tipo_cliente_venda = 'CNPJ' if len(doc_limpo) > 11 else 'CPF'
                    
                    regra_encontrada = None
                    for r in regras:
                        canal_vendedor = getattr(vendedor, 'canal', 'PAP') or 'PAP'
                        if r.plano.id == v.plano.id and r.tipo_cliente == tipo_cliente_venda and r.tipo_venda == canal_vendedor:
                            regra_encontrada = r
                            break
                    if regra_encontrada:
                        valor_item = float(regra_encontrada.valor_acelerado if bateu_meta else regra_encontrada.valor_base)
                    comissao_vendedor += valor_item
                comissao_total_geral += comissao_vendedor

            total_registradas_geral += qtd_registradas
            total_instaladas_geral += qtd_instaladas

        percentual_aproveitamento = 0.0
        if total_registradas_geral > 0:
            percentual_aproveitamento = (total_instaladas_geral / total_registradas_geral) * 100

        valor_comissao_final = comissao_total_geral if exibir_comissao else 0.0

        projecao_vendas = calcular_projecao(total_registradas_geral, peso_realizado_venda, peso_total_mes_venda)
        projecao_instaladas = calcular_projecao(total_instaladas_geral, peso_realizado_inst, peso_total_mes_inst)
        projecao_comissao = calcular_projecao(valor_comissao_final, peso_realizado_inst, peso_total_mes_inst)

        periodo_str = f"{data_inicio.strftime('%d/%m')} a {data_fim_date.strftime('%d/%m')}"

        dados = {
            "periodo": periodo_str,
            "meta": meta_display,
            "total_vendas": total_registradas_geral,
            "total_instaladas": total_instaladas_geral,
            "comissao_estimada": valor_comissao_final,
            "projecao_vendas": round(projecao_vendas, 1),
            "projecao_instaladas": round(projecao_instaladas, 1),
            "projecao_comissao": round(projecao_comissao, 2),
            "percentual_meta": round((total_registradas_geral / meta_display * 100), 1) if meta_display > 0 else 0,
            "percentual_aproveitamento": round(percentual_aproveitamento, 1),
            "status_bateu_meta": (total_registradas_geral >= meta_display) if meta_display > 0 else False,
            "exibir_comissao": exibir_comissao, 
            "status_detalhado": status_counts_geral,
            "detalhes_listas": detalhes_listas
        }

        if is_diretoria:
            fat_projetado = calcular_projecao(faturamento_operadora_real, peso_realizado_inst, peso_total_mes_inst)
            dados['diretoria_data'] = {
                'faturamento_real': round(faturamento_operadora_real, 2),
                'faturamento_projetado': round(fat_projetado, 2),
                'mix_velocidade': mix_velocidade,
                'mix_pagamento': mix_pagamento
            }

        return Response(dados)

class ClienteViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [CheckAPIPermission]
    resource_name = 'cliente'

    def get_queryset(self):
        queryset = Cliente.objects.annotate(
            vendas_count=Count('vendas', filter=Q(vendas__ativo=True))
        ).order_by('nome_razao_social')
        
        search = self.request.query_params.get('search')
        if search:
            filters = Q(nome_razao_social__icontains=search) | Q(cpf_cnpj__icontains=search)
            
            clean_search = ''.join(filter(str.isdigit, search))
            if len(clean_search) == 11:
                cpf_fmt = f"{clean_search[:3]}.{clean_search[3:6]}.{clean_search[6:9]}-{clean_search[9:]}"
                filters |= Q(cpf_cnpj__icontains=cpf_fmt)
            elif len(clean_search) == 14:
                cnpj_fmt = f"{clean_search[:2]}.{clean_search[2:5]}.{clean_search[5:8]}/{clean_search[8:12]}-{clean_search[12:]}"
                filters |= Q(cpf_cnpj__icontains=cnpj_fmt)
            
            queryset = queryset.filter(filters)
            
        return queryset

class ListaVendedoresView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        User = get_user_model()
        vendedores = User.objects.filter(is_active=True).values('id', 'username', 'autorizar_venda_automatica').order_by('username')
        return Response(list(vendedores))

# Em site-record/crm_app/views.py

class FolhaComissionamentoView(APIView):
    """GET comissionamento/folha/?ano=AAAA&mes=M&vendedor_id=opcional - Folha no formato Excel."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .comissao_folha_service import calcular_folha_mes
        hoje = timezone.now()
        try:
            ano = int(request.query_params.get('ano', hoje.year))
            mes = int(request.query_params.get('mes', hoje.month))
        except ValueError:
            ano = hoje.year
            mes = hoje.month
        vendedor_id = request.query_params.get('vendedor_id')
        if vendedor_id is not None:
            try:
                vendedor_id = int(vendedor_id)
            except ValueError:
                vendedor_id = None
        grupos_gestao = ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']
        use_effective_date = not is_member(request.user, grupos_gestao)
        dados = calcular_folha_mes(ano, mes, vendedor_id, use_effective_date_for_display=use_effective_date)
        return Response(dados)


class ComissionamentoView(APIView):
    """GET comissionamento/?ano=AAAA&mes=M — Relatório de comissões do mês. Delega ao ComissionamentoService."""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        hoje = timezone.now()
        try:
            ano = int(request.query_params.get("ano", hoje.year))
            mes = int(request.query_params.get("mes", hoje.month))
        except (TypeError, ValueError):
            ano = hoje.year
            mes = hoje.month

        from crm_app.services.comissionamento_service import gerar_relatorio_comissionamento

        dados = gerar_relatorio_comissionamento(ano, mes)
        return Response(dados)


def _fechar_pagamento_mes(ano, mes, total_pago=None):
    from .comissao_folha_service import (
        annotate_data_folha_comissao,
        calcular_folha_mes,
        get_vendas_ids_desconto_churn_mes,
    )

    status_pago = StatusCRM.objects.filter(tipo='Comissionamento', nome__iexact='PAGO').first()
    if not status_pago:
        status_pago = StatusCRM.objects.create(tipo='Comissionamento', nome='PAGO', cor='#198754')

    data_inicio = datetime(ano, mes, 1).date()
    if mes == 12:
        data_fim = datetime(ano + 1, 1, 1).date()
    else:
        data_fim = datetime(ano, mes + 1, 1).date()

    vendas_para_atualizar = annotate_data_folha_comissao(Venda.objects.filter(
        ativo=True,
        status_esteira__nome__iexact='INSTALADA',
    )).filter(
        data_folha_comissao__isnull=False,
        data_folha_comissao__gte=data_inicio,
        data_folha_comissao__lt=data_fim,
    ).exclude(status_comissionamento=status_pago)

    count = vendas_para_atualizar.count()
    vendas_para_atualizar.update(
        status_comissionamento=status_pago,
        data_pagamento_comissao=timezone.now().date()
    )

    ids_churn = get_vendas_ids_desconto_churn_mes(ano, mes)
    if ids_churn:
        Venda.objects.filter(id__in=ids_churn).update(desconto_churn_aplicado_em=ano * 100 + mes)

    folha = calcular_folha_mes(ano, mes)
    recebido_total = sum(float(v.get('resumo', {}).get('liquido', 0) or 0) for v in folha.get('vendedores', []))
    valor_pago_efetivo = float(total_pago) if total_pago not in (None, "") else recebido_total
    if valor_pago_efetivo <= 0:
        valor_pago_efetivo = recebido_total

    pagamento, _ = PagamentoComissao.objects.update_or_create(
        referencia_ano=ano,
        referencia_mes=mes,
        defaults={
            'total_pago_consultores': valor_pago_efetivo,
            'total_recebido_ciclo': recebido_total,
            'data_fechamento': timezone.now()
        }
    )

    itens = []
    for vd in folha.get('vendedores', []):
        vendedor_id = vd.get('vendedor_id')
        if not vendedor_id:
            continue
        liquido = float(vd.get('resumo', {}).get('liquido', 0) or 0)
        itens.append(
            PagamentoComissaoItem(
                pagamento=pagamento,
                vendedor_id=vendedor_id,
                valor_pago=liquido,
                valor_recebido_ciclo=liquido,
                enviado_whatsapp_em=timezone.now(),
            )
        )
    PagamentoComissaoItem.objects.filter(pagamento=pagamento).delete()
    if itens:
        PagamentoComissaoItem.objects.bulk_create(itens, ignore_conflicts=True)

    return {"mensagem": f"Fechamento realizado! {count} vendas atualizadas.", "vendas_atualizadas": count}

class FecharPagamentoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            ano = int(request.data.get('ano'))
            mes = int(request.data.get('mes'))
            total_pago = request.data.get('total_pago', 0)
            return Response(_fechar_pagamento_mes(ano, mes, total_pago=total_pago))
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ReabrirPagamentoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            ano = int(request.data.get('ano'))
            mes = int(request.data.get('mes'))

            status_pendente = StatusCRM.objects.filter(tipo='Comissionamento', nome__iexact='PENDENTE').first()
            if not status_pendente:
                return Response({"error": "Status 'PENDENTE' de comissionamento não encontrado."}, status=400)

            data_inicio = datetime(ano, mes, 1)
            if mes == 12: data_fim = datetime(ano + 1, 1, 1)
            else: data_fim = datetime(ano, mes + 1, 1)

            vendas_revertidas = Venda.objects.filter(
                status_esteira__nome__iexact='INSTALADA',
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
                status_comissionamento__nome__iexact='PAGO'
            ).update(
                status_comissionamento=status_pendente,
                data_pagamento_comissao=None
            )

            PagamentoComissao.objects.filter(referencia_ano=ano, referencia_mes=mes).delete()

            return Response({"mensagem": "Mês reaberto com sucesso!", "vendas_revertidas": vendas_revertidas})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class HistoricoPagamentoDetalheView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            ano = int(request.query_params.get('ano'))
            mes = int(request.query_params.get('mes'))
        except (TypeError, ValueError):
            return Response({"error": "Parâmetros ano/mes inválidos."}, status=400)

        pagamento = PagamentoComissao.objects.filter(referencia_ano=ano, referencia_mes=mes).first()
        if not pagamento:
            return Response({"itens": [], "periodo": f"{mes:02d}/{ano}"})

        itens = (
            PagamentoComissaoItem.objects.filter(pagamento=pagamento)
            .select_related('vendedor')
            .order_by('vendedor__username')
        )
        data = []
        for it in itens:
            pago = float(it.valor_pago or 0)
            recebido = float(it.valor_recebido_ciclo or 0)
            data.append({
                "vendedor_id": it.vendedor_id,
                "vendedor_nome": it.vendedor.username if it.vendedor else f"ID {it.vendedor_id}",
                "total_pago": pago,
                "total_recebido": recebido,
                "diferenca": recebido - pago,
                "enviado_whatsapp_em": it.enviado_whatsapp_em.isoformat() if it.enviado_whatsapp_em else None,
            })
        return Response({"itens": data, "periodo": f"{mes:02d}/{ano}"})

class GerarRelatorioPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ano = int(request.data.get('ano'))
        mes = int(request.data.get('mes'))
        consultores_ids = request.data.get('consultores', [])

        data_inicio = datetime(ano, mes, 1)
        if mes == 12: data_fim = datetime(ano + 1, 1, 1)
        else: data_fim = datetime(ano, mes + 1, 1)

        vendas = Venda.objects.filter(
            ativo=True,
            status_esteira__nome__iexact='INSTALADA',
            data_instalacao__gte=data_inicio,
            data_instalacao__lt=data_fim
        ).select_related('vendedor', 'cliente', 'plano', 'forma_pagamento', 'status_esteira')

        if consultores_ids:
            vendas = vendas.filter(vendedor_id__in=consultores_ids)

        anomes_str = f"{ano}{mes:02d}"
        churns = ImportacaoChurn.objects.filter(anomes_retirada=anomes_str)
        churn_map = {c.numero_pedido: c for c in churns}

        lista_final = []
        for v in vendas:
            eh_dacc = "SIM" if v.forma_pagamento and "DÉBITO" in v.forma_pagamento.nome.upper() else "NÃO"
            
            chave_busca = v.ordem_servico or "SEM_OS"
            dados_churn = churn_map.get(chave_busca)
            
            status_final = "ATIVO"
            dt_churn = "-"
            motivo_churn = "-"
            
            if dados_churn:
                status_final = "CHURN"
                dt_churn = dados_churn.dt_retirada.strftime('%d/%m/%Y') if dados_churn.dt_retirada else "S/D"
                motivo_churn = dados_churn.motivo_retirada or "Não informado"

            lista_final.append({
                'vendedor': v.vendedor.username.upper() if v.vendedor else 'SEM VENDEDOR',
                'cpf_cnpj': v.cliente.cpf_cnpj,
                'dacc': eh_dacc,
                'cliente': v.cliente.nome_razao_social.upper(),
                'plano': v.plano.nome.upper(),
                'dt_criacao': v.data_criacao.strftime('%d/%m/%Y'),
                'dt_instalacao': v.data_instalacao.strftime('%d/%m/%Y') if v.data_instalacao else "-",
                'os': v.ordem_servico or "-",
                'status_esteira': v.status_esteira.nome.upper() if v.status_esteira else "-",
                'dt_churn': dt_churn,
                'motivo_churn': motivo_churn,
                'situacao': status_final,
                'style_class': 'color: red; font-weight: bold;' if status_final == 'CHURN' else ''
            })

        lista_final.sort(key=itemgetter('vendedor', 'cliente'))

        html_string = f"""
        <html>
        <head>
            <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
            <style>
                @page {{ size: A4 landscape; margin: 1cm; }}
                body {{ font-family: Helvetica, sans-serif; font-size: 9px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th {{ background-color: #f2f2f2; border: 1px solid #ccc; padding: 4px; text-align: left; font-weight: bold; }}
                td {{ border: 1px solid #ccc; padding: 3px; }}
                h2 {{ text-align: center; color: #333; margin-bottom: 5px; }}
                .meta {{ text-align: center; font-size: 10px; margin-bottom: 15px; color: #666; }}
            </style>
        </head>
        <body>
            <h2>Extrato de Comissionamento - {mes}/{ano}</h2>
            <div class="meta">Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Total Vendas: {len(lista_final)}</div>
            <table>
                <thead>
                    <tr>
                        <th>VENDEDOR</th>
                        <th>CPF/CNPJ</th>
                        <th>DACC</th>
                        <th>CLIENTE</th>
                        <th>PLANO</th>
                        <th>DT VENDA</th>
                        <th>DT INST</th>
                        <th>O.S.</th>
                        <th>STATUS</th>
                        <th>DT CHURN</th>
                        <th>MOTIVO</th>
                        <th>SITUAÇÃO</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for item in lista_final:
            html_string += f"""
            <tr style="{item['style_class']}">
                <td>{item['vendedor']}</td>
                <td>{item['cpf_cnpj']}</td>
                <td>{item['dacc']}</td>
                <td>{item['cliente'][:25]}</td>
                <td>{item['plano']}</td>
                <td>{item['dt_criacao']}</td>
                <td>{item['dt_instalacao']}</td>
                <td>{item['os']}</td>
                <td>{item['status_esteira']}</td>
                <td>{item['dt_churn']}</td>
                <td>{item['motivo_churn'][:15]}</td>
                <td>{item['situacao']}</td>
            </tr>
            """
        
        html_string += """
                </tbody>
            </table>
        </body>
        </html>
        """

        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result, encoding='utf-8')
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="extrato_comissao_{mes}_{ano}.pdf"'
            return response
        return Response({"error": "Erro ao gerar PDF"}, status=500)

class EnviarExtratoEmailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            # 1. Receber dados
            ano = int(request.data.get('ano'))
            mes = int(request.data.get('mes'))
            consultores_ids = request.data.get('consultores', []) 
            email_destino_manual = request.data.get('email_destino') 

            if not consultores_ids:
                return Response({"error": "Nenhum consultor selecionado."}, status=400)

            User = get_user_model()
            data_inicio = datetime(ano, mes, 1)
            if mes == 12: data_fim = datetime(ano + 1, 1, 1)
            else: data_fim = datetime(ano, mes + 1, 1)

            # Carregar dados de churn para o mês
            anomes_str = f"{ano}{mes:02d}"
            churns = ImportacaoChurn.objects.filter(anomes_retirada=anomes_str)
            churn_map = {c.numero_pedido: c for c in churns}

            sucessos = 0
            erros = []

            # 2. Loop para processar cada consultor selecionado
            for c_id in consultores_ids:
                try:
                    consultor = User.objects.get(id=c_id)
                    
                    # Regra de E-mail: 
                    target_email = None
                    if len(consultores_ids) == 1 and email_destino_manual:
                        target_email = email_destino_manual
                    else:
                        target_email = getattr(consultor, 'email', None)

                    if not target_email:
                        erros.append(f"{consultor.username}: Sem e-mail cadastrado.")
                        continue

                    # Buscar Vendas do Consultor
                    vendas = Venda.objects.filter(
                        ativo=True,
                        vendedor_id=c_id,
                        status_esteira__nome__iexact='INSTALADA',
                        data_instalacao__gte=data_inicio,
                        data_instalacao__lt=data_fim
                    ).select_related('plano', 'forma_pagamento', 'cliente')

                    if not vendas.exists():
                        erros.append(f"{consultor.username}: Sem vendas instaladas no período.")
                        continue

                    # --- Montar Dados ---
                    lista_detalhada = []
                    for v in vendas:
                        eh_dacc = "SIM" if v.forma_pagamento and "DÉBITO" in v.forma_pagamento.nome.upper() else "NÃO"
                        chave_busca = v.ordem_servico or "SEM_OS"
                        dados_churn = churn_map.get(chave_busca)
                        
                        status_final = "ATIVO"
                        dt_churn = "-"
                        motivo_churn = "-"
                        
                        if dados_churn:
                            status_final = "CHURN"
                            dt_churn = dados_churn.dt_retirada.strftime('%d/%m/%Y') if dados_churn.dt_retirada else "S/D"
                            motivo_churn = dados_churn.motivo_retirada or "Não informado"

                        dt_pedido = v.data_pedido.strftime('%d/%m/%Y') if v.data_pedido else v.data_criacao.strftime('%d/%m/%Y')
                        dt_inst = v.data_instalacao.strftime('%d/%m/%Y') if v.data_instalacao else "-"

                        lista_detalhada.append({
                            'vendedor': v.vendedor.username.upper() if v.vendedor else 'SEM VENDEDOR',
                            'cpf_cnpj': v.cliente.cpf_cnpj,
                            'dacc': eh_dacc,
                            'cliente': v.cliente.nome_razao_social.upper(),
                            'plano': v.plano.nome.upper(),
                            'dt_pedido': dt_pedido,
                            'dt_inst': dt_inst,
                            'os': v.ordem_servico or "-",
                            'situacao': v.status_esteira.nome.upper() if v.status_esteira else "-",
                            'dt_churn': dt_churn,
                            'motivo_churn': motivo_churn,
                            'churn_ativo': status_final,
                            'color_style': 'color: red;' if status_final == 'CHURN' else ''
                        })
                    
                    lista_detalhada.sort(key=itemgetter('cliente'))

                    # --- Gerar HTML E-mail ---
                    html_content = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; font-size: 12px;">
                        <h2 style="color: #333;">Extrato de Comissionamento - {consultor.username.upper()}</h2>
                        <p>Referência: <strong>{mes}/{ano}</strong></p>
                        <table border="1" cellpadding="4" cellspacing="0" style="border-collapse: collapse; width: 100%; font-size: 10px;">
                            <thead style="background-color: #f2f2f2;">
                                <tr>
                                    <th>VENDEDOR</th><th>CPF/CNPJ</th><th>DACC</th><th>NOME CLIENTE</th><th>PLANO</th>
                                    <th>DT PEDIDO</th><th>DT INST</th><th>O.S</th><th>SITUAÇÃO</th>
                                    <th>DT CHURN</th><th>MOTIVO CHURN</th><th>CHURN_OU_ATIVO</th>
                                </tr>
                            </thead>
                            <tbody>
                    """
                    for item in lista_detalhada:
                        html_content += f"""
                            <tr style="{item['color_style']}">
                                <td>{item['vendedor']}</td>
                                <td>{item['cpf_cnpj']}</td>
                                <td>{item['dacc']}</td>
                                <td>{item['cliente'][:20]}</td>
                                <td>{item['plano']}</td>
                                <td>{item['dt_pedido']}</td>
                                <td>{item['dt_inst']}</td>
                                <td>{item['os']}</td>
                                <td>{item['situacao']}</td>
                                <td>{item['dt_churn']}</td>
                                <td>{item['motivo_churn'][:20]}</td>
                                <td><strong>{item['churn_ativo']}</strong></td>
                            </tr>
                        """
                    html_content += "</tbody></table><br><p>Segue anexo detalhado.</p></body></html>"

                    # --- Gerar PDF (Anexo) ---
                    html_pdf = f"""
                    <html><head><meta charset="utf-8">
                    <style>@page {{ size: A4 landscape; margin: 1cm; }} body {{ font-family: Helvetica; font-size: 9px; }} table {{ width: 100%; border-collapse: collapse; }} th, td {{ border: 1px solid #ccc; padding: 3px; }} th {{ background: #f2f2f2; }}</style>
                    </head><body>
                        <h2>Extrato {mes}/{ano} - {consultor.username.upper()}</h2>
                        <table><thead><tr>
                            <th>VENDEDOR</th><th>CPF/CNPJ</th><th>DACC</th><th>CLIENTE</th><th>PLANO</th>
                            <th>DT PEDIDO</th><th>DT INST</th><th>O.S</th><th>SITUAÇÃO</th>
                            <th>DT CHURN</th><th>MOTIVO</th><th>STATUS</th>
                        </tr></thead><tbody>
                    """
                    for item in lista_detalhada:
                        html_pdf += f"""<tr style="{item.get('color_style', '')}">
                            <td>{item['vendedor']}</td><td>{item['cpf_cnpj']}</td><td>{item['dacc']}</td>
                            <td>{item['cliente'][:25]}</td><td>{item['plano']}</td><td>{item['dt_pedido']}</td>
                            <td>{item['dt_inst']}</td><td>{item['os']}</td><td>{item['situacao']}</td>
                            <td>{item['dt_churn']}</td><td>{item['motivo_churn'][:15]}</td><td>{item['churn_ativo']}</td>
                        </tr>"""
                    html_pdf += "</tbody></table></body></html>"

                    pdf_buffer = BytesIO()
                    pisa_status = pisa.pisaDocument(BytesIO(html_pdf.encode("UTF-8")), pdf_buffer, encoding='utf-8')
                    
                    if pisa_status.err:
                        erros.append(f"{consultor.username}: Erro ao gerar PDF.")
                        continue

                    # --- Enviar ---
                    msg = EmailMultiAlternatives(
                        f"Extrato Comissionamento - {mes}/{ano} - {consultor.username}",
                        strip_tags(html_content),
                        settings.DEFAULT_FROM_EMAIL,
                        [target_email]
                    )
                    msg.attach_alternative(html_content, "text/html")
                    msg.attach(f"Extrato_{consultor.username}_{mes}_{ano}.pdf", pdf_buffer.getvalue(), 'application/pdf')
                    msg.send()
                    
                    sucessos += 1

                except Exception as e:
                    erros.append(f"Erro interno id {c_id}: {str(e)}")

            # 3. Retorno Final
            msg_final = f"Processo finalizado. Enviados: {sucessos}."
            if erros:
                msg_final += f" Falhas: {len(erros)}. Detalhes: {', '.join(erros)}"
            
            return Response({"mensagem": msg_final, "sucessos": sucessos, "erros": erros})

        except Exception as e:
            logger.error(f"Erro geral envio email: {e}")
            return Response({"error": str(e)}, status=500)

# Adicione este import no topo do arquivo se não tiver
import threading

class ImportacaoOsabView(APIView):
    permission_classes = [CheckAPIPermission]
    # Deve coincidir com ImportacaoOsab._meta.model_name → permissão crm_app.add_importacaoosab
    resource_name = 'importacaoosab'
    parser_classes = [MultiPartParser, FormParser]

    def _clean_key(self, key):
        if pd.isna(key) or key is None: return None
        return str(key).replace('.0', '').strip()
    
    def _normalize_pedido(self, pedido):
        """Normaliza pedido: mantém valor exato da planilha, apenas remove espaços e .0 se for número float"""
        if pd.isna(pedido) or pedido is None: return None
        pedido_str = str(pedido).strip()
        # Se terminar com .0 (conversão de float), remove apenas o .0, mas mantém zeros à esquerda
        if pedido_str.endswith('.0'):
            pedido_str = pedido_str[:-2]
        return pedido_str

    def _normalize_text(self, text):
        if not text: return ""
        text = str(text).upper().strip()
        import unicodedata
        return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

    def _status_osab_permitido_com_bloqueio(self, status_osab_raw, target_status_esteira):
        """Retorna True quando status pode atualizar mesmo com bloqueio OSAB marcado."""
        status_norm = self._normalize_text(status_osab_raw)
        target_norm = self._normalize_text(target_status_esteira.nome) if target_status_esteira else ""

        permitidos = {
            "INSTALADA",
            "INSTALADO",
            "CANCELADA",
            "CANCELADO",
            "INSTALADA OUTRO PDV",
            "NAO CONSTA NA OSAB",
        }
        return status_norm in permitidos or target_norm in permitidos

    def _enviar_mensagens_background(self, lista_mensagens):
        svc = WhatsAppService()
        for telefone, texto in lista_mensagens:
            try:
                svc.enviar_mensagem_texto(telefone, texto)
            except Exception as e:
                print(f"Erro envio background thread: {e}")

    def _sincronizar_seq_historico(self):
        """Garante que a sequence do histórico não esteja atrasada (evita PK duplicada)."""
        try:
            from django.db import connection
            from django.db.models import Max

            max_id = HistoricoAlteracaoVenda.objects.aggregate(max_id=Max('id')).get('max_id') or 0
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT setval(pg_get_serial_sequence(%s, 'id'), %s, true);",
                    ['crm_historico_alteracao_venda', max_id]
                )
        except Exception as e:
            print(f"Aviso: não foi possível sincronizar sequence do histórico: {e}")

    def _sincronizar_seq_osab(self):
        """Garante que a sequence da importação OSAB não esteja atrasada (evita PK duplicada)."""
        try:
            from django.db import connection
            from django.db.models import Max

            max_id = ImportacaoOsab.objects.aggregate(max_id=Max('id')).get('max_id') or 0
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT setval(pg_get_serial_sequence(%s, 'id'), %s, true);",
                    ['crm_importacao_osab', max_id]
                )
        except Exception as e:
            print(f"Aviso: não foi possível sincronizar sequence da OSAB: {e}")

    def _marcar_vendas_ausentes_na_osab(self, log_id, total_registros_import, osab_bot):
        """
        Após gravar a importação: vendas ativas com O.S. que não existem em ImportacaoOsab.
        - Só considera pedidos com data de abertura (ou criação, se sem abertura) **antes do dia atual**
          (base OSAB costuma refletir até o dia anterior).
        - Esteira contendo INSTALADA e não sendo já 'INSTALADA OUTRO PDV' -> INSTALADA OUTRO PDV.
        - Demais -> NÃO CONSTA NA OSAB.
        Não envia WhatsApp. Não executa se a planilha tiver 0 linhas (evita marcar o CRM inteiro).
        """
        from django.db import connection, transaction
        from django.db.models import Q
        from django.utils import timezone
        from crm_app.models import LogImportacaoOSAB, ImportacaoOsab, Venda, StatusCRM, HistoricoAlteracaoVenda

        out = {
            'crm_sem_osab_nao_consta': 0,
            'crm_sem_osab_outro_pdv': 0,
            'crm_sem_osab_ignoradas_mesmo_dia': 0,
        }
        if not total_registros_import:
            return out

        LogImportacaoOSAB.objects.filter(id=log_id).update(
            mensagem='Verificando vendas no CRM ausentes da base OSAB...'
        )

        osab_set = set()
        for doc in ImportacaoOsab.objects.values_list('documento', flat=True).iterator(chunk_size=8000):
            if doc is None:
                continue
            s = str(doc).strip()
            if not s:
                continue
            osab_set.add(s)
            n = self._normalize_pedido(doc)
            if n:
                osab_set.add(n)

        if not osab_set:
            return out

        st_nao = StatusCRM.objects.filter(tipo='Esteira', nome__iexact='NÃO CONSTA NA OSAB').first()
        if not st_nao:
            st_nao = StatusCRM.objects.filter(tipo='Esteira', nome__iexact='NAO CONSTA NA OSAB').first()
        st_outro = StatusCRM.objects.filter(tipo='Esteira', nome__iexact='INSTALADA OUTRO PDV').first()
        if not st_nao or not st_outro:
            print(
                "Aviso OSAB: cadastre os status Esteira 'NÃO CONSTA NA OSAB' e 'INSTALADA OUTRO PDV' no CRM."
            )
            return out

        hoje = timezone.localdate()
        vendas_atualizar = []
        historicos_criar = []

        qs = Venda.objects.filter(ativo=True).exclude(
            Q(ordem_servico__isnull=True) | Q(ordem_servico='')
        ).select_related('status_esteira')

        for venda in qs.iterator(chunk_size=2000):
            os_str = (venda.ordem_servico or '').strip()
            if not os_str:
                continue
            norm = self._normalize_pedido(os_str)
            if os_str in osab_set or (norm and norm in osab_set):
                continue

            if venda.data_abertura:
                da = venda.data_abertura
                ref_d = timezone.localtime(da).date() if timezone.is_aware(da) else da.date()
            else:
                dc = venda.data_criacao
                ref_d = timezone.localtime(dc).date() if timezone.is_aware(dc) else dc.date()
            if ref_d >= hoje:
                out['crm_sem_osab_ignoradas_mesmo_dia'] += 1
                continue

            se = venda.status_esteira
            if se:
                if se.id == st_nao.id or se.id == st_outro.id:
                    continue

            nome_se = (se.nome if se else '') or ''
            nome_u = nome_se.upper()
            if 'INSTALADA' in nome_u and 'OUTRO PDV' not in nome_u:
                target = st_outro
            else:
                target = st_nao

            if se and se.id == target.id:
                continue

            old_nome = se.nome if se else '(vazio)'
            venda.status_esteira = target
            vendas_atualizar.append(venda)
            historicos_criar.append(
                HistoricoAlteracaoVenda(
                    venda=venda,
                    usuario=osab_bot,
                    alteracoes={
                        'status_esteira': (
                            f"De '{old_nome}' para '{target.nome}' "
                            f"(O.S. não consta na base OSAB após importação)"
                        )
                    },
                )
            )
            if target.id == st_outro.id:
                out['crm_sem_osab_outro_pdv'] += 1
            else:
                out['crm_sem_osab_nao_consta'] += 1

        if not vendas_atualizar:
            return out

        try:
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                Venda.objects.bulk_update(vendas_atualizar, ['status_esteira'], batch_size=2000)
            if historicos_criar:
                with transaction.atomic():
                    with connection.cursor() as cursor:
                        cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                    self._sincronizar_seq_historico()
                    HistoricoAlteracaoVenda.objects.bulk_create(historicos_criar, batch_size=2000)
        except Exception as e:
            print(f"Erro ao marcar vendas sem OSAB: {e}")
            raise

        return out

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj: return Response({'error': 'Nenhum arquivo enviado.'}, status=400)
        
        opcao_front = str(request.data.get('enviar_whatsapp', 'true')).lower() == 'true'
        usuario = request.user
        pode_decidir = is_member(usuario, ['Diretoria', 'Admin'])
        flag_enviar_whatsapp = opcao_front if pode_decidir else True
        
        if not flag_enviar_whatsapp:
            print(f"--- Importação OSAB SILENCIOSA iniciada por: {usuario.username} ---")

        # Criar log antes de processar
        from crm_app.models import LogImportacaoOSAB
        log = LogImportacaoOSAB.objects.create(
            usuario=usuario,
            nome_arquivo=file_obj.name,
            status='PROCESSANDO',
            tamanho_arquivo=file_obj.size,
            enviar_whatsapp=flag_enviar_whatsapp
        )

        # Ler arquivo para memória ANTES de spawnar thread
        try:
            file_content = file_obj.read()
            file_name = file_obj.name
        except Exception as e:
            log.status = 'ERRO'
            log.mensagem_erro = f'Erro ao ler arquivo: {str(e)}'
            log.finalizado_em = timezone.now()
            log.calcular_duracao()
            log.save()
            return Response({'error': f'Erro leitura arquivo: {str(e)}'}, status=400)

        # Processar em background
        import threading
        thread = threading.Thread(
            target=self._processar_osab_interno,
            args=(log.id, file_content, file_name, flag_enviar_whatsapp)
        )
        thread.daemon = True
        thread.start()

        return Response({
            'success': True,
            'message': 'Importação OSAB iniciada! Processamento em andamento...',
            'log_id': log.id,
            'detalhes': 'O processamento está sendo executado em background. Atualize a página em alguns minutos para ver o resultado.',
        }, status=200)

    def _serialize_date_for_json(self, value):
        """Converte objetos date/datetime para string ISO para serialização JSON"""
        from datetime import date, datetime
        if value is None:
            return None
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value

    def _processar_osab_interno(self, log_id, file_content, file_name, flag_enviar_whatsapp):
        """Processamento OSAB em background thread"""
        from io import BytesIO
        from django.utils import timezone
        from crm_app.models import LogImportacaoOSAB
        
        try:
            log = LogImportacaoOSAB.objects.get(id=log_id)
            LogImportacaoOSAB.objects.filter(id=log_id).update(
                mensagem='Lendo arquivo OSAB...'
            )
            
            # Ler DataFrame do conteúdo
            try:
                file_buffer = BytesIO(file_content)
                # Usar openpyxl para forçar leitura de PEDIDO como texto (preserva zeros mesmo se salvo como número)
                if file_name.endswith('.xlsb'): 
                    # Para .xlsb, tentar usar dtype/converters para forçar PEDIDO como string
                    # Primeiro ler uma amostra para descobrir nome da coluna
                    file_buffer.seek(0)
                    df_sample = pd.read_excel(file_buffer, engine='pyxlsb', nrows=1)
                    file_buffer.seek(0)
                    # Encontrar coluna que parece ser PEDIDO (case-insensitive, antes da normalização)
                    pedido_col = None
                    for col in df_sample.columns:
                        col_normalizado = str(col).strip().upper().replace(' ', '_')
                        if col_normalizado == 'PEDIDO':
                            pedido_col = col  # Manter nome original da coluna
                            break
                    # Tentar usar dtype primeiro (se suportado), senão converters
                    if pedido_col:
                        try:
                            # Tentar dtype primeiro (pyxlsb pode suportar em algumas versões)
                            df = pd.read_excel(file_buffer, engine='pyxlsb', dtype={pedido_col: str})
                        except (TypeError, ValueError):
                            # Se dtype não funcionar, usar converters
                            try:
                                df = pd.read_excel(file_buffer, engine='pyxlsb', converters={pedido_col: lambda x: str(x) if pd.notna(x) else ''})
                            except:
                                # Se converters também falhar, ler normalmente
                                df = pd.read_excel(file_buffer, engine='pyxlsb')
                    else:
                        df = pd.read_excel(file_buffer, engine='pyxlsb')
                elif file_name.endswith('.xlsx'): 
                    # Para .xlsx, usar openpyxl para forçar PEDIDO como texto
                    try:
                        from openpyxl import load_workbook
                        file_buffer.seek(0)
                        wb = load_workbook(file_buffer, data_only=False, read_only=True)
                        ws = wb.active
                        
                        # Ler cabeçalhos
                        headers = [cell.value for cell in ws[1]]
                        # Encontrar índice da coluna PEDIDO
                        pedido_idx = None
                        for idx, header in enumerate(headers):
                            if header and str(header).strip().upper().replace(' ', '_') == 'PEDIDO':
                                pedido_idx = idx
                                break
                        
                        # Ler dados: para PEDIDO, forçar como string (preserva zeros)
                        data = []
                        for row in ws.iter_rows(min_row=2, values_only=False):
                            row_data = []
                            for idx, cell in enumerate(row):
                                if idx == pedido_idx and cell.value is not None:
                                    # Para PEDIDO: sempre converter para string (preserva zeros à esquerda)
                                    # Se célula tem formato texto (@) ou se é string, usar direto
                                    if cell.data_type == 's':
                                        row_data.append(str(cell.value))
                                    else:
                                        # Se for número, converter para string formatando sem perder zeros
                                        # Usar internal_value se disponível, senão value
                                        val = cell.value
                                        # Formatar como string inteiro (sem decimais)
                                        if isinstance(val, (int, float)):
                                            # Para números, converter para int primeiro para evitar .0
                                            if isinstance(val, float) and val.is_integer():
                                                val = int(val)
                                            row_data.append(str(val))
                                        else:
                                            row_data.append(str(val) if val is not None else '')
                                else:
                                    row_data.append(cell.value)
                            data.append(row_data)
                        
                        wb.close()
                        df = pd.DataFrame(data, columns=headers)
                    except Exception as e:
                        # Fallback para pandas normal se openpyxl falhar
                        file_buffer.seek(0)
                        df_sample = pd.read_excel(file_buffer, nrows=1)
                        file_buffer.seek(0)
                        pedido_col = None
                        for col in df_sample.columns:
                            if str(col).strip().upper().replace(' ', '_') == 'PEDIDO':
                                pedido_col = col
                                break
                        if pedido_col:
                            df = pd.read_excel(file_buffer, converters={pedido_col: lambda x: str(x) if pd.notna(x) else ''})
                        else:
                            df = pd.read_excel(file_buffer)
                elif file_name.endswith('.xls'): 
                    # Para .xls antigo, usar pandas normal com converters
                    file_buffer.seek(0)
                    df_sample = pd.read_excel(file_buffer, nrows=1)
                    file_buffer.seek(0)
                    pedido_col = None
                    for col in df_sample.columns:
                        if str(col).strip().upper().replace(' ', '_') == 'PEDIDO':
                            pedido_col = col
                            break
                    if pedido_col:
                        df = pd.read_excel(file_buffer, converters={pedido_col: lambda x: str(x) if pd.notna(x) else ''})
                    else:
                        df = pd.read_excel(file_buffer)
                else: 
                    raise ValueError('Formato inválido')
            except Exception as e:
                log.status = 'ERRO'
                log.mensagem_erro = f'Erro leitura arquivo: {str(e)}'
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
                return

                # 1. Normalização dos nomes das colunas
            df.columns = [str(col).strip().upper().replace(' ', '_') for col in df.columns]
            
            # 1.1 Validação de tipos de colunas esperadas
            colunas_esperadas_tipo = {
                'PRODUTO': 'TEXTO',
                'UF': 'TEXTO',
                'DT_REF': 'DATA',
                'PEDIDO': 'TEXTO',
                'SEGMENTO': 'TEXTO',
                'LOCALIDADE': 'TEXTO',
                'CELULA': 'TEXTO',
                'ID_BUNDLE': 'TEXTO',
                'TELEFONE': 'NÚMERO',
                'VELOCIDADE': 'TEXTO',
                'MATRICULA_VENDEDOR': 'TEXTO',
                'CLASSE_PRODUTO': 'TEXTO',
                'NOME_CNAL': 'TEXTO',
                'PDV_SAP': 'NÚMERO',
                'DESCRICAO': 'TEXTO',
                'DATA_ABERTURA': 'DATA E HORA',
                'DATA_FECHAMENTO': 'DATA E HORA',
                'SITUACAO': 'TEXTO',
                'CLASSIFICACAO': 'TEXTO',
                'DATA_AGENDAMENTO': 'DATA E HORA',
                'COD_PENDENCIA': 'NÚMERO',
                'DESC_PENDENCIA': 'TEXTO',
                'NUMERO_BA': 'TEXTO',
                'FG_VENDA_VALIDA': 'NÚMERO',
                'DESC_MOTIVO_ORDEM': 'TEXTO',
                'DESC_SUB_MOTIVO_ORDEM': 'TEXTO',
                'MEIO_PAGAMENTO': 'TEXTO',
                'CAMPANHA': 'TEXTO',
                'FLG_MEI': 'TEXTO',
                'NM_DIRETORIA': 'TEXTO',
                'NM_REGIONAL': 'TEXTO',
                'CD_REDE': 'NÚMERO',
                'GP_CANAL': 'TEXTO',
                'NM_PDV_REL': 'TEXTO',
                'GERENCIA': 'TEXTO',
                'NM_GC': 'TEXTO',
                'NR_ORDEM_ORIGINAL': 'TEXTO',
                'MOTIVO_CANCELAMENTO': 'TEXTO',
                'SUBMOTIVO_CANCELAMENTO': 'TEXTO',
            }
            # Log de colunas faltantes (apenas informativo, não bloqueia importação)
            colunas_faltantes = set(colunas_esperadas_tipo.keys()) - set(df.columns)
            if colunas_faltantes:
                log.mensagem_erro = f'Colunas esperadas não encontradas: {", ".join(sorted(colunas_faltantes))}. A importação continuará com as colunas disponíveis.'
                log.save()
            
            # Garantir que PEDIDO seja tratado como string para preservar zeros à esquerda
            if 'PEDIDO' in df.columns:
                # Converter para string (já foi lido como texto via openpyxl se .xlsx)
                df['PEDIDO'] = df['PEDIDO'].astype(str)
                # Remover 'nan' string (valores nulos do pandas convertidos para string)
                df['PEDIDO'] = df['PEDIDO'].replace('nan', '')
            
            # ==============================================================================
            # 2. PARSER DE DATA "INTELIGENTE" (Versão Corrigida para Conflito de Imports)
            # ==============================================================================
            def smart_date_parser(val):
                # Importação local com alias para evitar conflito com 'from datetime import datetime'
                import datetime as dt_sys 
                import pandas as pd # Garantir pandas aqui também

                # Se for nulo ou vazio
                if val is None or pd.isna(val) or val == '':
                    return None
                
                # Caso A: O Pandas/Engine já leu como objeto de data (datetime)
                # Verifica se é instância de datetime.datetime, datetime.date ou pd.Timestamp
                if isinstance(val, (dt_sys.datetime, dt_sys.date, pd.Timestamp)):
                    return val.date() if hasattr(val, 'date') else val

                # Caso B: É um número (Serial do Excel - ex: 45279.54)
                if isinstance(val, (float, int)):
                    try:
                        # Excel começa em 30/12/1899
                        return (dt_sys.datetime(1899, 12, 30) + dt_sys.timedelta(days=float(val))).date()
                    except:
                        return None

                # Caso C: É Texto (String) - Onde mora o perigo "04/12/2025  13:00:00"
                s_val = str(val).strip()
                
                # Tenta encontrar padrão DD/MM/AAAA via Regex (ignora o resto da string)
                # Usa 're' que deve estar importado no topo ou aqui
                import re 
                match_br = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s_val)
                if match_br:
                    d, m, y = match_br.groups()
                    try:
                        return dt_sys.date(int(y), int(m), int(d))
                    except ValueError:
                        pass # Data inválida matematicamente (ex: 30/02)

                # Tenta encontrar padrão AAAA-MM-DD
                match_iso = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s_val)
                if match_iso:
                    y, m, d = match_iso.groups()
                    try:
                        return dt_sys.date(int(y), int(m), int(d))
                    except ValueError:
                        pass

                return None
            # ==============================================================================

            # Aplica a função linha a linha nas colunas de data
            cols_data = ['DT_REF', 'DATA_ABERTURA', 'DATA_FECHAMENTO', 'DATA_AGENDAMENTO']
            for col in cols_data:
                if col in df.columns:
                    # O .apply é mais lento que vetorização, mas muito mais seguro para dados sujos
                    df[col] = df[col].apply(smart_date_parser)
            
            df = df.replace({np.nan: None, pd.NaT: None})

            total_registros = len(df)
            LogImportacaoOSAB.objects.filter(id=log_id).update(
                total_registros=total_registros,
                total_processadas=0,
                mensagem=f'Preparando dados... 0/{total_registros}'
            )

            # --- PREPARAÇÃO DO BANCO DE DADOS ---
            status_esteira_map = {s.nome.upper(): s for s in StatusCRM.objects.filter(tipo='Esteira')}
            status_tratamento_map = {s.nome.upper(): s for s in StatusCRM.objects.filter(tipo='Tratamento')}
            
            mapa_pagamentos = {}
            for fp in FormaPagamento.objects.filter(ativo=True):
                mapa_pagamentos[self._normalize_text(fp.nome)] = fp
            
            motivo_pendencia_map = {}
            for m in MotivoPendencia.objects.all():
                match = re.match(r'^(\d+)', m.nome.strip())
                if match: motivo_pendencia_map[match.group(1)] = m
            
            motivo_padrao_osab, _ = MotivoPendencia.objects.get_or_create(nome="VALIDAR OSAB", defaults={'tipo_pendencia': 'Operacional'})
            motivo_sem_agenda, _ = MotivoPendencia.objects.get_or_create(nome="APROVISIONAMENTO S/ DATA", defaults={'tipo_pendencia': 'Sistêmica'})

            # Obter pedidos mantendo valor exato da planilha (já convertido para string na linha 2408)
            lista_pedidos_raw = df['PEDIDO'].dropna().tolist() if 'PEDIDO' in df.columns else []
            # Normalizar pedidos mantendo valor exato (apenas remover .0 se for float convertido)
            lista_pedidos_limpos = set()
            for p in lista_pedidos_raw:
                p_str = str(p).strip()
                if p_str and p_str != 'nan':
                    # Se terminar com .0 (conversão de float para string), remove apenas o .0
                    if p_str.endswith('.0'):
                        p_str = p_str[:-2]
                    if p_str:  # Garantir que não está vazio após processamento
                        lista_pedidos_limpos.add(p_str)

            vendas_filtradas = Venda.objects.filter(
                ativo=True, 
                ordem_servico__in=lista_pedidos_limpos
            ).select_related('vendedor', 'status_esteira', 'status_tratamento')
            
            # Criar mapa usando pedido normalizado (mantém zeros à esquerda)
            vendas_map = {}
            for v in vendas_filtradas:
                os_normalizado = self._normalize_pedido(v.ordem_servico)
                if os_normalizado:
                    vendas_map[os_normalizado] = v
                # Também adicionar com a chave original para compatibilidade
                if v.ordem_servico:
                    vendas_map[v.ordem_servico] = v
            osab_bot = get_osab_bot_user()

            # Buscar OSAB existentes usando documentos normalizados
            osab_existentes = {}
            for obj in ImportacaoOsab.objects.filter(documento__in=lista_pedidos_limpos):
                doc_normalizado = self._normalize_pedido(obj.documento)
                if doc_normalizado:
                    osab_existentes[doc_normalizado] = obj
                # Também adicionar com documento original para compatibilidade
                if obj.documento:
                    osab_existentes[obj.documento] = obj

            osab_criar, osab_atualizar, vendas_atualizar, historicos_criar = [], [], [], []
            fila_mensagens_whatsapp = [] 

            coluna_map = {
                'PRODUTO': 'produto', 'UF': 'uf', 'DT_REF': 'dt_ref', 'PEDIDO': 'documento',
                'SEGMENTO': 'segmento', 'LOCALIDADE': 'localidade', 'CELULA': 'celula',
                'ID_BUNDLE': 'id_bundle', 'TELEFONE': 'telefone', 'VELOCIDADE': 'velocidade',
                'MATRICULA_VENDEDOR': 'matricula_vendedor', 'CLASSE_PRODUTO': 'classe_produto',
                'NOME_CNAL': 'nome_canal', 'PDV_SAP': 'pdv_sap', 'DESCRICAO': 'descricao',
                'DATA_ABERTURA': 'data_abertura', 'DATA_FECHAMENTO': 'data_fechamento',
                'SITUACAO': 'situacao', 'CLASSIFICACAO': 'classificacao',
                'DATA_AGENDAMENTO': 'data_agendamento', 'COD_PENDENCIA': 'cod_pendencia',
                'DESC_PENDENCIA': 'desc_pendencia', 'NUMERO_BA': 'numero_ba',
                'FG_VENDA_VALIDA': 'fg_venda_valida', 'DESC_MOTIVO_ORDEM': 'desc_motivo_ordem',
                'DESC_SUB_MOTIVO_ORDEM': 'desc_sub_motivo_ordem', 'MEIO_PAGAMENTO': 'meio_pagamento',
                'CAMPANHA': 'campanha', 'FLG_MEI': 'flg_mei', 'NM_DIRETORIA': 'nm_diretoria',
                'NM_REGIONAL': 'nm_regional', 'CD_REDE': 'cd_rede', 'GP_CANAL': 'gp_canal',
                'GERENCIA': 'gerencia', 'NM_GC': 'gc',
            }

            STATUS_MAP = {
                'CONCLUÍDO': 'INSTALADA', 'CONCLUIDO': 'INSTALADA', 'EXECUTADO': 'INSTALADA',
                'PENDÊNCIA CLIENTE': 'PENDENCIADA', 'PENDENCIA CLIENTE': 'PENDENCIADA',
                'PENDÊNCIA TÉCNICA': 'PENDENCIADA', 'PENDENCIA TECNICA': 'PENDENCIADA',
                'CANCELADO': 'CANCELADA', 'EM CANCELAMENTO': 'CANCELADA',
                'AGENDADO': 'AGENDADO', 'DRAFT': 'DRAFT', 
                'AGUARDANDO PAGAMENTO': 'AGUARDANDO PAGAMENTO'
            }
            
            report = {
                "status": "sucesso", "total_registros": len(df), "criados": 0, "atualizados": 0, 
                "vendas_encontradas": 0, "ja_corretos": 0, "erros": [], "logs_detalhados": [],
                "ignorados_dt_ref": 0, "bloqueados_flag_osab": 0, "arquivo_excel_b64": None
            }

            def _normalize_dt_ref(val):
                import datetime as dt_sys
                if val is None:
                    return None
                if isinstance(val, dt_sys.datetime):
                    return val.date()
                if isinstance(val, dt_sys.date):
                    return val
                return None

            # --- LOOP PRINCIPAL ---
            records = df.to_dict('records')
            progress_step = 5000

            for index, row in enumerate(records):
                log_item = {
                    "linha": index + 2,
                    "pedido": str(row.get('PEDIDO')),
                    "status_osab": str(row.get('SITUACAO')),
                    "dt_ref_planilha": self._serialize_date_for_json(row.get('DT_REF')),
                    "dt_ref_crm": None,
                    "consta_osab": "NAO",
                    "consta_crm": "NAO",
                    "resultado_osab": "",
                    "resultado_crm": "",
                    "detalhe": ""
                }
                try:
                    # A. ImportacaoOsab
                    dados_model = {}
                    for col_planilha, campo_model in coluna_map.items():
                        val = row.get(col_planilha)
                        if col_planilha == 'PEDIDO': 
                            # Manter valor exato da planilha (já está como string preservando zeros)
                            val = self._normalize_pedido(val)  # Apenas remove .0 se for float convertido
                        dados_model[campo_model] = val
                    
                    doc_chave = dados_model.get('documento')
                    if not doc_chave: 
                        log_item["resultado_osab"] = "IGNORADO"
                        log_item["resultado_crm"] = "IGNORADO"
                        report["logs_detalhados"].append(log_item)
                        continue

                    if doc_chave in osab_existentes:
                        obj = osab_existentes[doc_chave]
                        dt_ref_nova = _normalize_dt_ref(dados_model.get('dt_ref'))
                        dt_ref_atual = _normalize_dt_ref(getattr(obj, 'dt_ref', None))
                        log_item["consta_osab"] = "SIM"
                        log_item["dt_ref_crm"] = self._serialize_date_for_json(dt_ref_atual)
                        # Se a DT_REF nova for mais antiga (menor que), não atualiza nem altera a venda
                        # Se for igual ou maior, atualiza (permite atualização quando é igual)
                        if dt_ref_atual and (dt_ref_nova is None or dt_ref_nova < dt_ref_atual):
                            log_item["resultado_osab"] = "IGNORADO_DT_REF_ANTIGA"
                            log_item["resultado_crm"] = "IGNORADO_DT_REF_ANTIGA"
                            log_item["detalhe"] = f"DT_REF planilha ({dt_ref_nova}) < DT_REF CRM ({dt_ref_atual})"
                            report["ignorados_dt_ref"] += 1
                            report["logs_detalhados"].append(log_item)
                            continue

                        mudou = False
                        for k, v in dados_model.items():
                            if getattr(obj, k) != v:
                                setattr(obj, k, v)
                                mudou = True
                        if mudou:
                            osab_atualizar.append(obj)
                            log_item["resultado_osab"] = "ATUALIZADO_OSAB"
                        else:
                            log_item["resultado_osab"] = "SEM_MUDANCA_OSAB"
                    else:
                        osab_criar.append(ImportacaoOsab(**dados_model))
                        log_item["resultado_osab"] = "CRIADO_OSAB"

                    # B. Venda CRM
                    venda = vendas_map.get(doc_chave)
                    if not venda:
                        log_item["resultado_crm"] = "NAO_ENCONTRADO_CRM"
                        log_item["consta_crm"] = "NAO"
                        report["logs_detalhados"].append(log_item)
                        continue
                    
                    log_item["consta_crm"] = "SIM"
                    report["vendas_encontradas"] += 1
                    sit_osab_raw = str(row.get('SITUACAO', '')).strip().upper()
                    if sit_osab_raw in ["NONE", "NAN"]: sit_osab_raw = ""

                    target_status_esteira = None
                    target_status_tratamento = None
                    target_data_agenda = None
                    target_motivo_pendencia = None
                    
                    houve_alteracao = False
                    detalhes_hist = {}
                    msg_whatsapp_desta_venda = None

                    is_fraude = "PAYMENT_NOT_AUTHORIZED" in sit_osab_raw
                    if not sit_osab_raw or is_fraude:
                        pgto_raw = self._normalize_text(row.get('MEIO_PAGAMENTO'))
                        if "CARTAO" in pgto_raw:
                            st_reprovado = status_tratamento_map.get("REPROVADO CARTÃO DE CRÉDITO")
                            if st_reprovado: target_status_tratamento = st_reprovado
                    
                    elif sit_osab_raw == "EM APROVISIONAMENTO":
                        # Aqui usamos a data já limpa pelo parser inteligente
                        dt_ag = row.get('DATA_AGENDAMENTO') 
                        
                        # Validação simples: se existe, o parser já garantiu que é uma data válida
                        if dt_ag and dt_ag.year >= 2000:
                            target_status_esteira = status_esteira_map.get("AGENDADO")
                            target_data_agenda = dt_ag
                        else:
                            target_status_esteira = status_esteira_map.get("PENDENCIADA")
                            target_motivo_pendencia = motivo_sem_agenda
                    
                    else:
                        nome_est = STATUS_MAP.get(sit_osab_raw)
                        if not nome_est:
                            if sit_osab_raw.startswith("DRAFT"): nome_est = "DRAFT"
                            elif "AGUARDANDO PAGAMENTO" in sit_osab_raw: nome_est = "AGUARDANDO PAGAMENTO"
                            elif "REPROVADO" in sit_osab_raw: nome_est = "REPROVADO CARTÃO DE CRÉDITO"
                        if nome_est: target_status_esteira = status_esteira_map.get(nome_est)

                    if venda.bloquear_atualizacao_status_osab and not self._status_osab_permitido_com_bloqueio(
                        sit_osab_raw,
                        target_status_esteira,
                    ):
                        log_item["resultado_crm"] = "BLOQUEADO_FLAG_OSAB"
                        log_item["detalhe"] = (
                            "Atualizacao OSAB bloqueada por configuracao do pedido "
                            "(bloquear_atualizacao_status_osab=true)."
                        )
                        report["bloqueados_flag_osab"] = report.get("bloqueados_flag_osab", 0) + 1
                        historicos_criar.append(
                            HistoricoAlteracaoVenda(
                                venda=venda,
                                usuario=osab_bot,
                                alteracoes={
                                    "osab_bloqueado": (
                                        f"Tentativa bloqueada para status OSAB '{sit_osab_raw or '(vazio)'}'. "
                                        "Somente INSTALADA, CANCELADA, INSTALADA OUTRO PDV e NAO CONSTA NA OSAB "
                                        "podem atualizar quando o bloqueio estiver marcado."
                                    )
                                },
                            )
                        )
                        logger.info(
                            "[OSAB] Atualizacao bloqueada para venda=%s os=%s status_osab='%s'",
                            venda.id,
                            venda.ordem_servico,
                            sit_osab_raw or "",
                        )
                        report["logs_detalhados"].append(log_item)
                        continue

                    # --- 1. DATA DE ABERTURA ---
                    # Como usamos o parser inteligente, aqui já é date ou None
                    nova_data_abertura = row.get('DATA_ABERTURA')
                    if nova_data_abertura:
                        data_sistema = venda.data_abertura.date() if venda.data_abertura else None
                        if data_sistema != nova_data_abertura:
                            detalhes_hist['data_abertura'] = f"De '{data_sistema}' para '{nova_data_abertura}'"
                            venda.data_abertura = nova_data_abertura
                            houve_alteracao = True

                    # Aplica Alterações Status Tratamento
                    if target_status_tratamento and venda.status_tratamento != target_status_tratamento:
                        detalhes_hist['status_tratamento'] = f"De '{venda.status_tratamento}' para '{target_status_tratamento.nome}'"
                        venda.status_tratamento = target_status_tratamento
                        houve_alteracao = True

                    # Aplica Alterações Status Esteira
                    if target_status_esteira:
                        if venda.status_esteira != target_status_esteira:
                            detalhes_hist['status_esteira'] = f"De '{venda.status_esteira}' para '{target_status_esteira.nome}'"
                            venda.status_esteira = target_status_esteira
                            houve_alteracao = True
                            if 'PENDEN' not in target_status_esteira.nome.upper(): venda.motivo_pendencia = None
                        
                        nome_est_upper = target_status_esteira.nome.upper()

                        if 'INSTALADA' in nome_est_upper:
                            nova_dt = row.get('DATA_FECHAMENTO')
                            if nova_dt and nova_dt.year >= 2000:
                                data_inst_atual = venda.data_instalacao
                                if not data_inst_atual or data_inst_atual != nova_dt:
                                    detalhes_hist['data_instalacao'] = f"Nova: {nova_dt}"
                                    venda.data_instalacao = nova_dt
                                    houve_alteracao = True
                            
                            if houve_alteracao and venda.vendedor and venda.vendedor.tel_whatsapp:
                                dt_fmt = venda.data_instalacao.strftime('%d/%m') if venda.data_instalacao else "Hoje"
                                msg_whatsapp_desta_venda = (venda.vendedor.tel_whatsapp, f"✅ *VENDA INSTALADA (OSAB)*\n\n*Cliente:* {venda.cliente.nome_razao_social}\n*OS:* {venda.ordem_servico}\n*Data:* {dt_fmt}")

                        elif 'AGENDADO' in nome_est_upper:
                            nova_dt_ag = target_data_agenda or row.get('DATA_AGENDAMENTO')
                            if nova_dt_ag and nova_dt_ag.year >= 2000:
                                data_ag_atual = venda.data_agendamento
                                if not data_ag_atual or data_ag_atual != nova_dt_ag:
                                    detalhes_hist['data_agendamento'] = f"Nova: {nova_dt_ag}"
                                    venda.data_agendamento = nova_dt_ag
                                    houve_alteracao = True
                            
                            if houve_alteracao and venda.vendedor and venda.vendedor.tel_whatsapp:
                                dt_fmt = venda.data_agendamento.strftime('%d/%m') if venda.data_agendamento else "S/D"
                                msg_whatsapp_desta_venda = (venda.vendedor.tel_whatsapp, f"📅 *VENDA AGENDADA (OSAB)*\n\n*Cliente:* {venda.cliente.nome_razao_social}\n*OS:* {venda.ordem_servico}\n*Data:* {dt_fmt}")

                        elif 'PENDEN' in nome_est_upper:
                            novo_motivo = target_motivo_pendencia
                            if not novo_motivo:
                                # COD_PENDENCIA: match apenas pelo código numérico completo (ex.: 4 dígitos).
                                # Não usar prefixo de 2 dígitos — evita colidir 1234 vs 1256 (ambos "12").
                                cod_raw = row.get('COD_PENDENCIA', '')
                                cod_str = str(cod_raw).replace('.0', '').strip()
                                digits_only = re.sub(r'\D', '', cod_str)
                                novo_motivo = None
                                if digits_only:
                                    novo_motivo = motivo_pendencia_map.get(digits_only)
                                    # Lista oficial (ex.: pendencias_completas.csv) usa 4 dígitos com zeros à esquerda (0009-…).
                                    if not novo_motivo and len(digits_only) <= 4:
                                        novo_motivo = motivo_pendencia_map.get(
                                            digits_only.zfill(4)
                                        )
                                    if not novo_motivo and len(digits_only) >= 4:
                                        novo_motivo = motivo_pendencia_map.get(digits_only[:4])
                                if not novo_motivo:
                                    novo_motivo = motivo_padrao_osab
                            
                            if venda.motivo_pendencia_id != novo_motivo.id:
                                detalhes_hist['motivo_pendencia'] = f"Novo: {novo_motivo.nome}"
                                venda.motivo_pendencia = novo_motivo
                                houve_alteracao = True
                            
                            if houve_alteracao and venda.vendedor and venda.vendedor.tel_whatsapp:
                                msg_whatsapp_desta_venda = (venda.vendedor.tel_whatsapp, f"⚠️ *VENDA PENDENCIADA (OSAB)*\n\n*Cliente:* {venda.cliente.nome_razao_social}\n*OS:* {venda.ordem_servico}\n*Motivo:* {novo_motivo.nome}")

                    # Pagamento
                    pgto_osab_raw = row.get('MEIO_PAGAMENTO')
                    if pgto_osab_raw:
                        pgto_norm = self._normalize_text(pgto_osab_raw)
                        novo_fp = mapa_pagamentos.get(pgto_norm)
                        if not novo_fp:
                            for k, v in mapa_pagamentos.items():
                                if pgto_norm in k or k in pgto_norm: novo_fp = v; break
                        if novo_fp and (not venda.forma_pagamento or venda.forma_pagamento.id != novo_fp.id):
                            detalhes_hist['forma_pagamento'] = f"Novo: {novo_fp.nome}"
                            venda.forma_pagamento = novo_fp
                            houve_alteracao = True

                    # Conclusão
                    if houve_alteracao:
                        log_item["resultado_crm"] = "ATUALIZADO_CRM"
                        log_item["detalhe"] = "; ".join([f"{k}: {v}" for k, v in detalhes_hist.items()])
                        vendas_atualizar.append(venda)
                        historicos_criar.append(HistoricoAlteracaoVenda(venda=venda, usuario=osab_bot, alteracoes=detalhes_hist))
                        if msg_whatsapp_desta_venda and flag_enviar_whatsapp:
                            fila_mensagens_whatsapp.append(msg_whatsapp_desta_venda)
                    else:
                        log_item["resultado_crm"] = "SEM_MUDANCA_CRM"
                        report["ja_corretos"] += 1
                    
                    report["logs_detalhados"].append(log_item)

                except Exception as ex:
                    log_item["resultado_osab"] = log_item["resultado_osab"] or "ERRO"
                    log_item["resultado_crm"] = log_item["resultado_crm"] or "ERRO"
                    log_item["detalhe"] = str(ex)
                    report["erros"].append(f"L{index}: {ex}")
                    report["logs_detalhados"].append(log_item)

                # Atualizar progresso periodicamente
                if (index + 1) % progress_step == 0:
                    LogImportacaoOSAB.objects.filter(id=log_id).update(
                        total_processadas=index + 1,
                        mensagem=f'Processando registros... {index + 1}/{total_registros}'
                    )

            # --- 3. PERSISTÊNCIA ---
            LogImportacaoOSAB.objects.filter(id=log_id).update(
                total_processadas=total_registros,
                mensagem='Salvando resultados no banco...'
            )

            # Salvar em etapas para facilitar debug e evitar travar tudo num único transaction
            try:
                from django.db import connection
                if osab_criar:
                    LogImportacaoOSAB.objects.filter(id=log_id).update(
                        mensagem=f'Salvando OSAB: {len(osab_criar)} novos registros...'
                    )
                    with transaction.atomic():
                        with connection.cursor() as cursor:
                            cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                        self._sincronizar_seq_osab()
                        ImportacaoOsab.objects.bulk_create(osab_criar, batch_size=2000)

                if osab_atualizar:
                    LogImportacaoOSAB.objects.filter(id=log_id).update(
                        mensagem=f'Atualizando OSAB: {len(osab_atualizar)} registros...'
                    )
                    campos_osab = [f.name for f in ImportacaoOsab._meta.fields if f.name != 'id']
                    with transaction.atomic():
                        with connection.cursor() as cursor:
                            cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                        ImportacaoOsab.objects.bulk_update(osab_atualizar, campos_osab, batch_size=2000)

                if vendas_atualizar:
                    LogImportacaoOSAB.objects.filter(id=log_id).update(
                        mensagem=f'Atualizando vendas CRM: {len(vendas_atualizar)} registros...'
                    )
                    campos_venda = ['status_esteira', 'status_tratamento', 'data_instalacao', 'data_agendamento', 'forma_pagamento', 'motivo_pendencia', 'data_abertura']
                    with transaction.atomic():
                        with connection.cursor() as cursor:
                            cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                        Venda.objects.bulk_update(vendas_atualizar, campos_venda, batch_size=2000)

                if historicos_criar:
                    LogImportacaoOSAB.objects.filter(id=log_id).update(
                        mensagem=f'Salvando histórico: {len(historicos_criar)} registros...'
                    )
                    with transaction.atomic():
                        with connection.cursor() as cursor:
                            cursor.execute("SET LOCAL statement_timeout = '120000ms'")
                        self._sincronizar_seq_historico()
                        HistoricoAlteracaoVenda.objects.bulk_create(historicos_criar, batch_size=2000)

                try:
                    snap = self._marcar_vendas_ausentes_na_osab(log_id, total_registros, osab_bot)
                    report.update(snap)
                except Exception as e_snap:
                    print(f"Aviso: etapa CRM sem OSAB não concluída (importação OSAB segue válida): {e_snap}")

            except Exception as e:
                log.status = 'ERRO'
                log.mensagem_erro = f'Erro ao salvar no banco: {str(e)}'
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
                return

            atual_osab = len(vendas_atualizar)
            extra_nao = report.get('crm_sem_osab_nao_consta', 0)
            extra_outro = report.get('crm_sem_osab_outro_pdv', 0)
            report["atualizados"] = atual_osab + extra_nao + extra_outro
            report["atualizados_planilha_osab"] = atual_osab
            report["criados"] = len(osab_criar)

            # --- 4. ENVIO WHATSAPP ---
            if fila_mensagens_whatsapp and flag_enviar_whatsapp:
                import threading
                t = threading.Thread(target=self._enviar_mensagens_background, args=(fila_mensagens_whatsapp,))
                t.start()
                print(f"Iniciado envio de {len(fila_mensagens_whatsapp)} mensagens em background.")
            
            # Atualizar log com sucesso (usando update para garantir persistência)
            finalizado_agora = timezone.now()
            
            # Calcular duração
            log.refresh_from_db()
            if log.iniciado_em:
                duracao = int((finalizado_agora - log.iniciado_em).total_seconds())
            else:
                duracao = None
            
            LogImportacaoOSAB.objects.filter(id=log_id).update(
                status='SUCESSO',
                total_registros=report['total_registros'],
                total_processadas=report['total_registros'],  # Total processadas deve ser total_registros
                criados=report['criados'],
                atualizados=report['atualizados'],
                vendas_encontradas=report['vendas_encontradas'],
                ja_corretos=report['ja_corretos'],
                erros_count=len(report['erros']),
                mensagem=(
                    f"Processados {report['total_registros']} registros. "
                    f"{report['atualizados']} vendas atualizadas no CRM "
                    f"({report.get('atualizados_planilha_osab', report['atualizados'])} pela planilha; "
                    f"{report.get('crm_sem_osab_nao_consta', 0)} NÃO CONSTA OSAB; "
                    f"{report.get('crm_sem_osab_outro_pdv', 0)} INSTALADA OUTRO PDV; "
                    f"{report.get('bloqueados_flag_osab', 0)} bloqueadas pela flag OSAB)."
                ),
                detalhes_json=report,
                finalizado_em=finalizado_agora,
                duracao_segundos=duracao
            )
            
            print(f"✅ OSAB processado com sucesso - {report['atualizados']} vendas atualizadas")

        except Exception as e:
            # Atualizar log com erro
            print(f"❌ Erro no processamento OSAB: {str(e)}")
            import traceback
            traceback.print_exc()
            
            try:
                log.status = 'ERRO'
                log.mensagem_erro = str(e)
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
            except:
                print("Erro ao salvar log de erro")


class ImportacaoOsabDetailView(generics.RetrieveUpdateAPIView):
    queryset = ImportacaoOsab.objects.all()
    serializer_class = ImportacaoOsabSerializer
    permission_classes = [permissions.IsAuthenticated]


# --- Controle de TT's (vendedores sem venda há X dias) ---
from crm_app.controle_tts_service import controle_tts_listar_ordenado

# Compatibilidade com código que importava _controle_tts_listar_ordenado de views
_controle_tts_listar_ordenado = controle_tts_listar_ordenado


class ControleTTsAPIView(APIView):
    """GET: lista TTs dos últimos 2 meses na OSAB, com última venda válida e dias sem vender. Ordenado por dias sem vender decrescente."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['BackOffice', 'Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado. Apenas BackOffice, Diretoria ou Admin.'}, status=403)
        ontem = timezone.localdate() - timedelta(days=1)
        resultado = controle_tts_listar_ordenado()
        return Response({
            'data_referencia': ontem.isoformat(),
            'itens': resultado,
        })


class ControleTTsProximoAPIView(APIView):
    """GET: retorna o próximo TT da vez (primeiro da fila que ainda não foi marcado hoje)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['BackOffice', 'Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado.'}, status=403)
        lista = controle_tts_listar_ordenado()
        hoje = timezone.localdate()
        matriculas_marcadas_hoje = set(
            ControleTTDiaTratado.objects.filter(data=hoje).values_list('matricula_vendedor', flat=True)
        )
        lista_filtrada = [x for x in lista if x['matricula_vendedor'] not in matriculas_marcadas_hoje]
        proximo = lista_filtrada[0] if lista_filtrada else None
        return Response({'proximo': proximo})


class ControleTTTratadoAPIView(APIView):
    """GET: marcações (tratado) do mês. POST: marcar tratado para (matricula, data). DELETE: desmarcar."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['BackOffice', 'Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado.'}, status=403)

        mes_str = request.query_params.get('mes')  # YYYY-MM
        if mes_str:
            try:
                ano, mes = int(mes_str[:4]), int(mes_str[5:7])
                primeiro = date(ano, mes, 1)
                ultimo_dia = calendar.monthrange(ano, mes)[1]
                ultima_data = date(ano, mes, ultimo_dia)
            except (ValueError, IndexError):
                primeiro = timezone.localdate().replace(day=1)
                ultimo_dia = calendar.monthrange(primeiro.year, primeiro.month)[1]
                ultima_data = primeiro.replace(day=ultimo_dia)
        else:
            primeiro = timezone.localdate().replace(day=1)
            ultimo_dia = calendar.monthrange(primeiro.year, primeiro.month)[1]
            ultima_data = primeiro.replace(day=ultimo_dia)

        marcacoes = list(
            ControleTTDiaTratado.objects
            .filter(data__gte=primeiro, data__lte=ultima_data)
            .values_list('matricula_vendedor', 'data', 'tipo')
        )
        return Response({
            'mes': primeiro.strftime('%Y-%m'),
            'marcacoes': [{'matricula_vendedor': m, 'data': d.isoformat(), 'tipo': t or 'tratado'} for m, d, t in marcacoes],
        })

    def post(self, request):
        if not is_member(request.user, ['BackOffice', 'Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado.'}, status=403)

        mat = request.data.get('matricula_vendedor')
        data_str = request.data.get('data')  # YYYY-MM-DD
        tipo = request.data.get('tipo', 'tratado')
        if tipo not in (ControleTTDiaTratado.TIPO_TRATADO, ControleTTDiaTratado.TIPO_NAO_VENDAS):
            tipo = ControleTTDiaTratado.TIPO_TRATADO
        if not mat or not data_str:
            return Response({'error': 'matricula_vendedor e data são obrigatórios.'}, status=400)
        try:
            data_obj = datetime.strptime(data_str[:10], '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'data inválida. Use YYYY-MM-DD.'}, status=400)

        obj, created = ControleTTDiaTratado.objects.update_or_create(
            matricula_vendedor=mat.strip(),
            data=data_obj,
            defaults={'usuario': request.user, 'tipo': tipo},
        )
        return Response({'tipo': obj.tipo, 'created': created})

    def delete(self, request):
        if not is_member(request.user, ['BackOffice', 'Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado.'}, status=403)

        mat = request.query_params.get('matricula_vendedor')
        data_str = request.query_params.get('data')
        if not mat or not data_str:
            return Response({'error': 'matricula_vendedor e data são obrigatórios.'}, status=400)
        try:
            data_obj = datetime.strptime(data_str[:10], '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'data inválida.'}, status=400)

        deleted, _ = ControleTTDiaTratado.objects.filter(
            matricula_vendedor=mat.strip(),
            data=data_obj,
        ).delete()
        return Response({'tratado': False, 'deleted': deleted > 0})


def _normalizar_anomes_gross(val):
    """Normaliza ANOMES_GROSS para formato AAAAMM (ex.: '2025-07' -> '202507'). Aceita também número serial do Excel."""
    if val is None or pd.isna(val):
        return None
    # Número serial do Excel (ex.: 45658 = 1/12/2025)
    try:
        if isinstance(val, (int, float)) and 30000 <= float(val) <= 60000:
            from datetime import datetime, timedelta
            d = datetime(1899, 12, 30) + timedelta(days=int(float(val)))
            return d.strftime('%Y%m')
    except (ValueError, OSError):
        pass
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none', ''):
        return None
    # Remover separadores comuns
    s = s.replace('-', '').replace('/', '').replace(' ', '')
    # Se tem formato de data completa (YYYY-MM-DD), pegar só YYYYMM
    if len(s) >= 6 and s[:4].isdigit():
        if len(s) >= 8:  # YYYYMMDD ou YYYYMMDDHH...
            return s[:6]  # Pega YYYYMM
        elif len(s) == 6 and s.isdigit():
            return s  # Já está no formato AAAAMM
        elif len(s) == 7 and s[4] in ('-', '/'):  # YYYY-MM ou YYYY/MM
            return s[:4] + s[5:7]
    return s[:6] if len(s) >= 6 else None


class ImportacaoChurnView(APIView):
    permission_classes = [CheckAPIPermission]
    resource_name = 'importacao_churn'
    parser_classes = [MultiPartParser, FormParser]
    def get(self, request):
        queryset = ImportacaoChurn.objects.all().order_by('-id')
        serializer = ImportacaoChurnSerializer(queryset, many=True)
        return Response(serializer.data)
    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj: return Response({'error': 'Nenhum arquivo.'}, status=400)
        coluna_map = {'UF': 'uf', 'PRODUTO': 'produto', 'MATRICULA_VENDEDOR': 'matricula_vendedor', 'GV': 'gv', 'SAP_PRINCIPAL_FIM': 'sap_principal_fim', 'GESTAO': 'gestao', 'ST_REGIONAL': 'st_regional', 'GC': 'gc', 'NUMERO_PEDIDO': 'numero_pedido', 'NR_ORDEM': 'nr_ordem', 'DT_GROSS': 'dt_gross', 'ANOMES_GROSS': 'anomes_gross', 'DT_RETIRADA': 'dt_retirada', 'ANOMES_RETIRADA': 'anomes_retirada', 'GRUPO_UNIDADE': 'grupo_unidade', 'CODIGO_SAP': 'codigo_sap', 'MUNICIPIO': 'municipio', 'TIPO_RETIRADA': 'tipo_retirada', 'MOTIVO_RETIRADA': 'motivo_retirada', 'SUBMOTIVO_RETIRADA': 'submotivo_retirada', 'CLASSIFICACAO': 'classificacao', 'DESC_APELIDO': 'desc_apelido', 'NR_VELOCIDADE': 'nr_velocidade', 'VELOCIDADE': 'nr_velocidade'}
        try:
            if file_obj.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_obj)
            elif file_obj.name.endswith('.xlsb'):
                df = pd.read_excel(file_obj, engine='pyxlsb')
            elif file_obj.name.endswith('.csv'):
                df = pd.read_csv(file_obj)
            else:
                return Response({'error': 'Formato inválido. Use .xlsx, .xls, .xlsb ou .csv'}, status=400)
        except Exception as e: return Response({'error': str(e)}, status=400)
        df.columns = [str(col).strip().upper().replace(' ', '_') for col in df.columns]
        for f in ['DT_GROSS', 'DT_RETIRADA']:
            if f in df.columns: df[f] = pd.to_datetime(df[f], errors='coerce')
        df = df.replace({np.nan: None, pd.NaT: None})
        df.rename(columns=coluna_map, inplace=True)
        
        # Normalizar anomes_gross para formato AAAAMM
        if 'anomes_gross' in df.columns:
            df['anomes_gross'] = df['anomes_gross'].apply(_normalizar_anomes_gross)
        
        # Bulk operations optimization
        criados, atualizados, erros = 0, 0, []
        fields = {f.name for f in ImportacaoChurn._meta.get_fields() if f.name != 'id'}
        
        # Separar registros para criar e atualizar
        # Buscar existentes por numero_pedido OU nr_ordem
        pedidos_df = [row.get('numero_pedido') for _, row in df.iterrows() if row.get('numero_pedido')]
        nr_ordens_df = [row.get('nr_ordem') for _, row in df.iterrows() if row.get('nr_ordem')]
        
        existentes_pedido = {obj.numero_pedido: obj for obj in ImportacaoChurn.objects.filter(numero_pedido__in=pedidos_df) if obj.numero_pedido}
        existentes_nr_ordem = {obj.nr_ordem: obj for obj in ImportacaoChurn.objects.filter(nr_ordem__in=nr_ordens_df) if obj.nr_ordem}
        
        to_create = []
        to_update = []
        linhas_ignoradas = 0
        motivo_ignoradas = []
        
        for idx, row in df.iterrows():
            data = row.to_dict()
            pedido = data.get('numero_pedido')
            nr_ordem_val = data.get('nr_ordem')
            
            # Se não tem pedido nem nr_ordem, não tem como identificar unicamente
            if not pedido and not nr_ordem_val:
                linhas_ignoradas += 1
                motivo_ignoradas.append(f"Linha {idx+2}: sem numero_pedido e sem nr_ordem")
                continue
            
            filtered_data = {k: v for k, v in data.items() if k in fields}
            
            try:
                obj_existente = None
                chave_usada = None
                
                # Prioridade: buscar por numero_pedido, depois por nr_ordem
                if pedido and pedido in existentes_pedido:
                    obj_existente = existentes_pedido[pedido]
                    chave_usada = 'numero_pedido'
                elif nr_ordem_val and nr_ordem_val in existentes_nr_ordem:
                    obj_existente = existentes_nr_ordem[nr_ordem_val]
                    chave_usada = 'nr_ordem'
                
                if obj_existente:
                    # Atualizar existente
                    for k, v in filtered_data.items():
                        setattr(obj_existente, k, v)
                    to_update.append(obj_existente)
                else:
                    # Criar novo (pode ter numero_pedido=None se só tiver nr_ordem)
                    to_create.append(ImportacaoChurn(**filtered_data))
            except Exception as e:
                erros.append(f"Linha {idx+2}: {e}")
        
        # Executar bulk operations
        with transaction.atomic():
            if to_create:
                ImportacaoChurn.objects.bulk_create(to_create, batch_size=1000)
                criados = len(to_create)
            if to_update:
                ImportacaoChurn.objects.bulk_update(to_update, list(fields), batch_size=1000)
                atualizados = len(to_update)
        
        return Response({
            'status': 'sucesso',
            'total_registros': len(df),
            'criados': criados,
            'atualizados': atualizados,
            'linhas_ignoradas': linhas_ignoradas,
            'motivo_ignoradas': motivo_ignoradas[:10] if len(motivo_ignoradas) > 10 else motivo_ignoradas,  # Limitar a 10 exemplos
            'erros': erros
        }, status=200)

class ImportacaoChurnDetailView(generics.RetrieveUpdateAPIView):
    queryset = ImportacaoChurn.objects.all()
    serializer_class = ImportacaoChurnSerializer
    permission_classes = [permissions.IsAuthenticated]

class ImportacaoCicloPagamentoView(APIView):
    permission_classes = [CheckAPIPermission]
    resource_name = 'importacao_ciclo_pagamento'
    parser_classes = [MultiPartParser, FormParser]
    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj: return Response({'error': 'Arquivo não enviado.'}, status=400)
        coluna_map = {'ANO': 'ano', 'MES': 'mes', 'QUINZENA': 'quinzena', 'CICLO': 'ciclo', 'CICLO_COMPLEMENTAR': 'ciclo_complementar', 'EVENTO': 'evento', 'SUB_EVENTO': 'sub_evento', 'CANAL_DETALHADO': 'canal_detalhado', 'CANAL_AGRUPADO': 'canal_agrupado', 'SUB_CANAL': 'sub_canal', 'COD_SAP': 'cod_sap', 'COD_SAP_AGR': 'cod_sap_agr', 'PARCEIRO_AGR': 'parceiro_agr', 'UF_PARCEIRO_AGR': 'uf_parceiro_agr', 'FAMILIA': 'familia', 'PRODUTO': 'produto', 'OFERTA': 'oferta', 'PLANO_DETALHADO': 'plano_detalhado', 'CELULA': 'celula', 'METODO_PAGAMENTO': 'metodo_pagamento', 'CONTRATO': 'contrato', 'NUM_OS_PEDIDO_SIEBEL': 'num_os_pedido_siebel', 'ID_BUNDLE': 'id_bundle', 'DATA_ATV': 'data_atv', 'DATA_RETIRADA': 'data_retirada', 'QTD': 'qtd', 'COMISSAO_BRUTA': 'comissao_bruta', 'FATOR': 'fator', 'IQ': 'iq', 'VALOR_COMISSAO_FINAL': 'valor_comissao_final'}
        try:
            df = pd.read_excel(file_obj)
        except Exception as e: return Response({'error': str(e)}, status=400)
        df.columns = [str(col).strip().upper() for col in df.columns]
        df.rename(columns=coluna_map, inplace=True)
        for f in ['data_atv', 'data_retirada']:
            if f in df.columns: df[f] = pd.to_datetime(df[f], errors='coerce')
        for f in ['comissao_bruta', 'fator', 'iq', 'valor_comissao_final']:
            if f in df.columns:
                df[f] = df[f].astype(str).str.replace('R$', '', regex=False).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                df[f] = pd.to_numeric(df[f], errors='coerce')
        df = df.replace({np.nan: None, pd.NaT: None})
        
        # Bulk operations optimization
        criados, atualizados, erros = 0, 0, []
        fields = {f.name for f in CicloPagamento._meta.get_fields() if f.name != 'contrato'}
        
        # Separar registros para criar e atualizar
        contratos_df = [row.get('contrato') for _, row in df.iterrows() if row.get('contrato')]
        existentes = {obj.contrato: obj for obj in CicloPagamento.objects.filter(contrato__in=contratos_df)}
        
        to_create = []
        to_update = []
        
        for idx, row in df.iterrows():
            data = row.to_dict()
            contrato = data.get('contrato')
            if not contrato: continue
            
            filtered_data = {k: v for k, v in data.items() if k in fields or k == 'contrato'}
            
            try:
                if contrato in existentes:
                    obj = existentes[contrato]
                    for k, v in filtered_data.items():
                        if k != 'contrato':
                            setattr(obj, k, v)
                    to_update.append(obj)
                else:
                    to_create.append(CicloPagamento(**filtered_data))
            except Exception as e:
                erros.append(f"Linha {idx+2}: {e}")
        
        # Executar bulk operations
        with transaction.atomic():
            if to_create:
                CicloPagamento.objects.bulk_create(to_create, batch_size=1000, ignore_conflicts=True)
                criados = len(to_create)
            if to_update:
                CicloPagamento.objects.bulk_update(to_update, list(fields), batch_size=1000)
                atualizados = len(to_update)
        
        return Response({'total': len(df), 'criados': criados, 'atualizados': atualizados, 'erros': erros}, status=200)
        return Response({'total': len(df), 'criados': criados, 'atualizados': atualizados, 'erros': erros}, status=200)

class PerformanceVendasView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, *args, **kwargs):
        User = get_user_model()
        hoje = timezone.now().date()
        start_of_week = hoje - timedelta(days=hoje.weekday())
        start_of_month = hoje.replace(day=1)
        
        current_user = self.request.user

        # --- 1. DEFINIR QUEM VAI APARECER NO PAINEL ---
        # Filtra apenas usuários ativos e remove robôs/admins que não vendem
        base_users = User.objects.filter(is_active=True).exclude(username__in=['OSAB_IMPORT', 'admin', 'root'])

        if is_member(current_user, ['Diretoria', 'BackOffice', 'Admin', 'Auditoria', 'Qualidade']):
            users_to_process = base_users.select_related('supervisor')
        elif is_member(current_user, ['Supervisor']):
            # Supervisor vê a si mesmo e seus liderados
            users_to_process = base_users.filter(Q(id=current_user.id) | Q(supervisor=current_user)).select_related('supervisor')
        else:
            # Vendedor vê apenas a si mesmo (ou sua equipe, dependendo da regra. Aqui deixei restrito)
            users_to_process = base_users.filter(id=current_user.id).select_related('supervisor')

        # --- 2. AGREGAR VENDAS ---
        # Filtros: Vendas ativas, COM tratamento definido, SEM cancelamento na esteira
        base_filters = (
            Q(ativo=True) & 
            Q(status_tratamento__isnull=False) & 
            (Q(status_esteira__isnull=True) | ~Q(status_esteira__nome__iexact='CANCELADA'))
        )

        vendas = Venda.objects.filter(base_filters).values('vendedor_id').annotate(
            total_dia=Count('id', filter=Q(data_pedido__date=hoje)),
            total_mes=Count('id', filter=Q(data_pedido__date__gte=start_of_month)),
            total_mes_instalado=Count('id', filter=Q(data_pedido__date__gte=start_of_month, status_esteira__nome__iexact='Instalada')),
            # Semanal
            vendas_segunda=Count('id', filter=Q(data_pedido__date=start_of_week)),
            vendas_terca=Count('id', filter=Q(data_pedido__date=start_of_week + timedelta(days=1))),
            vendas_quarta=Count('id', filter=Q(data_pedido__date=start_of_week + timedelta(days=2))),
            vendas_quinta=Count('id', filter=Q(data_pedido__date=start_of_week + timedelta(days=3))),
            vendas_sexta=Count('id', filter=Q(data_pedido__date=start_of_week + timedelta(days=4))),
            vendas_sabado=Count('id', filter=Q(data_pedido__date=start_of_week + timedelta(days=5))),
        )
        
        vendas_por_vendedor = {v['vendedor_id']: v for v in vendas}

        # --- 3. MONTAR ESTRUTURA DOS TIMES ---
        teams = defaultdict(lambda: {
            'supervisor_name': 'Sem Supervisor', 
            'members': [], 
            'totals': {
                'daily': 0, 
                'weekly_breakdown': {'seg':0,'ter':0,'qua':0,'qui':0,'sex':0,'sab':0},
                'weekly_total': 0,
                'monthly': {'total': 0, 'instalados': 0}
            }
        })

        for user in users_to_process:
            supervisor_id = user.supervisor.id if user.supervisor else 'none'
            
            if supervisor_id != 'none' and user.supervisor:
                teams[supervisor_id]['supervisor_name'] = user.supervisor.username
            
            # Pega dados de vendas (ou 0 se não tiver)
            v_data = vendas_por_vendedor.get(user.id, {})
            
            dia_atual = v_data.get('total_dia', 0)
            weekly_total = sum([
                v_data.get('vendas_segunda', 0), v_data.get('vendas_terca', 0),
                v_data.get('vendas_quarta', 0), v_data.get('vendas_quinta', 0),
                v_data.get('vendas_sexta', 0), v_data.get('vendas_sabado', 0)
            ])
            total_mes = v_data.get('total_mes', 0)
            instalados_mes = v_data.get('total_mes_instalado', 0)
            
            aproveitamento_str = f"{(instalados_mes / total_mes * 100):.0f}%" if total_mes > 0 else "-"

            # Adiciona membro (MESMO SE TIVER 0 VENDAS)
            teams[supervisor_id]['members'].append({
                'name': user.username,
                'full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'daily': dia_atual,
                'weekly_breakdown': {
                    'seg': v_data.get('vendas_segunda', 0),
                    'ter': v_data.get('vendas_terca', 0),
                    'qua': v_data.get('vendas_quarta', 0),
                    'qui': v_data.get('vendas_quinta', 0),
                    'sex': v_data.get('vendas_sexta', 0),
                    'sab': v_data.get('vendas_sabado', 0)
                },
                'weekly_total': weekly_total,
                'monthly': {
                    'total': total_mes,
                    'instalados': instalados_mes,
                    'aproveitamento': aproveitamento_str
                }
            })

        # --- 4. CALCULAR TOTAIS E ORDENAR ---
        final_result = []
        
        for supervisor_id, team_data in teams.items():
            if not team_data['members']: continue

            # ORDENAÇÃO IMPORTANTE: Quem vendeu mais no dia fica em cima. Zeros em baixo.
            # Se empate no dia, desempata pelo mensal.
            team_data['members'] = sorted(
                team_data['members'], 
                key=lambda x: (x['daily'], x['monthly']['total']), 
                reverse=True
            )

            # Somar totais do time
            for member in team_data['members']:
                team_data['totals']['daily'] += member['daily']
                for day, count in member['weekly_breakdown'].items():
                    team_data['totals']['weekly_breakdown'][day] += count
                team_data['totals']['weekly_total'] += member['weekly_total']
                team_data['totals']['monthly']['total'] += member['monthly']['total']
                team_data['totals']['monthly']['instalados'] += member['monthly']['instalados']
            
            # Aproveitamento do Time
            t_total = team_data['totals']['monthly']['total']
            t_inst = team_data['totals']['monthly']['instalados']
            team_data['totals']['monthly']['aproveitamento'] = f"{(t_inst / t_total * 100):.0f}%" if t_total > 0 else "-"

            final_result.append(team_data)

        # Ordena as equipes pelo total de vendas do dia (Equipe que mais vendeu aparece primeiro)
        final_result = sorted(final_result, key=lambda x: x['totals']['daily'], reverse=True)
        
        return Response(final_result)

class ImportarKMLView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=400)

        if not file_obj.name.endswith('.kml'):
             return Response({'error': 'Por favor, envie um arquivo .kml (se for .kmz, descompacte antes).'}, status=400)

        try:
            tree = ET.parse(file_obj)
            root = tree.getroot()
            ns = {'kml': 'http://www.opengis.net/kml/2.2'}
            
            placemarks = root.findall('.//kml:Placemark', ns)
            if not placemarks:
                placemarks = root.findall('.//Placemark')

            areas_para_criar = []
            erros = 0
            
            # Compilar regex fora do loop para performance
            regex_cache = {}

            def extrair_rapido(chave, texto):
                if chave not in regex_cache:
                    regex_cache[chave] = re.compile(rf'<strong>{chave}:\s*</strong>(.*?)(?:<br>|$)', re.IGNORECASE)
                match = regex_cache[chave].search(texto)
                return match.group(1).strip() if match else None

            # Limpa tabela para evitar PK duplicada em reimportações
            AreaVenda.objects.all().delete()
            try:
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT setval(pg_get_serial_sequence('crm_app_areavenda', 'id'), COALESCE((SELECT MAX(id) FROM crm_app_areavenda), 1), false)"
                    )
            except Exception:
                pass  # Se não for Postgres ou falhar, segue com sequence atual

            for pm in placemarks:
                try:
                    def find_text(elem, tag):
                        res = elem.find(f'kml:{tag}', ns)
                        if res is None: res = elem.find(tag)
                        return res.text if res is not None else ""

                    nome = find_text(pm, 'name')
                    description = find_text(pm, 'description')
                    
                    coords_text = ""
                    poly = pm.find('.//kml:Polygon//kml:coordinates', ns)
                    if poly is None: poly = pm.find('.//Polygon//coordinates')
                    if poly is not None:
                        coords_text = poly.text.strip()
                    
                    # Pular se não tiver coordenadas (economiza tempo)
                    if not coords_text:
                        continue

                    obj = AreaVenda(
                        nome_kml=nome or "Sem Nome",
                        coordenadas=coords_text,
                        celula=extrair_rapido('Célula', description),
                        uf=extrair_rapido('UF', description),
                        municipio=extrair_rapido('Município', description),
                        estacao=extrair_rapido('Estação', description),
                        cluster=extrair_rapido('Cluster Célula', description),
                        status_venda=extrair_rapido('Status Venda Célula', description),
                        ocupacao=extrair_rapido(r'Ocup \(%\)', description),
                        atingimento_meta=extrair_rapido(r'Atingimento/Meta \(%\)', description)
                    )
                    
                    # Inteiros
                    for campo, chave in [('prioridade','Prioridade'), ('aging','Aging'), ('hc','HC'), ('hp','HP'), ('hp_viavel','HP Viável'), ('hp_viavel_total','HP Viável Total')]:
                        val = extrair_rapido(chave, description)
                        try: setattr(obj, campo, int(val) if val else 0)
                        except: setattr(obj, campo, 0)
                    
                    # Float
                    val_hc = extrair_rapido('HC Esperado', description)
                    if val_hc:
                        try: obj.hc_esperado = float(val_hc.replace(',', '.'))
                        except: obj.hc_esperado = 0.0
                    
                    areas_para_criar.append(obj)
                    
                except Exception as e:
                    # Não logar erro individual para não spammar log e atrasar
                    erros += 1

            # SALVAMENTO EM LOTE (BULK INSERT) - AQUI ESTÁ O GANHO DE PERFORMANCE
            # Divide em lotes de 1000 para não estourar memória do banco
            batch_size = 1000
            for i in range(0, len(areas_para_criar), batch_size):
                AreaVenda.objects.bulk_create(areas_para_criar[i:i + batch_size])

            return Response({
                'status': 'sucesso',
                'mensagem': f'Importação concluída! {len(areas_para_criar)} áreas importadas.',
            })

        except Exception as e:
            return Response({'error': f'Erro crítico KML: {str(e)}'}, status=500)
        
# --- IMPORTAÇÃO DFV (CSV) - VERSÃO PROFISSIONAL ---
class ImportarDFVView(APIView):
    """
    View para importação de arquivos DFV (Dados do Faturamento de Vendas).
    
    Esta view segue as melhores práticas:
    - Separação de responsabilidades (usa Service Layer)
    - Processamento assíncrono em background
    - Resposta imediata ao cliente
    - Tratamento robusto de erros
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        """
        Endpoint para upload e processamento de arquivo DFV.
        
        Returns:
            Response com log_id e status da importação iniciada
        """
        import logging
        import threading
        import tempfile
        import os
        
        logger = logging.getLogger(__name__)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'error': 'Arquivo CSV não enviado.'}, 
                status=400
            )
        
        try:
            from .models import LogImportacaoDFV
            from django.utils import timezone
            
            # Evitar múltiplas importações simultâneas (reduz lock e travamentos)
            log_em_andamento = LogImportacaoDFV.objects.filter(status='PROCESSANDO').order_by('-iniciado_em').first()
            if log_em_andamento:
                return Response(
                    {
                        'error': 'Já existe uma importação DFV em andamento. Aguarde finalizar para enviar outro arquivo.',
                        'log_id': log_em_andamento.id,
                        'status': 'PROCESSANDO'
                    },
                    status=409
                )
            
            # Criar log de importação ANTES de ler o arquivo
            log = LogImportacaoDFV.objects.create(
                nome_arquivo=file_obj.name,
                usuario=request.user,
                status='PROCESSANDO',
                tamanho_arquivo=0  # Será atualizado durante o processamento
            )
            
            logger.info(
                f"[DFV] Nova importação iniciada - Log ID: {log.id}, "
                f"Arquivo: {file_obj.name}, Usuário: {request.user.username}"
            )
            
            # Salvar arquivo em disco temporário para evitar uso excessivo de memória
            chunk_size = 1024 * 1024  # 1MB por chunk
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
            temp_path = temp_file.name
            total_bytes = 0

            logger.debug(f"[DFV] Salvando arquivo {file_obj.name} em disco temporário...")
            for chunk in file_obj.chunks(chunk_size):
                temp_file.write(chunk)
                total_bytes += len(chunk)
            temp_file.close()

            logger.info(
                f"[DFV] Arquivo salvo em disco: {file_obj.name} "
                f"({total_bytes / (1024*1024):.2f} MB)"
            )

            # Atualizar tamanho do arquivo no log
            LogImportacaoDFV.objects.filter(id=log.id).update(tamanho_arquivo=total_bytes)
            
            # Iniciar processamento em thread background
            def processar_dfv_async():
                """Wrapper para processamento assíncrono"""
                try:
                    from .services.dfv_import_service import DFVImportService
                    
                    service = DFVImportService(log_id=log.id)
                    service.process(None, file_obj.name, arquivo_path=temp_path)
                except Exception as e:
                    logger.error(
                        f"[DFV] Erro crítico no processamento assíncrono: {e}",
                        exc_info=True
                    )
                    # O serviço já atualiza o log em caso de erro
                finally:
                    try:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    except Exception as e:
                        logger.warning(f"[DFV] Não foi possível remover arquivo temporário: {e}")
            
            thread = threading.Thread(target=processar_dfv_async, daemon=True)
            thread.start()
            
            # Retornar imediatamente ao cliente
            return Response({
                'success': True,
                'log_id': log.id,
                'message': 'Importação DFV iniciada! O processamento continuará em segundo plano.',
                'status': 'PROCESSANDO',
                'background': True
            })
            
        except Exception as e:
            logger.error(
                f"[DFV] Erro ao iniciar importação: {e}",
                exc_info=True
            )
            return Response(
                {'error': f'Erro ao iniciar importação: {str(e)}'}, 
                status=500
            )


# --- IMPORTAÇÃO CNPJ RECEITA FEDERAL (ESTABELE) ---
class ImportarCNPJView(APIView):
    """
    Importa arquivos ESTABELE da Receita Federal (30 colunas, separador ;, sem cabeçalho).
    Processamento em streaming para suportar arquivos grandes.
    """
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        import threading
        import tempfile

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Arquivo não enviado.'}, status=400)

        nome = file_obj.name.upper()
        if 'ESTABELE' not in nome and not nome.endswith('.CSV'):
            return Response({
                'error': 'Arquivo deve ser do tipo ESTABELE (Receita Federal) ou CSV. '
                         'Ex: K3241.K03200Y0.D60214.ESTABELE'
            }, status=400)

        try:
            from .models import LogImportacaoEstabelecimentoCNPJ
            from .services.cnpj_estabele_import_service import processar_arquivo_estabele

            log_em_andamento = LogImportacaoEstabelecimentoCNPJ.objects.filter(
                status='PROCESSANDO'
            ).first()
            if log_em_andamento:
                return Response({
                    'error': 'Já existe uma importação CNPJ em andamento. Aguarde finalizar.',
                    'log_id': log_em_andamento.id,
                }, status=409)

            log = LogImportacaoEstabelecimentoCNPJ.objects.create(
                nome_arquivo=file_obj.name,
                usuario=request.user,
                status='PROCESSANDO',
                tamanho_arquivo=0,
            )

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.ESTABELE')
            temp_path = temp_file.name
            total_bytes = 0
            for chunk in file_obj.chunks(chunk_size=1024 * 1024):
                temp_file.write(chunk)
                total_bytes += len(chunk)
            temp_file.close()

            LogImportacaoEstabelecimentoCNPJ.objects.filter(id=log.id).update(
                tamanho_arquivo=total_bytes
            )

            aplicar_filtros = request.data.get('aplicar_filtros') in (True, 'true', '1')
            cnae = request.data.get('cnae_fiscal') or None
            municipio = request.data.get('codigo_municipio') or None
            situacao = request.data.get('situacao_cadastral') or None

            def processar_async():
                try:
                    processar_arquivo_estabele(
                        log_id=log.id,
                        arquivo_path=temp_path,
                        aplicar_filtros=aplicar_filtros,
                        cnae_fiscal=cnae,
                        codigo_municipio=municipio,
                        situacao_cadastral=situacao,
                    )
                finally:
                    try:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    except Exception:
                        pass

            threading.Thread(target=processar_async, daemon=True).start()

            return Response({
                'success': True,
                'log_id': log.id,
                'message': 'Importação CNPJ iniciada! O processamento continuará em segundo plano.',
                'status': 'PROCESSANDO',
                'background': True,
            })
        except Exception as e:
            logger.exception("[CNPJ] Erro ao iniciar importação")
            return Response({'error': str(e)}, status=500)


class LogsImportacaoCNPJView(APIView):
    """Lista logs de importação CNPJ Receita Federal"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import LogImportacaoEstabelecimentoCNPJ
        limit = int(request.query_params.get('limit', 20))
        logs = LogImportacaoEstabelecimentoCNPJ.objects.all().order_by('-iniciado_em')[:limit]
        data = []
        for log in logs:
            data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'tamanho_arquivo': log.tamanho_arquivo,
                'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_linhas': log.total_linhas,
                'total_importadas': log.total_importadas,
                'total_erros': log.total_erros,
                'mensagem': log.mensagem,
                'mensagem_erro': log.mensagem_erro,
                'usuario_nome': log.usuario.username if log.usuario else None,
            })
        return Response({'success': True, 'logs': data})


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def enviar_comissao_whatsapp(request):
    """
    Envia resumo de comissão (card) por WhatsApp aos consultores selecionados.
    Valida entrada e delega ao ComissaoWhatsAppService.
    """
    try:
        ano = int(request.data.get("ano"))
        mes = int(request.data.get("mes"))
    except (TypeError, ValueError):
        return Response({"error": "ano e mes numéricos são obrigatórios."}, status=400)

    consultores_ids = request.data.get("consultores", [])
    if not consultores_ids:
        return Response(
            {"error": "Nenhum consultor selecionado."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        from crm_app.services.comissao_whatsapp_service import (
            enviar_comissao_whatsapp_consultores,
        )

        sucessos, erros = enviar_comissao_whatsapp_consultores(
            ano=ano,
            mes=mes,
            consultores_ids=[int(cid) for cid in consultores_ids],
        )
        return Response(
            {
                "mensagem": f"Processamento concluído. Sucessos: {sucessos}. Falhas: {len(erros)}",
                "detalhes_erro": erros,
            }
        )
    except Exception as e:
        logger.exception("enviar_comissao_whatsapp: %s", e)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def enviar_folha_extrato_whatsapp(request):
    """
    Envia para WhatsApp um único PDF (folha + extrato), para um ou vários vendedores.
    Body:
      { "ano": int, "mes": int, "vendedor_id": int? , "vendedores_ids": [int]? , "fechar_mes": bool? }.
    Destino: tel_whatsapp (WhatsApp 1 principal) do usuário.
    """
    try:
        ano = request.data.get('ano')
        mes = request.data.get('mes')
        vendedor_id = request.data.get('vendedor_id')
        vendedores_ids = request.data.get('vendedores_ids') or []
        fechar_mes = bool(request.data.get('fechar_mes', True))
        if ano is None or mes is None:
            return Response(
                {"error": "Envie ano e mes."},
                status=status.HTTP_400_BAD_REQUEST
            )
        ano, mes = int(ano), int(mes)

        ids_envio = []
        if isinstance(vendedores_ids, list):
            for x in vendedores_ids:
                try:
                    ids_envio.append(int(x))
                except (TypeError, ValueError):
                    continue
        if vendedor_id is not None:
            try:
                ids_envio.append(int(vendedor_id))
            except (TypeError, ValueError):
                pass
        ids_envio = sorted(set(ids_envio))
        if not ids_envio:
            return Response({"error": "Selecione ao menos um vendedor."}, status=status.HTTP_400_BAD_REQUEST)

        User = get_user_model()
        consultores = {u.id: u for u in User.objects.filter(id__in=ids_envio)}

        from .comissao_folha_service import calcular_folha_mes
        folha = calcular_folha_mes(ano, mes, use_effective_date_for_display=True)
        map_folha = {int(x.get('vendedor_id')): x for x in folha.get('vendedores', []) if x.get('vendedor_id') is not None}
        periodo = folha.get('periodo', f"{mes:02d}/{ano}")
        svc = WhatsAppService()
        from crm_app.comissao_folha_whatsapp_pdf import montar_html_folha_e_extrato_pdf

        enviados = []
        erros = []
        for vid in ids_envio:
            consultor = consultores.get(vid)
            if not consultor:
                erros.append(f"ID {vid}: vendedor não encontrado")
                continue
            telefone = getattr(consultor, 'tel_whatsapp', None) or (consultor.telefone if hasattr(consultor, 'telefone') else None)
            if not telefone or not str(telefone).strip():
                erros.append(f"{consultor.username}: sem WhatsApp cadastrado")
                continue
            vendedor_data = map_folha.get(vid)
            if not vendedor_data:
                erros.append(f"{consultor.username}: sem dados de folha no período")
                continue

            html_string = montar_html_folha_e_extrato_pdf(vendedor_data, periodo)
            pdf_buffer = BytesIO()
            pisa_status = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), pdf_buffer, encoding="utf-8")
            if pisa_status.err:
                erros.append(f"{consultor.username}: erro ao gerar PDF")
                continue
            pdf_bytes = pdf_buffer.getvalue()
            pdf_b64 = base64.b64encode(pdf_bytes).decode("utf-8")
            nome_seguro = "".join(
                c if c.isalnum() or c in ("-", "_") else "_"
                for c in str(vendedor_data.get("vendedor_nome") or "vendedor")
            ).strip("_") or "vendedor"
            nome_pdf = f"Folha_Comissao_{nome_seguro}_{mes}_{ano}.pdf"
            resp_pdf = svc.enviar_pdf_b64(
                telefone,
                pdf_b64,
                nome_arquivo=nome_pdf,
                caption=f"Folha de comissão {periodo} (resumo + extrato)",
            )
            if resp_pdf is None or (isinstance(resp_pdf, dict) and resp_pdf.get('error')):
                erros.append(f"{consultor.username}: falha no envio WhatsApp")
                continue
            enviados.append(vid)

        if not enviados:
            return Response(
                {"error": "Nenhum envio foi concluído.", "erros": erros},
                status=status.HTTP_502_BAD_GATEWAY
            )

        if fechar_mes:
            total_pago_mes = sum(float(v.get('resumo', {}).get('liquido', 0) or 0) for v in folha.get('vendedores', []))
            try:
                _fechar_pagamento_mes(ano, mes, total_pago=total_pago_mes)
            except Exception as e:
                return Response(
                    {
                        "ok": False,
                        "mensagem": f"Envios concluídos para {len(enviados)} vendedor(es), mas falhou ao fechar o mês.",
                        "enviados": enviados,
                        "erros": erros,
                        "erro_fechamento": str(e),
                    },
                    status=status.HTTP_207_MULTI_STATUS,
                )

        return Response({
            "ok": True,
            "mensagem": f"PDF enviado para {len(enviados)} vendedor(es)." + (" Mês conferido/pago/fechado." if fechar_mes else ""),
            "enviados": enviados,
            "erros": erros,
        })
    except Exception as e:
        logger.exception("enviar_folha_extrato_whatsapp: %s", e)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def exportar_folha_extrato_pdf(request):
    """
    Gera download de PDF (folha + extrato no mesmo arquivo), no mesmo layout do WhatsApp.
    Body: { "ano": int, "mes": int, "vendedor_id": int?, "vendedores_ids": [int]? }.
    Um vendedor: retorna application/pdf. Vários: retorna application/zip com um PDF por vendedor.
    Não envia WhatsApp e não fecha o mês.
    """
    import zipfile
    from urllib.parse import quote

    try:
        ano = request.data.get('ano')
        mes = request.data.get('mes')
        vendedor_id = request.data.get('vendedor_id')
        vendedores_ids = request.data.get('vendedores_ids') or []
        if ano is None or mes is None:
            return Response(
                {"error": "Envie ano e mes."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ano, mes = int(ano), int(mes)

        ids_envio = []
        if isinstance(vendedores_ids, list):
            for x in vendedores_ids:
                try:
                    ids_envio.append(int(x))
                except (TypeError, ValueError):
                    continue
        if vendedor_id is not None:
            try:
                ids_envio.append(int(vendedor_id))
            except (TypeError, ValueError):
                pass
        ids_envio = sorted(set(ids_envio))
        if not ids_envio:
            return Response(
                {"error": "Selecione ao menos um vendedor."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        User = get_user_model()
        consultores = {u.id: u for u in User.objects.filter(id__in=ids_envio)}

        from .comissao_folha_service import calcular_folha_mes
        from crm_app.comissao_folha_whatsapp_pdf import montar_html_folha_e_extrato_pdf

        folha = calcular_folha_mes(ano, mes, use_effective_date_for_display=True)
        map_folha = {
            int(x.get('vendedor_id')): x
            for x in folha.get('vendedores', [])
            if x.get('vendedor_id') is not None
        }
        periodo = folha.get('periodo', f"{mes:02d}/{ano}")

        def _nome_arquivo_seguro(nome_base: str) -> str:
            s = "".join(
                c if c.isalnum() or c in ("-", "_") else "_"
                for c in str(nome_base or "vendedor")
            ).strip("_") or "vendedor"
            return s

        pdfs_ok = []
        erros = []
        for vid in ids_envio:
            consultor = consultores.get(vid)
            if not consultor:
                erros.append(f"ID {vid}: vendedor não encontrado")
                continue
            vendedor_data = map_folha.get(vid)
            if not vendedor_data:
                erros.append(f"{consultor.username}: sem dados de folha no período")
                continue
            html_string = montar_html_folha_e_extrato_pdf(vendedor_data, periodo)
            pdf_buffer = BytesIO()
            pisa_status = pisa.pisaDocument(
                BytesIO(html_string.encode("UTF-8")), pdf_buffer, encoding="utf-8"
            )
            if pisa_status.err:
                erros.append(f"{consultor.username}: erro ao gerar PDF")
                continue
            nome_pdf = (
                f"Folha_Comissao_{_nome_arquivo_seguro(vendedor_data.get('vendedor_nome'))}_{mes}_{ano}.pdf"
            )
            pdfs_ok.append((nome_pdf, pdf_buffer.getvalue()))

        if not pdfs_ok:
            return Response(
                {
                    "error": "Nenhum PDF foi gerado.",
                    "detalhes": erros,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(pdfs_ok) == 1:
            nome, conteudo = pdfs_ok[0]
            resp = HttpResponse(conteudo, content_type="application/pdf")
            resp["Content-Disposition"] = (
                f'attachment; filename="{nome}"; filename*=UTF-8\'\'{quote(nome)}'
            )
            return resp

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for nome, conteudo in pdfs_ok:
                zf.writestr(nome, conteudo)
        zip_bytes = zip_buffer.getvalue()
        zip_nome = f"Folha_Comissao_{mes:02d}_{ano}_varios_vendedores.zip"
        resp = HttpResponse(zip_bytes, content_type="application/zip")
        resp["Content-Disposition"] = (
            f'attachment; filename="{zip_nome}"; filename*=UTF-8\'\'{quote(zip_nome)}'
        )
        if erros:
            resp["X-Export-Warnings"] = "; ".join(erros[:5])
        return resp
    except Exception as e:
        logger.exception("exportar_folha_extrato_pdf: %s", e)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def exportar_comissionamento_resumo_excel(request):
    """
    Exporta resumo de comissionamento em XLSX para o mês/ano informado.
    Colunas: vendedor, comissão bruta, comissão líquida e descontos por tipo.
    """
    from decimal import Decimal

    try:
        ano = request.data.get('ano')
        mes = request.data.get('mes')
        if ano is None or mes is None:
            return Response(
                {"error": "Envie ano e mes."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ano, mes = int(ano), int(mes)

        from .comissao_folha_service import calcular_folha_mes

        folha = calcular_folha_mes(ano, mes, use_effective_date_for_display=True)
        vendedores = sorted(
            folha.get('vendedores', []),
            key=lambda x: (x.get('vendedor_nome') or '').upper()
        )

        colunas_desconto = [
            "DESCONTO_BOLETO",
            "DESCONTO_ANTECIPACAO_INSTALACAO",
            "DESCONTO_ADIANT_CNPJ",
            "DESCONTO_ADIANT_COMISSAO",
            "DESCONTO_CHURN_M0",
            "DESCONTO_CHURN_M1",
            "DESCONTO_OUTROS",
        ]

        tipo_para_coluna = {
            "folha_boleto_vendas": "DESCONTO_BOLETO",
            "boleto": "DESCONTO_BOLETO",
            "folha_antecipacao_instalacao": "DESCONTO_ANTECIPACAO_INSTALACAO",
            "antecipacao_instalacao": "DESCONTO_ANTECIPACAO_INSTALACAO",
            "adiant_cnpj": "DESCONTO_ADIANT_CNPJ",
            "adiant_comissao": "DESCONTO_ADIANT_COMISSAO",
            "churn_m0": "DESCONTO_CHURN_M0",
            "churn_m1": "DESCONTO_CHURN_M1",
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comissionamento"

        headers = [
            "VENDEDOR",
            "COMISSAO_BRUTA",
            "COMISSAO_LIQUIDA",
            *colunas_desconto,
            "TOTAL_DESCONTOS",
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        dinheiro_fmt = "R$ #,##0.00"

        for vendedor in vendedores:
            resumo = vendedor.get('resumo', {}) or {}
            bruto = Decimal(str(resumo.get('comissao_total_geral') or 0))
            liquido = Decimal(str(resumo.get('liquido') or 0))
            descontos_total = Decimal(str(resumo.get('total_descontos') or 0))

            descontos_por_coluna = {c: Decimal("0") for c in colunas_desconto}
            for d in (resumo.get('detalhes_descontos') or []):
                tipo = str((d.get('tipo_exibicao') or '')).strip().lower()
                coluna = tipo_para_coluna.get(tipo, "DESCONTO_OUTROS")
                valor = Decimal(str(d.get('valor') or 0))
                descontos_por_coluna[coluna] += valor

            ws.append([
                vendedor.get('vendedor_nome') or f"ID {vendedor.get('vendedor_id')}",
                float(bruto),
                float(liquido),
                float(descontos_por_coluna["DESCONTO_BOLETO"]),
                float(descontos_por_coluna["DESCONTO_ANTECIPACAO_INSTALACAO"]),
                float(descontos_por_coluna["DESCONTO_ADIANT_CNPJ"]),
                float(descontos_por_coluna["DESCONTO_ADIANT_COMISSAO"]),
                float(descontos_por_coluna["DESCONTO_CHURN_M0"]),
                float(descontos_por_coluna["DESCONTO_CHURN_M1"]),
                float(descontos_por_coluna["DESCONTO_OUTROS"]),
                float(descontos_total),
            ])

        primeira_linha_dados = 2
        ultima_linha_dados = ws.max_row
        linha_total = ultima_linha_dados + 1
        ws.cell(row=linha_total, column=1, value="TOTAL GERAL")
        ws.cell(row=linha_total, column=1).font = Font(bold=True)
        ws.cell(row=linha_total, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        if ultima_linha_dados >= primeira_linha_dados:
            for col in range(2, len(headers) + 1):
                letra = get_column_letter(col)
                ws.cell(row=linha_total, column=col, value=f"=SUM({letra}{primeira_linha_dados}:{letra}{ultima_linha_dados})")
                ws.cell(row=linha_total, column=col).font = Font(bold=True)
                ws.cell(row=linha_total, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=len(headers)):
            for cell in row:
                cell.number_format = dinheiro_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.column_dimensions["A"].width = 34
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"comissionamento_resumo_{mes:02d}_{ano}.xlsx"
        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
        return resp
    except Exception as e:
        logger.exception("exportar_comissionamento_resumo_excel: %s", e)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def enviar_resultado_campanha_whatsapp(request):
    try:
        campanha_id = request.data.get('campanha_id')
        vendedores_dados = request.data.get('vendedores_dados', [])
        
        if not campanha_id or not vendedores_dados:
            return Response({"error": "Dados incompletos."}, status=400)

        campanha = Campanha.objects.get(id=campanha_id)
        User = get_user_model()
        svc = WhatsAppService() 
        
        # 1. Carregar TODAS as regras da campanha ordenadas (Menor -> Maior)
        # Exemplo: [Meta 20, Meta 40, Meta 60]
        regras_ordenadas = list(campanha.regras_meta.all().order_by('meta'))
        
        # Define a meta máxima absoluta da campanha
        meta_maxima_absoluta = regras_ordenadas[-1].meta if regras_ordenadas else campanha.meta_vendas

        sucessos = 0
        erros = []
        periodo_str = f"{campanha.data_inicio.strftime('%d/%m')} até {campanha.data_fim.strftime('%d/%m')}"
        
        for item in vendedores_dados:
            vendedor_id = item.get('vendedor_id')
            try:
                vendedor = User.objects.get(id=vendedor_id)
                if not vendedor.tel_whatsapp:
                    erros.append(f"{vendedor.username}: Sem Zap.")
                    continue

                vendas_validas = int(item.get('vendas_validas', 0))
                
                # --- LÓGICA DE RECALCULO DE FAIXAS (Backend Source of Truth) ---
                meta_atual = 0
                premio_atual = 0.0
                proxima_meta_obj = None
                
                if regras_ordenadas:
                    # Percorre a escada para descobrir onde o vendedor está
                    for regra in regras_ordenadas:
                        if vendas_validas >= regra.meta:
                            # Bateu essa faixa, atualiza o status atual
                            meta_atual = regra.meta
                            premio_atual = float(regra.valor_premio)
                        else:
                            # Se as vendas são menores que esta regra, esta é a PRÓXIMA meta
                            proxima_meta_obj = regra
                            break # Encontramos o próximo degrau, paramos de procurar
                else:
                    # Fallback para campanhas legadas (sem faixas múltiplas)
                    if vendas_validas >= campanha.meta_vendas:
                        meta_atual = campanha.meta_vendas
                        premio_atual = float(campanha.valor_premio)
                    else:
                        # Se não bateu a única meta, ela vira a próxima
                        proxima_meta_obj = type('obj', (object,), {'meta': campanha.meta_vendas, 'valor_premio': campanha.valor_premio})

                # Formatador de Moeda
                def fmt_real(val): return f"R$ {float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                # --- CONSTRUÇÃO DO TEXTO INCENTIVADOR ---
                primeiro_nome = vendedor.first_name.split()[0].title() if vendedor.first_name else "Campeão(a)"
                
                msg = f"🚀 *PERFORMANCE: {campanha.nome}*\n"
                msg += f"🗓️ {periodo_str}\n\n"
                msg += f"Olá, *{primeiro_nome}*! Aqui está sua parcial:\n\n"
                msg += f"📊 *Vendas Válidas:* {vendas_validas}\n"

                # CENÁRIO 1: ESTÁ NO MEIO DA JORNADA (Bateu uma, falta outra)
                if premio_atual > 0 and proxima_meta_obj:
                    faltam = int(proxima_meta_obj.meta) - vendas_validas
                    diferenca_grana = float(proxima_meta_obj.valor_premio) - premio_atual
                    
                    msg += f"✅ *Meta Batida:* {meta_atual} vendas\n"
                    msg += f"💰 *JÁ GARANTIDO:* {fmt_real(premio_atual)}\n\n"
                    msg += f"🔥 *VOCÊ ESTÁ QUASE LÁ!*\n"
                    msg += f"Faltam só *{faltam} vendas* para o próximo nível ({proxima_meta_obj.meta})!\n"
                    msg += f"🚀 *Acelera!* Se bater essa meta, seu prêmio sobe para *{fmt_real(proxima_meta_obj.valor_premio)}*.\n"
                    msg += f"(Isso é *{fmt_real(diferenca_grana)} a mais* no seu bolso! 💵)"

                # CENÁRIO 2: LENDÁRIO (Bateu a última faixa disponível)
                elif premio_atual > 0 and not proxima_meta_obj:
                    msg += f"🏆 *LENDÁRIO! VOCÊ ZEROU A CAMPANHA!*\n"
                    msg += f"✅ Atingiu o topo máximo de {meta_atual} vendas.\n"
                    msg += f"💰 *PRÊMIO MÁXIMO:* {fmt_real(premio_atual)}\n\n"
                    msg += f"⭐ Parabéns pela performance incrível! Você é referência."

                # CENÁRIO 3: INICIANTE (Não bateu a primeira faixa ainda)
                else:
                    # A próxima meta é a primeira de todas
                    meta_alvo = proxima_meta_obj.meta if proxima_meta_obj else meta_maxima_absoluta
                    valor_alvo = proxima_meta_obj.valor_premio if proxima_meta_obj else 0.0
                    faltam = int(meta_alvo) - vendas_validas
                    
                    msg += f"⚠️ *Status:* Em busca da primeira meta!\n"
                    msg += f"🎯 *Alvo:* {meta_alvo} vendas\n"
                    msg += f"⚡ *FALTAM APENAS {faltam} VENDAS!*\n\n"
                    msg += f"💰 Bata essa meta e garanta *{fmt_real(valor_alvo)}*.\n"
                    msg += f"💪 *É totalmente possível!* Foque nos clientes pendentes e vamos buscar esse resultado!"

                msg += "\n\n_Atualizado em: " + timezone.localtime().strftime('%d/%m às %H:%M') + "_"

                # --- GERAÇÃO DA IMAGEM ---
                # Prepara os dados limpos para o gerador de imagem
                dados_card = {
                    'vendedor': primeiro_nome,
                    'vendas': vendas_validas,
                    'premio_atual': premio_atual,
                    'meta_atual': meta_atual,
                    'prox_meta': proxima_meta_obj.meta if proxima_meta_obj else None,
                    'prox_premio': proxima_meta_obj.valor_premio if proxima_meta_obj else None,
                    'campanha': campanha.nome
                }
                
                img_b64 = svc.gerar_card_campanha_b64(dados_card)

                # --- ENVIO ---
                if img_b64:
                    svc.enviar_imagem_b64(vendedor.tel_whatsapp, img_b64, caption=msg)
                else:
                    svc.enviar_mensagem_texto(vendedor.tel_whatsapp, msg)
                
                sucessos += 1
                
            except Exception as e:
                erros.append(f"Erro ID {vendedor_id}: {e}")
                logger.error(f"Erro envio campanha: {e}")

        return Response({"mensagem": f"Enviados: {sucessos}. Erros: {len(erros)}", "erros": erros})

    except Exception as e:
        return Response({"error": str(e)}, status=500)

# --- NOVA API DE PAINEL DE PERFORMANCE ---


def _perf_grupos_gestao():
    return ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']


def _aplicar_filtro_vendedor_ativo_perf(users_qs, user, vendedor_ativo_raw):
    """
    Gestão: query param vendedor_ativo=ativos|inativos|todos (default todos = inclui inativos).
    Demais perfis: sempre apenas usuários ativos.
    """
    if not is_member(user, _perf_grupos_gestao()):
        return users_qs.filter(is_active=True)
    raw = (vendedor_ativo_raw or 'todos').strip().lower()
    if raw == 'ativos':
        return users_qs.filter(is_active=True)
    if raw == 'inativos':
        return users_qs.filter(is_active=False)
    return users_qs


def _perf_grupos_export_excel():
    return ['Diretoria', 'Admin', 'BackOffice']


def _perf_grupos_filtro_data_livre():
    return ['Diretoria', 'Admin', 'BackOffice']


def _perf_ultimo_dia_mes(year, month):
    return date(year, month, calendar.monthrange(year, month)[1])


def _perf_parse_cluster_atual(val):
    if not val:
        return None
    s = str(val).strip().upper()
    if s == 'CLUSTER_1':
        return 1
    if s == 'CLUSTER_2':
        return 2
    if s == 'CLUSTER_3':
        return 3
    return None


def _perf_cluster_sugerido_por_soma(total):
    """Soma instaladas: mês do filtro + mês civil anterior. C1 >= 50, C2 entre 30 e 49, C3 < 30."""
    if total >= 50:
        return 1
    if total >= 30:
        return 2
    return 3


def _perf_movimento_cluster(atual_num, sug_num):
    if atual_num is None or sug_num is None:
        return '—', '—'
    ca, cs = 'C%d' % atual_num, 'C%d' % sug_num
    if sug_num < atual_num:
        return 'Sobe', '%s→%s' % (ca, cs)
    if sug_num > atual_num:
        return 'Desce', '%s→%s' % (ca, cs)
    return 'Mantém', '%s=%s' % (ca, cs)


def _perf_filtro_data_efetiva_instalacao_intervalo(d_ini, d_fim):
    return (
        (Q(vendas__data_instalacao_fisica__isnull=False)
         & Q(vendas__data_instalacao_fisica__gte=d_ini)
         & Q(vendas__data_instalacao_fisica__lte=d_fim))
        | (
            Q(vendas__data_instalacao_fisica__isnull=True)
            & Q(vendas__data_instalacao__gte=d_ini)
            & Q(vendas__data_instalacao__lte=d_fim)
        )
    )


def _filtro_data_efetiva_instalacao_intervalo_venda(d_ini, d_fim):
    return (
        (Q(data_instalacao_fisica__isnull=False)
         & Q(data_instalacao_fisica__gte=d_ini)
         & Q(data_instalacao_fisica__lte=d_fim))
        | (
            Q(data_instalacao_fisica__isnull=True)
            & Q(data_instalacao__gte=d_ini)
            & Q(data_instalacao__lte=d_fim)
        )
    )


def _perf_resolver_periodos_performance(request, user, hoje_d):
    livre = is_member(user, _perf_grupos_filtro_data_livre())
    aviso = None
    hoje_ref = hoje_d
    inicio_semana_def = hoje_d - timedelta(days=hoje_d.weekday())
    if livre:
        ss = (request.query_params.get('semana_segunda') or '').strip()
        if ss:
            try:
                seg = datetime.strptime(ss, '%Y-%m-%d').date()
                if seg.weekday() != 0:
                    seg = seg - timedelta(days=seg.weekday())
                inicio_semana = seg
            except ValueError:
                aviso = 'semana_segunda inválida; usando semana atual.'
                inicio_semana = inicio_semana_def
        else:
            inicio_semana = inicio_semana_def
    else:
        inicio_semana = inicio_semana_def
    fim_semana = inicio_semana + timedelta(days=5)
    dias_semana = [inicio_semana + timedelta(days=i) for i in range(6)]
    mr = (request.query_params.get('mes_referencia') or '').strip()
    y_m, m_m = hoje_d.year, hoje_d.month
    if livre:
        if mr:
            try:
                parts = mr.split('-')
                if len(parts) != 2:
                    raise ValueError
                y_m, m_m = int(parts[0]), int(parts[1])
                if m_m < 1 or m_m > 12:
                    raise ValueError
            except ValueError:
                aviso = aviso or 'mes_referencia inválida; usando mês atual.'
                y_m, m_m = hoje_d.year, hoje_d.month
    else:
        if mr:
            try:
                parts = mr.split('-')
                y_m, m_m = int(parts[0]), int(parts[1])
                cy, cm = hoje_d.year, hoje_d.month
                if cm == 1:
                    allowed = {(cy, cm), (cy - 1, 12)}
                else:
                    allowed = {(cy, cm), (cy, cm - 1)}
                if (y_m, m_m) not in allowed:
                    aviso = aviso or 'Somente mês atual ou anterior; usando mês atual.'
                    y_m, m_m = cy, cm
            except (ValueError, IndexError):
                y_m, m_m = hoje_d.year, hoje_d.month
    inicio_mes = date(y_m, m_m, 1)
    fim_mes = _perf_ultimo_dia_mes(y_m, m_m)
    return {
        'hoje_ref': hoje_ref,
        'inicio_semana': inicio_semana,
        'fim_semana': fim_semana,
        'dias_semana': dias_semana,
        'inicio_mes': inicio_mes,
        'fim_mes': fim_mes,
        'aviso': aviso,
        'mes_referencia': '%04d-%02d' % (y_m, m_m),
        'semana_segunda': inicio_semana.isoformat(),
    }


def _perf_add_months(base_month, delta_months):
    total = (base_month.year * 12 + (base_month.month - 1)) + delta_months
    y = total // 12
    m = (total % 12) + 1
    return date(y, m, 1)


def _perf_parse_mes_ref(valor, fallback):
    txt = (valor or '').strip()
    if not txt:
        return fallback
    try:
        y, m = txt.split('-')
        return date(int(y), int(m), 1)
    except (ValueError, TypeError):
        return fallback


def _perf_semana_bucket_dia_mes(dia_mes):
    if dia_mes <= 7:
        return 'S1'
    if dia_mes <= 14:
        return 'S2'
    if dia_mes <= 21:
        return 'S3'
    return 'S4'


def _perf_montar_payload_gestao(users, inicio_mes_ref, request):
    meses_hist = request.query_params.get('gestao_meses')
    try:
        meses_hist = int(meses_hist)
    except (TypeError, ValueError):
        meses_hist = 6
    meses_hist = max(1, min(12, meses_hist))

    mes_comp = _perf_parse_mes_ref(
        request.query_params.get('gestao_comp_mes'),
        _perf_add_months(inicio_mes_ref, -1),
    )
    tipo_metrica = (request.query_params.get('gestao_tipo') or 'BRUTA').strip().upper()
    if tipo_metrica not in ('BRUTA', 'INSTALADA'):
        tipo_metrica = 'BRUTA'

    meses_ref = [_perf_add_months(inicio_mes_ref, -idx) for idx in range(meses_hist)]
    todos_meses = meses_ref + [mes_comp]
    mes_min = min(todos_meses)
    mes_max = max(todos_meses)
    fim_mes_max = _perf_ultimo_dia_mes(mes_max.year, mes_max.month)

    filtro_base = (
        Q(vendas__ativo=True)
        & ~Q(vendas__ordem_servico='')
        & Q(vendas__ordem_servico__isnull=False)
        & Q(vendas__status_tratamento__nome__iexact='CADASTRADA')
    )
    filtro_cc = (
        Q(vendas__forma_pagamento__nome__icontains='CREDIT')
        | Q(vendas__forma_pagamento__nome__icontains='CRÉDIT')
        | (Q(vendas__forma_pagamento__nome__icontains='CARTA') & ~Q(vendas__forma_pagamento__nome__icontains='DEBIT'))
    )
    filtro_inst = Q(vendas__status_esteira__nome__iexact='INSTALADA')

    users_base = list(
        users.values('id', 'username', 'canal', 'cluster', 'meta_comissao').order_by('username')
    )
    user_index = {u['id']: u for u in users_base}

    mes_user_stats = {}
    for mref in meses_ref:
        fim_m = _perf_ultimo_dia_mes(mref.year, mref.month)
        filtro_ab = Q(vendas__data_abertura__date__gte=mref) & Q(vendas__data_abertura__date__lte=fim_m)
        filtro_inst_mes = _perf_filtro_data_efetiva_instalacao_intervalo(mref, fim_m)
        qs = users.annotate(
            total_vendas=Count('vendas', filter=filtro_base & filtro_ab),
            total_cc=Count('vendas', filter=filtro_base & filtro_ab & filtro_cc),
            instaladas=Count('vendas', filter=filtro_base & filtro_inst_mes & filtro_inst),
            instaladas_cc=Count('vendas', filter=filtro_base & filtro_inst_mes & filtro_inst & filtro_cc),
        ).values('id', 'total_vendas', 'total_cc', 'instaladas', 'instaladas_cc')
        mes_user_stats[mref.strftime('%Y-%m')] = {int(r['id']): r for r in qs}

    def _get_user_stat(ym, uid):
        return (mes_user_stats.get(ym) or {}).get(uid) or {'total_vendas': 0, 'total_cc': 0, 'instaladas': 0, 'instaladas_cc': 0}

    def _valor_principal(stat):
        return int(stat['instaladas'] or 0) if tipo_metrica == 'INSTALADA' else int(stat['total_vendas'] or 0)

    def _valor_cc(stat):
        return int(stat['instaladas_cc'] or 0) if tipo_metrica == 'INSTALADA' else int(stat['total_cc'] or 0)

    labels = []
    for idx, mref in enumerate(meses_ref):
        ym = mref.strftime('%Y-%m')
        labels.append({'key': ym, 'label': f"M-{idx}", 'mes': ym})

    ym_ref = inicio_mes_ref.strftime('%Y-%m')
    ym_comp = mes_comp.strftime('%Y-%m')

    rows_vendedor = []
    for u in users_base:
        serie = []
        for lb in labels:
            st = _get_user_stat(lb['key'], u['id'])
            serie.append(_valor_principal(st))
        st_ref = _get_user_stat(ym_ref, u['id'])
        st_comp = _get_user_stat(ym_comp, u['id'])
        total_ref = _valor_principal(st_ref)
        total_comp = _valor_principal(st_comp)
        cc_ref = _valor_cc(st_ref)
        inst_ref = int(st_ref['instaladas'] or 0)
        var_abs = total_ref - total_comp
        var_pct = (var_abs / total_comp * 100.0) if total_comp > 0 else (100.0 if total_ref > 0 else 0.0)
        pct_cc_ref = (cc_ref / total_ref * 100.0) if total_ref > 0 else 0.0
        aprov_ref = (inst_ref / total_ref * 100.0) if total_ref > 0 else 0.0
        meta = int(u.get('meta_comissao') or 0)
        ating = (total_ref / meta * 100.0) if meta > 0 else 0.0
        rows_vendedor.append({
            'grupo': u['username'].upper(),
            'canal': u.get('canal') or '-',
            'cluster': u.get('cluster') or '-',
            'meta': meta,
            'serie': serie,
            'ref_total': total_ref,
            'comp_total': total_comp,
            'var_abs': var_abs,
            'var_pct': round(var_pct, 2),
            'pct_cc_ref': round(pct_cc_ref, 2),
            'aprov_ref': round(aprov_ref, 2),
            'atingimento_meta': round(ating, 2),
        })

    agrup_canal_cluster = {}
    for u in users_base:
        key = f"{(u.get('canal') or '-').upper()}|{(u.get('cluster') or '-').upper()}"
        if key not in agrup_canal_cluster:
            agrup_canal_cluster[key] = {
                'canal': (u.get('canal') or '-').upper(),
                'cluster': (u.get('cluster') or '-').upper(),
                'meta': 0,
                'uid': [],
            }
        agrup_canal_cluster[key]['meta'] += int(u.get('meta_comissao') or 0)
        agrup_canal_cluster[key]['uid'].append(int(u['id']))

    rows_canal_cluster = []
    for _, grp in sorted(agrup_canal_cluster.items(), key=lambda item: (item[1]['canal'], item[1]['cluster'])):
        serie = []
        for lb in labels:
            ym = lb['key']
            total = sum(_valor_principal(_get_user_stat(ym, uid)) for uid in grp['uid'])
            serie.append(total)
        ref_total = sum(_valor_principal(_get_user_stat(ym_ref, uid)) for uid in grp['uid'])
        comp_total = sum(_valor_principal(_get_user_stat(ym_comp, uid)) for uid in grp['uid'])
        cc_ref = sum(_valor_cc(_get_user_stat(ym_ref, uid)) for uid in grp['uid'])
        inst_ref = sum(int(_get_user_stat(ym_ref, uid)['instaladas'] or 0) for uid in grp['uid'])
        var_abs = ref_total - comp_total
        var_pct = (var_abs / comp_total * 100.0) if comp_total > 0 else (100.0 if ref_total > 0 else 0.0)
        pct_cc_ref = (cc_ref / ref_total * 100.0) if ref_total > 0 else 0.0
        aprov_ref = (inst_ref / ref_total * 100.0) if ref_total > 0 else 0.0
        meta = int(grp['meta'] or 0)
        ating = (ref_total / meta * 100.0) if meta > 0 else 0.0
        rows_canal_cluster.append({
            'grupo': f"{grp['canal']} / {grp['cluster']}",
            'canal': grp['canal'],
            'cluster': grp['cluster'],
            'meta': meta,
            'serie': serie,
            'ref_total': ref_total,
            'comp_total': comp_total,
            'var_abs': var_abs,
            'var_pct': round(var_pct, 2),
            'pct_cc_ref': round(pct_cc_ref, 2),
            'aprov_ref': round(aprov_ref, 2),
            'atingimento_meta': round(ating, 2),
        })

    base_semanais = (
        Venda.objects.filter(
            ativo=True,
            vendedor_id__in=list(user_index.keys()),
        )
        .exclude(ordem_servico__isnull=True)
        .exclude(ordem_servico='')
        .filter(status_tratamento__nome__iexact='CADASTRADA')
        .select_related('forma_pagamento')
        .only('data_abertura', 'data_instalacao', 'data_instalacao_fisica', 'status_esteira__nome', 'forma_pagamento__nome')
    )
    if tipo_metrica == 'INSTALADA':
        base_semanais = base_semanais.filter(status_esteira__nome__iexact='INSTALADA')
    vendas_semanais = base_semanais
    semanas = {
        'ref': {'S1': {'total': 0, 'cc': 0}, 'S2': {'total': 0, 'cc': 0}, 'S3': {'total': 0, 'cc': 0}, 'S4': {'total': 0, 'cc': 0}},
        'comp': {'S1': {'total': 0, 'cc': 0}, 'S2': {'total': 0, 'cc': 0}, 'S3': {'total': 0, 'cc': 0}, 'S4': {'total': 0, 'cc': 0}},
    }
    for v in vendas_semanais:
        if tipo_metrica == 'INSTALADA':
            d_ab = v.data_instalacao_fisica or v.data_instalacao
        else:
            if not v.data_abertura:
                continue
            d_ab = timezone.localtime(v.data_abertura).date()
        if not d_ab:
            continue
        if d_ab < mes_min or d_ab > fim_mes_max:
            continue
        ym = d_ab.strftime('%Y-%m')
        alvo = 'ref' if ym == ym_ref else ('comp' if ym == ym_comp else None)
        if not alvo:
            continue
        bucket = _perf_semana_bucket_dia_mes(d_ab.day)
        semanas[alvo][bucket]['total'] += 1
        nome_fp = (getattr(getattr(v, 'forma_pagamento', None), 'nome', '') or '').upper()
        if ('CREDIT' in nome_fp) or ('CRÉDIT' in nome_fp) or (('CARTA' in nome_fp) and ('DEBIT' not in nome_fp)):
            semanas[alvo][bucket]['cc'] += 1

    semanal_cmp = []
    for b in ['S1', 'S2', 'S3', 'S4']:
        t_ref = semanas['ref'][b]['total']
        t_comp = semanas['comp'][b]['total']
        v_abs = t_ref - t_comp
        v_pct = (v_abs / t_comp * 100.0) if t_comp > 0 else (100.0 if t_ref > 0 else 0.0)
        semanal_cmp.append({
            'semana': b,
            'ref_total': t_ref,
            'comp_total': t_comp,
            'var_abs': v_abs,
            'var_pct': round(v_pct, 2),
        })

    return {
        'tipo_metrica': tipo_metrica,
        'meses_historico': meses_hist,
        'mes_ref': ym_ref,
        'mes_comp': ym_comp,
        'labels': labels,
        'rows_vendedor': rows_vendedor,
        'rows_canal_cluster': rows_canal_cluster,
        'semanal_comparativo': semanal_cmp,
    }


class PainelPerformanceView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        User = get_user_model()
        user = request.user
        agora_local = timezone.localtime(timezone.now())
        hoje = agora_local.date()
        periodo = _perf_resolver_periodos_performance(request, user, hoje)
        hoje_ref = periodo['hoje_ref']
        inicio_semana = periodo['inicio_semana']
        fim_semana = periodo['fim_semana']
        dias_semana = periodo['dias_semana']
        inicio_mes = periodo['inicio_mes']
        fim_mes = periodo['fim_mes']

        users = User.objects.exclude(username__in=['OSAB_IMPORT', 'admin', 'root'])
        grupos_gestao = _perf_grupos_gestao()
        if is_member(user, grupos_gestao):
            pass
        elif is_member(user, ['Supervisor']):
            users = users.filter(Q(supervisor=user) | Q(id=user.id))
        else:
            users = users.filter(id=user.id)

        filtro_canal = request.query_params.get('canal')
        if filtro_canal:
            users = users.filter(canal__iexact=filtro_canal)
        filtro_cluster = request.query_params.get('cluster')
        if filtro_cluster:
            users = users.filter(cluster=filtro_cluster)

        users = _aplicar_filtro_vendedor_ativo_perf(users, user, request.query_params.get('vendedor_ativo'))

        filtro_os_sem_reemissao = (
            Q(vendas__ativo=True)
            & ~Q(vendas__ordem_servico='')
            & Q(vendas__ordem_servico__isnull=False)
            & Q(vendas__status_tratamento__nome__iexact='CADASTRADA')
            & Q(vendas__reemissao=False)
        )
        filtro_os_com_reemissao = (
            Q(vendas__ativo=True)
            & ~Q(vendas__ordem_servico='')
            & Q(vendas__ordem_servico__isnull=False)
            & Q(vendas__status_tratamento__nome__iexact='CADASTRADA')
        )
        filtro_cc = (
            Q(vendas__forma_pagamento__nome__icontains='CREDIT')
            | Q(vendas__forma_pagamento__nome__icontains='CRÉDIT')
            | (Q(vendas__forma_pagamento__nome__icontains='CARTA') & ~Q(vendas__forma_pagamento__nome__icontains='DEBIT'))
        )
        filtro_inst = Q(vendas__status_esteira__nome__iexact='INSTALADA')
        filtro_abertura_no_mes = (
            Q(vendas__data_abertura__date__gte=inicio_mes)
            & Q(vendas__data_abertura__date__lte=fim_mes)
        )
        # Instaladas no mês: data física se preenchida, senão data instalação OSAB
        filtro_data_inst_mes = _perf_filtro_data_efetiva_instalacao_intervalo(inicio_mes, fim_mes)

        filtro_semana_abertura = (
            Q(vendas__data_abertura__date__gte=inicio_semana)
            & Q(vendas__data_abertura__date__lte=fim_semana)
        )

        qs_hoje = users.annotate(
            vendas_total=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=hoje_ref)),
            vendas_cc=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=hoje_ref) & filtro_cc)
        ).values('username', 'canal', 'cluster', 'vendas_total', 'vendas_cc').order_by('username')

        lista_hoje = []
        for u in qs_hoje:
            total = u['vendas_total']
            cc = u['vendas_cc']
            pct = (cc / total * 100) if total > 0 else 0
            lista_hoje.append({
                'vendedor': u['username'].upper(),
                'canal': u['canal'],
                'cluster': u.get('cluster', '-'),
                'total': total,
                'cc': cc,
                'pct_cc': round(pct, 2)
            })

        qs_semana = users.annotate(
            seg=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[0])),
            ter=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[1])),
            qua=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[2])),
            qui=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[3])),
            sex=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[4])),
            sab=Count('vendas', filter=filtro_os_sem_reemissao & Q(vendas__data_abertura__date=dias_semana[5])),
            total_semana=Count('vendas', filter=filtro_os_sem_reemissao & filtro_semana_abertura),
            total_cc=Count('vendas', filter=filtro_os_sem_reemissao & filtro_semana_abertura & filtro_cc)
        ).values('username', 'cluster', 'seg', 'ter', 'qua', 'qui', 'sex', 'sab', 'total_semana', 'total_cc').order_by('username')

        lista_semana = []
        for u in qs_semana:
            total = u['total_semana']
            pct = (u['total_cc'] / total * 100) if total > 0 else 0
            lista_semana.append({
                'vendedor': u['username'].upper(),
                'cluster': u.get('cluster', '-'),
                'dias': [u['seg'], u['ter'], u['qua'], u['qui'], u['sex'], u['sab']],
                'total': total,
                'cc': u['total_cc'],
                'pct_cc': round(pct, 2)
            })

        qs_mes = users.annotate(
            total_vendas=Count('vendas', filter=filtro_os_com_reemissao & filtro_abertura_no_mes),
            instaladas=Count('vendas', filter=filtro_os_com_reemissao & filtro_data_inst_mes & filtro_inst),
            total_cc=Count('vendas', filter=filtro_os_com_reemissao & filtro_abertura_no_mes & filtro_cc),
            instaladas_cc=Count('vendas', filter=filtro_os_com_reemissao & filtro_data_inst_mes & filtro_inst & filtro_cc),
            pendenciadas=Count('vendas', filter=filtro_os_com_reemissao & filtro_abertura_no_mes & Q(vendas__status_esteira__nome__icontains='PENDEN')),
            agendadas=Count('vendas', filter=filtro_os_com_reemissao & filtro_abertura_no_mes & Q(vendas__status_esteira__nome__iexact='AGENDADO')),
            canceladas=Count('vendas', filter=filtro_os_com_reemissao & filtro_abertura_no_mes & Q(vendas__status_esteira__nome__icontains='CANCELAD'))
        ).values(
            'username', 'cluster', 'total_vendas', 'instaladas', 'total_cc', 'instaladas_cc', 'pendenciadas', 'agendadas', 'canceladas'
        ).order_by('username')

        # Aval. Cluster: soma instaladas no MÊS DO FILTRO + no MÊS CIVIL IMEDIATAMENTE ANTERIOR
        # (ex.: filtro mar/2026 → fev/2026 + mar/2026; não jan+fev)
        m_ref_start = inicio_mes
        m_ref_end = fim_mes
        m_ant_end = m_ref_start - timedelta(days=1)
        m_ant_start = m_ant_end.replace(day=1)
        base_inst_m = filtro_os_com_reemissao & filtro_inst
        f_mes_ref = base_inst_m & _perf_filtro_data_efetiva_instalacao_intervalo(m_ref_start, m_ref_end)
        f_mes_ant = base_inst_m & _perf_filtro_data_efetiva_instalacao_intervalo(m_ant_start, m_ant_end)
        qs_cluster = users.annotate(
            inst_aval_mes_ref=Count('vendas', filter=f_mes_ref),
            inst_aval_mes_ant=Count('vendas', filter=f_mes_ant),
        ).values('username', 'cluster', 'inst_aval_mes_ref', 'inst_aval_mes_ant')
        map_cluster = {r['username']: r for r in qs_cluster}

        lista_mes = []
        for u in qs_mes:
            tot = u['total_vendas']
            inst = u['instaladas']
            pct_cc_total = (u['total_cc'] / tot * 100) if tot > 0 else 0
            pct_cc_inst = (u['instaladas_cc'] / inst * 100) if inst > 0 else 0
            aproveitamento = (inst / tot * 100) if tot > 0 else 0
            nome_display = u['username']
            ev = map_cluster.get(nome_display) or {}
            i_ref = int(ev.get('inst_aval_mes_ref') or 0)
            i_ant = int(ev.get('inst_aval_mes_ant') or 0)
            soma_inst = i_ant + i_ref
            sug = _perf_cluster_sugerido_por_soma(soma_inst)
            atual = _perf_parse_cluster_atual(u.get('cluster'))
            mov, trans = _perf_movimento_cluster(atual, sug)
            if trans == '—' and atual is not None:
                aval_txt = '%s (soma %d)' % (mov, soma_inst)
            elif trans == '—':
                aval_txt = 'soma %d → sugerido C%d' % (soma_inst, sug)
            else:
                aval_txt = '%s (%s) soma %d' % (mov, trans, soma_inst)

            lista_mes.append({
                'vendedor': nome_display.upper(),
                'cluster': u.get('cluster', '-'),
                'total': tot,
                'instaladas': inst,
                'cc_total': u['total_cc'],
                'cc_inst': u['instaladas_cc'],
                'pct_cc_total': round(pct_cc_total, 2),
                'pct_cc_inst': round(pct_cc_inst, 2),
                'aproveitamento': round(aproveitamento, 2),
                'pend': u['pendenciadas'],
                'agend': u['agendadas'],
                'canc': u['canceladas'],
                'avaliacao_cluster': aval_txt,
                'soma_instaladas_m1_m2': soma_inst,
                'aval_inst_mes_anterior': i_ant,
                'aval_inst_mes_referencia': i_ref,
                'cluster_sugerido': sug,
            })

        total_hoje = sum(i['total'] for i in lista_hoje)
        total_semana = sum(i['total'] for i in lista_semana)
        total_mes = sum(i['total'] for i in lista_mes)
        payload_gestao = _perf_montar_payload_gestao(users, inicio_mes, request)

        payload = {
            'hoje': lista_hoje,
            'semana': lista_semana,
            'mes': lista_mes,
            'gestao': payload_gestao,
            'totais': {'hoje': total_hoje, 'semana': total_semana, 'mes': total_mes},
            'periodo': {
                'mes_referencia': periodo['mes_referencia'],
                'semana_segunda': periodo['semana_segunda'],
                'semana_fim': fim_semana.isoformat(),
            },
            'pode_exportar_excel': is_member(user, _perf_grupos_export_excel()),
        }
        if periodo.get('aviso'):
            payload['aviso'] = periodo['aviso']
        return Response(payload)


# --- CORREÇÃO APLICADA: VIEW DE PÁGINA NORMAL (SEM API_VIEW) ---
def page_painel_performance(request):
    return render(request, 'painel_performance.html')
class ExportarPerformanceExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        User = get_user_model()
        user = request.user
        if not is_member(user, _perf_grupos_export_excel()):
            return Response(
                {'detail': 'Exportação permitida apenas para BackOffice, Diretoria e Admin.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        agora_local = timezone.localtime(timezone.now())
        hoje = agora_local.date()
        p = _perf_resolver_periodos_performance(request, user, hoje)
        hoje_ref = p['hoje_ref']
        inicio_semana = p['inicio_semana']
        fim_semana = p['fim_semana']
        inicio_mes = p['inicio_mes']
        fim_mes = p['fim_mes']

        # 2. Mesmo conjunto de vendedores do PainelPerformanceView (evita vendas de usuário inativo/bot fora do painel)
        grupos_gestao = _perf_grupos_gestao()
        users_export = User.objects.exclude(username__in=['OSAB_IMPORT', 'admin', 'root'])
        if is_member(user, grupos_gestao):
            pass
        elif is_member(user, ['Supervisor']):
            users_export = users_export.filter(Q(supervisor=user) | Q(id=user.id))
        else:
            users_export = users_export.filter(id=user.id)

        filtro_canal = request.query_params.get('canal')
        if filtro_canal:
            users_export = users_export.filter(canal__iexact=filtro_canal)
        filtro_cluster = request.query_params.get('cluster')
        if filtro_cluster:
            users_export = users_export.filter(cluster=filtro_cluster)

        users_export = _aplicar_filtro_vendedor_ativo_perf(users_export, user, request.query_params.get('vendedor_ativo'))

        vendas = (
            Venda.objects.filter(ativo=True, vendedor_id__in=users_export.values_list('id', flat=True))
            .select_related('vendedor', 'cliente', 'plano', 'forma_pagamento', 'status_esteira')
        )

        # Período selecionado no botão (define nome do arquivo e prioridade das abas)
        periodo = (request.query_params.get('periodo') or 'HOJE').strip().upper()
        if periodo not in ('HOJE', 'SEMANAL', 'MENSAL', 'GESTAO'):
            periodo = 'HOJE'

        if periodo == 'GESTAO':
            payload_gestao = _perf_montar_payload_gestao(users_export, inicio_mes, request)
            visao = (request.query_params.get('gestao_visao') or 'VENDEDOR').strip().upper()
            linhas_base = payload_gestao['rows_canal_cluster'] if visao == 'CANAL_CLUSTER' else payload_gestao['rows_vendedor']
            labels = payload_gestao['labels']
            dados_gestao = []
            for r in linhas_base:
                linha = {
                    'Grupo': r.get('grupo'),
                    'Canal': r.get('canal'),
                    'Cluster': r.get('cluster'),
                    'Meta': r.get('meta', 0),
                    'Mês Ref': payload_gestao.get('mes_ref'),
                    'Total Ref': r.get('ref_total', 0),
                    'Mês Comparação': payload_gestao.get('mes_comp'),
                    'Total Comparação': r.get('comp_total', 0),
                    'Var. Abs': r.get('var_abs', 0),
                    'Var. %': r.get('var_pct', 0),
                    '% CC Ref': r.get('pct_cc_ref', 0),
                    'Aprov. Ref %': r.get('aprov_ref', 0),
                    'Atingimento Meta %': r.get('atingimento_meta', 0),
                }
                for idx, lb in enumerate(labels):
                    linha[lb['label']] = (r.get('serie') or [])[idx] if idx < len(r.get('serie') or []) else 0
                dados_gestao.append(linha)

            dados_semanal = []
            for s in payload_gestao.get('semanal_comparativo') or []:
                dados_semanal.append({
                    'Semana': s.get('semana'),
                    f"Total {payload_gestao.get('mes_ref')}": s.get('ref_total', 0),
                    f"Total {payload_gestao.get('mes_comp')}": s.get('comp_total', 0),
                    'Var. Abs': s.get('var_abs', 0),
                    'Var. %': s.get('var_pct', 0),
                })

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pd.DataFrame(dados_gestao or [{'Status': 'Sem dados para visão selecionada'}]).to_excel(
                    writer, sheet_name='Gestao', index=False
                )
                pd.DataFrame(dados_semanal or [{'Status': 'Sem dados semanais para comparação'}]).to_excel(
                    writer, sheet_name='Gestao Semanas', index=False
                )
            output.seek(0)
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Performance_Gestao_{hoje_ref}.xlsx"'
            return response

        # 3. Preparar os 3 DataFrames (consultores veem data efetiva na coluna Data Instalação)
        grupos_gestao = _perf_grupos_gestao()
        use_data_efetiva = not is_member(user, grupos_gestao)

        # Regra de reemissão:
        # - Hoje e Semana: exclui reemissão
        # - Mês: inclui reemissão
        vendas_sem_reemissao = vendas.filter(reemissao=False)

        # --- ABA 1: HOJE ---
        vendas_hoje = vendas_sem_reemissao.filter(data_criacao__date=hoje_ref)
        dados_hoje = self._montar_dados(vendas_hoje, use_data_efetiva)
        
        # --- ABA 2: SEMANA (segunda a sábado) ---
        vendas_semana = vendas_sem_reemissao.filter(
            data_criacao__date__gte=inicio_semana,
            data_criacao__date__lte=fim_semana,
        )
        dados_semana = self._montar_dados(vendas_semana, use_data_efetiva)
        
        # --- ABA 3: MÊS ---
        filtro_instalacao_mes = (
            (Q(data_instalacao_fisica__isnull=False) & Q(data_instalacao_fisica__gte=inicio_mes) & Q(data_instalacao_fisica__lte=fim_mes))
            | (Q(data_instalacao_fisica__isnull=True) & Q(data_instalacao__gte=inicio_mes) & Q(data_instalacao__lte=fim_mes))
        )

        vendas_mes = vendas.filter(
            Q(data_criacao__date__gte=inicio_mes, data_criacao__date__lte=fim_mes)
            | (Q(status_esteira__nome__iexact='INSTALADA') & filtro_instalacao_mes)
        ).distinct()
        dados_mes = self._montar_dados(vendas_mes, use_data_efetiva, inicio_mes=inicio_mes, fim_mes=fim_mes)

        # 4. Gerar o Excel
        planilhas = [
            ('Hoje', dados_hoje),
            ('Semana Atual', dados_semana),
            ('Mês Atual', dados_mes),
        ]
        ordem_periodo = {'HOJE': 'Hoje', 'SEMANAL': 'Semana Atual', 'MENSAL': 'Mês Atual'}
        aba_prioritaria = ordem_periodo[periodo]
        planilhas = sorted(planilhas, key=lambda item: 0 if item[0] == aba_prioritaria else 1)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for nome_aba, dados_aba in planilhas:
                pd.DataFrame(dados_aba).to_excel(writer, sheet_name=nome_aba, index=False)
            
        output.seek(0)
        
        # 5. Retornar Arquivo
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Performance_{periodo.title()}_{hoje_ref}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _montar_dados(self, queryset, use_data_efetiva=False, inicio_mes=None, fim_mes=None):
        dados = []
        for v in queryset:
            # Converter para horário local para validação visual
            dt_criacao_local = timezone.localtime(v.data_criacao).strftime('%d/%m/%Y %H:%M:%S') if v.data_criacao else '-'
            # Consultores: data efetiva (física se preenchida); gestão: data OSAB
            dt_inst = (v.data_instalacao_fisica or v.data_instalacao) if use_data_efetiva else v.data_instalacao
            dt_inst_str = dt_inst.strftime('%d/%m/%Y') if dt_inst else '-'
            linha = {
                'ID Venda': v.id,
                'Data Criação (Local)': dt_criacao_local,
                'Vendedor': v.vendedor.username.upper() if v.vendedor else '-',
                'Canal': v.vendedor.canal if v.vendedor else '-',
                'Cliente': v.cliente.nome_razao_social if v.cliente else '-',
                'CPF/CNPJ': v.cliente.cpf_cnpj if v.cliente else '-',
                'Plano': v.plano.nome if v.plano else '-',
                'Forma Pagamento': v.forma_pagamento.nome if v.forma_pagamento else '-',
                'Status Esteira': v.status_esteira.nome if v.status_esteira else '-',
                'Data Instalação': dt_inst_str,
                'Data Física (no cliente)': v.data_instalacao_fisica.strftime('%d/%m/%Y') if v.data_instalacao_fisica else '-',
                'OS': v.ordem_servico or '-',
                'Reemissão': 'Sim' if getattr(v, 'reemissao', False) else 'Não',
            }
            if inicio_mes:
                fim_recorte = fim_mes or inicio_mes
                data_criacao_local_date = timezone.localtime(v.data_criacao).date() if v.data_criacao else None
                instalada_no_mes = bool(
                    dt_inst and inicio_mes <= dt_inst <= fim_recorte
                    and v.status_esteira
                    and str(v.status_esteira.nome or '').strip().upper() == 'INSTALADA'
                )
                if data_criacao_local_date and inicio_mes <= data_criacao_local_date <= fim_recorte:
                    origem_mes = 'Venda do mês'
                elif instalada_no_mes:
                    origem_mes = 'Instalada no mês (venda anterior)'
                else:
                    origem_mes = 'Fora do recorte mensal'
                linha['Origem Mensal'] = origem_mes
            dados.append(linha)
        if not dados:
            return [{'Status': 'Sem vendas neste período'}]
        return dados
    
# 1. API para listar/criar grupos na tela
class GrupoDisparoViewSet(viewsets.ModelViewSet):
    queryset = GrupoDisparo.objects.filter(ativo=True)
    serializer_class = GrupoDisparoSerializer
    permission_classes = [permissions.IsAuthenticated]

# 2. API que recebe a Imagem (Base64) e manda pro Z-API
# Ou gera a imagem no servidor (gerar_server=True) para unificar o estilo com o robô automático
class EnviarImagemPerformanceView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        chat_id = request.data.get('chat_id')
        imagem_b64 = request.data.get('imagem_b64')
        titulo = request.data.get('titulo', 'Performance')
        gerar_server = request.data.get('gerar_server', False)

        if not chat_id:
            return Response({'error': 'Destino (chat_id) é obrigatório.'}, status=400)

        # Aceitar múltiplos destinos (vírgula ou ponto e vírgula)
        raw = str(chat_id).replace(";", ",")
        destinos = [d.strip() for d in raw.split(",") if d.strip()]

        # Modo 1: Gerar imagem no servidor (Pillow) - unifica estilo com robô automático
        if gerar_server:
            try:
                imagem_b64 = self._gerar_imagem_servidor(request)
                if not imagem_b64:
                    return Response({'error': 'Falha ao gerar imagem.'}, status=500)
                titulo = request.data.get('titulo', titulo)
            except Exception as e:
                return Response({'error': str(e)}, status=500)
        else:
            # Modo 2: Usar imagem enviada (html2canvas)
            if not imagem_b64:
                return Response({'error': 'Envie imagem_b64 ou use gerar_server=true.'}, status=400)
            if "base64," in imagem_b64:
                imagem_b64 = imagem_b64.split("base64,")[1]

        try:
            svc = WhatsAppService()
            caption = f"📊 *{titulo}* \nGerado em: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
            sucessos = 0
            erros = []
            for dest in destinos:
                try:
                    resp = svc.enviar_imagem_base64_direto(dest, imagem_b64, caption)
                    if resp:
                        sucessos += 1
                    else:
                        erros.append(f"{dest}: Z-API não confirmou")
                except Exception as e:
                    erros.append(f"{dest}: {str(e)[:80]}")
            if sucessos == 0 and erros:
                return Response({'error': '; '.join(erros[:3])}, status=500)
            return Response({
                'status': 'sucesso',
                'enviados': sucessos,
                'total': len(destinos),
                'erros': erros[:5] if erros else []
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    def _gerar_imagem_servidor(self, request):
        """Gera imagem com Pillow usando os mesmos dados e layout do robô automático."""
        from types import SimpleNamespace
        from crm_app.tasks import _filtro_cc
        User = get_user_model()
        user = request.user
        canal = request.data.get('canal', '')
        cluster = request.data.get('cluster', '')
        tipo = (request.data.get('tipo') or 'HOJE').upper()
        if tipo not in ('HOJE', 'SEMANAL', 'MENSAL'):
            tipo = 'HOJE'

        qpd = {}
        for k in ('semana_segunda', 'mes_referencia'):
            v = request.data.get(k)
            if v is not None and str(v).strip() != '':
                qpd[k] = v

        class _PerfQP:
            def __init__(self, d):
                self._d = d

            def get(self, key, default=None):
                return self._d.get(key, default)

        fake_req = SimpleNamespace(query_params=_PerfQP(qpd))
        agora_local = timezone.localtime(timezone.now())
        hoje = agora_local.date()
        p = _perf_resolver_periodos_performance(fake_req, user, hoje)
        hoje_ref = p['hoje_ref']
        inicio_semana = p['inicio_semana']
        fim_semana = p['fim_semana']
        inicio_mes = p['inicio_mes']
        fim_mes = p['fim_mes']

        users = User.objects.exclude(username__in=['OSAB_IMPORT', 'admin', 'root'])
        grupos_gestao = _perf_grupos_gestao()
        if is_member(user, grupos_gestao):
            pass
        elif is_member(user, ['Supervisor']):
            users = users.filter(Q(supervisor=user) | Q(id=user.id))
        else:
            users = users.filter(id=user.id)

        if canal and str(canal).upper() != 'TODOS':
            users = users.filter(canal__iexact=canal)
        if cluster and str(cluster).strip():
            users = users.filter(cluster__iexact=cluster.strip())

        users = _aplicar_filtro_vendedor_ativo_perf(users, user, request.data.get('vendedor_ativo'))

        filtro_os_sem_reemissao = (
            Q(vendas__ativo=True)
            & ~Q(vendas__ordem_servico='')
            & Q(vendas__ordem_servico__isnull=False)
            & Q(vendas__status_tratamento__nome__iexact='CADASTRADA')
            & Q(vendas__reemissao=False)
        )
        filtro_os_com_reemissao = (
            Q(vendas__ativo=True)
            & ~Q(vendas__ordem_servico='')
            & Q(vendas__ordem_servico__isnull=False)
            & Q(vendas__status_tratamento__nome__iexact='CADASTRADA')
        )
        filtro_cc = _filtro_cc()
        filtro_inst = Q(vendas__status_esteira__nome__iexact='INSTALADA')

        lista_dados = []
        t_total = 0
        t_cc = 0
        t_instaladas = 0
        titulo_extra = " Hoje" if tipo == "HOJE" else (" Semanal" if tipo == "SEMANAL" else " Mensal")

        if tipo == 'HOJE':
            filtro = filtro_os_sem_reemissao & Q(vendas__data_abertura__date=hoje_ref)
            qs = users.annotate(
                total=Count('vendas', filter=filtro),
                cc=Count('vendas', filter=filtro & filtro_cc)
            ).order_by('username')
            for u in qs:
                pct = int((u.cc / u.total * 100)) if u.total > 0 else 0
                lista_dados.append({
                    'nome': u.username.upper(),
                    'cluster': getattr(u, 'cluster', None) or '-',
                    'canal': getattr(u, 'canal', None) or '-',
                    'total': u.total,
                    'cc': u.cc,
                    'pct': f"{pct}%"
                })
                t_total += u.total
                t_cc += u.cc
        elif tipo == 'SEMANAL':
            filtro_sem = (
                Q(vendas__data_abertura__date__gte=inicio_semana)
                & Q(vendas__data_abertura__date__lte=fim_semana)
            )
            qs = users.annotate(
                total_semana=Count('vendas', filter=filtro_os_sem_reemissao & filtro_sem),
                total_cc=Count('vendas', filter=filtro_os_sem_reemissao & filtro_sem & filtro_cc)
            ).order_by('username').values('username', 'cluster', 'canal', 'total_semana', 'total_cc')
            for u in qs:
                tot = u['total_semana']
                cc = u['total_cc']
                pct = int((cc / tot * 100)) if tot > 0 else 0
                lista_dados.append({
                    'nome': u['username'].upper(),
                    'cluster': u.get('cluster') or '-',
                    'canal': u.get('canal') or '-',
                    'total': tot,
                    'cc': cc,
                    'pct': f"{pct}%"
                })
                t_total += tot
                t_cc += cc
        else:  # MENSAL
            filtro_ab_m = (
                Q(vendas__data_abertura__date__gte=inicio_mes)
                & Q(vendas__data_abertura__date__lte=fim_mes)
            )
            filtro_inst_m = _perf_filtro_data_efetiva_instalacao_intervalo(inicio_mes, fim_mes)
            qs = users.annotate(
                total_vendas=Count('vendas', filter=filtro_os_com_reemissao & filtro_ab_m),
                instaladas=Count('vendas', filter=filtro_os_com_reemissao & filtro_inst_m & filtro_inst),
                total_cc=Count('vendas', filter=filtro_os_com_reemissao & filtro_ab_m & filtro_cc)
            ).order_by('username').values('username', 'cluster', 'canal', 'total_vendas', 'instaladas', 'total_cc')
            for u in qs:
                tot = u['total_vendas']
                inst = u['instaladas']
                cc = u['total_cc']
                pct = int((cc / tot * 100)) if tot > 0 else 0
                aprov = int((inst / tot * 100)) if tot > 0 else 0
                lista_dados.append({
                    'nome': u['username'].upper(),
                    'cluster': u.get('cluster') or '-',
                    'canal': u.get('canal') or '-',
                    'total': tot,
                    'instaladas': inst,
                    'aprov': f"{aprov}%",
                    'cc': cc,
                    'pct': f"{pct}%"
                })
                t_total += tot
                t_cc += cc
                t_instaladas += inst

        pct_geral = int((t_cc / t_total * 100)) if t_total > 0 else 0
        aprov_geral = int((t_instaladas / t_total * 100)) if t_total > 0 and tipo == 'MENSAL' else 0
        payload = {
            'titulo': f"Performance -{titulo_extra.strip()}",
            'data': hoje_ref.strftime('%d/%m/%Y'),
            'lista': lista_dados,
            'totais': {
                'total': t_total,
                'cc': t_cc,
                'pct': f"{pct_geral}%",
                'instaladas': t_instaladas if tipo == 'MENSAL' else None,
                'aprov': f"{aprov_geral}%" if tipo == 'MENSAL' else None,
            },
            'tipo': tipo,
        }
        svc = WhatsAppService()
        img_b64 = svc.gerar_imagem_performance_b64(payload)
        if img_b64 and "base64," in img_b64:
            img_b64 = img_b64.split("base64,")[1]
        return img_b64
        
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def listar_grupos_whatsapp_api(request):
    """
    Consulta a Z-API e retorna os grupos para seleção.
    """
    try:
        svc = WhatsAppService()
        grupos_zapi = svc.listar_grupos()
        
        # Filtra e formata para o front
        lista_formatada = []
        
        # Garante que seja uma lista (algumas versoes retornam {response: [...]})
        if isinstance(grupos_zapi, dict) and 'response' in grupos_zapi:
            grupos_zapi = grupos_zapi['response']
        
        if isinstance(grupos_zapi, list):
            for g in grupos_zapi:
                if isinstance(g, dict):
                    # Tenta capturar o ID de várias formas possíveis
                    g_id = g.get('id') or g.get('chatId') or g.get('phone')
                    
                    # Tenta capturar o Nome
                    g_name = g.get('name') or g.get('subject') or g.get('contactName') or 'Sem Nome'

                    # Só adiciona se tiver ID válido
                    if g_id:
                        lista_formatada.append({
                            'id': g_id,
                            'name': g_name
                        })
        
        return Response(lista_formatada)
    except Exception as e:
        logger.error(f"Erro view listar grupos: {e}")
        return Response({'error': str(e)}, status=500)


class ViaCepProxyView(APIView):
    """Proxy para consulta de CEP evitando CORS no frontend."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, cep):
        cep_limpo = re.sub(r'\D', '', str(cep or ''))
        if len(cep_limpo) != 8:
            return Response({'error': 'CEP inválido.'}, status=400)

        url = f"https://viacep.com.br/ws/{cep_limpo}/json/"
        try:
            resp = requests.get(url, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            return Response(data)
        except requests.exceptions.Timeout:
            logger.warning("[ViaCEP] Timeout ao consultar CEP %s", cep_limpo)
            return Response({'error': 'Tempo esgotado ao consultar CEP. Tente novamente.'}, status=502)
        except requests.exceptions.RequestException as e:
            logger.warning("[ViaCEP] Erro de rede ao consultar CEP %s: %s", cep_limpo, e)
            return Response({'error': f'Erro ao consultar CEP: {str(e)}'}, status=502)
        except (ValueError, KeyError) as e:
            logger.warning("[ViaCEP] Resposta inválida para CEP %s: %s", cep_limpo, e)
            return Response({'error': 'Resposta inválida do serviço de CEP.'}, status=502)
        except Exception as e:
            logger.exception("[ViaCEP] Erro inesperado ao consultar CEP %s: %s", cep_limpo, e)
            return Response({'error': f'Erro ao consultar CEP: {str(e)}'}, status=502)


class NominatimProxyView(APIView):
    """Proxy para consulta Nominatim evitando CORS no frontend."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            import json
            import urllib.parse
            import urllib.request

            q = request.query_params.get('q')
            postalcode = request.query_params.get('postalcode')
            if not q and not postalcode:
                return Response({'error': 'Parâmetro ausente.'}, status=400)

            params = {
                'format': 'json',
                'limit': '1',
                'countrycodes': 'br',
            }
            if q:
                params['q'] = q
            if postalcode:
                params['postalcode'] = postalcode

            url = f"https://nominatim.openstreetmap.org/search?{urllib.parse.urlencode(params)}"
            req = urllib.request.Request(
                url,
                headers={'User-Agent': 'RecordPap/1.0'}
            )
            with urllib.request.urlopen(req, timeout=8) as response:
                data = json.loads(response.read().decode('utf-8'))
            return Response(data)
        except Exception as e:
            return Response({'error': f'Erro ao consultar Nominatim: {str(e)}'}, status=502)
    

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def relatorio_resultado_campanha(request, campanha_id):
    try:
        campanha = Campanha.objects.get(id=campanha_id)
        # Ordena faixas da MAIOR para a MENOR para achar a atingida mais fácil
        faixas_premiacao = campanha.regras_meta.all().order_by('-meta') 
        
        # --- CORREÇÃO DE LÓGICA DE DATAS ---
        if campanha.tipo_meta == 'LIQUIDA':
            # REGRA NOVA: Se a meta é INSTALADA (Líquida), olhamos a Data de Instalação.
            # Não importa se vendeu mês passado, se instalou dentro da campanha, conta.
            filtros = Q(
                data_instalacao__gte=campanha.data_inicio,
                data_instalacao__lte=campanha.data_fim,
                status_esteira__nome__iexact='INSTALADA',
                ativo=True
            )
        else:
            # REGRA PADRÃO: Se a meta é VENDA (Bruta), olhamos a Data de Criação.
            filtros = Q(
                data_criacao__date__gte=campanha.data_inicio,
                data_criacao__date__lte=campanha.data_fim,
                ativo=True
            )
        # -----------------------------------

        if campanha.canal_alvo != 'TODOS':
            filtros &= Q(vendedor__canal=campanha.canal_alvo)

        planos_validos = campanha.planos_elegiveis.all()
        if planos_validos.exists(): filtros &= Q(plano__in=planos_validos)
        
        pgtos_validos = campanha.formas_pagamento_elegiveis.all()
        if pgtos_validos.exists(): filtros &= Q(forma_pagamento__in=pgtos_validos)

        vendas = Venda.objects.filter(filtros).values(
            'vendedor__id', 'vendedor__first_name', 'vendedor__last_name', 'vendedor__username'
        ).annotate(total_vendas=Count('id')).order_by('-total_vendas')

        resultado = []
        ranking = 1
        # Meta máxima apenas para desenhar a barra de progresso visual
        meta_maxima_visual = campanha.meta_vendas if campanha.meta_vendas > 0 else 1

        for v in vendas:
            nome_vendedor = f"{v['vendedor__first_name']} {v['vendedor__last_name']}".strip() or v['vendedor__username']
            qtd = v['total_vendas']
            
            premio_receber = 0.0
            meta_alcancada = 0 
            
            # 1. Lógica de Faixas (Escalonada)
            if faixas_premiacao.exists():
                for faixa in faixas_premiacao:
                    if qtd >= faixa.meta:
                        premio_receber = float(faixa.valor_premio)
                        meta_alcancada = faixa.meta 
                        break 
            # 2. Lógica Simples (Legado)
            else:
                if qtd >= campanha.meta_vendas:
                    premio_receber = float(campanha.valor_premio)
                    meta_alcancada = campanha.meta_vendas

            percentual = round((qtd / meta_maxima_visual) * 100, 1)
            
            resultado.append({
                'ranking': ranking,
                'vendedor_id': v['vendedor__id'],
                'vendedor': nome_vendedor,
                'vendas_validas': qtd,
                'meta': meta_maxima_visual, 
                'meta_alcancada': meta_alcancada, 
                'percentual': percentual,
                'atingiu_meta': premio_receber > 0,
                'premio_receber': premio_receber
            })
            ranking += 1

        return Response({
            'campanha': campanha.nome,
            'periodo': f"{campanha.data_inicio.strftime('%d/%m/%Y')} a {campanha.data_fim.strftime('%d/%m/%Y')}",
            'tipo': campanha.get_tipo_meta_display(),
            'ranking': resultado
        })

    except Campanha.DoesNotExist: return Response({'error': 'Campanha não encontrada'}, status=404)
    except Exception as e: return Response({'error': str(e)}, status=500)
    
class LancamentoFinanceiroViewSet(viewsets.ModelViewSet):
    serializer_class = LancamentoFinanceiroSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = LancamentoFinanceiro.objects.select_related('usuario', 'criado_por').all().order_by('-data')
        ano = self.request.query_params.get('ano')
        mes = self.request.query_params.get('mes')
        if ano and mes:
            try:
                ano, mes = int(ano), int(mes)
                if 1 <= mes <= 12:
                    qs = qs.filter(data__year=ano, data__month=mes)
            except (TypeError, ValueError):
                pass
        return qs

    def list(self, request, *args, **kwargs):
        # Quando filtrar por ano/mês, retornar lista completa (sem paginação) para exibir tudo no "Adiantamentos e Descontos"
        if request.query_params.get('ano') and request.query_params.get('mes'):
            qs = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(qs, many=True)
            return Response(serializer.data)
        return super().list(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        import logging
        logger = logging.getLogger(__name__)
        logger.info('[LancamentoFinanceiro] POST create: usuario=%s tipo=%s data=%s valor=%s', request.data.get('usuario'), request.data.get('tipo'), request.data.get('data'), request.data.get('valor'))
        tipo = (request.data.get('tipo') or '').strip().upper()
        usuario_id = request.data.get('usuario')
        if tipo == 'ADIANTAMENTO_CNPJ' and usuario_id:
            try:
                usuario = Usuario.objects.get(id=usuario_id)
            except Usuario.DoesNotExist:
                return Response({'detail': 'Colaborador inválido.'}, status=status.HTTP_400_BAD_REQUEST)
            if getattr(usuario, 'recebe_adiantamento_cnpj', True) is False:
                return Response({'detail': 'Este usuário não está habilitado para receber Adiantamento de CNPJ.'}, status=status.HTTP_400_BAD_REQUEST)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        # Salva automaticamente quem criou o registro (segurança/auditoria)
        serializer.save(criado_por=self.request.user)

    @action(detail=False, methods=['get'], url_path='vendas-instaladas-mes')
    def vendas_instaladas_mes(self, request):
        """Lista vendas instaladas de um vendedor em um mês. ?tipo_cliente=CNPJ para só vendas CNPJ.
        A coluna 'Comissão est.' usa a faixa 'Adiantamento' em REGRAS_FAIXAS quando existir (valores fixos por plano 500MB/700/1GB)."""
        from datetime import datetime
        import re
        vendedor_id = request.query_params.get('vendedor_id')
        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        tipo_cliente = (request.query_params.get('tipo_cliente') or '').strip().upper()
        if not vendedor_id or not ano or not mes:
            return Response({'error': 'Parâmetros vendedor_id, ano e mes são obrigatórios.'}, status=400)
        try:
            ano, mes = int(ano), int(mes)
            if not (1 <= mes <= 12):
                return Response({'error': 'Mês inválido.'}, status=400)
            data_inicio = datetime(ano, mes, 1).date()
            if mes == 12:
                data_fim = datetime(ano + 1, 1, 1).date()
            else:
                data_fim = datetime(ano, mes + 1, 1).date()
        except (TypeError, ValueError):
            return Response({'error': 'ano/mes inválidos.'}, status=400)
        vendas = Venda.objects.filter(
            vendedor_id=vendedor_id,
            ativo=True,
            status_esteira__nome__iexact='INSTALADA',
            data_instalacao__gte=data_inicio,
            data_instalacao__lt=data_fim,
        ).select_related('plano', 'cliente').order_by('data_instalacao', 'id')
        # Faixa com finalidade ADIANTAMENTO define o valor fixo da "Comissão est." (500MB/700/1GB)
        faixa_adiantamento = RegraComissaoFaixa.objects.filter(
            finalidade='ADIANTAMENTO'
        ).first()
        if not faixa_adiantamento:
            faixa_adiantamento = RegraComissaoFaixa.objects.filter(
                faixa_nome__iexact='Adiantamento'
            ).first()
        only_cnpj = (tipo_cliente == 'CNPJ')
        hoje_sp = timezone.localtime(timezone.now())
        semana_ini = (hoje_sp - timedelta(days=hoje_sp.weekday())).date()
        semana_fim = semana_ini + timedelta(days=7)
        lista = []
        for v in vendas:
            if only_cnpj and v.cliente:
                digits = re.sub(r'\D', '', v.cliente.cpf_cnpj or '')
                if len(digits) != 14:  # CNPJ tem 14 dígitos
                    continue
            is_cnpj_venda = v.cliente and len(re.sub(r'\D', '', v.cliente.cpf_cnpj or '')) == 14
            valor_est = self._valor_comissao_estimado_venda(v, faixa_adiantamento, is_cnpj_venda)
            status_adiantamento = 'ELEGIVEL'
            motivo_nao_elegivel = ''
            if v.flag_adiant_cnpj:
                status_adiantamento = 'ADIANTAMENTO_CNPJ_REALIZADO'
                motivo_nao_elegivel = 'Já adiantado'
            elif not getattr(v.vendedor, 'recebe_adiantamento_cnpj', True):
                status_adiantamento = 'NAO_ELEGIVEL'
                motivo_nao_elegivel = 'Vendedor não habilitado para adiantamento de CNPJ'
            elif not is_cnpj_venda:
                status_adiantamento = 'NAO_ELEGIVEL'
                motivo_nao_elegivel = 'Cliente não é CNPJ'
            elif not (v.data_criacao and semana_ini <= v.data_criacao.date() <= semana_fim):
                status_adiantamento = 'NAO_ELEGIVEL'
                motivo_nao_elegivel = 'Venda fora da semana atual'
            elif not (v.data_instalacao and semana_ini <= v.data_instalacao <= semana_fim):
                status_adiantamento = 'NAO_ELEGIVEL'
                motivo_nao_elegivel = 'Instalação fora da semana atual'
            lista.append({
                'id': v.id,
                'cliente_nome': v.cliente.nome_razao_social if v.cliente else '',
                'plano_nome': v.plano.nome if v.plano else '',
                'data_instalacao': v.data_instalacao.isoformat()[:10] if v.data_instalacao else '',
                'ordem_servico': v.ordem_servico or '',
                'valor_comissao_estimado': round(valor_est, 2),
                'status_adiantamento': status_adiantamento,
                'motivo_nao_elegivel': motivo_nao_elegivel,
                'adiantamento_cnpj_realizado_em': v.adiantamento_cnpj_realizado_em.isoformat() if v.adiantamento_cnpj_realizado_em else None,
            })
        return Response(lista)

    @staticmethod
    def _banda_plano(nome_plano):
        """Retorna '500mb', '700mb' ou '1gb' conforme o nome do plano, ou None."""
        if not nome_plano:
            return None
        n = (nome_plano or '').strip().lower()
        if '1g' in n or '1 g' in n:
            return '1gb'
        if '700' in n:
            return '700mb'
        if '500' in n:
            return '500mb'
        return None

    @staticmethod
    def _valor_comissao_estimado_venda(venda, faixa_adiantamento, is_cnpj):
        """Usa a faixa 'Adiantamento' (REGRAS_FAIXAS) para valor fixo; senão fallback em plano.comissao_base."""
        if faixa_adiantamento:
            nome_plano = venda.plano.nome if venda.plano else ''
            banda = LancamentoFinanceiroViewSet._banda_plano(nome_plano)
            if banda and is_cnpj:
                if banda == '500mb' and faixa_adiantamento.valor_500mb_cnpj is not None:
                    return float(faixa_adiantamento.valor_500mb_cnpj)
                if banda == '700mb' and faixa_adiantamento.valor_700mb_cnpj is not None:
                    return float(faixa_adiantamento.valor_700mb_cnpj)
                if banda == '1gb' and faixa_adiantamento.valor_1gb_cnpj is not None:
                    return float(faixa_adiantamento.valor_1gb_cnpj)
            if banda and not is_cnpj:
                if banda == '500mb' and faixa_adiantamento.valor_500mb_pap is not None:
                    return float(faixa_adiantamento.valor_500mb_pap)
                if banda == '700mb' and faixa_adiantamento.valor_700mb_pap is not None:
                    return float(faixa_adiantamento.valor_700mb_pap)
                if banda == '1gb' and faixa_adiantamento.valor_1gb_pap is not None:
                    return float(faixa_adiantamento.valor_1gb_pap)
        return float(venda.plano.comissao_base) if venda.plano and venda.plano.comissao_base is not None else 0


# --- PAINEL DO AGENTE FINANCEIRO (SEGUNDA-FEIRA) ---

class PainelSegundaAPIView(APIView):
    """GET ?semana=YYYY-MM-DD (segunda-feira). Retorna tabela por usuário. Acesso: Diretoria e Admin."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser and not is_member(request.user, ['Diretoria', 'Admin']):
            return Response({'error': 'Acesso negado. Apenas Diretoria e Admin.'}, status=403)
        semana_str = request.query_params.get('semana')
        if not semana_str:
            return Response({'error': 'Parâmetro semana (YYYY-MM-DD, segunda-feira) é obrigatório.'}, status=400)
        try:
            from datetime import datetime
            data_param = datetime.strptime(semana_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Formato de data inválido. Use YYYY-MM-DD.'}, status=400)
        from datetime import timedelta
        from .painel_segunda_service import gerar_painel_semana, get_semana_seg_sab
        seg, sab = get_semana_seg_sab(data_param)
        # Semana a exibir: segunda a sábado ANTERIORES à segunda selecionada
        semana_inicio = seg - timedelta(days=7)
        semana_fim = semana_inicio + timedelta(days=5)
        dados = gerar_painel_semana(semana_inicio)
        return Response({
            'semana_inicio': semana_inicio.isoformat(),
            'semana_fim': semana_fim.isoformat(),
            'dados': dados,
        })


# --- VIEWS PARA CONFIRMAÇÃO E REVERSÃO DE DESCONTOS ---

class PendenciasDescontoView(APIView):
    """Pendências de desconto/adiantamento. GET ?ano=&mes= para filtrar por mês da instalação (somente mês/ano)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from datetime import datetime
        vendas = Venda.objects.filter(
            ativo=True,
            status_esteira__nome__iexact='INSTALADA'
        ).select_related('vendedor', 'forma_pagamento', 'cliente')

        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        if ano and mes:
            try:
                ano, mes = int(ano), int(mes)
                if 1 <= mes <= 12:
                    data_inicio = datetime(ano, mes, 1).date()
                    if mes == 12:
                        data_fim = datetime(ano + 1, 1, 1).date()
                    else:
                        data_fim = datetime(ano, mes + 1, 1).date()
                    vendas = vendas.filter(
                        data_instalacao__gte=data_inicio,
                        data_instalacao__lt=data_fim,
                    )
            except (TypeError, ValueError):
                pass

        pendencias = []

        for v in vendas:
            consultor = v.vendedor
            if not consultor: continue

            doc = v.cliente.cpf_cnpj if v.cliente else ''
            doc_limpo = ''.join(filter(str.isdigit, doc))
            eh_cnpj = len(doc_limpo) > 11

            if eh_cnpj and not v.flag_adiant_cnpj:
                val = float(consultor.adiantamento_cnpj or 0)
                if val > 0: pendencias.append(self._montar_obj(v, 'CNPJ', val, 'Adiantamento CNPJ'))

            if v.forma_pagamento and 'BOLETO' in v.forma_pagamento.nome.upper() and not v.flag_desc_boleto:
                val = float(consultor.desconto_boleto or 0)
                if val > 0: pendencias.append(self._montar_obj(v, 'BOLETO', val, 'Desconto Boleto'))

            if v.inclusao and not v.flag_desc_viabilidade:
                val = float(consultor.desconto_inclusao_viabilidade or 0)
                if val > 0: pendencias.append(self._montar_obj(v, 'VIABILIDADE', val, 'Desconto Viabilidade'))

            if v.antecipou_instalacao and not v.flag_desc_antecipacao:
                val = float(consultor.desconto_instalacao_antecipada or 0)
                if val > 0: pendencias.append(self._montar_obj(v, 'ANTECIPACAO', val, 'Desconto Antecipação'))

        vendas_sab_cancel = Venda.objects.filter(
            ativo=True,
            status_esteira__nome__icontains='CANCEL',
            adiantamento_sabado_marcado=True,
            flag_desc_adiantamento_sabado=False,
        ).select_related('vendedor', 'forma_pagamento', 'cliente')

        ano_q = request.query_params.get('ano')
        mes_q = request.query_params.get('mes')
        if ano_q and mes_q:
            try:
                ay, my = int(ano_q), int(mes_q)
                if 1 <= my <= 12:
                    data_inicio_c = datetime(ay, my, 1).date()
                    if my == 12:
                        data_fim_c = datetime(ay + 1, 1, 1).date()
                    else:
                        data_fim_c = datetime(ay, my + 1, 1).date()
                    vendas_sab_cancel = vendas_sab_cancel.filter(
                        data_ultima_alteracao__date__gte=data_inicio_c,
                        data_ultima_alteracao__date__lt=data_fim_c,
                    )
            except (TypeError, ValueError):
                pass

        for v in vendas_sab_cancel:
            val = float(v.adiantamento_sabado_valor or 0)
            if val <= 0:
                continue
            consultor = v.vendedor
            if not consultor:
                continue
            pendencias.append(
                self._montar_obj(
                    v, 'ADIANT_SABADO', val, 'Desconto adiantamento sábado (não instalado / cancelado)'
                )
            )

        return Response(pendencias)

    def _montar_obj(self, venda, tipo_codigo, valor, titulo):
        ref = None
        if getattr(venda, 'data_ultima_alteracao', None):
            ref = venda.data_ultima_alteracao.date()
        return {
            'venda_id': venda.id,
            'data_instalacao': venda.data_instalacao,
            'data_referencia': ref,
            'cliente': venda.cliente.nome_razao_social,
            'cliente_cpf': venda.cliente.cpf_cnpj,
            'os': venda.ordem_servico or "",
            'vendedor_id': venda.vendedor.id,
            'vendedor_nome': venda.vendedor.username,
            'tipo_codigo': tipo_codigo,
            'titulo': titulo,
            'valor': valor
        }


class ConfirmarDescontosEmMassaView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        data_lancamento = request.data.get('data')
        itens = request.data.get('itens', []) 

        if not data_lancamento or not itens:
            return Response({'error': 'Dados inválidos.'}, status=400)

        agrupado = defaultdict(lambda: {'total': 0.0, 'ids_vendas': [], 'tipos': set()})
        updates_flags = {'CNPJ': [], 'BOLETO': [], 'VIABILIDADE': [], 'ANTECIPACAO': [], 'ADIANT_SABADO': []}

        for item in itens:
            v_id = item['venda_id']
            tipo = item['tipo_codigo']
            valor = float(item['valor'])
            vend_id = item['vendedor_id']

            agrupado[vend_id]['total'] += valor
            agrupado[vend_id]['ids_vendas'].append(v_id) # Guarda ID como int
            agrupado[vend_id]['tipos'].add(tipo)
            
            if tipo in updates_flags: updates_flags[tipo].append(v_id)

        try:
            with transaction.atomic():
                lancamentos_criados = []
                for vend_id, dados in agrupado.items():
                    tipos_lista = list(dados['tipos'])
                    tipos_str = ", ".join(tipos_lista)
                    qtd = len(dados['ids_vendas'])
                    
                    categ = 'ADIANTAMENTO_CNPJ' if 'CNPJ' in dados['tipos'] and len(dados['tipos']) == 1 else 'DESCONTO'
                    
                    # SALVA OS IDs NO JSON 'METADADOS' PARA PERMITIR REVERSÃO
                    metadados = {
                        "origem": "automatico",
                        "ids_vendas": dados['ids_vendas'],
                        "tipos_processados": tipos_lista
                    }

                    lancamentos_criados.append(LancamentoFinanceiro(
                        usuario_id=vend_id,
                        tipo=categ,
                        data=data_lancamento,
                        valor=dados['total'],
                        quantidade_vendas=qtd,
                        descricao=f"Processamento Auto: {tipos_str}",
                        metadados=metadados,
                        criado_por=request.user
                    ))
                
                if lancamentos_criados:
                    LancamentoFinanceiro.objects.bulk_create(lancamentos_criados)

                if updates_flags['CNPJ']: Venda.objects.filter(id__in=updates_flags['CNPJ']).update(flag_adiant_cnpj=True)
                if updates_flags['BOLETO']: Venda.objects.filter(id__in=updates_flags['BOLETO']).update(flag_desc_boleto=True)
                if updates_flags['VIABILIDADE']: Venda.objects.filter(id__in=updates_flags['VIABILIDADE']).update(flag_desc_viabilidade=True)
                if updates_flags['ANTECIPACAO']: Venda.objects.filter(id__in=updates_flags['ANTECIPACAO']).update(flag_desc_antecipacao=True)
                if updates_flags['ADIANT_SABADO']: Venda.objects.filter(id__in=updates_flags['ADIANT_SABADO']).update(flag_desc_adiantamento_sabado=True)

            return Response({'mensagem': f'{len(itens)} descontos processados com sucesso!'})

        except Exception as e:
            return Response({'error': str(e)}, status=500)

class HistoricoDescontosAutoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Lista lançamentos financeiros automáticos recentes
        # Filtra onde metadados contém "origem": "automatico"
        # SQLite não suporta filtro JSON nativo complexo em Django antigo, então filtramos na descrição ou trazemos tudo
        lancamentos = LancamentoFinanceiro.objects.filter(
            descricao__startswith="Processamento Auto"
        ).select_related('usuario', 'criado_por').order_by('-data_criacao')[:50] # Últimos 50

        data = []
        for l in lancamentos:
            data.append({
                'id': l.id,
                'data_lancamento': l.data,
                'vendedor': l.usuario.username,
                'tipo': l.tipo,
                'descricao': l.descricao,
                'valor': l.valor,
                'criado_por': l.criado_por.username if l.criado_por else 'Sistema',
                'criado_em': l.data_criacao
            })
        return Response(data)

class ReverterDescontoMassaView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        lancamento_id = request.data.get('id')
        if not lancamento_id:
            return Response({'error': 'ID inválido.'}, status=400)

        try:
            lancamento = LancamentoFinanceiro.objects.get(id=lancamento_id)
            
            # Verifica se é um lançamento automático
            if not lancamento.descricao.startswith("Processamento Auto"):
                 return Response({'error': 'Apenas lançamentos automáticos podem ser revertidos por aqui.'}, status=400)

            # --- CORREÇÃO PARA REGISTROS ANTIGOS ---
            # Se não tiver metadados (registros criados durante os testes anteriores),
            # permitimos excluir apenas o financeiro para limpar a tela.
            if not lancamento.metadados:
                lancamento.delete()
                return Response({'mensagem': 'Registro antigo excluído do financeiro! Nota: As flags nas vendas NÃO foram revertidas (dados de vínculo ausentes).'})
            # ---------------------------------------

            meta = lancamento.metadados
            ids_vendas = meta.get('ids_vendas', [])
            tipos = meta.get('tipos_processados', [])

            with transaction.atomic():
                if ids_vendas:
                    # Reverte as flags nas vendas originais
                    vendas_afetadas = Venda.objects.filter(id__in=ids_vendas)
                    updates = {}
                    
                    if 'CNPJ' in tipos: updates['flag_adiant_cnpj'] = False
                    if 'BOLETO' in tipos: updates['flag_desc_boleto'] = False
                    if 'VIABILIDADE' in tipos: updates['flag_desc_viabilidade'] = False
                    if 'ANTECIPACAO' in tipos: updates['flag_desc_antecipacao'] = False
                    if 'ADIANT_SABADO' in tipos: updates['flag_desc_adiantamento_sabado'] = False
                    
                    if updates:
                        vendas_afetadas.update(**updates)

                # Apaga o lançamento financeiro
                lancamento.delete()

            return Response({'mensagem': 'Reversão concluída com sucesso! As vendas voltaram para a lista de pendências.'})

        except LancamentoFinanceiro.DoesNotExist:
            return Response({'error': 'Lançamento não encontrado.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class AdiantamentosEsteiraView(APIView):
    """
    GET ?ano=&mes=
    Lista lançamentos de adiantamento de comissão e de CNPJ criados na esteira (Instaladas),
    e opcionalmente marcações CNPJ sem lançamento (legado antes do fluxo financeiro).
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from datetime import datetime, time

        ano = request.query_params.get('ano')
        mes = request.query_params.get('mes')
        if not ano or not mes:
            return Response({'error': 'Informe ano e mes.'}, status=400)
        try:
            ano, mes = int(ano), int(mes)
            if not (1 <= mes <= 12):
                raise ValueError
        except (TypeError, ValueError):
            return Response({'error': 'ano/mes inválidos.'}, status=400)

        data_inicio = datetime(ano, mes, 1).date()
        if mes == 12:
            data_fim = datetime(ano + 1, 1, 1).date()
        else:
            data_fim = datetime(ano, mes + 1, 1).date()

        tz = timezone.get_current_timezone()
        start_dt = timezone.make_aware(datetime.combine(data_inicio, time.min), tz)
        end_dt = timezone.make_aware(datetime.combine(data_fim, time.min), tz)

        lancs = (
            LancamentoFinanceiro.objects.filter(
                tipo='ADIANTAMENTO_COMISSAO',
                data__gte=data_inicio,
                data__lt=data_fim,
            )
            .select_related('usuario', 'criado_por')
            .order_by('-data', '-data_criacao')
        )

        adiantamento_comissao = []
        for l in lancs:
            meta = l.metadados if isinstance(l.metadados, dict) else {}
            if meta.get('origem') != 'instaladas_comissao':
                continue
            uname = l.usuario.username if l.usuario else ''
            nome = (
                (l.usuario.get_full_name() or '').strip() or uname
                if l.usuario
                else ''
            )
            criador = ''
            if l.criado_por:
                criador = (l.criado_por.get_full_name() or '').strip() or (
                    l.criado_por.username or ''
                )
            adiantamento_comissao.append(
                {
                    'id': l.id,
                    'data': l.data.isoformat(),
                    'vendedor_id': l.usuario_id,
                    'vendedor_nome': nome,
                    'valor': float(l.valor or 0),
                    'quantidade_vendas': l.quantidade_vendas or 0,
                    'descricao': l.descricao,
                    'venda_ids': list(meta.get('venda_ids') or []),
                    'criado_por': criador,
                    'criado_em': l.data_criacao.isoformat() if l.data_criacao else None,
                }
            )

        adiantamento_sabado_agendados = []
        for l in lancs:
            meta = l.metadados if isinstance(l.metadados, dict) else {}
            if meta.get('origem') != 'esteira_sabado_agendados':
                continue
            uname = l.usuario.username if l.usuario else ''
            nome = (
                (l.usuario.get_full_name() or '').strip() or uname
                if l.usuario
                else ''
            )
            criador = ''
            if l.criado_por:
                criador = (l.criado_por.get_full_name() or '').strip() or (
                    l.criado_por.username or ''
                )
            adiantamento_sabado_agendados.append(
                {
                    'id': l.id,
                    'data': l.data.isoformat(),
                    'vendedor_id': l.usuario_id,
                    'vendedor_nome': nome,
                    'valor': float(l.valor or 0),
                    'quantidade_vendas': l.quantidade_vendas or 0,
                    'descricao': l.descricao,
                    'venda_ids': list(meta.get('venda_ids') or []),
                    'valores_por_venda_id': meta.get('valores_por_venda_id') or {},
                    'criado_por': criador,
                    'criado_em': l.data_criacao.isoformat() if l.data_criacao else None,
                }
            )

        lancs_cnpj = (
            LancamentoFinanceiro.objects.filter(
                tipo='ADIANTAMENTO_CNPJ',
                data__gte=data_inicio,
                data__lt=data_fim,
            )
            .select_related('usuario', 'criado_por')
            .order_by('-data', '-data_criacao')
        )
        adiantamento_cnpj = []
        venda_ids_com_lancamento = set()
        for l in lancs_cnpj:
            meta = l.metadados if isinstance(l.metadados, dict) else {}
            if meta.get('origem') != 'instaladas_cnpj':
                continue
            ids = [int(x) for x in (meta.get('venda_ids') or []) if x is not None]
            venda_ids_com_lancamento.update(ids)
            uname = l.usuario.username if l.usuario else ''
            nome = (
                (l.usuario.get_full_name() or '').strip() or uname
                if l.usuario
                else ''
            )
            criador = ''
            if l.criado_por:
                criador = (l.criado_por.get_full_name() or '').strip() or (
                    l.criado_por.username or ''
                )
            adiantamento_cnpj.append(
                {
                    'id': l.id,
                    'data': l.data.isoformat(),
                    'vendedor_id': l.usuario_id,
                    'vendedor_nome': nome,
                    'valor': float(l.valor or 0),
                    'quantidade_vendas': l.quantidade_vendas or 0,
                    'descricao': l.descricao,
                    'venda_ids': ids,
                    'criado_por': criador,
                    'criado_em': l.data_criacao.isoformat() if l.data_criacao else None,
                }
            )

        vendas_cnpj = (
            Venda.objects.filter(
                ativo=True,
                flag_adiant_cnpj=True,
                adiantamento_cnpj_realizado_em__gte=start_dt,
                adiantamento_cnpj_realizado_em__lt=end_dt,
            )
            .select_related('vendedor', 'cliente')
            .order_by('-adiantamento_cnpj_realizado_em')
        )

        marcacoes_cnpj_esteira = []
        for v in vendas_cnpj:
            if v.id in venda_ids_com_lancamento:
                continue
            val = 0.0
            if v.vendedor:
                val = float(getattr(v.vendedor, 'adiantamento_cnpj', None) or 0)
            vn = ''
            if v.vendedor:
                vn = (v.vendedor.get_full_name() or '').strip() or (
                    v.vendedor.username or ''
                )
            marcacoes_cnpj_esteira.append(
                {
                    'venda_id': v.id,
                    'vendedor_nome': vn,
                    'cliente': v.cliente.nome_razao_social if v.cliente else '',
                    'os': v.ordem_servico or '',
                    'data_marcacao': v.adiantamento_cnpj_realizado_em.isoformat()
                    if v.adiantamento_cnpj_realizado_em
                    else None,
                    'valor_parametro_cnpj': val,
                }
            )

        return Response(
            {
                'periodo': f'{mes:02d}/{ano}',
                'adiantamento_comissao': adiantamento_comissao,
                'adiantamento_sabado_agendados': adiantamento_sabado_agendados,
                'adiantamento_cnpj': adiantamento_cnpj,
                'marcacoes_cnpj_esteira': marcacoes_cnpj_esteira,
            }
        )


class VerificarPermissaoGestaoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Reutiliza sua função is_member
        eh_gestao = is_member(request.user, ['Diretoria', 'Admin'])
        return Response({'eh_gestao': eh_gestao})

# Em site-record/crm_app/views.py

class ImportacaoLegadoView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    # --- 1. DOWNLOAD DO MODELO (GET) ---
    def get(self, request):
        """
        Gera e baixa a planilha modelo atualizada (V4).
        """
        headers = {
            # DADOS DA VENDA
            'DATA_VENDA': ['01/01/2024'],
            'DATA_ABERTURA': ['01/01/2024 08:00'],
            'DATA_INSTALACAO': ['05/01/2024'],
            'LOGIN_VENDEDOR': ['joao.silva'],
            'NOME_PLANO': ['Internet 600 Mega'],
            'FORMA_PAGAMENTO': ['Boleto'],
            'OS': ['0123456'], # Exemplo com zero à esquerda
            
            # TRÂMITE / STATUS
            'STATUS_ESTEIRA': ['Instalada'], 
            'STATUS_TRATAMENTO': ['Fechado'], 
            'MOTIVO_PENDENCIA': [''], 
            'DATA_AGENDAMENTO': [''], 
            'PERIODO_AGENDAMENTO': [''], 
            
            # DADOS DO CLIENTE
            'CPF_CNPJ_CLIENTE': ['12345678900'],
            'NOME_CLIENTE': ['Fulano de Tal'],
            'DATA_NASCIMENTO': ['15/05/1990'],
            'NOME_MAE': ['Maria da Silva'],
            
            # CONTATO
            'TELEFONE_1': ['31999998888'],
            'TELEFONE_2': [''],
            'EMAIL_CLIENTE': ['cliente@email.com'],

            # ENDEREÇO
            'CEP': ['30000000'],
            'NUMERO': ['123'],
            'COMPLEMENTO': ['Apto 101'],
            'PONTO_REFERENCIA': ['Próximo ao Mercado'],
            'LOGRADOURO': [''], 
            'BAIRRO': [''],     
            'CIDADE': [''],     
            'UF': [''],         
            
            'OBSERVACOES': ['Importação Histórico']
        }
        
        df = pd.DataFrame(headers)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Modelo Legado')
            worksheet = writer.sheets['Modelo Legado']
            
            for idx, col in enumerate(df.columns):
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = 20

        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="modelo_legado_v5.xlsx"'
        return response

    # --- 2. IMPORTAÇÃO E PROCESSAMENTO (POST) - ASYNC ---
    def post(self, request):
        """
        Recebe arquivo Excel e inicia processamento assíncrono.
        Retorna imediatamente com log_id para acompanhamento.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        # 1. Criar log inicial
        log = LogImportacaoLegado.objects.create(
            usuario=request.user,
            nome_arquivo=file_obj.name,
            status='PROCESSANDO'
        )

        # 2. Ler arquivo para memória
        try:
            file_content = file_obj.read()
            file_name = file_obj.name
        except Exception as e:
            log.status = 'ERRO'
            log.mensagem_erro = f'Erro ao ler arquivo: {str(e)}'
            log.finalizado_em = timezone.now()
            log.save()
            return Response({'error': f'Erro ao ler arquivo: {str(e)}'}, status=400)

        # 3. Iniciar thread de processamento
        thread = threading.Thread(
            target=self._processar_legado_interno,
            args=(log.id, file_content, file_name)
        )
        thread.daemon = True
        thread.start()

        # 4. Retornar imediatamente
        # Mantém padrão dos outros imports: retorna sucesso imediato + log_id
        return Response({
            'success': True,
            'message': 'Processamento iniciado! Acompanhe o progresso na aba Histórico.',
            'status': 'PROCESSANDO',
            'background': True,
            'log_id': log.id
        }, status=200)

    def _processar_legado_interno(self, log_id, file_content, file_name):
        """Processa importação em background"""
        try:
            log = LogImportacaoLegado.objects.get(id=log_id)
            
            # Ler Excel em memória com dtype=str para preservar zeros
            try:
                df = pd.read_excel(BytesIO(file_content), dtype=str)
            except Exception as e:
                log.status = 'ERRO'
                log.mensagem_erro = f'Erro ao ler Excel: {str(e)}'
                log.finalizado_em = timezone.now()
                log.save()
                return

            # Normaliza colunas
            df.columns = [str(c).strip().upper() for c in df.columns]
            df = df.replace({np.nan: None, 'nan': None, 'NaN': None, 'None': None})
            
            log.total_linhas = len(df)
            log.save()
            
            # --- CACHES PARA PERFORMANCE ---
            users_map = {u.username.upper(): u for u in get_user_model().objects.all()}
            for u in get_user_model().objects.all():
                if u.email: users_map[u.email.upper()] = u

            planos_map = {p.nome.upper(): p for p in Plano.objects.all()}
            pgto_map = {fp.nome.upper(): fp for fp in FormaPagamento.objects.all()}
            
            status_esteira_map = {s.nome.upper(): s for s in StatusCRM.objects.filter(tipo='Esteira')}
            status_tratamento_map = {s.nome.upper(): s for s in StatusCRM.objects.filter(tipo='Tratamento')}
            
            motivo_map = {}
            for m in MotivoPendencia.objects.all():
                motivo_map[m.nome.upper().strip()] = m
                parts = m.nome.split('-')
                if len(parts) > 1: motivo_map[parts[0].strip()] = m

            cache_viacep = {}
            vendas_para_criar = []
            logs_erro = []
            vendas_criadas = 0
            clientes_criados = 0

            # --- FUNÇÕES AUXILIARES ---
            def parse_dt(val):
                if not val: return None
                try: return pd.to_datetime(val, dayfirst=True, errors='coerce').date()
                except: return None

            def parse_periodo(val):
                if not val: return None
                v = str(val).upper()
                if 'MANH' in v: return 'MANHA'
                if 'TARDE' in v: return 'TARDE'
                if 'NOITE' in v: return 'NOITE'
                return None

            def parse_dt_time(val):
                """Retorna datetime timezone-aware ou None. Aceita 'DD/MM/YYYY' ou 'DD/MM/YYYY HH:MM'."""
                if not val: return None
                try:
                    dt = pd.to_datetime(val, dayfirst=True, errors='coerce')
                    if pd.isna(dt): return None
                    if hasattr(dt, 'to_pydatetime'):
                        dt = dt.to_pydatetime()
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt)
                    return dt
                except Exception:
                    return None

            def consultar_cep_otimizado(cep_input):
                if not cep_input: return {}
                cep_limpo = str(cep_input).replace('-', '').replace('.', '').strip()
                if len(cep_limpo) != 8: return {}
                
                if cep_limpo in cache_viacep: return cache_viacep[cep_limpo]
                
                try:
                    resp = requests.get(f"https://viacep.com.br/ws/{cep_limpo}/json/", timeout=2)
                    if resp.status_code == 200:
                        dados = resp.json()
                        if 'erro' not in dados:
                            res = {
                                'logradouro': dados.get('logradouro', '').upper(),
                                'bairro': dados.get('bairro', '').upper(),
                                'cidade': dados.get('localidade', '').upper(),
                                'uf': dados.get('uf', '').upper()
                            }
                            cache_viacep[cep_limpo] = res
                            return res
                except: pass
                
                cache_viacep[cep_limpo] = {}
                return {}

            # --- LOOP DE PROCESSAMENTO ---
            for index, row in df.iterrows():
                linha = index + 2
                try:
                    # 1. CLIENTE & CPF
                    cpf_raw = str(row.get('CPF_CNPJ_CLIENTE', ''))
                    cpf_limpo = ''.join(filter(str.isdigit, cpf_raw))
                    
                    if cpf_limpo and len(cpf_limpo) < 11:
                        cpf_limpo = cpf_limpo.zfill(11)
                    
                    if not cpf_limpo or len(cpf_limpo) < 11:
                        logs_erro.append(f"Linha {linha}: CPF inválido ou vazio ({cpf_raw}).")
                        continue

                    nome_cli = str(row.get('NOME_CLIENTE', 'Cliente Importado')).upper()
                    
                    cliente, criou_cliente = Cliente.objects.get_or_create(
                        cpf_cnpj=cpf_limpo, defaults={'nome_razao_social': nome_cli}
                    )
                    if criou_cliente:
                        clientes_criados += 1
                    
                    mudou_cliente = False
                    if row.get('TELEFONE_1'): 
                        cliente.telefone1 = str(row.get('TELEFONE_1'))
                        mudou_cliente = True
                    if row.get('TELEFONE_2'): 
                        cliente.telefone2 = str(row.get('TELEFONE_2'))
                        mudou_cliente = True
                    if row.get('EMAIL_CLIENTE'): 
                        cliente.email = str(row.get('EMAIL_CLIENTE'))
                        mudou_cliente = True
                    if row.get('DATA_NASCIMENTO'): 
                        cliente.data_nascimento = parse_dt(row.get('DATA_NASCIMENTO'))
                        mudou_cliente = True
                    if row.get('NOME_MAE'): 
                        cliente.nome_mae = str(row.get('NOME_MAE')).upper()
                        mudou_cliente = True
                    
                    if mudou_cliente:
                        cliente.save()

                    # 2. VENDEDOR
                    login_vend = str(row.get('LOGIN_VENDEDOR', '')).upper().strip()
                    vendedor = users_map.get(login_vend)

                    # 3. PLANO & PAGAMENTO
                    nome_plano = str(row.get('NOME_PLANO', '')).upper().strip()
                    plano = planos_map.get(nome_plano)
                    
                    nome_pgto = str(row.get('FORMA_PAGAMENTO', '')).upper().strip()
                    pgto = pgto_map.get(nome_pgto)

                    # 4. STATUS
                    nome_est = str(row.get('STATUS_ESTEIRA', '')).upper().strip()
                    status_esteira = status_esteira_map.get(nome_est)
                    
                    if not status_esteira and nome_est:
                        if 'INSTAL' in nome_est or 'CONCLU' in nome_est: 
                            status_esteira = status_esteira_map.get('INSTALADA')
                        elif 'PENDEN' in nome_est: 
                            status_esteira = status_esteira_map.get('PENDENCIADA')
                        elif 'AGEND' in nome_est: 
                            status_esteira = status_esteira_map.get('AGENDADO')
                        elif 'CANCEL' in nome_est: 
                            status_esteira = status_esteira_map.get('CANCELADA')

                    motivo_pend = None
                    if row.get('MOTIVO_PENDENCIA'):
                        motivo_pend = motivo_map.get(str(row.get('MOTIVO_PENDENCIA')).upper().strip())

                    status_tratamento = status_tratamento_map.get(str(row.get('STATUS_TRATAMENTO', '')).upper().strip())
                    if not status_tratamento:
                        if status_esteira and status_esteira.nome.upper() in ['INSTALADA', 'CANCELADA']:
                            status_tratamento = status_tratamento_map.get('FECHADO')
                        else:
                            status_tratamento = status_tratamento_map.get('SEM TRATAMENTO')

                    # 5. DATAS
                    dt_venda = parse_dt(row.get('DATA_VENDA'))
                    if not dt_venda: dt_venda = timezone.now().date()

                    dt_abertura = parse_dt_time(row.get('DATA_ABERTURA'))
                    dt_inst = parse_dt(row.get('DATA_INSTALACAO'))
                    dt_agend = parse_dt(row.get('DATA_AGENDAMENTO'))

                    if status_esteira and status_esteira.nome.upper() == 'INSTALADA' and not dt_inst:
                        dt_inst = dt_venda

                    # 6. ENDEREÇO
                    cep_raw = str(row.get('CEP', '')).replace('-', '').replace('.', '').strip()
                    cep_final = cep_raw[:9]
                    
                    logradouro = str(row.get('LOGRADOURO', '')).upper()
                    bairro = str(row.get('BAIRRO', '')).upper()
                    cidade = str(row.get('CIDADE', '')).upper()
                    uf = str(row.get('UF', '')).upper()
                    
                    if cep_final and (not logradouro or not bairro):
                        dados_api = consultar_cep_otimizado(cep_final)
                        if dados_api:
                            logradouro = dados_api.get('logradouro', '')
                            bairro = dados_api.get('bairro', '')
                            cidade = dados_api.get('cidade', '')
                            uf = dados_api.get('uf', '')

                    # 7. O.S. (Preservando Zero à Esquerda)
                    os_raw = str(row.get('OS', '')).strip()
                    if os_raw.endswith('.0'): os_raw = os_raw[:-2]

                    # --- INSTANCIAÇÃO ---
                    venda = Venda(
                        cliente=cliente,
                        vendedor=vendedor,
                        plano=plano,
                        forma_pagamento=pgto,
                        
                        status_esteira=status_esteira,
                        status_tratamento=status_tratamento,
                        motivo_pendencia=motivo_pend,
                        
                        data_pedido=dt_venda,
                        data_abertura=dt_abertura,
                        data_instalacao=dt_inst,
                        data_agendamento=dt_agend,
                        periodo_agendamento=parse_periodo(row.get('PERIODO_AGENDAMENTO')),
                        
                        data_criacao=dt_venda, 
                        
                        cep=cep_final,
                        logradouro=logradouro[:255],
                        numero_residencia=str(row.get('NUMERO', ''))[:20],
                        complemento=str(row.get('COMPLEMENTO', '')).upper()[:100],
                        bairro=bairro[:100],
                        cidade=cidade[:100],
                        estado=uf[:2],
                        ponto_referencia=str(row.get('PONTO_REFERENCIA', '')).upper()[:255],

                        ordem_servico=os_raw,
                        observacoes=str(row.get('OBSERVACOES', 'Importação Legado'))[:500],
                        ativo=True
                    )
                    
                    vendas_para_criar.append(venda)
                    vendas_criadas += 1

                except Exception as e:
                    logs_erro.append(f"Linha {linha}: {str(e)}")

            # --- GRAVAÇÃO EM LOTE ---
            if vendas_para_criar:
                with transaction.atomic():
                    Venda.objects.bulk_create(vendas_para_criar, batch_size=1000)

            # Atualizar log
            log.total_processadas = vendas_criadas
            log.vendas_criadas = vendas_criadas
            log.clientes_criados = clientes_criados
            log.erros_count = len(logs_erro)
            log.finalizado_em = timezone.now()
            
            if logs_erro:
                log.status = 'PARCIAL' if vendas_criadas > 0 else 'ERRO'
                log.mensagem_erro = '\n'.join(logs_erro[:50])
                log.mensagem = f'{vendas_criadas} vendas importadas, {len(logs_erro)} erros'
            else:
                log.status = 'SUCESSO'
                log.mensagem = f'{vendas_criadas} vendas importadas com sucesso!'
            
            log.detalhes_json = {
                'clientes_criados': clientes_criados,
                'vendas_criadas': vendas_criadas,
                'erros': logs_erro[:100]
            }
            log.calcular_duracao()
            log.save()

        except Exception as e:
            try:
                log = LogImportacaoLegado.objects.get(id=log_id)
                log.status = 'ERRO'
                log.mensagem_erro = str(e)
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
            except:
                pass
# Adicione ou certifique-se que esta classe existe
class ConfigurarAutomacaoView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin']):
            return Response({"error": "Sem permissão"}, status=403)

        acao = request.data.get('acao')
        
        if acao == 'listar':
            regras = AgendamentoDisparo.objects.all().order_by('prioridade', 'id').values()
            return Response(list(regras))
            
        elif acao == 'salvar':
            d = request.data.get('dados')
            modo_envio = (d.get('modo_envio') or 'INTERVALO').strip().upper()
            if modo_envio not in ['INTERVALO', 'ESPECIFICO']:
                return Response({'error': 'Modo de envio inválido.'}, status=400)

            tipo = (d.get('tipo') or '').strip().upper()
            if tipo not in ['HORARIO', 'SEMANAL']:
                return Response({'error': 'Frequência inválida.'}, status=400)

            tipo_relatorio = (d.get('tipo_relatorio') or 'HOJE').strip().upper() or 'HOJE'
            if tipo_relatorio not in ['HOJE', 'SEMANAL', 'MENSAL']:
                return Response({'error': 'Tipo de relatório inválido.'}, status=400)

            horarios_especificos = d.get('horarios_especificos') or []
            dias_semana = d.get('dias_semana') or []

            if modo_envio == 'ESPECIFICO':
                if not isinstance(horarios_especificos, list) or not horarios_especificos:
                    return Response({'error': 'Selecione ao menos um horário específico.'}, status=400)

                horarios_validos = []
                vistos = set()
                for h in horarios_especificos:
                    hs = str(h or '').strip()
                    try:
                        hh_str, mm_str = hs.split(':')
                        hh = int(hh_str)
                        mm = int(mm_str)
                    except Exception:
                        return Response({'error': f'Horário inválido: {hs}'}, status=400)
                    if hh < 8 or hh > 22 or mm < 0 or mm > 59:
                        return Response({'error': f'Horário fora da faixa permitida (08:00-22:59): {hs}'}, status=400)
                    normalizado = f"{hh:02d}:{mm:02d}"
                    if normalizado not in vistos:
                        vistos.add(normalizado)
                        horarios_validos.append(normalizado)
                horarios_especificos = sorted(horarios_validos)

                if tipo == 'SEMANAL':
                    if not isinstance(dias_semana, list) or not dias_semana:
                        return Response({'error': 'Selecione ao menos um dia da semana para frequência semanal.'}, status=400)
                    dias_validos = []
                    vistos_dias = set()
                    for dia in dias_semana:
                        try:
                            di = int(dia)
                        except Exception:
                            return Response({'error': f'Dia da semana inválido: {dia}'}, status=400)
                        if di < 0 or di > 6:
                            return Response({'error': f'Dia da semana fora da faixa 0-6: {dia}'}, status=400)
                        if di not in vistos_dias:
                            vistos_dias.add(di)
                            dias_validos.append(di)
                    dias_semana = sorted(dias_validos)
                else:
                    dias_semana = []
            else:
                horarios_especificos = []
                dias_semana = []

            status_destinatarios = (d.get('status_destinatarios') or 'somente_ativos').strip().lower()
            if status_destinatarios not in ['somente_ativos', 'somente_inativos', 'todos']:
                return Response({'error': 'Status dos destinatários inválido.'}, status=400)

            prioridade_raw = d.get('prioridade')
            if prioridade_raw in [None, '']:
                return Response({'error': 'Prioridade é obrigatória.'}, status=400)
            try:
                prioridade = int(prioridade_raw)
            except (TypeError, ValueError):
                return Response({'error': 'Prioridade deve ser um número inteiro.'}, status=400)
            if prioridade <= 0:
                return Response({'error': 'Prioridade deve ser maior que zero.'}, status=400)

            defaults = {
                'nome': d['nome'], 'tipo': tipo,
                'canal_alvo': d['canal_alvo'],
                'cluster_alvo': (d.get('cluster_alvo') or '').strip() or '',
                'destinatarios': d['destinatarios'],
                'ativo': d['ativo'],
                'modo_envio': modo_envio,
                'intervalo_minutos': int(d.get('intervalo_minutos', 60) or 60) if modo_envio == 'INTERVALO' else None,
                'hora_fim': int(d.get('hora_fim', 19) or 19) if modo_envio == 'INTERVALO' else None,
                'horarios_especificos': horarios_especificos,
                'dias_semana': dias_semana,
                'controle_disparos': {},
                'tipo_relatorio': tipo_relatorio,
                'status_destinatarios': status_destinatarios,
                'prioridade': prioridade,
            }
            if d.get('id'):
                AgendamentoDisparo.objects.filter(id=d['id']).update(**defaults)
            else:
                AgendamentoDisparo.objects.create(**defaults)
            return Response({'ok': True})
            
        elif acao == 'excluir':
            try:
                AgendamentoDisparo.objects.filter(id=request.data.get('id')).delete()
            except Exception as e:
                from django.db.utils import ProgrammingError
                if isinstance(e, ProgrammingError) and 'logenvioperformance' in str(e).lower():
                    return Response({
                        'error': 'Tabela de histórico de envio não existe. Execute no servidor: python manage.py migrate crm_app'
                    }, status=503)
                raise
            return Response({'ok': True})

        elif acao == 'historico':
            from django.utils import timezone
            from django.db.utils import ProgrammingError
            try:
                hoje = timezone.localtime(timezone.now()).date()
                logs = LogEnvioPerformance.objects.filter(
                    data_hora__date=hoje
                ).order_by('-data_hora')[:100]
                return Response([{
                    'id': l.id,
                    'regra_nome': l.regra_nome,
                    'data_hora': l.data_hora.isoformat() if l.data_hora else None,
                    'sucesso': l.sucesso,
                    'total_destinos': l.total_destinos,
                    'sucessos': l.sucessos,
                    'falhas': l.falhas,
                    'detalhe': l.detalhe or '',
                } for l in logs])
            except ProgrammingError:
                return Response([])

        return Response({'error': 'Ação inválida'}, status=400)
class CdoiCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        try:
            resumo_msg = None
            data = request.POST
            files = request.FILES
            
            nome_condominio = data.get('nome_condominio') or data.get('cliente')
            if not nome_condominio:
                 return Response({"error": "Nome do condomínio é obrigatório."}, status=400)

            # Upload OneDrive (Mantido igual)
            uploader = OneDriveUploader()
            clean_name = str(nome_condominio).replace('/', '-').strip()
            folder_name = f"{clean_name}_{data.get('cep', '')}"
            
            link_carta = None
            if 'arquivo_carta' in files:
                f = files['arquivo_carta']
                link_carta = uploader.upload_file(f, folder_name, f"CARTA_{f.name}")

            link_fachada = None
            if 'arquivo_fachada' in files:
                f = files['arquivo_fachada']
                link_fachada = uploader.upload_file(f, folder_name, f"FACHADA_{f.name}")

            destinatarios_config = []
            try:
                regra = RegraAutomacao.objects.filter(evento_gatilho='NOVO_CDOI').first()
                if regra and isinstance(regra.destinos_numeros, list):
                    destinatarios_config = [str(x).strip() for x in regra.destinos_numeros if str(x).strip()]
            except Exception:
                destinatarios_config = []

            cdoi = CdoiSolicitacao.objects.create(
                nome_condominio=nome_condominio,
                nome_sindico=data.get('nome_sindico'),
                contato_sindico=data.get('contato'),
                cep=data.get('cep'),
                logradouro=data.get('logradouro'),
                numero=data.get('numero'),
                bairro=data.get('bairro'),
                cidade=data.get('cidade'),
                uf=data.get('uf'),
                latitude=data.get('latitude'),
                longitude=data.get('longitude'),
                infraestrutura_tipo=data.get('infraestrutura'),
                possui_shaft_dg=(data.get('possui_shaft') == 'on'),
                total_hps=data.get('total_hps_final') or 0,
                pre_venda_minima=data.get('prevenda_final') or 0,
                link_carta_sindico=link_carta,
                link_fotos_fachada=link_fachada,
                destinatarios_resumo=",".join(destinatarios_config) if destinatarios_config else None,
                criado_por=request.user,
                status="SEM_TRATAMENTO" # Status inicial novo
            )

            # Salvar Blocos (Melhorado com logs e tratamento de erro)
            blocos_json = data.get('dados_blocos_json')
            import logging
            logger = logging.getLogger(__name__)
            
            if blocos_json:
                try:
                    blocos = json.loads(blocos_json)
                    logger.info(f"[CDOI] Salvando {len(blocos)} blocos para {cdoi.nome_condominio} (ID: {cdoi.id})")
                    
                    blocos_criados = 0
                    for b in blocos:
                        try:
                            CdoiBloco.objects.create(
                                solicitacao=cdoi,
                                nome_bloco=b.get('nome', ''),
                                andares=int(b.get('andares', 0)),
                                unidades_por_andar=int(b.get('aptos', 0)),
                                total_hps_bloco=int(b.get('total', 0))
                            )
                            blocos_criados += 1
                            logger.debug(f"[CDOI] Bloco criado: {b.get('nome')} - {b.get('andares')} andares, {b.get('aptos')} aptos")
                        except Exception as e_bloco_individual:
                            logger.error(f"[CDOI] Erro ao criar bloco individual {b}: {e_bloco_individual}", exc_info=True)
                    
                    logger.info(f"[CDOI] {blocos_criados}/{len(blocos)} blocos salvos com sucesso para {cdoi.nome_condominio}")
                    
                    # Verifica se todos foram salvos
                    if blocos_criados != len(blocos):
                        logger.warning(f"[CDOI] ATENCAO: Apenas {blocos_criados} de {len(blocos)} blocos foram salvos!")
                    
                except json.JSONDecodeError as e_json:
                    logger.error(f"[CDOI] Erro ao decodificar JSON de blocos: {e_json}")
                    logger.error(f"[CDOI] JSON recebido: {blocos_json}")
                except Exception as e_blocos:
                    logger.error(f"[CDOI] Erro ao salvar blocos: {e_blocos}", exc_info=True)
            else:
                logger.warning(f"[CDOI] Nenhum dado de blocos recebido para {cdoi.nome_condominio} (ID: {cdoi.id})")

            # --- ITEM 4: WHATSAPP DESCOMENTADO ---
            try:
                if cdoi.contato_sindico:
                     # Instancia o serviço (certifique-se que as credenciais no .env estão ok)
                     svc = WhatsAppService()
                     msg_text = f"Olá, sua solicitação de CDOI para *{cdoi.nome_condominio}* foi recebida com sucesso! ID: {cdoi.id}. Status: Sem Tratamento."
                     # Removemos o # para ativar
                     svc.enviar_mensagem_texto(cdoi.contato_sindico, msg_text)
            except Exception as e_zap:
                print(f"Erro ao enviar Zap CDOI: {e_zap}")

            # Enviar resumo para acionador e destinatários extras
            try:
                def _limpar_tel(tel):
                    tel_limpo = re.sub(r'\D', '', str(tel or ''))
                    return tel_limpo if tel_limpo else None

                blocos_data = []
                if blocos_json:
                    try:
                        blocos_data = json.loads(blocos_json)
                    except Exception:
                        blocos_data = []

                blocos_txt = "\n".join(
                    [f"- {b.get('nome','')} | {b.get('andares','')} andares | {b.get('aptos','')} aptos/andar | {b.get('total','')} HPs"
                     for b in blocos_data]
                ) or "- (não informado)"

                resumo = (
                    f"✅ *Resumo CDOI*\n"
                    f"ID: {cdoi.id}\n"
                    f"Condomínio: {cdoi.nome_condominio}\n"
                    f"CEP: {cdoi.cep}\n"
                    f"Endereço: {cdoi.logradouro}, {cdoi.numero} - {cdoi.bairro}\n"
                    f"Cidade/UF: {cdoi.cidade}-{cdoi.uf}\n"
                    f"Infraestrutura: {cdoi.infraestrutura_tipo}\n"
                    f"Shaft/DG: {'Sim' if cdoi.possui_shaft_dg else 'Não'}\n"
                    f"Total HPs: {cdoi.total_hps}\n"
                    f"Pré-venda (10%): {cdoi.pre_venda_minima}\n"
                    f"Síndico: {cdoi.nome_sindico}\n"
                    f"Contato: {cdoi.contato_sindico}\n"
                    f"Latitude: {cdoi.latitude or '-'}\n"
                    f"Longitude: {cdoi.longitude or '-'}\n"
                    f"Blocos:\n{blocos_txt}"
                )
                resumo_msg = resumo

                destinatarios = []
                tel_user = getattr(request.user, 'tel_whatsapp', None)
                tel_user = _limpar_tel(tel_user)
                if tel_user:
                    destinatarios.append(tel_user)

                extras = []
                if destinatarios_config:
                    extras.extend(destinatarios_config)
                extras_raw = data.get('destinatarios_resumo') or ''
                extras.extend([t.strip() for t in re.split(r'[;,\\n]', str(extras_raw)) if t.strip()])
                for t in extras:
                    tel_extra = _limpar_tel(t)
                    if tel_extra and tel_extra not in destinatarios:
                        destinatarios.append(tel_extra)

                if destinatarios:
                    svc = WhatsAppService()
                    for tel in destinatarios:
                        svc.enviar_mensagem_texto(tel, resumo)
            except Exception as e_zap_resumo:
                print(f"Erro ao enviar resumo CDOI: {e_zap_resumo}")

            return Response({'mensagem': f'Solicitação enviada! ID: {cdoi.id}', 'resumo': resumo_msg}, status=200)

        except Exception as e:
            print(f"Erro CDOI: {e}")
            return Response({'error': f"Erro ao processar: {str(e)}"}, status=500)

# --- 2. LISTAGEM (Meus Acionamentos) ---
class CdoiListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        # Perfis que veem tudo
        grupos_gestao = ['Diretoria', 'Admin', 'BackOffice']
        eh_gestao = is_member(user, grupos_gestao)

        if eh_gestao:
            queryset = CdoiSolicitacao.objects.all().order_by('-data_criacao')
        else:
            # Usuário comum vê apenas os seus
            queryset = CdoiSolicitacao.objects.filter(criado_por=user).order_by('-data_criacao')

        data = []
        for item in queryset.select_related('criado_por'):
            # Usa username ao invés do nome completo
            criado_por_nome = item.criado_por.username if item.criado_por else '-'
            
            data.append({
                'id': item.id,
                'nome': item.nome_condominio,
                'cidade': f"{item.cidade}-{item.uf}",
                'logradouro': item.logradouro,
                'numero': item.numero,
                'bairro': item.bairro,
                'nome_sindico': item.nome_sindico,
                'contato': item.contato_sindico,
                'total_hps': item.total_hps,
                'prevenda': item.pre_venda_minima,
                'status': item.get_status_display(),
                'status_cod': item.status,
                'observacao': item.observacao or "",
                'data': item.data_criacao.strftime('%d/%m/%Y'),
                'link_fotos_fachada': item.link_fotos_fachada or "",
                'link_carta_sindico': item.link_carta_sindico or "",
                'can_edit': eh_gestao, # Flag para o frontend saber se libera edição
                'criado_por_id': item.criado_por.id if item.criado_por else None,
                'criado_por_nome': criado_por_nome
            })
        
        return Response(data)


class CnpjEstabelecimentosCdoiView(APIView):
    """Lista estabelecimentos CNPJ (base Receita Federal) com filtros múltiplos por CNAE, município e bairro."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import ImportacaoEstabelecimentoCNPJ, DFV

        codigos_municipio = request.query_params.getlist('codigo_municipio') or []
        codigos_municipio = [c.strip() for c in codigos_municipio if c and c.strip()]
        ufs = request.query_params.getlist('uf') or []
        ufs = [u.strip().upper()[:2] for u in ufs if u and u.strip()]
        bairros = request.query_params.getlist('bairro') or []
        bairros = [b.strip() for b in bairros if b and b.strip()]
        cnaes = request.query_params.getlist('cnae') or []
        cnaes = [c.strip().zfill(7) for c in cnaes if c and c.strip()]
        if not cnaes:
            cnaes = ['8112500']
        page = int(request.query_params.get('page', 1))
        limit = min(int(request.query_params.get('limit', 50)), 500)
        format_type = request.query_params.get('format', 'json')

        qs = ImportacaoEstabelecimentoCNPJ.objects.filter(situacao_cadastral='02')
        if cnaes:
            qs = qs.filter(cnae_fiscal__in=cnaes)
        if codigos_municipio:
            qs = qs.filter(codigo_municipio__in=codigos_municipio)
        if ufs:
            qs = qs.filter(uf__in=ufs)
        if bairros:
            qs = qs.filter(bairro__in=bairros)

        total = qs.count()
        # Usar 'export' em vez de 'format' para evitar 404 da content negotiation do DRF
        format_type = (request.query_params.get('export') or request.query_params.get('format') or 'json').strip().lower()

        def _normalize_numero(num):
            """Normaliza número/fachada: parte antes de '(' apenas dígitos (ex: '391 (BL 2)' -> '391')."""
            if num is None:
                return None
            s = str(num).strip().split("(")[0].strip()
            digits = ''.join(c for c in s if c.isdigit())
            return digits if digits else None

        def _build_dfv_map_cep_fachada(pares_cep_num):
            """
            Monta mapa (cep_limpo, num_limpo) -> tipo_viabilidade a partir da base DFV.
            pares_cep_num: set de (cep_limpo, num_limpo) com cep 8 dígitos e num normalizado.
            Também retorna dfv_map_cep_only para fallback quando não acha por CEP+fachada.
            """
            if not pares_cep_num:
                return {}, {}
            ceps_norm = {p[0] for p in pares_cep_num if p[0]}
            dfv_map_cep_num = {}
            dfv_map_cep_only = {}
            try:
                from django.db.models.functions import Replace
                from django.db.models import Value
                qs_dfv = DFV.objects.annotate(
                    cep_limpo=Replace(Replace('cep', Value('-'), Value('')), Value(' '), Value(''))
                ).filter(cep_limpo__in=ceps_norm)
                for dfv_row in qs_dfv.values_list('cep_limpo', 'num_fachada', 'tipo_viabilidade'):
                    cl = (dfv_row[0] or '').strip()
                    num_f = _normalize_numero(dfv_row[1])
                    tv = (dfv_row[2] or '').strip()
                    if not cl:
                        continue
                    key = (cl, num_f)
                    if key not in dfv_map_cep_num or ('VIAVEL' in tv.upper() or 'VIÁVEL' in tv.upper()):
                        dfv_map_cep_num[key] = tv or '-'
                    if cl not in dfv_map_cep_only or ('VIAVEL' in tv.upper() or 'VIÁVEL' in tv.upper()):
                        dfv_map_cep_only[cl] = tv or '-'
            except Exception:
                for row in DFV.objects.filter(cep__in=ceps_norm).values_list('cep', 'num_fachada', 'tipo_viabilidade'):
                    cl = ''.join(c for c in (row[0] or '') if c.isdigit())[:8]
                    num_f = _normalize_numero(row[1])
                    tv = (row[2] or '').strip()
                    if not cl:
                        continue
                    key = (cl, num_f)
                    if key not in dfv_map_cep_num or ('VIAVEL' in tv.upper() or 'VIÁVEL' in tv.upper()):
                        dfv_map_cep_num[key] = tv or '-'
                    if cl not in dfv_map_cep_only or ('VIAVEL' in tv.upper() or 'VIÁVEL' in tv.upper()):
                        dfv_map_cep_only[cl] = tv or '-'
            return dfv_map_cep_num, dfv_map_cep_only

        def _get_retorno_viab(cep_limpo, numero, dfv_map_cep_num, dfv_map_cep_only):
            num_limpo = _normalize_numero(numero)
            if cep_limpo:
                key = (cep_limpo, num_limpo)
                if key in dfv_map_cep_num:
                    return dfv_map_cep_num[key]
                return dfv_map_cep_only.get(cep_limpo, '-')
            return '-'

        if format_type in ('csv', 'xlsx', 'excel'):
            import logging
            _log = logging.getLogger(__name__)
            _log.info('Exportação CNPJ iniciada: %s registros (formato=%s)', total, format_type)
            try:
                rows_list = list(qs.order_by('bairro', 'nome_fantasia').values_list(
                    'cnpj_completo', 'nome_fantasia', 'logradouro', 'numero', 'bairro', 'cep', 'uf', 'codigo_municipio',
                    'ddd_telefone_1', 'telefone_1', 'email', 'cnae_fiscal', 'situacao_cadastral'
                )[:50000])
                pares_cep_num = set()
                for row in rows_list:
                    c = ''.join(x for x in (row[5] or '') if x.isdigit())[:8]
                    n = _normalize_numero(row[3])
                    if c:
                        pares_cep_num.add((c, n))
                dfv_map_cep_num, dfv_map_cep_only = _build_dfv_map_cep_fachada(pares_cep_num)
                from .ibge_municipios import get_nome_municipio_por_codigo
                from .services.cep_lookup import get_municipio_por_cep

                if format_type == 'xlsx' or format_type == 'excel':
                    import io
                    import openpyxl
                    from django.http import HttpResponse
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = 'Estabelecimentos'
                    headers = [
                        'CNPJ', 'Nome Fantasia', 'Logradouro', 'Numero', 'Bairro', 'CEP', 'UF', 'Cod.Municipio', 'Município',
                        'Telefone', 'Email', 'CNAE', 'Situacao', 'Retorno Viabilidade (DFV)'
                    ]
                    ws.append(headers)
                    cache_cep_municipio = {}
                    viacep_limit = 500  # evita milhares de chamadas API em exportações grandes
                    for row in rows_list:
                        cep_limpo = ''.join(x for x in (row[5] or '') if x.isdigit())[:8]
                        retorno_viab = _get_retorno_viab(cep_limpo, row[3], dfv_map_cep_num, dfv_map_cep_only)
                        nome_mun = get_nome_municipio_por_codigo(row[7], uf=row[6]) or ''
                        if not nome_mun and cep_limpo and len(cache_cep_municipio) < viacep_limit:
                            nome_mun = get_municipio_por_cep(cep_limpo, cache=cache_cep_municipio) or ''
                        ddd, tel = row[8], row[9]
                        telefone = f"{ddd or ''}{tel or ''}".strip()
                        ws.append(list(row[:8]) + [nome_mun] + [telefone] + list(row[10:13]) + [retorno_viab])
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    _log.info('Exportação CNPJ Excel concluída: %s linhas', len(rows_list))
                    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="cnpj_estabelecimentos_cidade_bairro.xlsx"'
                    return response
                else:
                    import csv
                    from django.http import HttpResponse
                    response = HttpResponse(content_type='text/csv; charset=utf-8')
                    response['Content-Disposition'] = 'attachment; filename="cnpj_estabelecimentos_cidade_bairro.csv"'
                    response.write('\ufeff')
                    writer = csv.writer(response, delimiter=';')
                    writer.writerow([
                        'CNPJ', 'Nome Fantasia', 'Logradouro', 'Numero', 'Bairro', 'CEP', 'UF', 'Cod.Municipio', 'Município',
                        'Telefone', 'Email', 'CNAE', 'Situacao', 'Retorno Viabilidade (DFV)'
                    ])
                    cache_cep_municipio = {}
                    viacep_limit = 500
                    for row in rows_list:
                        cep_limpo = ''.join(x for x in (row[5] or '') if x.isdigit())[:8]
                        retorno_viab = _get_retorno_viab(cep_limpo, row[3], dfv_map_cep_num, dfv_map_cep_only)
                        nome_mun = get_nome_municipio_por_codigo(row[7], uf=row[6]) or ''
                        if not nome_mun and cep_limpo and len(cache_cep_municipio) < viacep_limit:
                            nome_mun = get_municipio_por_cep(cep_limpo, cache=cache_cep_municipio) or ''
                        ddd, tel = row[8], row[9]
                        out = list(row[:8]) + [nome_mun] + [f"{ddd or ''}{tel or ''}".strip()] + list(row[10:13]) + [retorno_viab]
                        writer.writerow([(c or '') for c in out])
                    return response
            except Exception as e:
                import logging
                logging.getLogger(__name__).exception('Erro ao exportar CNPJ estabelecimentos')
                return Response({'error': str(e), 'detail': 'Erro ao gerar arquivo. Tente novamente.'}, status=500)

        start = (page - 1) * limit
        rows = qs.order_by('bairro', 'nome_fantasia')[start:start + limit]
        pares_cep_num = set()
        for r in rows:
            c = ''.join(x for x in (r.cep or '') if x.isdigit())[:8]
            n = _normalize_numero(r.numero)
            if c:
                pares_cep_num.add((c, n))
        dfv_map_cep_num, dfv_map_cep_only = _build_dfv_map_cep_fachada(pares_cep_num)
        from .ibge_municipios import get_nome_municipio_por_codigo
        from .services.cep_lookup import get_municipio_por_cep
        cache_cep_municipio = {}
        data = []
        for r in rows:
            cep_limpo = ''.join(x for x in (r.cep or '') if x.isdigit())[:8]
            retorno_viab = _get_retorno_viab(cep_limpo, r.numero, dfv_map_cep_num, dfv_map_cep_only)
            nome_mun = get_nome_municipio_por_codigo(r.codigo_municipio, uf=r.uf) or ''
            if not nome_mun and cep_limpo:
                nome_mun = get_municipio_por_cep(cep_limpo, cache=cache_cep_municipio) or ''
            data.append({
                'cnpj': r.cnpj_completo or '',
                'nome_fantasia': r.nome_fantasia or '',
                'logradouro': r.logradouro or '',
                'numero': r.numero or '',
                'bairro': r.bairro or '',
                'cep': r.cep or '',
                'uf': r.uf or '',
                'codigo_municipio': r.codigo_municipio or '',
                'nome_municipio': nome_mun,
                'telefone': (r.ddd_telefone_1 or '') + (r.telefone_1 or ''),
                'email': r.email or '',
                'cnae': r.cnae_fiscal or '',
                'retorno_viabilidade': retorno_viab,
            })
        return Response({'results': data, 'total': total, 'page': page, 'limit': limit})


class CnpjCnaesCdoiView(APIView):
    """Lista CNAEs distintos da base CNPJ para multi-select (CDOI)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import ImportacaoEstabelecimentoCNPJ
        from django.db.models import Count
        qs = (
            ImportacaoEstabelecimentoCNPJ.objects.filter(cnae_fiscal__isnull=False)
            .exclude(cnae_fiscal='')
            .filter(situacao_cadastral='02')
            .values('cnae_fiscal')
            .annotate(qtd=Count('id'))
            .order_by('cnae_fiscal')
        )
        lista = [{'cnae': x['cnae_fiscal'], 'qtd': x['qtd']} for x in qs[:200]]
        return Response({'cnaes': lista})


class CnpjBairrosCdoiView(APIView):
    """Lista bairros distintos da base CNPJ para multi-select (CDOI). Opcional: filtrar por codigo_municipio/uf."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import ImportacaoEstabelecimentoCNPJ
        from django.db.models import Count
        qs = (
            ImportacaoEstabelecimentoCNPJ.objects.filter(situacao_cadastral='02')
            .exclude(bairro__isnull=True)
            .exclude(bairro='')
        )
        codigos = request.query_params.getlist('codigo_municipio')
        codigos = [c.strip() for c in codigos if c and c.strip()]
        ufs_param = request.query_params.getlist('uf')
        ufs_param = [u.strip().upper()[:2] for u in ufs_param if u and u.strip()]
        if codigos:
            qs = qs.filter(codigo_municipio__in=codigos)
        if ufs_param:
            qs = qs.filter(uf__in=ufs_param)
        qs = qs.values('bairro').annotate(qtd=Count('id')).order_by('bairro')[:1000]
        lista = [{'bairro': x['bairro'], 'qtd': x['qtd']} for x in qs]
        return Response({'bairros': lista})


class CnpjMunicipiosCdoiView(APIView):
    """Lista códigos de município distintos da base CNPJ para dropdown (CDOI). Inclui nome do município via IBGE."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import ImportacaoEstabelecimentoCNPJ
        from django.db.models import Count
        from .ibge_municipios import get_nome_municipio_por_codigo
        qs = (
            ImportacaoEstabelecimentoCNPJ.objects.filter(codigo_municipio__isnull=False)
            .exclude(codigo_municipio='')
            .values('codigo_municipio', 'uf')
            .annotate(qtd=Count('id'))
            .order_by('uf', 'codigo_municipio')
        )
        lista = []
        for x in qs[:500]:
            cod = x['codigo_municipio']
            nome = get_nome_municipio_por_codigo(cod, uf=x['uf'])
            lista.append({
                'codigo_municipio': cod,
                'uf': x['uf'],
                'qtd': x['qtd'],
                'nome_municipio': nome or '',
            })
        return Response({'municipios': lista})


class CnpjUfsCdoiView(APIView):
    """Lista UFs distintas da base CNPJ para multi-select com pesquisa (CDOI)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from .models import ImportacaoEstabelecimentoCNPJ
        from django.db.models import Count
        qs = (
            ImportacaoEstabelecimentoCNPJ.objects.filter(situacao_cadastral='02')
            .exclude(uf__isnull=True)
            .exclude(uf='')
            .values('uf')
            .annotate(qtd=Count('id'))
            .order_by('uf')
        )
        lista = [{'uf': x['uf'], 'qtd': x['qtd']} for x in qs]
        return Response({'ufs': lista})


class CdoiDashboardView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum

        queryset = CdoiSolicitacao.objects.all()
        total_acionamentos = queryset.count()
        total_hps = queryset.aggregate(total=Sum('total_hps')).get('total') or 0
        total_prevenda = queryset.aggregate(total=Sum('pre_venda_minima')).get('total') or 0

        por_status = (
            queryset.values('status')
            .annotate(qtd=Count('id'))
            .order_by('-qtd')
        )

        # Cruzar CEP + fachada (numero) com vendas
        pares = []
        for item in queryset.values('cep', 'numero'):
            cep_limpo = re.sub(r'\D', '', str(item.get('cep') or ''))
            numero = str(item.get('numero') or '').strip()
            if cep_limpo and numero:
                pares.append((cep_limpo, numero))

        vendas_realizadas = 0
        if pares:
            ceps = list({p[0] for p in pares})
            pares_set = set(pares)
            vendas = Venda.objects.filter(cep__in=ceps).values('cep', 'numero_residencia')
            for v in vendas:
                cep_limpo = re.sub(r'\D', '', str(v.get('cep') or ''))
                numero = str(v.get('numero_residencia') or '').strip()
                if (cep_limpo, numero) in pares_set:
                    vendas_realizadas += 1

        return Response({
            'success': True,
            'total_acionamentos': total_acionamentos,
            'total_hps': total_hps,
            'total_prevenda': total_prevenda,
            'vendas_realizadas': vendas_realizadas,
            'por_status': list(por_status),
        })

# --- 3. EDIÇÃO DE STATUS (Apenas Gestão) ---
class CdoiUpdateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request, pk):
        user = request.user
        if not is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({"error": "Acesso negado."}, status=403)

        try:
            cdoi = CdoiSolicitacao.objects.get(pk=pk)
            # Usa select_related/prefetch_related para otimizar a consulta
            blocos_queryset = cdoi.blocos.all().order_by('nome_bloco')
            blocos = [
                {
                    'nome': b.nome_bloco,
                    'andares': b.andares,
                    'aptos': b.unidades_por_andar,
                    'total': b.total_hps_bloco,
                }
                for b in blocos_queryset
            ]
            
            # Log para debug detalhado
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"[CDOI] Editando {cdoi.nome_condominio} (ID: {pk}) - {len(blocos)} blocos encontrados")
            
            # Log detalhado se não houver blocos mas houver total_hps
            if len(blocos) == 0 and cdoi.total_hps > 0:
                logger.warning(
                    f"[CDOI] ATENCAO: Condominio {cdoi.nome_condominio} (ID: {pk}) "
                    f"tem total_hps={cdoi.total_hps} mas nenhum bloco cadastrado!"
                )
                logger.warning(f"[CDOI] Isso indica que os blocos nao foram salvos durante a criacao.")

            # Monta nome do criador
            criado_por_nome = '-'
            if cdoi.criado_por:
                if cdoi.criado_por.first_name or cdoi.criado_por.last_name:
                    criado_por_nome = f"{cdoi.criado_por.first_name or ''} {cdoi.criado_por.last_name or ''}".strip()
                else:
                    criado_por_nome = cdoi.criado_por.username or '-'
            
            data = {
                'id': cdoi.id,
                'nome_condominio': cdoi.nome_condominio,
                'cep': cdoi.cep,
                'logradouro': cdoi.logradouro or '',
                'numero': cdoi.numero,
                'bairro': cdoi.bairro or '',
                'cidade': cdoi.cidade or '',
                'uf': cdoi.uf or '',
                'nome_sindico': cdoi.nome_sindico,
                'contato': cdoi.contato_sindico,
                'infraestrutura': cdoi.infraestrutura_tipo,
                'possui_shaft': cdoi.possui_shaft_dg,
                'latitude': cdoi.latitude or '',
                'longitude': cdoi.longitude or '',
                'blocos': blocos,
                'criado_por_id': cdoi.criado_por.id if cdoi.criado_por else None,
                'criado_por_nome': criado_por_nome,
                'link_fotos_fachada': cdoi.link_fotos_fachada or '',
                'link_carta_sindico': cdoi.link_carta_sindico or '',
            }
            return Response(data)
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Solicitação não encontrada."}, status=404)

    def patch(self, request, pk):
        user = request.user
        if not is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({"error": "Acesso negado."}, status=403)

        try:
            def _to_int(value, default=0):
                if value is None:
                    return default
                if isinstance(value, str) and not value.strip():
                    return default
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return default

            cdoi = CdoiSolicitacao.objects.get(pk=pk)
            # JSON (modal de status) e multipart/form usam request.data; POST puro não inclui JSON.
            data = request.data
            files = request.FILES
            import logging
            logger = logging.getLogger(__name__)

            # Processar upload de arquivos se fornecidos
            if files:
                from crm_app.onedrive_service import OneDriveUploader
                uploader = OneDriveUploader()
                clean_name = str(cdoi.nome_condominio).replace('/', '-').strip()
                folder_name = f"{clean_name}_{cdoi.cep or ''}"
                
                if 'arquivo_carta' in files:
                    f = files['arquivo_carta']
                    link_carta = uploader.upload_file(f, folder_name, f"CARTA_{f.name}")
                    cdoi.link_carta_sindico = link_carta
                    logger.info(f"[CDOI] Arquivo carta atualizado para {cdoi.nome_condominio} (ID: {pk})")
                
                if 'arquivo_fachada' in files:
                    f = files['arquivo_fachada']
                    link_fachada = uploader.upload_file(f, folder_name, f"FACHADA_{f.name}")
                    cdoi.link_fotos_fachada = link_fachada
                    logger.info(f"[CDOI] Arquivo fachada atualizado para {cdoi.nome_condominio} (ID: {pk})")

            # Campos pontuais
            if data.get('status'):
                cdoi.status = data.get('status')
            if 'observacao' in data:
                cdoi.observacao = data.get('observacao')
            if data.get('nome_condominio'):
                cdoi.nome_condominio = data.get('nome_condominio')
            
            # Atualizar blocos se fornecidos
            blocos_json = data.get('dados_blocos_json') or data.get('input_blocos_json')
            if blocos_json:
                try:
                    # Remove blocos antigos
                    cdoi.blocos.all().delete()
                    logger.info(f"[CDOI] Blocos antigos removidos para {cdoi.nome_condominio} (ID: {pk})")
                    
                    # Cria novos blocos
                    blocos = json.loads(blocos_json)
                    blocos_criados = 0
                    for b in blocos:
                        try:
                            CdoiBloco.objects.create(
                                solicitacao=cdoi,
                                nome_bloco=b.get('nome', ''),
                                andares=_to_int(b.get('andares', 0)),
                                unidades_por_andar=_to_int(b.get('aptos', 0)),
                                total_hps_bloco=_to_int(b.get('total', 0))
                            )
                            blocos_criados += 1
                        except Exception as e_bloco:
                            logger.error(f"[CDOI] Erro ao criar bloco na edicao: {e_bloco}")
                    
                    logger.info(f"[CDOI] {blocos_criados} blocos atualizados para {cdoi.nome_condominio} (ID: {pk})")
                except json.JSONDecodeError as e_json:
                    logger.error(f"[CDOI] Erro ao decodificar JSON de blocos na edicao: {e_json}")
                except Exception as e_blocos:
                    logger.error(f"[CDOI] Erro ao atualizar blocos: {e_blocos}", exc_info=True)
            if data.get('cep'):
                cdoi.cep = data.get('cep')
            if data.get('numero'):
                cdoi.numero = data.get('numero')
            if data.get('nome_sindico'):
                cdoi.nome_sindico = data.get('nome_sindico')
            if data.get('contato'):
                cdoi.contato_sindico = data.get('contato')
            if data.get('infraestrutura'):
                cdoi.infraestrutura_tipo = data.get('infraestrutura')
            if 'possui_shaft' in data:
                val = data.get('possui_shaft')
                if isinstance(val, str):
                    cdoi.possui_shaft_dg = val.lower() in ['on', 'true', '1']
                else:
                    cdoi.possui_shaft_dg = bool(val)
            if data.get('latitude'):
                cdoi.latitude = data.get('latitude')
            if data.get('longitude'):
                cdoi.longitude = data.get('longitude')
            
            # Permite alterar quem acionou (criado_por)
            if 'criado_por_id' in data and data.get('criado_por_id'):
                try:
                    from usuarios.models import Usuario
                    novo_criador = Usuario.objects.get(id=data.get('criado_por_id'))
                    cdoi.criado_por = novo_criador
                except Usuario.DoesNotExist:
                    pass  # Ignora se usuário não existir

            # Atualiza total_hps e pre_venda se fornecidos
            if 'total_hps_final' in data:
                cdoi.total_hps = _to_int(data.get('total_hps_final'), cdoi.total_hps or 0)
            if 'prevenda_final' in data:
                cdoi.pre_venda_minima = _to_int(data.get('prevenda_final'), cdoi.pre_venda_minima or 0)
            
            cdoi.save()
            return Response({"mensagem": "Atualizado com sucesso!"})
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Solicitação não encontrada."}, status=404)

    def put(self, request, pk):
        user = request.user
        if not is_member(user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({"error": "Acesso negado."}, status=403)

        try:
            cdoi = CdoiSolicitacao.objects.get(pk=pk)
            data = request.data

            # Atualiza campos principais
            cdoi.nome_condominio = data.get('nome_condominio', cdoi.nome_condominio)
            cdoi.nome_sindico = data.get('nome_sindico', cdoi.nome_sindico)
            cdoi.contato_sindico = data.get('contato', cdoi.contato_sindico)
            cdoi.cep = data.get('cep', cdoi.cep)
            cdoi.logradouro = data.get('logradouro', cdoi.logradouro)
            cdoi.numero = data.get('numero', cdoi.numero)
            cdoi.bairro = data.get('bairro', cdoi.bairro)
            cdoi.cidade = data.get('cidade', cdoi.cidade)
            cdoi.uf = data.get('uf', cdoi.uf)
            cdoi.latitude = data.get('latitude', cdoi.latitude)
            cdoi.longitude = data.get('longitude', cdoi.longitude)
            cdoi.infraestrutura_tipo = data.get('infraestrutura', cdoi.infraestrutura_tipo)
            # checkbox pode vir como 'on' ou 'true'
            possui_shaft_val = data.get('possui_shaft', None)
            if isinstance(possui_shaft_val, str):
                cdoi.possui_shaft_dg = possui_shaft_val.lower() in ['on', 'true', '1']
            elif isinstance(possui_shaft_val, bool):
                cdoi.possui_shaft_dg = possui_shaft_val

            # Totais
            try:
                cdoi.total_hps = int(data.get('total_hps_final') or data.get('input_total_hps') or cdoi.total_hps or 0)
            except Exception:
                pass
            try:
                cdoi.pre_venda_minima = int(data.get('prevenda_final') or data.get('input_prevenda') or cdoi.pre_venda_minima or 0)
            except Exception:
                pass

            cdoi.save()

            # Atualiza blocos se enviados
            blocos_json = data.get('dados_blocos_json') or data.get('input_blocos_json')
            if blocos_json:
                try:
                    import json as _json
                    blocos = _json.loads(blocos_json)
                    # Remove blocos existentes e recria
                    cdoi.blocos.all().delete()
                    for b in blocos:
                        CdoiBloco.objects.create(
                            solicitacao=cdoi,
                            nome_bloco=b.get('nome'),
                            andares=int(b.get('andares') or 0),
                            unidades_por_andar=int(b.get('aptos') or 0),
                            total_hps_bloco=int(b.get('total') or 0)
                        )
                except Exception as e:
                    return Response({"error": f"Erro ao atualizar blocos: {e}"}, status=400)

            return Response({"mensagem": "Solicitação atualizada com sucesso!"})
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Solicitação não encontrada."}, status=404)

    def delete(self, request, pk):
        user = request.user
        if not is_member(user, ['Diretoria', 'Admin']):
            return Response({"error": "Acesso negado."}, status=403)

        try:
            cdoi = CdoiSolicitacao.objects.get(pk=pk)
            # Exclusão lógica: marca como CANCELADA
            cdoi.status = 'CANCELADA'
            cdoi.save()
            return Response({"mensagem": "Solicitação marcada como cancelada."}, status=200)
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Solicitação não encontrada."}, status=404)

# --- AQUI ESTÁ A CORREÇÃO: A FUNÇÃO DEVE FICAR FORA DA CLASSE ---
# Note que não tem espaço antes do 'def'

def page_cdoi_novo(request):
    """View simples para abrir a página no navegador"""
    can_config = is_member(request.user, ['Admin', 'Diretoria']) if request.user.is_authenticated else False
    return render(request, 'cdoi_form.html', {'can_config': can_config})


def prevenda_publica_landing(request, codigo):
    """View pública para renderizar a landing page de pré-vendas"""
    try:
        link_publico = LinkPublicoPreVenda.objects.select_related('acionamento').get(
            codigo_unico=codigo,
            ativo=True
        )
        
        # Gera URL completa para QR Code
        url_completa = link_publico.get_url_publica(request)
        
        context = {
            'link': link_publico,
            'acionamento': link_publico.acionamento,
            'url_completa': url_completa,
            'codigo': codigo
        }
        
        return render(request, 'prevenda_publica.html', context)
        
    except LinkPublicoPreVenda.DoesNotExist:
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound('<h1>Link não encontrado ou inativo</h1>')


# =============================================================================
# PRÉ-VENDAS (Formulário Simples)
# =============================================================================

# --- NOVAS VIEWS PARA PRÉ-VENDAS PÚBLICAS ---

class GerarLinkPublicoPreVendaView(APIView):
    """API para gerar/criar link público único por acionamento CDOI"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, cdoi_id):
        try:
            # Busca o acionamento CDOI
            acionamento = CdoiSolicitacao.objects.get(pk=cdoi_id)
            
            # Verifica se já existe um link ativo para este acionamento
            link_existente = LinkPublicoPreVenda.objects.filter(
                acionamento=acionamento,
                ativo=True
            ).first()
            
            # Upload de imagem/banner se fornecido
            imagem_banner = None
            if 'imagem_banner' in request.FILES:
                from crm_app.onedrive_service import OneDriveUploader
                import logging
                logger = logging.getLogger(__name__)
                uploader = OneDriveUploader()
                clean_name = acionamento.nome_condominio.replace('/', '-').strip()
                folder_name = f"PREVENDAS_{clean_name}"
                f = request.FILES['imagem_banner']
                # Processar imagem: redimensionar e comprimir antes do upload
                try:
                    from PIL import Image
                    import io
                    from django.core.files.uploadedfile import InMemoryUploadedFile
                    import sys
                    
                    # Ler imagem
                    img = Image.open(f)
                    
                    # Converter para RGB se necessário (para remover transparência de PNG)
                    if img.mode in ('RGBA', 'LA', 'P'):
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'P':
                            img = img.convert('RGBA')
                        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                        img = background
                    
                    # Redimensionar se muito grande (máximo 1200px de largura, mantém proporção)
                    MAX_WIDTH = 1200
                    if img.width > MAX_WIDTH:
                        ratio = MAX_WIDTH / img.width
                        new_height = int(img.height * ratio)
                        try:
                            img = img.resize((MAX_WIDTH, new_height), Image.Resampling.LANCZOS)
                        except AttributeError:
                            # Compatibilidade com versões antigas do Pillow
                            img = img.resize((MAX_WIDTH, new_height), Image.LANCZOS)
                    
                    # Salvar em buffer com compressão
                    output = io.BytesIO()
                    img.save(output, format='JPEG', quality=85, optimize=True)
                    output.seek(0)
                    
                    # Criar novo arquivo em memória
                    filename = f.name.rsplit('.', 1)[0] + '.jpg'  # Mudar extensão para .jpg
                    f_processed = InMemoryUploadedFile(
                        output, None, filename, 'image/jpeg', sys.getsizeof(output), None
                    )
                    f = f_processed
                except ImportError:
                    # Se PIL não estiver disponível, usar imagem original
                    import logging
                    logger_prevenda = logging.getLogger(__name__)
                    logger_prevenda.warning("[Pré-venda] Pillow não disponível, usando imagem original sem processamento")
                except Exception as e:
                    import logging
                    logger_prevenda = logging.getLogger(__name__)
                    logger_prevenda.warning(f"[Pré-venda] Erro ao processar imagem: {e}, usando original")
                
                # Usa upload_file_and_get_download_url para obter URL de download direto (funciona melhor em mobile)
                try:
                    import logging
                    logger_prevenda = logging.getLogger(__name__)
                    imagem_banner = uploader.upload_file_and_get_download_url(f, folder_name, f"BANNER_{f.name}")
                    # Valida se a URL parece ser uma URL de download direto (não SharePoint webUrl)
                    if imagem_banner and ('sharepoint.com/_layouts' in imagem_banner.lower() or 'onedrive.live.com' in imagem_banner.lower()):
                        logger_prevenda.warning(f"[Pré-venda] URL parece ser SharePoint webUrl (pode causar CORS): {imagem_banner[:100]}")
                        # Não rejeitar, mas avisar - pode funcionar em alguns casos
                    logger_prevenda.info(f"[Pré-venda] Banner uploaded com sucesso: {imagem_banner[:80] if imagem_banner else 'None'}...")
                except Exception as e:
                    import logging
                    logger_prevenda = logging.getLogger(__name__)
                    logger_prevenda.error(f"[Pré-venda] Erro ao fazer upload do banner: {e}")
                    return Response({'error': f'Erro ao fazer upload da imagem: {str(e)}'}, status=500)
            elif request.POST.get('imagem_banner_url'):
                imagem_banner = request.POST.get('imagem_banner_url').strip()
            
            if link_existente:
                # Se já existe, atualiza a imagem se fornecida
                if imagem_banner:
                    link_existente.imagem_banner = imagem_banner
                    link_existente.save(update_fields=['imagem_banner'])
                url_publica = link_existente.get_url_publica(request)
                return Response({
                    'mensagem': 'Link já existe para este acionamento' + (' (imagem atualizada)' if imagem_banner else ''),
                    'link': url_publica,
                    'codigo': link_existente.codigo_unico,
                    'id': link_existente.id
                }, status=200)
            
            # Cria novo link público
            link_publico = LinkPublicoPreVenda.objects.create(
                acionamento=acionamento,
                imagem_banner=imagem_banner,
                criado_por=request.user
            )
            
            url_publica = link_publico.get_url_publica(request)
            
            return Response({
                'mensagem': 'Link público criado com sucesso!',
                'link': url_publica,
                'codigo': link_publico.codigo_unico,
                'id': link_publico.id
            }, status=201)
            
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Acionamento CDOI não encontrado."}, status=404)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao gerar link público: {e}", exc_info=True)
            return Response({'error': f"Erro ao processar: {str(e)}"}, status=500)


class PreVendaPublicaFormView(APIView):
    """API PÚBLICA (sem autenticação) para receber formulário da landing page"""
    permission_classes = [permissions.AllowAny]

    def post(self, request, codigo):
        try:
            # Busca o link público pelo código
            link_publico = LinkPublicoPreVenda.objects.get(codigo_unico=codigo, ativo=True)
            
            data = request.data
            
            nome_cliente = data.get('nome_cliente', '').strip()
            telefone_whatsapp = data.get('telefone_whatsapp', '').strip()
            email = data.get('email', '').strip()
            bloco = data.get('bloco', '').strip()
            apartamento = data.get('apartamento', '').strip()
            
            # Validações
            if not nome_cliente:
                return Response({"error": "Nome é obrigatório."}, status=400)
            if not telefone_whatsapp:
                return Response({"error": "Telefone/WhatsApp é obrigatório."}, status=400)
            
            # Obtém IP de origem
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_origem = x_forwarded_for.split(',')[0]
            else:
                ip_origem = request.META.get('REMOTE_ADDR')
            
            # Cria a pré-venda
            prevenda = PreVenda.objects.create(
                link_publico=link_publico,
                nome_cliente=nome_cliente,
                telefone_whatsapp=telefone_whatsapp,
                email=email if email else None,
                bloco=bloco if bloco else None,
                apartamento=apartamento if apartamento else None,
                ip_origem=ip_origem
            )
            
            return Response({
                'mensagem': 'Cadastro realizado com sucesso! Entraremos em contato em breve.',
                'id': prevenda.id
            }, status=201)
            
        except LinkPublicoPreVenda.DoesNotExist:
            return Response({"error": "Link não encontrado ou inativo."}, status=404)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao processar pré-venda pública: {e}", exc_info=True)
            return Response({'error': f"Erro ao processar: {str(e)}"}, status=500)


class PreVendasPorAcionamentoView(APIView):
    """API para listar pré-vendas de um acionamento CDOI específico"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, cdoi_id):
        try:
            acionamento = CdoiSolicitacao.objects.get(pk=cdoi_id)
            
            # Busca o link público do acionamento
            link_publico = LinkPublicoPreVenda.objects.filter(
                acionamento=acionamento,
                ativo=True
            ).first()
            
            if not link_publico:
                return Response({
                    'prevendas': [],
                    'total': 0,
                    'link_existe': False
                })
            
            # Busca todas as pré-vendas do link
            prevendas = PreVenda.objects.filter(
                link_publico=link_publico
            ).order_by('-data_cadastro')
            
            data = []
            for item in prevendas:
                data.append({
                    'id': item.id,
                    'nome_cliente': item.nome_cliente,
                    'telefone_whatsapp': item.telefone_whatsapp,
                    'email': item.email or '',
                    'bloco': item.bloco or '',
                    'apartamento': item.apartamento or '',
                    'data_cadastro': item.data_cadastro.strftime('%d/%m/%Y %H:%M')
                })
            
            return Response({
                'prevendas': data,
                'total': len(data),
                'link_existe': True,
                'link_codigo': link_publico.codigo_unico,
                'link_url': link_publico.get_url_publica(request)
            })
            
        except CdoiSolicitacao.DoesNotExist:
            return Response({"error": "Acionamento não encontrado."}, status=404)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao listar pré-vendas do acionamento: {e}", exc_info=True)
            return Response({'error': str(e)}, status=500)




# =============================================================================
# IMPORTAÇÃO DE AGENDAMENTOS E TAREFAS
# =============================================================================

class ImportacaoAgendamentoView(APIView):
    permission_classes = [CheckAPIPermission]
    resource_name = 'importacao_agendamento'
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        """
        Recebe arquivo Excel/XLSB e inicia processamento assíncrono.
        Retorna imediatamente com log_id para acompanhamento.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=400)

        # 1. Criar log inicial
        log = LogImportacaoAgendamento.objects.create(
            usuario=request.user,
            nome_arquivo=file_obj.name,
            status='PROCESSANDO'
        )

        # 2. Ler arquivo para memória
        try:
            file_content = file_obj.read()
            file_name = file_obj.name
        except Exception as e:
            log.status = 'ERRO'
            log.mensagem_erro = f'Erro ao ler arquivo: {str(e)}'
            log.finalizado_em = timezone.now()
            log.save()
            return Response({'error': f'Erro ao ler arquivo: {str(e)}'}, status=400)

        # 3. Iniciar thread de processamento
        thread = threading.Thread(
            target=self._processar_agendamento_interno,
            args=(log.id, file_content, file_name)
        )
        thread.daemon = True
        thread.start()

        # 4. Retornar imediatamente
        return Response({
            'success': True,
            'status': 'PROCESSANDO',
            'background': True,
            'log_id': log.id,
            'message': 'Processamento iniciado! Acompanhe o progresso na aba Histórico.'
        }, status=200)

    def _processar_agendamento_interno(self, log_id, file_content, file_name):
        """Processa importação em background"""
        try:
            log = LogImportacaoAgendamento.objects.get(id=log_id)
            
            # Verificar se foi cancelado antes de começar
            log.refresh_from_db()
            if log.status == 'CANCELADO':
                return
            
            # Ler Excel/XLSB em memória (preservando zeros à esquerda em nr_ordem_venda)
            try:
                file_buffer = BytesIO(file_content)
                if file_name.endswith('.xlsb'):
                    # Para .xlsb, tentar usar dtype/converters para forçar nr_ordem_venda como string
                    file_buffer.seek(0)
                    df_sample = pd.read_excel(file_buffer, engine='pyxlsb', nrows=1)
                    file_buffer.seek(0)
                    # Encontrar coluna que parece ser NR_ORDEM_VENDA (case-insensitive, antes da normalização)
                    nr_ordem_venda_col = None
                    for col in df_sample.columns:
                        col_normalizado = str(col).strip().upper().replace(' ', '_')
                        if col_normalizado == 'NR_ORDEM_VENDA':
                            nr_ordem_venda_col = col
                            break
                    # Tentar usar dtype primeiro (se suportado), senão converters
                    if nr_ordem_venda_col:
                        try:
                            df = pd.read_excel(file_buffer, engine='pyxlsb', dtype={nr_ordem_venda_col: str})
                        except (TypeError, ValueError):
                            try:
                                df = pd.read_excel(file_buffer, engine='pyxlsb', converters={nr_ordem_venda_col: lambda x: str(x) if pd.notna(x) else ''})
                            except:
                                df = pd.read_excel(file_buffer, engine='pyxlsb')
                    else:
                        df = pd.read_excel(file_buffer, engine='pyxlsb')
                elif file_name.endswith('.xlsx'):
                    # Para .xlsx, usar openpyxl para forçar nr_ordem_venda como texto
                    try:
                        from openpyxl import load_workbook
                        file_buffer.seek(0)
                        wb = load_workbook(file_buffer, data_only=False, read_only=True)
                        ws = wb.active
                        
                        # Ler cabeçalhos
                        headers = [cell.value for cell in ws[1]]
                        # Encontrar índice da coluna NR_ORDEM_VENDA
                        nr_ordem_venda_idx = None
                        for idx, header in enumerate(headers):
                            if header and str(header).strip().upper().replace(' ', '_') == 'NR_ORDEM_VENDA':
                                nr_ordem_venda_idx = idx
                                break
                        
                        # Ler dados: para NR_ORDEM_VENDA, forçar como string (preserva zeros)
                        data = []
                        for row in ws.iter_rows(min_row=2, values_only=False):
                            row_data = []
                            for idx, cell in enumerate(row):
                                if idx == nr_ordem_venda_idx and cell.value is not None:
                                    # Para NR_ORDEM_VENDA: sempre converter para string (preserva zeros à esquerda)
                                    # Se célula tem formato texto (@) ou se é string, usar direto
                                    if cell.data_type == 's':
                                        row_data.append(str(cell.value))
                                    else:
                                        # Se for número, converter para string formatando sem perder zeros
                                        val = cell.value
                                        # Formatar como string inteiro (sem decimais)
                                        if isinstance(val, (int, float)):
                                            # Para números, converter para int primeiro para evitar .0
                                            if isinstance(val, float) and val.is_integer():
                                                val = int(val)
                                            row_data.append(str(val))
                                        else:
                                            row_data.append(str(val) if val is not None else '')
                                else:
                                    row_data.append(cell.value)
                            data.append(row_data)
                        
                        wb.close()
                        df = pd.DataFrame(data, columns=headers)
                    except Exception as e_openpyxl:
                        # Se openpyxl falhar, ler normalmente
                        file_buffer.seek(0)
                        df = pd.read_excel(file_buffer)
                elif file_name.endswith('.xls'):
                    df = pd.read_excel(file_buffer)
                else:
                    log.status = 'ERRO'
                    log.mensagem_erro = 'Formato inválido. Envie .xlsx, .xls ou .xlsb'
                    log.finalizado_em = timezone.now()
                    log.save()
                    return
            except Exception as e:
                log.status = 'ERRO'
                log.mensagem_erro = f'Erro ao ler arquivo: {str(e)}'
                log.finalizado_em = timezone.now()
                log.save()
                return

            # Normaliza nomes das colunas
            df.columns = [str(col).strip().lower() for col in df.columns]

            log.total_linhas = len(df)
            log.save()

            # Converte datas
            campos_data = ['dt_abertura_ba', 'dt_execucao_particao', 'dt_agendamento']
            campos_datetime = ['dt_inicio_agendamento', 'dt_fim_agendamento', 'dt_inicio_execucao_real', 'dt_fim_execucao_real']

            for campo in campos_data:
                if campo in df.columns:
                    df[campo] = pd.to_datetime(df[campo], errors='coerce')

            for campo in campos_datetime:
                if campo in df.columns:
                    df[campo] = pd.to_datetime(df[campo], errors='coerce')

            # Processa linhas
            registros_criados = 0
            registros_atualizados = 0
            nao_encontrados = 0
            erros = []

            for idx, row in df.iterrows():
                # Verificar se foi cancelado (a cada 100 linhas para não sobrecarregar)
                if idx % 100 == 0:
                    log.refresh_from_db()
                    if log.status == 'CANCELADO':
                        # Marcar como cancelado e sair
                        log.finalizado_em = timezone.now()
                        log.calcular_duracao()
                        log.mensagem_erro = 'Processo cancelado pelo usuário durante o processamento.'
                        log.mensagem = f'Importação cancelada. {registros_criados} registros foram processados antes do cancelamento.'
                        log.save()
                        return
                try:
                    # Construir dicionário de dados, mas evitando campos que contenham 'id' no nome
                    dados = {}
                    campos_permitidos = [
                        'sg_uf', 'nm_municipio', 'indicador', 'cd_nrba', 'st_ba', 'cd_encerramento',
                        'desc_observacao', 'desc_macro_atividade', 'ds_atividade', 'nr_ordem', 
                        'nr_ordem_venda', 'anomes', 'cd_sap_original', 'cd_rede', 'nm_pdv_rel',
                        'rede', 'gp_canal', 'sg_gerencia', 'nm_gc'
                    ]
                    for campo in campos_permitidos:
                        if campo in df.columns:
                            valor = row.get(campo)
                            dados[campo] = str(valor) if pd.notna(valor) else None

                    # Campos de data - conversão melhorada para evitar datas inválidas
                    for campo in campos_data:
                        if campo in df.columns:
                            valor = row[campo]
                            # Verifica se é um Timestamp válido (não NaT)
                            if isinstance(valor, pd.Timestamp) and pd.notna(valor):
                                dados[campo] = valor.date()
                            else:
                                dados[campo] = None

                    # Campos de datetime - conversão melhorada para evitar datas inválidas (1970-01-01)
                    for campo in campos_datetime:
                        if campo in df.columns:
                            valor = row[campo]
                            # Verifica se é um Timestamp válido (não NaT) antes de converter
                            if isinstance(valor, pd.Timestamp) and pd.notna(valor):
                                dt_obj = valor.to_pydatetime()
                                # Validação adicional: não permitir datas muito antigas (antes de 1900)
                                if dt_obj.year >= 1900:
                                    dados[campo] = dt_obj
                                else:
                                    dados[campo] = None
                            else:
                                dados[campo] = None

                    # Garantir que campos de ID/chave primária não sejam passados (proteção contra coluna 'id' na planilha)
                    # Remove qualquer campo relacionado a ID que possa existir na planilha (case-insensitive)
                    campos_para_remover = ['id', 'pk', 'ID', 'PK', 'Id', 'Pk', 'iD', 'pK']
                    for campo in campos_para_remover:
                        if campo in dados:
                            del dados[campo]
                    
                    # Remover também campos que são auto-gerados pelo Django
                    campos_auto = {'id', 'criado_em', 'atualizado_em'}
                    for campo_auto in campos_auto:
                        if campo_auto in dados:
                            del dados[campo_auto]
                    
                    # Normalizar nr_ordem_venda (remover espaços e garantir que não seja vazio)
                    nr_ordem_venda_val = dados.get('nr_ordem_venda')
                    if nr_ordem_venda_val:
                        nr_ordem_venda_val = str(nr_ordem_venda_val).strip()
                        if not nr_ordem_venda_val:
                            nr_ordem_venda_val = None
                        else:
                            # Atualizar o valor normalizado no dicionário para garantir consistência
                            dados['nr_ordem_venda'] = nr_ordem_venda_val
                    
                    # Lógica: Se nr_ordem_venda existir, SEMPRE atualizar (sem depender de comparação de datas)
                    if nr_ordem_venda_val:
                        try:
                            # Buscar registro existente por nr_ordem_venda
                            registro_existente = ImportacaoAgendamento.objects.filter(nr_ordem_venda=nr_ordem_venda_val).first()
                            
                            if registro_existente:
                                # Sempre atualizar quando o registro existe
                                for key, value in dados.items():
                                    setattr(registro_existente, key, value)
                                registro_existente.save()
                                registros_atualizados += 1
                            else:
                                # Não existe, cria novo registro
                                ImportacaoAgendamento.objects.create(**dados)
                                registros_criados += 1
                        except Exception as e_inner:
                            # Se houver erro na busca/atualização, tenta criar normalmente
                            try:
                                ImportacaoAgendamento.objects.create(**dados)
                                registros_criados += 1
                            except:
                                raise e_inner
                    else:
                        # Se não tem nr_ordem_venda, cria normalmente
                        ImportacaoAgendamento.objects.create(**dados)
                        registros_criados += 1

                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

            # Atualizar log
            log.total_processadas = registros_criados
            log.agendamentos_criados = registros_criados
            log.agendamentos_atualizados = registros_atualizados
            log.nao_encontrados = nao_encontrados
            log.erros_count = len(erros)
            log.finalizado_em = timezone.now()
            
            if erros:
                log.status = 'PARCIAL' if registros_criados > 0 else 'ERRO'
                log.mensagem_erro = '\n'.join(erros[:50])
                log.mensagem = f'{registros_criados} registros importados, {len(erros)} erros'
            else:
                log.status = 'SUCESSO'
                log.mensagem = f'{registros_criados} agendamentos importados com sucesso!'
            
            log.detalhes_json = {
                'registros_criados': registros_criados,
                'erros': erros[:100]
            }
            log.save()

        except Exception as e:
            try:
                log = LogImportacaoAgendamento.objects.get(id=log_id)
                log.status = 'ERRO'
                log.mensagem_erro = str(e)
                log.finalizado_em = timezone.now()
                log.save()
            except:
                pass


class ImportacaoRecompraView(APIView):
    permission_classes = [CheckAPIPermission]
    resource_name = 'importacao_recompra'
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        from .models import ImportacaoRecompra, LogImportacaoRecompra
        
        file_obj = request.FILES.get('arquivo')
        if not file_obj:
            return Response({'success': False, 'error': 'Arquivo não enviado'}, status=400)

        # Criar log inicial
        log = LogImportacaoRecompra.objects.create(
            usuario=request.user,
            nome_arquivo=file_obj.name,
            status='PROCESSANDO',
            tamanho_arquivo=file_obj.size
        )

        # Processar em thread background
        file_content = file_obj.read()
        file_name = file_obj.name
        
        def processar_recompra_async():
            self._processar_recompra_interno(log.id, file_content, file_name)
        
        thread = threading.Thread(target=processar_recompra_async, daemon=True)
        thread.start()

        return Response({
            'success': True,
            'log_id': log.id,
            'message': 'Importação Recompra iniciada! O processamento continuará em segundo plano.',
            'status': 'PROCESSANDO',
            'background': True
        })

    def _processar_recompra_interno(self, log_id, file_content, file_name):
        """Processa Recompra em background thread"""
        from .models import ImportacaoRecompra, LogImportacaoRecompra
        from django.utils import timezone
        from io import BytesIO
        
        log = LogImportacaoRecompra.objects.get(id=log_id)
        inicio = timezone.now()
        
        try:
            # Ler arquivo
            file_obj = BytesIO(file_content)
            
            if file_name.endswith('.xlsb'):
                df = pd.read_excel(file_obj, engine='pyxlsb')
            elif file_name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_obj)
            else:
                raise ValueError('Formato inválido')

            df.columns = [str(col).strip() for col in df.columns]
            campos_data = ['dt_venda_particao', 'dt_encerramento', 'dt_inicio_ativo']

            for campo in campos_data:
                if campo in df.columns:
                    df[campo] = pd.to_datetime(df[campo], errors='coerce')

            coluna_map = {
                'ds_anomes': 'ds_anomes',
                'dt_venda_particao': 'dt_venda_particao',
                'dt_encerramento': 'dt_encerramento',
                'nr_ordem': 'nr_ordem',
                'st_ordem': 'st_ordem',
                'nm_seg': 'nm_seg',
                'sg_uf': 'sg_uf',
                'cd_sap_pdv': 'cd_sap_pdv',
                'cd_tr_vdd': 'cd_tr_vdd',
                'nr_cep': 'nr_cep',
                'nm_municipio': 'nm_municipio',
                'nm_bairro': 'nm_bairro',
                'resultado': 'resultado',
                'dt_inicio_ativo': 'dt_inicio_ativo',
                'nr_cep_base': 'nr_cep_base',
                'nr_complemento1_base': 'nr_complemento1_base',
                'nr_complemento2_base': 'nr_complemento2_base',
                'nr_complemento3_base': 'nr_complemento3_base',
                'nm_diretoria': 'nm_diretoria',
                'nm_regional': 'nm_regional',
                'cd_rede': 'cd_rede',
                'gp_canal': 'gp_canal',
                'nm_pdv_rel': 'nm_pdv_rel',
                'GERENCIA': 'GERENCIA',
                'nm_gc': 'nm_gc',
                'REDE': 'REDE',
            }

            registros_criados = 0
            erros = []

            for idx, row in df.iterrows():
                try:
                    dados = {}
                    for col_arquivo, col_model in coluna_map.items():
                        if col_arquivo in df.columns:
                            valor = row.get(col_arquivo)
                            if pd.isna(valor) or valor == '':
                                dados[col_model] = None
                            else:
                                if col_model in campos_data:
                                    try:
                                        dados[col_model] = pd.to_datetime(valor).date()
                                    except:
                                        dados[col_model] = None
                                else:
                                    dados[col_model] = str(valor).strip()
                        else:
                            dados[col_model] = None

                    ImportacaoRecompra.objects.create(**dados)
                    registros_criados += 1

                except Exception as e:
                    erros.append(f"Linha {idx + 2}: {str(e)}")

            # Atualizar log
            log.total_linhas = len(df)
            log.total_processadas = registros_criados
            log.registros_criados = registros_criados
            log.erros_count = len(erros)
            log.finalizado_em = timezone.now()
            
            if erros:
                log.status = 'PARCIAL' if registros_criados > 0 else 'ERRO'
                log.mensagem_erro = '\n'.join(erros[:50])
                log.mensagem = f'{registros_criados} registros importados, {len(erros)} erros'
            else:
                log.status = 'SUCESSO'
                log.mensagem = f'{registros_criados} registros importados com sucesso!'
            
            log.calcular_duracao()
            log.detalhes_json = {
                'registros_criados': registros_criados,
                'erros': erros[:100]
            }
            log.save()

        except Exception as e:
            try:
                log.refresh_from_db()
                log.status = 'ERRO'
                log.mensagem_erro = str(e)
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
            except:
                pass


# =============================================================================
# BÔNUS M-10 & FPD - VIEWS
# =============================================================================

from django.db.models import Q, Count, Sum, Avg
from django.http import HttpResponse
import pandas as pd
from datetime import datetime
from .models import SafraM10, ContratoM10, FaturaM10


def page_bonus_m10(request):
    """View para renderizar a página HTML"""
    return render(request, 'bonus_m10.html')


def page_boas_vindas(request):
    """View para a ferramenta Boas Vindas (gestão à vista)."""
    return render(request, 'boas-vindas.html')


def page_validacao_fpd(request):
    """View para renderizar a página de validação de importações FPD"""
    return render(request, 'validacao-fpd.html')


def page_validacao_churn(request):
    """View para renderizar a página de validação de importações CHURN"""
    return render(request, 'validacao-churn.html')


def page_validacao_osab(request):
    """View para renderizar a página de validação de importações OSAB"""
    return render(request, 'validacao-osab.html')


def page_validacao_agendamento(request):
    """View para renderizar a página de validação de importações Agendamento"""
    return render(request, 'validacao-agendamento.html')


def page_validacao_legado(request):
    """View para renderizar a página de validação de importações Legado"""
    return render(request, 'validacao-legado.html')


def page_validacao_dfv(request):
    """View para renderizar a página de validação de importações DFV"""
    return render(request, 'validacao-dfv.html')


def page_validacao_recompra(request):
    """View para renderizar a página de validação de importações Recompra"""
    return render(request, 'validacao-recompra.html')


def page_record_apoia(request):
    """View para renderizar a página HTML do Record Apoia"""
    return render(request, 'record_apoia.html')


def page_conhecimento_ia(request):
    """View para a página de upload de documentos (PDF/Excel/PPT) que alimentam a IA do bot."""
    return render(request, 'conhecimento_ia.html')


class SafraM10ListView(APIView):
    """Lista todas as safras disponíveis"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        safras = SafraM10.objects.all().order_by('-mes_referencia')
        data = []
        # Mapeamento de meses em português
        meses_pt = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
            5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
            9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }
        
        pode_ver_valor_bonus = is_member(request.user, ['Diretoria', 'Admin'])
        for s in safras:
            mes_nome = meses_pt.get(s.mes_referencia.month, s.mes_referencia.strftime('%m'))
            mes_formatado = f"{mes_nome}/{s.mes_referencia.year}"
            data.append({
                'id': s.id,
                'mes_referencia': s.mes_referencia.isoformat(),
                'mes_referencia_formatado': mes_formatado,
                'total_instalados': s.total_instalados,
                'total_ativos': s.total_ativos,
                'total_elegivel_bonus': s.total_elegivel_bonus,
                'valor_bonus_total': float(s.valor_bonus_total) if pode_ver_valor_bonus else 0,
            })
        return Response(data)


def _map_status_fpd(status_raw: str) -> str:
    """Mapeia status do FPD para rótulos amigáveis segundo mapeamento do usuário:
    - Paga → Paga
    - Paga_aguardando_repasse → Paga
    - Aguardando_arrecadacao → Não Pago
    - Ajustada → Paga
    - Erro_nao_recobravel → Não Pago
    """
    if not status_raw:
        return '-'
    mapping = {
        'paga': 'Paga',
        'paga_aguardando_repasse': 'Paga',
        'aguardando_arrecadacao': 'Não Pago',
        'ajustada': 'Paga',
        'erro_nao_recobravel': 'Não Pago',
    }
    key = status_raw.lower()
    return mapping.get(key, status_raw.replace('_', ' '))


def _safra_to_data_range(safra_param):
    """Converte parâmetro safra (ID SafraM10 ou YYYY-MM) em (data_inicio, data_fim) do mês.
    data_fim é exclusivo (primeiro dia do mês seguinte). Retorna (None, None) se inválido.
    """
    if not safra_param:
        return None, None
    try:
        pk = int(safra_param)
        s = SafraM10.objects.filter(pk=pk).first()
        if s:
            di = s.mes_referencia
            df = di + relativedelta(months=1)
            return di, df
    except (ValueError, TypeError):
        pass
    if isinstance(safra_param, str) and len(safra_param) == 7 and safra_param[4] == '-':
        try:
            ano, mes = int(safra_param[:4]), int(safra_param[5:7])
            di = date(ano, mes, 1)
            df = di + relativedelta(months=1)
            return di, df
        except (ValueError, TypeError):
            pass
    return None, None


def _recalcular_totais_safra_m10(safra_str):
    """GAP 3: Recalcula total_instalados, total_ativos, total_elegivel_bonus e valor_bonus_total da SafraM10.
    safra_str: formato YYYY-MM (mês de instalação).
    """
    data_inicio, data_fim = _safra_to_data_range(safra_str)
    if data_inicio is None or data_fim is None:
        return
    from django.db.models import Count, Q
    total_instalados = ContratoM10.objects.filter(
        data_instalacao__gte=data_inicio,
        data_instalacao__lt=data_fim,
    ).count()
    total_ativos = ContratoM10.objects.filter(
        data_instalacao__gte=data_inicio,
        data_instalacao__lt=data_fim,
        status_contrato='ATIVO',
    ).count()
    # Elegíveis: ativo, sem downgrade, todas as faturas cadastradas pagas (ou fallback FPD paga)
    contratos_safra = ContratoM10.objects.filter(
        data_instalacao__gte=data_inicio,
        data_instalacao__lt=data_fim,
    ).annotate(
        total_faturas=Count('faturas', distinct=True),
        faturas_pagas=Count('faturas', filter=Q(faturas__status='PAGO'), distinct=True),
    )
    elegiveis = 0
    for c in contratos_safra:
        total_f = c.total_faturas or 0
        pagas = c.faturas_pagas or 0
        if total_f == 0 and c.status_fatura_fpd and str(c.status_fatura_fpd).lower().startswith('paga'):
            total_f, pagas = 1, 1
        if (c.status_contrato == 'ATIVO' and not c.teve_downgrade and
                total_f > 0 and pagas == total_f):
            elegiveis += 1
    valor_bonus_total = elegiveis * 150
    SafraM10.objects.filter(mes_referencia=data_inicio).update(
        total_instalados=total_instalados,
        total_ativos=total_ativos,
        total_elegivel_bonus=elegiveis,
        valor_bonus_total=valor_bonus_total,
    )


class VendedoresM10View(APIView):
    """Lista vendedores que têm contratos na M-10 (opcionalmente filtrado por safra)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        safra_param = request.GET.get('safra')
        queryset = ContratoM10.objects.select_related('vendedor').filter(vendedor__isnull=False)
        if safra_param:
            data_inicio, data_fim = _safra_to_data_range(safra_param)
            if data_inicio is not None and data_fim is not None:
                queryset = queryset.filter(
                    data_instalacao__gte=data_inicio,
                    data_instalacao__lt=data_fim,
                )

        # Deduplica por vendedor_id para evitar repetição na lista
        vistos = {}
        for v in queryset.values('vendedor_id', 'vendedor__username', 'vendedor__first_name', 'vendedor__last_name'):
            vid = v['vendedor_id']
            if vid not in vistos and vid is not None:
                nome = f"{v.get('vendedor__first_name', '')} {v.get('vendedor__last_name', '')}".strip()
                vistos[vid] = {
                    'id': vid,
                    'username': v.get('vendedor__username') or '-',
                    'nome': nome or v.get('vendedor__username') or '-',
                }

        data = sorted(vistos.values(), key=lambda x: (x['username'] or '').lower())
        return Response(data)


class DashboardM10View(APIView):
    """Dashboard com estatísticas M-10"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            safra_id = request.GET.get('safra')
            if not safra_id:
                return Response({'error': 'Safra não informada'}, status=400)

            try:
                safra = SafraM10.objects.get(id=safra_id)
            except SafraM10.DoesNotExist:
                return Response({'error': 'Safra não encontrada'}, status=404)

            # Regra: safra = mês da data de instalação. Ao filtrar pelo mês selecionado,
            # exibir todos os contratos cuja data_instalacao está naquele mês.
            data_inicio = safra.mes_referencia
            data_fim = data_inicio + relativedelta(months=1)

            queryset = ContratoM10.objects.filter(
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
            ).select_related('vendedor')
            
            vendedor = request.GET.get('vendedor')
            if vendedor:
                queryset = queryset.filter(vendedor_id=vendedor)
            
            status = request.GET.get('status')
            if status:
                queryset = queryset.filter(status_contrato=status)
            
            elegivel = request.GET.get('elegivel')
            if elegivel:
                queryset = queryset.filter(elegivel_bonus=(elegivel == 'true'))

            busca = request.GET.get('q')
            if busca:
                busca_digits = re.sub(r'\D', '', busca)
                filtros_busca = (
                    Q(numero_contrato__icontains=busca)
                    | Q(numero_contrato_definitivo__icontains=busca)
                    | Q(cliente_nome__icontains=busca)
                    | Q(ordem_servico__icontains=busca)
                )
                if busca_digits:
                    filtros_busca |= Q(cpf_cliente__icontains=busca_digits)
                queryset = queryset.filter(filtros_busca)

            # Anotações de faturas para calcular elegibilidade dinâmica
            queryset = queryset.annotate(
                total_faturas=Count('faturas', distinct=True),
                faturas_pagas=Count('faturas', filter=Q(faturas__status='PAGO'), distinct=True),
            )

            # Aplica fallback: se não houver faturas cadastradas mas status FPD for Paga, conta como 1/1
            contratos_list = list(queryset)
            for c in contratos_list:
                if c.total_faturas == 0 and c.status_fatura_fpd and str(c.status_fatura_fpd).lower().startswith('paga'):
                    c.total_faturas = 1
                    c.faturas_pagas = 1

            # Mapa da fatura 1 por contrato (prioridade para exibir vencimento/status na lista)
            ids_contratos = [c.id for c in contratos_list]
            faturas1_map = {
                f.contrato_id: f
                for f in FaturaM10.objects.filter(contrato_id__in=ids_contratos, numero_fatura=1)
            }
            status_display_map = dict(FaturaM10.STATUS_CHOICES)

            # Estatísticas
            total = len(contratos_list)
            ativos = len([c for c in contratos_list if c.status_contrato == 'ATIVO'])
            elegiveis = len([
                c for c in contratos_list
                if c.status_contrato == 'ATIVO'
                and not c.teve_downgrade
                and (c.total_faturas or 0) > 0
                and c.total_faturas == c.faturas_pagas
            ])
            valor_total = elegiveis * 150  # R$ 150 por contrato elegível
            # Valor total do bônus só visível para Diretoria e Admin
            pode_ver_valor_bonus = is_member(request.user, ['Diretoria', 'Admin'])
            if not pode_ver_valor_bonus:
                valor_total = 0

            taxa_permanencia = round((ativos / total * 100) if total > 0 else 0, 1)

            # Paginação - Limita a 100 registros por página
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 100))
            start = (page - 1) * page_size
            end = start + page_size

            # Contratos para tabela (apenas dados essenciais)
            contratos_data = []
            for c in contratos_list[start:end]:
                is_elegivel = (
                    c.status_contrato == 'ATIVO'
                    and (c.total_faturas or 0) > 0
                    and c.faturas_pagas == c.total_faturas
                    and not c.teve_downgrade
                )
                data_instalacao_fmt = c.data_instalacao.strftime('%d/%m/%Y') if c.data_instalacao else '-'
                f1 = faturas1_map.get(c.id)
                venc_fatura1 = f1.data_vencimento.strftime('%d/%m/%Y') if f1 and f1.data_vencimento else None
                pagto_fatura1 = f1.data_pagamento.strftime('%d/%m/%Y') if f1 and f1.data_pagamento else None
                status_fatura1 = f1.status if f1 else None
                status_fatura1_display = status_display_map.get(status_fatura1, '-') if status_fatura1 else None
                # FPD permanece fallback se não houver fatura 1
                venc_exib = venc_fatura1 or (c.data_vencimento_fpd.strftime('%d/%m/%Y') if c.data_vencimento_fpd else '-')
                pagto_exib = pagto_fatura1 or (c.data_pagamento_fpd.strftime('%d/%m/%Y') if c.data_pagamento_fpd else '-')
                # Sempre mostrar status da fatura 1 se existir, senão FPD
                status_fpd_exib = status_fatura1 or (c.status_fatura_fpd or '-')
                # Aplicar mapeamento em ambos os casos (fatura 1 ou FPD)
                if status_fatura1_display:
                    status_fpd_display = status_fatura1_display
                else:
                    status_fpd_display = _map_status_fpd(c.status_fatura_fpd)
                contratos_data.append({
                    'id': c.id,
                    'numero_contrato': c.numero_contrato,
                    'numero_contrato_definitivo': c.numero_contrato_definitivo or '-',
                    'cliente_nome': c.cliente_nome,
                    'cpf_cliente': c.cpf_cliente,
                    'vendedor_nome': c.vendedor.username if c.vendedor else '-',
                    'data_instalacao': data_instalacao_fmt,
                    'plano_atual': c.plano_atual,
                    'status': c.status_contrato,
                    'status_display': c.get_status_contrato_display(),
                    'faturas_pagas': c.faturas_pagas,
                    'total_faturas': c.total_faturas,
                    'elegivel': is_elegivel,
                    # Dados de fatura 1 (prioritário) ou FPD como fallback
                    'status_fatura_fpd': status_fpd_exib,
                    'status_fatura_fpd_display': status_fpd_display,
                    'data_vencimento_fpd': venc_exib,
                    'data_pagamento_fpd': pagto_exib,
                    'valor_fatura_fpd': float(c.valor_fatura_fpd) if c.valor_fatura_fpd else 0,
                    'nr_dias_atraso_fpd': c.nr_dias_atraso_fpd or 0,
                })

            total_pages = (total + page_size - 1) // page_size

            return Response({
                'total': total,
                'ativos': ativos,
                'elegiveis': elegiveis,
                'valor_total': valor_total,
                'pode_ver_valor_bonus': pode_ver_valor_bonus,
                'taxa_permanencia': taxa_permanencia,
                'contratos': contratos_data,
                'page': page,
                'total_pages': total_pages,
                'page_size': page_size,
            })

        except Exception as e:
            logger.exception('Erro ao gerar dashboard M-10')
            return Response({'error': 'Erro ao carregar dashboard M-10', 'detail': str(e)}, status=500)


class DashboardFPDView(APIView):
    """Dashboard com estatísticas FPD"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            mes_str = request.GET.get('mes')  # formato: 2025-07
            
            # Filtrar faturas número 1 (primeira fatura)
            queryset = FaturaM10.objects.filter(numero_fatura=1)
            
            if mes_str:
                try:
                    ano, mes = mes_str.split('-')
                    queryset = queryset.filter(data_vencimento__year=ano, data_vencimento__month=mes)
                except ValueError:
                    return Response({'error': 'Parâmetro mes inválido. Use YYYY-MM.'}, status=400)
            
            status_filtro = request.GET.get('status')
            if status_filtro:
                queryset = queryset.filter(status=status_filtro)
            
            vendedor = request.GET.get('vendedor')
            if vendedor:
                queryset = queryset.filter(contrato__vendedor_id=vendedor)

            # Estatísticas
            total_geradas = queryset.count()
            total_pagas = queryset.filter(status='PAGO').count()
            total_aberto = total_geradas - total_pagas  # contas em aberto = total - pagas
            # Taxa FPD = volume de contas em aberto / total de contas (em %)
            taxa_fpd = round((total_aberto / total_geradas * 100) if total_geradas > 0 else 0, 1)

            # Faturas para tabela
            faturas_data = []
            for f in queryset.select_related('contrato', 'contrato__vendedor'):
                data_venc = f.data_vencimento.strftime('%d/%m/%Y') if f.data_vencimento else '-'
                data_pag = f.data_pagamento.strftime('%d/%m/%Y') if f.data_pagamento else None
                faturas_data.append({
                    'id': f.id,
                    'contrato_id': f.contrato.id,
                    'numero_contrato': f.contrato.numero_contrato,
                    'cliente_nome': f.contrato.cliente_nome,
                    'vendedor_nome': f.contrato.vendedor.username if f.contrato.vendedor else '-',
                    'data_vencimento': data_venc,
                    'valor': float(f.valor) if f.valor is not None else 0,
                    'data_pagamento': data_pag,
                    'status': f.status,
                    'status_display': f.get_status_display(),
                })

            return Response({
                'total_geradas': total_geradas,
                'total_pagas': total_pagas,
                'total_aberto': total_aberto,
                'taxa_fpd': taxa_fpd,
                'faturas': faturas_data,
            })

        except Exception as e:
            logger.exception('Erro ao gerar dashboard FPD')
            return Response({'error': 'Erro ao carregar dashboard FPD', 'detail': str(e)}, status=500)


class PopularSafraM10View(APIView):
    """Popula SafraM10 e ContratoM10 a partir da tabela Venda baseado em data_instalacao"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Admin', 'BackOffice', 'Diretoria']):
            return Response({'error': 'Sem permissão'}, status=403)

        mes_referencia = request.data.get('mes_referencia')  # Formato: '2025-07'
        
        if not mes_referencia:
            return Response({'error': 'mes_referencia é obrigatório (formato: YYYY-MM)'}, status=400)

        try:
            # Converte para data (primeiro dia do mês)
            ano, mes = mes_referencia.split('-')
            data_inicio = datetime(int(ano), int(mes), 1).date()
            # Próximo mês para o range
            if mes == '12':
                data_fim = datetime(int(ano) + 1, 1, 1).date()
            else:
                data_fim = datetime(int(ano), int(mes) + 1, 1).date()
            
            # Busca ou cria a Safra M-10
            safra, safra_criada = SafraM10.objects.get_or_create(
                mes_referencia=data_inicio,
                defaults={
                    'total_instalados': 0,
                    'total_ativos': 0,
                    'total_elegivel_bonus': 0,
                    'valor_bonus_total': 0,
                }
            )

            safra_str = mes_referencia  # YYYY-MM, usado para contagens e filtros

            # Busca Vendas com data_instalacao no mês de referência E status INSTALADA.
            # Ordena por data_criacao (mais antiga primeiro): quando a mesma O.S. tem várias vendas
            # (ex.: criada junho vs julho, ambas instaladas em julho), preferimos a de criação
            # fora do mês, que antes ficava "não considerada".
            vendas = Venda.objects.filter(
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
                data_instalacao__isnull=False,
                ativo=True,
                status_esteira__nome__iexact='INSTALADA'
            ).order_by('data_criacao').select_related('cliente', 'vendedor', 'status_esteira', 'plano')

            contratos_criados = 0
            contratos_duplicados = 0

            for venda in vendas:
                numero_contrato = venda.ordem_servico or f"VENDA_{venda.id}"
                # Evita duplicata: por O.S. ou por numero_contrato (VENDA_X quando sem O.S.)
                if venda.ordem_servico:
                    contrato_existe = ContratoM10.objects.filter(ordem_servico=venda.ordem_servico).exists()
                else:
                    contrato_existe = ContratoM10.objects.filter(numero_contrato=numero_contrato).exists()

                if contrato_existe:
                    contratos_duplicados += 1
                    continue

                # Cria novo ContratoM10 (safra = mês da instalação)
                contrato = ContratoM10.objects.create(
                    safra=safra_str,
                    venda=venda,
                    numero_contrato=numero_contrato,
                    ordem_servico=venda.ordem_servico,
                    cliente_nome=venda.cliente.nome_razao_social if venda.cliente else 'N/D',
                    cpf_cliente=venda.cliente.cpf_cnpj if venda.cliente else '',
                    vendedor=venda.vendedor,
                    data_instalacao=venda.data_instalacao,
                    plano_original=venda.plano.nome if venda.plano else 'N/D',
                    plano_atual=venda.plano.nome if venda.plano else 'N/D',
                    valor_plano=venda.plano.valor if venda.plano else 0,
                    status_contrato='ATIVO',
                    observacao=f"Importado de Venda #{venda.id}"
                )
                contratos_criados += 1

            # Conta por data_instalacao no mês (igual ao dashboard)
            total_contratos_safra = ContratoM10.objects.filter(
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
            ).count()
            total_ativos = ContratoM10.objects.filter(
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
                status_contrato='ATIVO',
            ).count()

            safra.total_instalados = total_contratos_safra
            safra.total_ativos = total_ativos
            safra.save()
            # GAP 3: Recalcular elegíveis e valor do bônus da safra
            _recalcular_totais_safra_m10(safra_str)

            return Response({
                'message': f'Safra {mes_referencia} populada com sucesso!',
                'safra_id': safra.id,
                'contratos_criados': contratos_criados,
                'contratos_duplicados': contratos_duplicados,
                'total_contratos_safra': total_contratos_safra,
            })

        except ValueError as e:
            return Response({'error': f'Formato de data inválido: {str(e)}'}, status=400)
        except Exception as e:
            return Response({'error': f'Erro ao popular safra: {str(e)}'}, status=500)


class ContratoM10DetailView(APIView):
    """Detalhes de um contrato com suas faturas"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            contrato = ContratoM10.objects.get(pk=pk)
            faturas = []
            for f in contrato.faturas.all().order_by('numero_fatura'):
                data_venc = f.data_vencimento.isoformat() if f.data_vencimento else None
                data_pag = f.data_pagamento.isoformat() if f.data_pagamento else None
                faturas.append({
                    'id': f.id,
                    'numero_fatura': f.numero_fatura,
                    'numero_fatura_operadora': f.numero_fatura_operadora or '',
                    'valor': float(f.valor) if f.valor is not None else 0,
                    'data_vencimento': data_venc,
                    'data_pagamento': data_pag or '',
                    'status': f.status,
                })
            
            return Response({
                'id': contrato.id,
                'numero_contrato': contrato.numero_contrato,
                'cliente_nome': contrato.cliente_nome,
                'cpf_cliente': contrato.cpf_cliente,
                'data_vencimento_fpd': contrato.data_vencimento_fpd.isoformat() if contrato.data_vencimento_fpd else None,
                'data_pagamento_fpd': contrato.data_pagamento_fpd.isoformat() if contrato.data_pagamento_fpd else None,
                'status_fatura_fpd': contrato.status_fatura_fpd or '',
                'valor_fatura_fpd': float(contrato.valor_fatura_fpd) if contrato.valor_fatura_fpd else 0,
                'faturas': faturas,
            })
        except ContratoM10.DoesNotExist:
            return Response({'error': 'Contrato não encontrado'}, status=404)
        except Exception as e:
            logger.exception('Erro ao recuperar contrato M-10')
            return Response({'error': 'Erro ao carregar contrato', 'detail': str(e)}, status=500)


class ImportarFPDView(APIView):
    """Importa planilha FPD da operadora e faz crossover com ContratoM10 por O.S"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request):
        if not is_member(request.user, ['Admin', 'BackOffice', 'Diretoria']):
            return Response({'error': 'Sem permissão'}, status=403)

        arquivo = request.FILES.get('file')
        if not arquivo:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        from .models import ImportacaoFPD, LogImportacaoFPD
        from django.utils import timezone
        from django.db import transaction
        
        # Criar log de importação
        log = LogImportacaoFPD.objects.create(
            nome_arquivo=arquivo.name,
            usuario=request.user,
            status='PROCESSANDO'
        )
        
        # Ler arquivo em memória para passar para thread
        arquivo_bytes = arquivo.read()
        arquivo_nome = arquivo.name
        user_id = request.user.id
        
        # Iniciar processamento em thread background
        def processar_fpd_async():
            self._processar_fpd_interno(log.id, arquivo_bytes, arquivo_nome, user_id)
        
        thread = threading.Thread(target=processar_fpd_async, daemon=True)
        thread.start()
        
        # Retornar imediatamente ao cliente
        return Response({
            'success': True,
            'log_id': log.id,
            'message': 'Importação FPD iniciada! O processamento continuará em segundo plano. Atualize a página em alguns minutos para ver o resultado.',
            'status': 'PROCESSANDO',
            'background': True
        })
    
    def _processar_fpd_interno(self, log_id, arquivo_bytes, arquivo_nome, user_id):
        """Processa FPD em background thread"""
        from .models import ImportacaoFPD, LogImportacaoFPD
        from django.utils import timezone
        from django.db import transaction
        from io import BytesIO
        
        # Recuperar log e usuário
        log = LogImportacaoFPD.objects.get(id=log_id)
        User = get_user_model()
        usuario = User.objects.get(id=user_id)
        # Recuperar log e usuário
        log = LogImportacaoFPD.objects.get(id=log_id)
        User = get_user_model()
        usuario = User.objects.get(id=user_id)
        
        inicio = timezone.now()
        os_nao_encontradas = []
        erros_detalhados = []

        try:
            # Criar objeto BytesIO do arquivo
            arquivo_io = BytesIO(arquivo_bytes)
            
            # Lê arquivo Excel/CSV
            # IMPORTANTE: Ler colunas numéricas como STRING para preservar leading zeros
            dtype_spec = {
                'ID_CONTRATO': str,      # Força leitura como texto
                'NR_FATURA': str,        # Força leitura como texto  
                'NR_ORDEM': str,         # Força leitura como texto
            }
            
            if arquivo_nome.endswith('.csv'):
                df = pd.read_csv(arquivo_io, dtype=dtype_spec)
            elif arquivo_nome.endswith('.xlsb'):
                try:
                    df = pd.read_excel(arquivo_io, engine='pyxlsb', dtype=dtype_spec)
                except Exception as e:
                    log.status = 'ERRO'
                    log.mensagem_erro = f'Formato .xlsb não suportado: {str(e)}'
                    log.finalizado_em = timezone.now()
                    log.calcular_duracao()
                    log.save()
                    return
            else:
                df = pd.read_excel(arquivo_io, dtype=dtype_spec)

            # Normalizar nomes de colunas para minúsculas E remover espaços extras
            df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
            
            # GAP 7: Validar presença da coluna obrigatória NR_ORDEM
            if 'nr_ordem' not in df.columns:
                colunas_encontradas = ', '.join(sorted(df.columns[:15]))
                log.status = 'ERRO'
                log.mensagem_erro = (
                    'Coluna NR_ORDEM não encontrada no arquivo. '
                    'A planilha FPD deve ter uma coluna "NR_ORDEM" (ou "Nr Ordem"). '
                    f'Colunas encontradas: {colunas_encontradas}' + ('...' if len(df.columns) > 15 else '')
                )
                log.finalizado_em = timezone.now()
                log.calcular_duracao()
                log.save()
                return
            
            log.total_linhas = len(df)
            log.save(update_fields=['total_linhas'])
            registros_nao_encontrados = 0
            registros_importacoes_fpd = 0
            registros_atualizados = 0
            registros_pulados = 0
            valor_total = 0
            data_importacao_agora = timezone.now()
            
            # Otimização: pre-carregar contratos em memória para evitar N queries
            # Criar dicionário com múltiplas variações de chaves para melhor matching
            contratos_dict = {}
            # Pre-carregar faturas para evitar N+1 queries
            contratos_list = list(ContratoM10.objects.prefetch_related('faturas').all())
            # Criar dicionário de faturas por contrato para acesso rápido
            faturas_por_contrato = {}
            for c in contratos_list:
                if c.ordem_servico:
                    os = str(c.ordem_servico).strip()
                    if not os:
                        continue
                    # Indexar por múltiplas variações para facilitar busca
                    os_sem_zeros = os.lstrip('0') or '0'  # Proteger contra string vazia
                    variacoes = [
                        os,
                        os_sem_zeros,
                        os.zfill(8) if len(os) <= 8 else os,  # 8 dígitos (compatível com churn)
                        f'OS-{os}',
                        f'OS-{os_sem_zeros}',
                        f'OS-{os.zfill(8)}' if len(os) <= 8 else None,
                    ]
                    variacoes = [v for v in variacoes if v]
                    for variacao in variacoes:
                        if variacao and variacao not in contratos_dict:
                            contratos_dict[variacao] = c
                # Pre-carregar faturas do contrato em dicionário para acesso rápido
                faturas_contrato = list(c.faturas.all())
                faturas_por_contrato[c.id] = {f.numero_fatura: f for f in faturas_contrato}
            
            # Otimização: pre-carregar ImportacaoFPD em memória também
            # Indexar apenas por nr_ordem (atualizar registro existente com mesmo nr_ordem)
            importacoes_dict = {}
            for imp in ImportacaoFPD.objects.all().order_by('-atualizada_em'):
                if imp.nr_ordem and imp.nr_ordem not in importacoes_dict:
                    importacoes_dict[imp.nr_ordem] = imp
            
            # Listas para bulk operations (reduz queries drasticamente)
            faturas_para_criar = []
            faturas_para_atualizar = []
            importacoes_para_criar = []
            importacoes_para_atualizar = []
            # GAP 1+2: contratos que tiveram fatura 1 criada/atualizada (para atualizar ContratoM10 e elegibilidade)
            contratos_afetados_ids = set()

            with transaction.atomic():  # Garantir atomicidade
                for idx, row in df.iterrows():
                    try:
                        # Busca por O.S (nr_ordem - com coluna normalizada para minúsculas)
                        nr_ordem_raw = row.get('nr_ordem', '')
                        
                        # Converter para string mas MANTER ZEROS à esquerda
                        if pd.isna(nr_ordem_raw):
                            registros_pulados += 1
                            continue
                        
                        nr_ordem = str(nr_ordem_raw).strip()
                        
                        # Se for número, remover ".0" se existir (vem do pandas quando lê números do Excel)
                        if nr_ordem.replace('.', '').replace('-', '').isdigit():
                            nr_ordem = nr_ordem.split('.')[0]
                        
                        if not nr_ordem or nr_ordem == 'nan':
                            registros_pulados += 1
                            continue

                        # Tenta encontrar contrato por ordem_servico com variações (lookup em memória)
                        # Tenta múltiplas variações para melhor matching
                        contrato = None
                        variacoes_nr_ordem = [
                            nr_ordem,                          # Exato (como veio)
                            nr_ordem.zfill(8),                # Com zeros à esquerda (8 dígitos)
                            nr_ordem.lstrip('0') or '0',       # Sem zeros à esquerda
                            f'OS-{nr_ordem}',                  # Com prefixo OS-
                            f'OS-{nr_ordem.zfill(8)}',         # Prefixo OS- com zeros
                            f'OS-{nr_ordem.lstrip("0") or "0"}', # Prefixo OS- sem zeros
                        ]
                        for variacao in variacoes_nr_ordem:
                            if variacao in contratos_dict:
                                contrato = contratos_dict[variacao]
                                break
                        if contrato:
                            
                            # ID_CONTRATO e NR_FATURA já vêm como STRING do pandas (dtype=str)
                            nr_contrato = str(row.get('id_contrato', '')).strip()
                            if nr_contrato and nr_contrato != 'nan':
                                contrato.numero_contrato_definitivo = nr_contrato
                                # Save será feito em bulk ao final
                            
                            # Extrai dados FPD
                            # Já vêm como STRING do pandas, preservando zeros
                            id_contrato = str(row.get('id_contrato', '')).strip()
                            dt_venc = row.get('dt_venc_orig')
                            dt_pgto = row.get('dt_pagamento')
                            status_str = str(row.get('ds_status_fatura', 'NAO_PAGO')).upper()
                            nr_fatura = str(row.get('nr_fatura', '')).strip()
                            vl_fatura = row.get('vl_fatura', 0)
                            nr_dias_atraso = row.get('nr_dias_atraso', 0)
                            
                            # Normalizar status usando mapeamento padronizado
                            status = normalizar_status_fpd(status_str)

                            # Extrair e converter datas - Excel armazena como números serial
                            dt_venc = row.get('dt_venc_orig')
                            if pd.notna(dt_venc):
                                # Se for número, converter de serial Excel
                                if isinstance(dt_venc, (int, float)):
                                    dt_venc_date = (pd.Timestamp("1900-01-01") + pd.Timedelta(days=dt_venc - 2)).date()
                                else:
                                    dt_venc_date = pd.to_datetime(dt_venc).date()
                            else:
                                dt_venc_date = timezone.now().date()

                            dt_pgto = row.get('dt_pagamento')
                            if pd.notna(dt_pgto):
                                # Se for número, converter de serial Excel
                                if isinstance(dt_pgto, (int, float)):
                                    dt_pgto_date = (pd.Timestamp("1900-01-01") + pd.Timedelta(days=dt_pgto - 2)).date()
                                else:
                                    dt_pgto_date = pd.to_datetime(dt_pgto).date()
                            else:
                                dt_pgto_date = None

                            # Se houver data de vencimento, define safra de FPD por mês de vencimento
                            if dt_venc_date:
                                safra_fpd_mes = dt_venc_date.replace(day=1)
                                safra_fpd, _ = SafraM10.objects.get_or_create(
                                    mes_referencia=safra_fpd_mes,
                                    defaults={'total_instalados': 0, 'total_ativos': 0}
                                )

                            # Preparar fatura para bulk update/create
                            vl_fatura_float = float(vl_fatura) if pd.notna(vl_fatura) else 0
                            nr_dias_atraso_int = int(nr_dias_atraso) if pd.notna(nr_dias_atraso) else 0
                        
                            # Verificar se fatura já existe (lookup em memória via dicionário pré-carregado)
                            fatura_existente = faturas_por_contrato.get(contrato.id, {}).get(1)
                            
                            if fatura_existente:
                                fatura_existente.numero_fatura_operadora = nr_fatura
                                fatura_existente.valor = vl_fatura_float
                                fatura_existente.data_vencimento = dt_venc_date
                                fatura_existente.data_pagamento = dt_pgto_date
                                fatura_existente.dias_atraso = nr_dias_atraso_int
                                fatura_existente.status = status
                                fatura_existente.id_contrato_fpd = id_contrato
                                fatura_existente.dt_pagamento_fpd = dt_pgto_date
                                fatura_existente.ds_status_fatura_fpd = status_str
                                fatura_existente.data_importacao_fpd = data_importacao_agora
                                faturas_para_atualizar.append(fatura_existente)
                                contratos_afetados_ids.add(contrato.id)
                            else:
                                faturas_para_criar.append(FaturaM10(
                                    contrato=contrato,
                                    numero_fatura=1,
                                    numero_fatura_operadora=nr_fatura,
                                    valor=vl_fatura_float,
                                    data_vencimento=dt_venc_date,
                                    data_pagamento=dt_pgto_date,
                                    dias_atraso=nr_dias_atraso_int,
                                    status=status,
                                    id_contrato_fpd=id_contrato,
                                    dt_pagamento_fpd=dt_pgto_date,
                                    ds_status_fatura_fpd=status_str,
                                    data_importacao_fpd=data_importacao_agora
                                ))
                                contratos_afetados_ids.add(contrato.id)

                            # Preparar ImportacaoFPD para bulk
                            # Buscar por nr_ordem (atualizar se existir, criar se não existir)
                            importacao_existente = importacoes_dict.get(nr_ordem)
                            
                            if importacao_existente:
                                # Atualizar registro existente com novos dados da planilha
                                importacao_existente.id_contrato = id_contrato
                                importacao_existente.nr_fatura = nr_fatura  # Atualiza também o nr_fatura
                                importacao_existente.dt_venc_orig = dt_venc_date
                                importacao_existente.dt_pagamento = dt_pgto_date
                                importacao_existente.nr_dias_atraso = nr_dias_atraso_int
                                importacao_existente.ds_status_fatura = status_str
                                importacao_existente.vl_fatura = vl_fatura_float
                                importacao_existente.contrato_m10 = contrato
                                importacoes_para_atualizar.append(importacao_existente)
                                registros_atualizados += 1
                            else:
                                # Criar novo registro
                                importacoes_para_criar.append(ImportacaoFPD(
                                    nr_ordem=nr_ordem,
                                    nr_fatura=nr_fatura,
                                    id_contrato=id_contrato,
                                    dt_venc_orig=dt_venc_date,
                                    dt_pagamento=dt_pgto_date,
                                    nr_dias_atraso=nr_dias_atraso_int,
                                    ds_status_fatura=status_str,
                                    vl_fatura=vl_fatura_float,
                                    contrato_m10=contrato
                                ))
                                registros_importacoes_fpd += 1
                            
                            valor_total += vl_fatura_float
                        else:  # Contrato não encontrado
                            # Se não encontrou contrato M10, salva mesmo assim sem vínculo
                            # O usuário pode fazer matching depois
                            
                            # Extrai dados FPD mesmo sem contrato
                            # Já vêm como STRING do pandas, preservando zeros
                            id_contrato = str(row.get('id_contrato', '')).strip()
                            dt_venc = row.get('dt_venc_orig')
                            dt_pgto = row.get('dt_pagamento')
                            status_str = str(row.get('ds_status_fatura', 'NAO_PAGO')).upper()
                            nr_fatura = str(row.get('nr_fatura', '')).strip()
                            vl_fatura = row.get('vl_fatura', 0)
                            nr_dias_atraso = row.get('nr_dias_atraso', 0)
                            
                            # Normalizar status usando mapeamento padronizado
                            status = normalizar_status_fpd(status_str)

                            # Extrair e converter datas - Excel armazena como números serial
                            if pd.notna(dt_venc):
                                # Se for número, converter de serial Excel
                                if isinstance(dt_venc, (int, float)):
                                    dt_venc_date = (pd.Timestamp("1900-01-01") + pd.Timedelta(days=dt_venc - 2)).date()
                                else:
                                    dt_venc_date = pd.to_datetime(dt_venc).date()
                            else:
                                dt_venc_date = timezone.now().date()

                            if pd.notna(dt_pgto):
                                # Se for número, converter de serial Excel
                                if isinstance(dt_pgto, (int, float)):
                                    dt_pgto_date = (pd.Timestamp("1900-01-01") + pd.Timedelta(days=dt_pgto - 2)).date()
                                else:
                                    dt_pgto_date = pd.to_datetime(dt_pgto).date()
                            else:
                                dt_pgto_date = None

                            # Converte valores
                            vl_fatura_float = float(vl_fatura) if pd.notna(vl_fatura) else 0
                            nr_dias_atraso_int = int(nr_dias_atraso) if pd.notna(nr_dias_atraso) else 0
                        
                            # Preparar ImportacaoFPD sem contrato para bulk
                            # Buscar por nr_ordem (atualizar se existir, criar se não existir)
                            importacao_sem_contrato = importacoes_dict.get(nr_ordem)
                            
                            if importacao_sem_contrato:
                                # Atualizar registro existente com novos dados da planilha
                                importacao_sem_contrato.id_contrato = id_contrato
                                importacao_sem_contrato.nr_fatura = nr_fatura  # Atualiza também o nr_fatura
                                importacao_sem_contrato.dt_venc_orig = dt_venc_date
                                importacao_sem_contrato.dt_pagamento = dt_pgto_date
                                importacao_sem_contrato.nr_dias_atraso = nr_dias_atraso_int
                                importacao_sem_contrato.ds_status_fatura = status_str
                                importacao_sem_contrato.vl_fatura = vl_fatura_float
                                importacao_sem_contrato.contrato_m10 = None
                                importacoes_para_atualizar.append(importacao_sem_contrato)
                                registros_nao_encontrados += 1
                            else:
                                # Criar novo registro
                                importacoes_para_criar.append(ImportacaoFPD(
                                    nr_ordem=nr_ordem,
                                    nr_fatura=nr_fatura,
                                    id_contrato=id_contrato,
                                    dt_venc_orig=dt_venc_date,
                                    dt_pagamento=dt_pgto_date,
                                    nr_dias_atraso=nr_dias_atraso_int,
                                    ds_status_fatura=status_str,
                                    vl_fatura=vl_fatura_float,
                                    contrato_m10=None
                                ))
                                registros_importacoes_fpd += 1
                            
                            valor_total += vl_fatura_float
                            if len(os_nao_encontradas) < 20:
                                os_nao_encontradas.append(f"{nr_ordem} (sem contrato)")
                            continue
                    
                    except Exception as e:
                        erros_detalhados.append(f"Linha {idx+2}: {str(e)}")
                        if len(erros_detalhados) <= 10:
                            log.detalhes_json['erros'] = erros_detalhados
                
                # Executar bulk operations (reduz milhares de queries para dezenas)
                if faturas_para_criar:
                    FaturaM10.objects.bulk_create(faturas_para_criar, batch_size=500)
                if faturas_para_atualizar:
                    FaturaM10.objects.bulk_update(faturas_para_atualizar, [
                        'numero_fatura_operadora', 'valor', 'data_vencimento',
                        'data_pagamento', 'dias_atraso', 'status', 'id_contrato_fpd',
                        'dt_pagamento_fpd', 'ds_status_fatura_fpd', 'data_importacao_fpd'
                    ], batch_size=500)
                
                if importacoes_para_criar:
                    ImportacaoFPD.objects.bulk_create(importacoes_para_criar, batch_size=500)
                if importacoes_para_atualizar:
                    ImportacaoFPD.objects.bulk_update(importacoes_para_atualizar, [
                        'id_contrato', 'nr_fatura', 'dt_venc_orig', 'dt_pagamento', 'nr_dias_atraso',
                        'ds_status_fatura', 'vl_fatura', 'contrato_m10'
                    ], batch_size=500)
                
                # Salvar contratos atualizados em bulk
                contratos_para_atualizar = [c for c in contratos_dict.values() if c.numero_contrato_definitivo]
                if contratos_para_atualizar:
                    ContratoM10.objects.bulk_update(contratos_para_atualizar, ['numero_contrato_definitivo'], batch_size=500)

                # GAP 1+2: Atualizar campos FPD no ContratoM10 e recalcular elegibilidade nos contratos afetados
                if contratos_afetados_ids:
                    faturas1 = FaturaM10.objects.filter(
                        contrato_id__in=contratos_afetados_ids,
                        numero_fatura=1
                    ).select_related('contrato')
                    contratos_fpd_atualizar = []
                    for fatura in faturas1:
                        c = fatura.contrato
                        c.data_vencimento_fpd = fatura.data_vencimento
                        c.data_pagamento_fpd = fatura.data_pagamento
                        c.status_fatura_fpd = fatura.ds_status_fatura_fpd or (fatura.status if fatura.status else None)
                        c.valor_fatura_fpd = fatura.valor
                        c.nr_dias_atraso_fpd = fatura.dias_atraso or 0
                        c.data_ultima_sincronizacao_fpd = data_importacao_agora
                        contratos_fpd_atualizar.append(c)
                    if contratos_fpd_atualizar:
                        ContratoM10.objects.bulk_update(contratos_fpd_atualizar, [
                            'data_vencimento_fpd', 'data_pagamento_fpd', 'status_fatura_fpd',
                            'valor_fatura_fpd', 'nr_dias_atraso_fpd', 'data_ultima_sincronizacao_fpd'
                        ], batch_size=500)
                    for cid in contratos_afetados_ids:
                        try:
                            contrato = ContratoM10.objects.get(pk=cid)
                            contrato.calcular_elegibilidade()
                        except ContratoM10.DoesNotExist:
                            pass
                    # GAP 3: Recalcular totais das safras afetadas
                    safras_afetadas = set(
                        ContratoM10.objects.filter(id__in=contratos_afetados_ids).values_list('safra', flat=True)
                    )
                    safras_afetadas.discard(None)
                    safras_afetadas.discard('')
                    for safra_str in safras_afetadas:
                        _recalcular_totais_safra_m10(safra_str)

            # Finalizar log
            log.finalizado_em = timezone.now()
            log.calcular_duracao()
            # Total processadas = novos criados + atualizados + sem contrato criados/atualizados
            log.total_processadas = registros_importacoes_fpd + registros_atualizados
            log.total_erros = len(erros_detalhados)
            log.total_contratos_nao_encontrados = registros_nao_encontrados
            log.total_valor_importado = valor_total
            log.exemplos_nao_encontrados = ', '.join(os_nao_encontradas[:10]) if os_nao_encontradas else None
            
            # Debug log
            print(f"DEBUG Final: Pulados={registros_pulados} | Criados={registros_importacoes_fpd} | Atualizados={registros_atualizados} | Sem M10={registros_nao_encontrados}")
            
            if registros_pulados == log.total_linhas:
                # TODAS as linhas foram puladas!
                log.status = 'ERRO'
                log.mensagem_erro = f'Todas as {registros_pulados} linhas foram puladas (NR_ORDEM vazio ou inválido). Verificar formato do arquivo.'
            elif registros_nao_encontrados > 0 and registros_atualizados == 0:
                # Todos os registros foram salvos sem contrato (sem vincular ao M10)
                log.status = 'PARCIAL'
                log.mensagem_erro = f'{registros_nao_encontrados} registros FPD importados sem vínculo M10 (O.S não encontrados na base ContratoM10). Você pode fazer matching depois.'
            elif registros_nao_encontrados > 0:
                # Alguns registros com contrato, alguns sem
                log.status = 'PARCIAL'
                log.mensagem_erro = f'{registros_atualizados} registros vinculados a contratos M10, {registros_nao_encontrados} importados sem vínculo (O.S não encontradas). Pode fazer matching depois.'
            else:
                log.status = 'SUCESSO'
            
            log.save()

            # FIM DO PROCESSAMENTO - retorno já foi enviado antes via HTTP
            
        except Exception as e:
            log.status = 'ERRO'
            log.mensagem_erro = str(e)
            log.finalizado_em = timezone.now()
            log.calcular_duracao()
            log.save()


class LogsImportacaoFPDView(APIView):
    """Lista logs de importações FPD"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoFPD
        
        # Buscar últimos 20 logs do usuário (se não for admin, só seus próprios)
        if is_member(request.user, ['Admin', 'Diretoria']):
            logs = LogImportacaoFPD.objects.all().order_by('-iniciado_em')[:20]
        else:
            logs = LogImportacaoFPD.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'status_display': log.get_status_display(),
                'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_linhas': log.total_linhas,
                'total_processadas': log.total_processadas,
                'total_contratos_nao_encontrados': log.total_contratos_nao_encontrados,
                'total_valor_importado': str(log.total_valor_importado) if log.total_valor_importado else '0.00',
                'mensagem_erro': log.mensagem_erro,
                'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
                'exemplos_nao_encontrados': log.exemplos_nao_encontrados,
            })
        
        return Response({
            'success': True,
            'logs': logs_data
        })


class LogsImportacaoOSABView(APIView):
    """Lista logs de importações OSAB"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoOSAB
        
        # Buscar últimos 20 logs (todos os usuários podem ver todos os logs OSAB)
        if is_member(request.user, ['Admin', 'Diretoria']):
            logs = LogImportacaoOSAB.objects.all().order_by('-iniciado_em')[:20]
        else:
            logs = LogImportacaoOSAB.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'status_display': log.get_status_display(),
                'iniciado_em': log.iniciado_em.strftime('%d/%m/%Y %H:%M:%S') if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.strftime('%d/%m/%Y %H:%M:%S') if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_registros': log.total_registros,
                'total_processadas': log.total_processadas,
                'criados': log.criados,
                'atualizados': log.atualizados,
                'vendas_encontradas': log.vendas_encontradas,
                'ja_corretos': log.ja_corretos,
                'erros_count': log.erros_count,
                'mensagem': log.mensagem,
                'mensagem_erro': log.mensagem_erro,
                'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
                'enviar_whatsapp': log.enviar_whatsapp,
                'download_url': f"/api/crm/logs-osab/{log.id}/relatorio/",
            })
        
        return Response({
            'success': True,
            'logs': logs_data
        })


class DownloadRelatorioOSABView(APIView):
    """Gera relatório Excel da importação OSAB"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, log_id):
        from io import BytesIO
        from django.http import HttpResponse
        from crm_app.models import LogImportacaoOSAB, ImportacaoOsab
        import traceback

        try:
            log = LogImportacaoOSAB.objects.filter(id=log_id).first()
            if not log:
                return Response({'error': 'Log não encontrado.'}, status=404)

            if not is_member(request.user, ['Admin', 'Diretoria']) and log.usuario != request.user:
                return Response({'error': 'Sem permissão para acessar este relatório.'}, status=403)

            report = log.detalhes_json or {}
            logs = report.get('logs_detalhados') or []
            if not logs:
                return Response({'error': 'Relatório ainda não disponível. O relatório será gerado após a conclusão da importação.'}, status=404)

            df = pd.DataFrame(logs)
            
            # Criar campo 'resultado' simplificado baseado no resultado_crm (prioritário)
            if 'resultado_crm' in df.columns and 'resultado_osab' in df.columns:
                # Usar resultado_crm como base, se vazio usar resultado_osab
                df['resultado'] = df['resultado_crm'].fillna('')
                mask_vazio = df['resultado'] == ''
                df.loc[mask_vazio, 'resultado'] = df.loc[mask_vazio, 'resultado_osab'].fillna('')
                # Simplificar alguns valores
                df['resultado'] = df['resultado'].replace({
                    'SEM_MUDANCA_CRM': 'SEM_MUDANCA',
                    'ATUALIZADO_CRM': 'ATUALIZADO',
                    'IGNORADO_DT_REF_ANTIGA': 'IGNORADO_DT_REF',
                    'NAO_ENCONTRADO_CRM': 'NAO_ENCONTRADO_CRM',
                })
            elif 'resultado_crm' in df.columns:
                df['resultado'] = df['resultado_crm'].fillna('')
            elif 'resultado_osab' in df.columns:
                df['resultado'] = df['resultado_osab'].fillna('')
            else:
                df['resultado'] = ''
            
            # Colunas simplificadas para o relatório principal
            colunas_principais = [
                'linha', 'pedido', 'status_osab', 'resultado', 'detalhe'
            ]
            
            # Garantir que todas as colunas existam
            for col in colunas_principais:
                if col not in df.columns:
                    df[col] = ''
            
            df_principal = df[colunas_principais].copy()
            
            # Colunas completas para análise (manter para outras abas)
            colunas_completas = [
                'linha', 'pedido', 'status_osab', 'dt_ref_planilha', 'dt_ref_crm',
                'consta_osab', 'consta_crm', 'resultado_osab', 'resultado_crm', 'detalhe'
            ]
            for col in colunas_completas:
                if col not in df.columns:
                    df[col] = ''
            df_completo = df[colunas_completas]

            resumo = [
                {'metrica': 'Total registros planilha', 'valor': report.get('total_registros', len(df))},
                {'metrica': 'Vendas encontradas CRM', 'valor': report.get('vendas_encontradas', 0)},
                {'metrica': 'Atualizados CRM', 'valor': report.get('atualizados', 0)},
                {'metrica': 'Criados OSAB', 'valor': report.get('criados', 0)},
                {'metrica': 'Ignorados DT_REF', 'valor': report.get('ignorados_dt_ref', 0)},
                {'metrica': 'Erros', 'valor': len(report.get('erros', []))},
            ]
            df_resumo = pd.DataFrame(resumo)
            
            # Tratar caso de DataFrame vazio
            if len(df_completo) > 0:
                df_contagem_osab = df_completo['resultado_osab'].value_counts().reset_index()
                df_contagem_osab.columns = ['resultado_osab', 'quantidade']
                df_contagem_resultado = df_principal['resultado'].value_counts().reset_index()
                df_contagem_resultado.columns = ['resultado', 'quantidade']
                df_planilha_nao_crm = df_completo[df_completo['resultado_crm'] == 'NAO_ENCONTRADO_CRM'].copy()
            else:
                df_contagem_osab = pd.DataFrame(columns=['resultado_osab', 'quantidade'])
                df_contagem_resultado = pd.DataFrame(columns=['resultado', 'quantidade'])
                df_planilha_nao_crm = pd.DataFrame(columns=colunas_completas)

            pedidos_planilha = set(df_principal['pedido'].dropna().astype(str)) if len(df_principal) > 0 else set()
            qs_crm = ImportacaoOsab.objects.exclude(documento__in=pedidos_planilha).values(
                'documento', 'dt_ref', 'uf', 'localidade', 'produto'
            )
            df_crm_nao_planilha = pd.DataFrame(list(qs_crm))
            if df_crm_nao_planilha.empty:
                df_crm_nao_planilha = pd.DataFrame(columns=['documento', 'dt_ref', 'uf', 'localidade', 'produto'])

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                df_contagem_resultado.to_excel(writer, sheet_name='Resumo_Resultados', index=False)
                df_contagem_osab.to_excel(writer, sheet_name='Resumo_OSAB', index=False)
                df_principal.to_excel(writer, sheet_name='Detalhado', index=False)  # Aba principal simplificada
                df_completo.to_excel(writer, sheet_name='Completo', index=False)  # Aba com todos os campos
                df_planilha_nao_crm.to_excel(writer, sheet_name='Planilha_nao_CRM', index=False)
                df_crm_nao_planilha.to_excel(writer, sheet_name='CRM_nao_Planilha', index=False)

            output.seek(0)
            # Limpar nome do arquivo: remover extensões e caracteres especiais
            nome_limpo = log.nome_arquivo.rsplit('.', 1)[0]  # Remove extensão (.xlsb, .xlsx, etc)
            nome_limpo = nome_limpo.replace(' ', '_').replace('-', '_')
            # Remover underscores repetidos e no final
            import re
            nome_limpo = re.sub(r'_+', '_', nome_limpo)  # Remove underscores repetidos
            nome_limpo = nome_limpo.rstrip('_')  # Remove underscores no final
            filename = f"Relatorio_OSAB_{log.id}_{nome_limpo}.xlsx"
            # Garantir que filename não termina com underscore antes da extensão
            filename = re.sub(r'_+\.xlsx$', '.xlsx', filename)  # Remove underscore(s) antes de .xlsx
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro ao gerar relatório OSAB (log_id={log_id}): {str(e)}\n{traceback.format_exc()}")
            return Response({'error': f'Erro ao gerar relatório: {str(e)}'}, status=500)


class AnaliseComparacaoOSABView(APIView):
    """Análise comparativa entre planilha OSAB e banco de dados (sem importar)"""
    permission_classes = [CheckAPIPermission]
    resource_name = 'importacaoosab'
    parser_classes = [MultiPartParser, FormParser]

    def _clean_key(self, key):
        import pandas as pd
        if pd.isna(key) or key is None: return None
        return str(key).replace('.0', '').strip()

    def _normalize_dt_ref(self, val):
        import datetime as dt_sys
        if val is None:
            return None
        if isinstance(val, dt_sys.datetime):
            return val.date()
        if isinstance(val, dt_sys.date):
            return val
        return None

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=400)

        try:
            from io import BytesIO
            from crm_app.models import ImportacaoOsab
            import pandas as pd
            import numpy as np

            # Ler arquivo
            file_buffer = BytesIO(file_obj.read())
            file_name = file_obj.name
            
            if file_name.endswith('.xlsb'):
                df = pd.read_excel(file_buffer, engine='pyxlsb')
            elif file_name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_buffer)
            else:
                return Response({'error': 'Formato inválido. Use .xlsx, .xls ou .xlsb'}, status=400)

            # Normalizar colunas
            df.columns = [str(col).strip().upper().replace(' ', '_') for col in df.columns]
            
            if 'PEDIDO' not in df.columns:
                return Response({'error': 'Coluna PEDIDO não encontrada na planilha'}, status=400)

            # Limpar pedidos
            lista_pedidos_limpos = []
            for p in df['PEDIDO'].dropna():
                p_limpo = self._clean_key(p)
                if p_limpo:
                    lista_pedidos_limpos.append(p_limpo)

            if not lista_pedidos_limpos:
                return Response({'error': 'Nenhum pedido válido encontrado na planilha'}, status=400)

            # Buscar registros no banco
            osab_existentes = {
                obj.documento: obj for obj in ImportacaoOsab.objects.filter(documento__in=lista_pedidos_limpos)
            }

            # Análise
            total_planilha = len(df)
            total_banco = len(osab_existentes)
            
            analise_detalhada = []
            stats = {
                'existe_banco': 0,
                'nao_existe_banco': 0,
                'dt_ref_planilha_none': 0,
                'dt_ref_banco_none': 0,
                'seria_ignorado': 0,
                'seria_processado': 0,
                'dt_ref_planilha_maior': 0,
                'dt_ref_planilha_igual': 0,
                'dt_ref_planilha_menor': 0,
            }

            # Processar parser de data (mesma lógica da importação)
            def smart_date_parser(val):
                import datetime as dt_sys
                import pandas as pd
                if val is None or pd.isna(val) or val == '':
                    return None
                if isinstance(val, (dt_sys.datetime, dt_sys.date, pd.Timestamp)):
                    return val.date() if hasattr(val, 'date') else val
                if isinstance(val, (float, int)):
                    try:
                        return (dt_sys.datetime(1899, 12, 30) + dt_sys.timedelta(days=float(val))).date()
                    except:
                        return None
                s_val = str(val).strip()
                import re
                match_br = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s_val)
                if match_br:
                    d, m, y = match_br.groups()
                    try:
                        return dt_sys.date(int(y), int(m), int(d))
                    except ValueError:
                        pass
                match_iso = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s_val)
                if match_iso:
                    y, m, d = match_iso.groups()
                    try:
                        return dt_sys.date(int(y), int(m), int(d))
                    except ValueError:
                        pass
                return None

            # Processar DT_REF da planilha
            if 'DT_REF' in df.columns:
                df['DT_REF_PARSED'] = df['DT_REF'].apply(smart_date_parser)
            else:
                df['DT_REF_PARSED'] = None

            # Analisar cada linha
            for index, row in df.iterrows():
                pedido = self._clean_key(row.get('PEDIDO'))
                if not pedido:
                    continue

                dt_ref_planilha = row.get('DT_REF_PARSED')
                obj_banco = osab_existentes.get(pedido)
                
                item_analise = {
                    'linha': index + 2,
                    'pedido': pedido,
                    'dt_ref_planilha': str(dt_ref_planilha) if dt_ref_planilha else None,
                    'existe_banco': obj_banco is not None,
                    'dt_ref_banco': str(obj_banco.dt_ref) if obj_banco and obj_banco.dt_ref else None,
                    'seria_ignorado': False,
                    'motivo': ''
                }

                if obj_banco:
                    stats['existe_banco'] += 1
                    dt_ref_banco = self._normalize_dt_ref(obj_banco.dt_ref)
                    
                    if dt_ref_banco is None:
                        stats['dt_ref_banco_none'] += 1
                    if dt_ref_planilha is None:
                        stats['dt_ref_planilha_none'] += 1
                    
                    # Aplicar mesma lógica da importação
                    if dt_ref_banco and (dt_ref_planilha is None or dt_ref_planilha <= dt_ref_banco):
                        stats['seria_ignorado'] += 1
                        item_analise['seria_ignorado'] = True
                        if dt_ref_planilha is None:
                            item_analise['motivo'] = 'DT_REF planilha é None'
                        elif dt_ref_planilha <= dt_ref_banco:
                            if dt_ref_planilha == dt_ref_banco:
                                stats['dt_ref_planilha_igual'] += 1
                                item_analise['motivo'] = f'DT_REF planilha ({dt_ref_planilha}) == DT_REF banco ({dt_ref_banco})'
                            else:
                                stats['dt_ref_planilha_menor'] += 1
                                item_analise['motivo'] = f'DT_REF planilha ({dt_ref_planilha}) < DT_REF banco ({dt_ref_banco})'
                    else:
                        stats['seria_processado'] += 1
                        if dt_ref_planilha and dt_ref_banco and dt_ref_planilha > dt_ref_banco:
                            stats['dt_ref_planilha_maior'] += 1
                            item_analise['motivo'] = f'DT_REF planilha ({dt_ref_planilha}) > DT_REF banco ({dt_ref_banco}) - SERIA PROCESSADO'
                else:
                    stats['nao_existe_banco'] += 1
                    stats['seria_processado'] += 1
                    item_analise['motivo'] = 'Não existe no banco - seria criado'

                analise_detalhada.append(item_analise)

            return Response({
                'resumo': {
                    'total_planilha': total_planilha,
                    'total_banco': total_banco,
                    'existe_banco': stats['existe_banco'],
                    'nao_existe_banco': stats['nao_existe_banco'],
                    'seria_ignorado': stats['seria_ignorado'],
                    'seria_processado': stats['seria_processado'],
                    'dt_ref_planilha_none': stats['dt_ref_planilha_none'],
                    'dt_ref_banco_none': stats['dt_ref_banco_none'],
                    'dt_ref_planilha_maior': stats['dt_ref_planilha_maior'],
                    'dt_ref_planilha_igual': stats['dt_ref_planilha_igual'],
                    'dt_ref_planilha_menor': stats['dt_ref_planilha_menor'],
                },
                'detalhes': analise_detalhada[:100],  # Limitar a 100 primeiros para não sobrecarregar
                'total_detalhes': len(analise_detalhada)
            })

        except Exception as e:
            import traceback
            return Response({'error': f'Erro na análise: {str(e)}\n{traceback.format_exc()}'}, status=500)


class LimparImportacaoOSABView(APIView):
    """Limpa todos os registros da tabela ImportacaoOsab"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        from crm_app.models import ImportacaoOsab
        from rest_framework.permissions import IsAdminUser
        
        # Apenas admin pode limpar
        if not request.user.is_staff:
            return Response({'error': 'Apenas administradores podem limpar a tabela OSAB.'}, status=403)
        
        try:
            count = ImportacaoOsab.objects.count()
            ImportacaoOsab.objects.all().delete()
            return Response({
                'mensagem': f'Tabela ImportacaoOsab limpa com sucesso.',
                'registros_removidos': count
            })
        except Exception as e:
            return Response({'error': f'Erro ao limpar tabela: {str(e)}'}, status=500)


class CancelarImportacaoOSABView(APIView):
    """Cancela uma importação OSAB em processamento."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, log_id):
        from django.utils import timezone
        from crm_app.models import LogImportacaoOSAB

        log = LogImportacaoOSAB.objects.filter(id=log_id).first()
        if not log:
            return Response({'error': 'Log não encontrado.'}, status=404)

        if not is_member(request.user, ['Admin', 'Diretoria']) and log.usuario != request.user:
            return Response({'error': 'Sem permissão para cancelar este log.'}, status=403)

        if log.status != 'PROCESSANDO':
            return Response({'error': 'A importação já foi finalizada.'}, status=400)

        LogImportacaoOSAB.objects.filter(id=log_id).update(
            status='ERRO',
            mensagem_erro='Cancelado manualmente pelo usuário.',
            finalizado_em=timezone.now()
        )
        log.calcular_duracao()

        return Response({'success': True, 'message': 'Importação cancelada.'})


class LogsImportacaoDFVView(APIView):
    """Lista logs de importações DFV"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoDFV
        
        # Buscar últimos 20 logs
        if is_member(request.user, ['Admin', 'Diretoria']):
            logs = LogImportacaoDFV.objects.all().order_by('-iniciado_em')[:20]
        else:
            logs = LogImportacaoDFV.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'iniciado_em': log.iniciado_em.strftime('%d/%m/%Y %H:%M:%S') if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.strftime('%d/%m/%Y %H:%M:%S') if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_registros': log.total_registros,
                'total_processadas': log.total_processadas,
                'sucesso': log.sucesso,
                'erros': log.erros,
                'total_valor_importado': str(log.total_valor_importado) if log.total_valor_importado else '0.00',
                'mensagem': log.mensagem,
                'mensagem_erro': log.mensagem_erro,
                'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
            })
        
        return Response({
            'success': True,
            'logs': logs_data
        })


class LogsImportacaoRecompraView(APIView):
    """Lista logs de importações Recompra"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoRecompra
        from django.db import ProgrammingError
        
        try:
            # Buscar últimos 20 logs
            if is_member(request.user, ['Admin', 'Diretoria']):
                logs = LogImportacaoRecompra.objects.all().order_by('-iniciado_em')[:20]
            else:
                logs = LogImportacaoRecompra.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
            
            logs_data = []
            for log in logs:
                logs_data.append({
                    'id': log.id,
                    'nome_arquivo': log.nome_arquivo,
                    'status': log.status,
                    'status_display': log.get_status_display(),
                    'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                    'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                    'duracao_segundos': log.duracao_segundos,
                    'total_linhas': log.total_linhas,
                    'total_processadas': log.total_processadas,
                    'registros_criados': log.registros_criados,
                    'erros_count': log.erros_count,
                    'mensagem': log.mensagem,
                    'mensagem_erro': log.mensagem_erro,
                    'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
                })
            
            return Response({
                'success': True,
                'logs': logs_data
            })
        except ProgrammingError as e:
            # Tabela não existe ainda - migração não aplicada
            return Response({
                'success': False,
                'error': 'Tabela não existe. Por favor, execute: python manage.py migrate',
                'logs': []
            }, status=503)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e),
                'logs': []
            }, status=500)


class LogsImportacaoLegadoView(APIView):
    """Lista logs de importações Legado"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoLegado
        
        # Buscar últimos 20 logs
        if is_member(request.user, ['Admin', 'Diretoria']):
            logs = LogImportacaoLegado.objects.all().order_by('-iniciado_em')[:20]
        else:
            logs = LogImportacaoLegado.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'status_display': log.get_status_display(),
                'iniciado_em': log.iniciado_em.strftime('%d/%m/%Y %H:%M:%S') if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.strftime('%d/%m/%Y %H:%M:%S') if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_linhas': log.total_linhas,
                'total_processadas': log.total_processadas,
                'vendas_criadas': log.vendas_criadas,
                'vendas_atualizadas': log.vendas_atualizadas,
                'clientes_criados': log.clientes_criados,
                'erros_count': log.erros_count,
                'mensagem': log.mensagem,
                'mensagem_erro': log.mensagem_erro,
                'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
            })
        
        return Response({
            'success': True,
            'logs': logs_data
        })


class CancelarImportacaoAgendamentoView(APIView):
    """Cancela uma importação de Agendamento em andamento"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, log_id):
        try:
            log = LogImportacaoAgendamento.objects.get(id=log_id)
            
            # Verificar se pode cancelar (apenas PROCESSANDO)
            if log.status != 'PROCESSANDO':
                return Response({
                    'success': False,
                    'error': f'Esta importação não pode ser cancelada. Status atual: {log.get_status_display()}'
                }, status=400)
            
            # Marcar como cancelado
            log.status = 'CANCELADO'
            log.finalizado_em = timezone.now()
            log.calcular_duracao()
            log.mensagem_erro = 'Processo cancelado pelo usuário.'
            log.mensagem = 'Importação cancelada pelo usuário antes da conclusão.'
            log.save()
            
            # Mensagem informativa sobre o que pode ter acontecido
            mensagem_info = (
                "Importação cancelada. Possíveis motivos para processos travados:\n"
                "- Arquivo muito grande pode demorar para processar\n"
                "- Problemas de conexão com o banco de dados\n"
                "- Erro não capturado no código de processamento\n"
                "- Thread foi interrompida pelo servidor\n"
                "- Processo lento devido a muitas operações no banco"
            )
            
            return Response({
                'success': True,
                'message': 'Importação cancelada com sucesso.',
                'info': mensagem_info
            })
            
        except LogImportacaoAgendamento.DoesNotExist:
            return Response({
                'success': False,
                'error': 'Log de importação não encontrado.'
            }, status=404)
        except Exception as e:
            return Response({
                'success': False,
                'error': f'Erro ao cancelar importação: {str(e)}'
            }, status=500)


class LogsImportacaoAgendamentoView(APIView):
    """Lista logs de importações Agendamento"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import LogImportacaoAgendamento
        
        # Buscar últimos 20 logs
        if is_member(request.user, ['Admin', 'Diretoria']):
            logs = LogImportacaoAgendamento.objects.all().order_by('-iniciado_em')[:20]
        else:
            logs = LogImportacaoAgendamento.objects.filter(usuario=request.user).order_by('-iniciado_em')[:20]
        
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'nome_arquivo': log.nome_arquivo,
                'status': log.status,
                'status_display': log.get_status_display(),
                'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                'duracao_segundos': log.duracao_segundos,
                'total_linhas': log.total_linhas,
                'total_processadas': log.total_processadas,
                'agendamentos_criados': log.agendamentos_criados,
                'agendamentos_atualizados': log.agendamentos_atualizados,
                'nao_encontrados': log.nao_encontrados,
                'erros_count': log.erros_count,
                'mensagem': log.mensagem,
                'mensagem_erro': log.mensagem_erro,
                'usuario': log.usuario.username if log.usuario else 'Sistema',
                'usuario_nome': log.usuario.get_full_name() if log.usuario else 'Sistema',
            })
        
        return Response({
            'success': True,
            'logs': logs_data
        })


class ImportarChurnView(APIView):
    """Importa base de churn (cancelamentos) e faz crossover com ContratoM10 por O.S"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        from .models import ImportacaoChurn, ContratoM10, LogImportacaoChurn
        
        if not is_member(request.user, ['Admin', 'BackOffice', 'Diretoria']):
            return Response({'error': 'Sem permissão'}, status=403)

        arquivo = request.FILES.get('file')
        if not arquivo:
            return Response({'error': 'Arquivo não enviado'}, status=400)

        # Criar log de importação
        log = LogImportacaoChurn.objects.create(
            nome_arquivo=arquivo.name,
            tamanho_arquivo=getattr(arquivo, 'size', 0) or 0,
            usuario=request.user,
            status='PROCESSANDO'
        )
        inicio = datetime.now()

        try:
            if arquivo.name.endswith('.csv'):
                df = pd.read_csv(arquivo, dtype={'PEDIDO': str, 'NR_ORDEM': str, 'NUMERO_PEDIDO': str})
            elif arquivo.name.endswith('.xlsb'):
                try:
                    df = pd.read_excel(arquivo, engine='pyxlsb', dtype={'PEDIDO': str, 'NR_ORDEM': str, 'NUMERO_PEDIDO': str})
                except Exception as e:
                    log.status = 'ERRO'
                    log.mensagem_erro = (
                        'Formato .xlsb não suportado ou falha ao ler. '
                        'Use arquivo .xlsx ou .csv. Se precisar usar .xlsb, instale: pip install pyxlsb'
                    )
                    log.finalizado_em = datetime.now()
                    log.save()
                    return Response({
                        'error': log.mensagem_erro
                    }, status=400)
            else:
                df = pd.read_excel(arquivo, dtype={'PEDIDO': str, 'NR_ORDEM': str, 'NUMERO_PEDIDO': str})

            # Normalizar nomes das colunas (ex.: "ANOMES GROSS" -> "ANOMES_GROSS", "NR ORDEM" -> "NR_ORDEM")
            df.columns = df.columns.str.strip().str.upper().str.replace(' ', '_', regex=False)
            
            log.total_linhas = len(df)
            log.save()

            cancelados = 0
            criados_churn = 0
            atualizados_churn = 0
            reativados = 0
            nao_encontrados = 0
            erros = 0
            
            # Coletar todas as O.S que aparecem no CHURN
            ordens_no_churn = set()

            # GAP 4: Dicionário de ContratoM10 por várias formas da O.S. (como no FPD)
            contratos_churn_dict = {}
            for c in ContratoM10.objects.exclude(ordem_servico__isnull=True).exclude(ordem_servico=''):
                os_val = str(c.ordem_servico).strip()
                os_sem_zeros = os_val.lstrip('0') or '0'
                for key in [os_val, os_sem_zeros, os_val.zfill(8), f'OS-{os_val}', f'OS-{os_sem_zeros}', f'OS-{os_val.zfill(8)}']:
                    if key and key not in contratos_churn_dict:
                        contratos_churn_dict[key] = c

            for _, row in df.iterrows():
                try:
                    # Busca por O.S - pode vir como NR_ORDEM ou NUMERO_PEDIDO
                    # Prioridade: NR_ORDEM se existir, senão NUMERO_PEDIDO
                    nr_ordem_raw = row.get('NR_ORDEM', '')
                    if pd.isna(nr_ordem_raw) or str(nr_ordem_raw).strip() == '':
                        # Tentar NUMERO_PEDIDO como fallback
                        numero_pedido_raw = row.get('NUMERO_PEDIDO', '')
                        if pd.notna(numero_pedido_raw) and str(numero_pedido_raw).strip():
                            nr_ordem_raw = numero_pedido_raw
                        else:
                            continue  # Pula se não tiver nenhum dos dois
                    
                    nr_ordem = str(nr_ordem_raw).strip().zfill(8)  # Padronizar com 8 dígitos
                    nr_ordem_sem_zeros = nr_ordem.lstrip('0') or '0'
                    ordens_no_churn.add(nr_ordem)
                    ordens_no_churn.add(nr_ordem_sem_zeros)

                    # Salvar registro na ImportacaoChurn
                    try:
                        # Usar PEDIDO como principal, com fallback para NUMERO_PEDIDO
                        numero_pedido_raw = row.get('PEDIDO', '') or row.get('NUMERO_PEDIDO', '')
                        numero_pedido_val = str(numero_pedido_raw).strip() if pd.notna(numero_pedido_raw) and str(numero_pedido_raw).strip() else None
                        
                        # Se numero_pedido for None, usar nr_ordem como chave alternativa
                        # Mas como numero_pedido tem unique=True, não podemos usar None
                        # Vamos usar uma chave composta ou tratar de outra forma
                        if numero_pedido_val:
                            obj, created = ImportacaoChurn.objects.update_or_create(
                                numero_pedido=numero_pedido_val,
                                defaults={
                                    'nr_ordem': nr_ordem,
                                    'uf': str(row.get('UF', ''))[:2] if pd.notna(row.get('UF')) else None,
                                    'produto': str(row.get('PRODUTO', '')) if pd.notna(row.get('PRODUTO')) else None,
                                    'matricula_vendedor': str(row.get('MATRICULA_VENDEDOR', '')) if pd.notna(row.get('MATRICULA_VENDEDOR')) else None,
                                    'gv': str(row.get('GV', '')) if pd.notna(row.get('GV')) else None,
                                    'sap_principal_fim': str(row.get('SAP_PRINCIPAL_FIM', '')) if pd.notna(row.get('SAP_PRINCIPAL_FIM')) else None,
                                    'gestao': str(row.get('GESTAO', '')) if pd.notna(row.get('GESTAO')) else None,
                                    'st_regional': str(row.get('ST_REGIONAL', '')) if pd.notna(row.get('ST_REGIONAL')) else None,
                                    'gc': str(row.get('GC', '')) if pd.notna(row.get('GC')) else None,
                                    'dt_gross': pd.to_datetime(row.get('DT_GROSS')).date() if pd.notna(row.get('DT_GROSS')) else None,
                                    'anomes_gross': _normalizar_anomes_gross(row.get('ANOMES_GROSS', '')),
                                    'dt_retirada': pd.to_datetime(row.get('DT_RETIRADA')).date() if pd.notna(row.get('DT_RETIRADA')) else None,
                                    'anomes_retirada': str(row.get('ANOMES_RETIRADA', '')) if pd.notna(row.get('ANOMES_RETIRADA')) else None,
                                    'grupo_unidade': str(row.get('GRUPO_UNIDADE', '')) if pd.notna(row.get('GRUPO_UNIDADE')) else None,
                                    'codigo_sap': str(row.get('CODIGO_SAP', '')) if pd.notna(row.get('CODIGO_SAP')) else None,
                                    'municipio': str(row.get('MUNICIPIO', '')) if pd.notna(row.get('MUNICIPIO')) else None,
                                    'tipo_retirada': str(row.get('TIPO_RETIRADA', '')) if pd.notna(row.get('TIPO_RETIRADA')) else None,
                                    'motivo_retirada': str(row.get('MOTIVO_RETIRADA', '')) if pd.notna(row.get('MOTIVO_RETIRADA')) else None,
                                    'submotivo_retirada': str(row.get('SUBMOTIVO_RETIRADA', '')) if pd.notna(row.get('SUBMOTIVO_RETIRADA')) else None,
                                    'classificacao': str(row.get('CLASSIFICACAO', '')) if pd.notna(row.get('CLASSIFICACAO')) else None,
                                    'desc_apelido': str(row.get('DESC_APELIDO', '')) if pd.notna(row.get('DESC_APELIDO')) else None,
                                    'nr_velocidade': str(row.get('NR_VELOCIDADE', '') or row.get('VELOCIDADE', ''))[:50] if pd.notna(row.get('NR_VELOCIDADE', row.get('VELOCIDADE'))) else None,
                                }
                            )
                            if created:
                                criados_churn += 1
                            else:
                                atualizados_churn += 1
                        else:
                            # Se numero_pedido for None, usar nr_ordem como alternativa
                            # Buscar por nr_ordem se numero_pedido não estiver disponível
                            try:
                                obj_existente = ImportacaoChurn.objects.get(nr_ordem=nr_ordem)
                                # Atualizar campos
                                for campo, valor in {
                                    'uf': str(row.get('UF', ''))[:2] if pd.notna(row.get('UF')) else None,
                                    'produto': str(row.get('PRODUTO', '')) if pd.notna(row.get('PRODUTO')) else None,
                                    'matricula_vendedor': str(row.get('MATRICULA_VENDEDOR', '')) if pd.notna(row.get('MATRICULA_VENDEDOR')) else None,
                                    'gv': str(row.get('GV', '')) if pd.notna(row.get('GV')) else None,
                                    'sap_principal_fim': str(row.get('SAP_PRINCIPAL_FIM', '')) if pd.notna(row.get('SAP_PRINCIPAL_FIM')) else None,
                                    'gestao': str(row.get('GESTAO', '')) if pd.notna(row.get('GESTAO')) else None,
                                    'st_regional': str(row.get('ST_REGIONAL', '')) if pd.notna(row.get('ST_REGIONAL')) else None,
                                    'gc': str(row.get('GC', '')) if pd.notna(row.get('GC')) else None,
                                    'dt_gross': pd.to_datetime(row.get('DT_GROSS')).date() if pd.notna(row.get('DT_GROSS')) else None,
                                    'anomes_gross': _normalizar_anomes_gross(row.get('ANOMES_GROSS', '')),
                                    'dt_retirada': pd.to_datetime(row.get('DT_RETIRADA')).date() if pd.notna(row.get('DT_RETIRADA')) else None,
                                    'anomes_retirada': str(row.get('ANOMES_RETIRADA', '')) if pd.notna(row.get('ANOMES_RETIRADA')) else None,
                                    'grupo_unidade': str(row.get('GRUPO_UNIDADE', '')) if pd.notna(row.get('GRUPO_UNIDADE')) else None,
                                    'codigo_sap': str(row.get('CODIGO_SAP', '')) if pd.notna(row.get('CODIGO_SAP')) else None,
                                    'municipio': str(row.get('MUNICIPIO', '')) if pd.notna(row.get('MUNICIPIO')) else None,
                                    'tipo_retirada': str(row.get('TIPO_RETIRADA', '')) if pd.notna(row.get('TIPO_RETIRADA')) else None,
                                    'motivo_retirada': str(row.get('MOTIVO_RETIRADA', '')) if pd.notna(row.get('MOTIVO_RETIRADA')) else None,
                                    'submotivo_retirada': str(row.get('SUBMOTIVO_RETIRADA', '')) if pd.notna(row.get('SUBMOTIVO_RETIRADA')) else None,
                                    'classificacao': str(row.get('CLASSIFICACAO', '')) if pd.notna(row.get('CLASSIFICACAO')) else None,
                                    'desc_apelido': str(row.get('DESC_APELIDO', '')) if pd.notna(row.get('DESC_APELIDO')) else None,
                                }.items():
                                    setattr(obj_existente, campo, valor)
                                obj_existente.save()
                                atualizados_churn += 1
                            except ImportacaoChurn.DoesNotExist:
                                # Criar novo registro sem numero_pedido (permitido pelo modelo)
                                ImportacaoChurn.objects.create(
                                    numero_pedido=None,
                                    nr_ordem=nr_ordem,
                                    uf=str(row.get('UF', ''))[:2] if pd.notna(row.get('UF')) else None,
                                    produto=str(row.get('PRODUTO', '')) if pd.notna(row.get('PRODUTO')) else None,
                                    matricula_vendedor=str(row.get('MATRICULA_VENDEDOR', '')) if pd.notna(row.get('MATRICULA_VENDEDOR')) else None,
                                    gv=str(row.get('GV', '')) if pd.notna(row.get('GV')) else None,
                                    sap_principal_fim=str(row.get('SAP_PRINCIPAL_FIM', '')) if pd.notna(row.get('SAP_PRINCIPAL_FIM')) else None,
                                    gestao=str(row.get('GESTAO', '')) if pd.notna(row.get('GESTAO')) else None,
                                    st_regional=str(row.get('ST_REGIONAL', '')) if pd.notna(row.get('ST_REGIONAL')) else None,
                                    gc=str(row.get('GC', '')) if pd.notna(row.get('GC')) else None,
                                    dt_gross=pd.to_datetime(row.get('DT_GROSS')).date() if pd.notna(row.get('DT_GROSS')) else None,
                                    anomes_gross=_normalizar_anomes_gross(row.get('ANOMES_GROSS', '')),
                                    dt_retirada=pd.to_datetime(row.get('DT_RETIRADA')).date() if pd.notna(row.get('DT_RETIRADA')) else None,
                                    anomes_retirada=str(row.get('ANOMES_RETIRADA', '')) if pd.notna(row.get('ANOMES_RETIRADA')) else None,
                                    grupo_unidade=str(row.get('GRUPO_UNIDADE', '')) if pd.notna(row.get('GRUPO_UNIDADE')) else None,
                                    codigo_sap=str(row.get('CODIGO_SAP', '')) if pd.notna(row.get('CODIGO_SAP')) else None,
                                    municipio=str(row.get('MUNICIPIO', '')) if pd.notna(row.get('MUNICIPIO')) else None,
                                    tipo_retirada=str(row.get('TIPO_RETIRADA', '')) if pd.notna(row.get('TIPO_RETIRADA')) else None,
                                    motivo_retirada=str(row.get('MOTIVO_RETIRADA', '')) if pd.notna(row.get('MOTIVO_RETIRADA')) else None,
                                    submotivo_retirada=str(row.get('SUBMOTIVO_RETIRADA', '')) if pd.notna(row.get('SUBMOTIVO_RETIRADA')) else None,
                                    classificacao=str(row.get('CLASSIFICACAO', '')) if pd.notna(row.get('CLASSIFICACAO')) else None,
                                    desc_apelido=str(row.get('DESC_APELIDO', '')) if pd.notna(row.get('DESC_APELIDO')) else None,
                                    nr_velocidade=str(row.get('NR_VELOCIDADE', '') or row.get('VELOCIDADE', ''))[:50] if pd.notna(row.get('NR_VELOCIDADE', row.get('VELOCIDADE'))) else None,
                                )
                                criados_churn += 1
                    except Exception as e:
                        erros += 1
                        print(f"Erro ao salvar ImportacaoChurn: {e}")

                    # Atualizar status do contrato M10 para CANCELADO (GAP 4: match por várias formas de O.S.)
                    try:
                        contrato = None
                        for key in [nr_ordem, nr_ordem_sem_zeros, str(nr_ordem_raw).strip(), f'OS-{nr_ordem}', f'OS-{nr_ordem_sem_zeros}']:
                            if key and key in contratos_churn_dict:
                                contrato = contratos_churn_dict[key]
                                break
                        if contrato is None:
                            nao_encontrados += 1
                            # Salvar ImportacaoChurn já foi feito acima; segue para próxima linha
                            continue
                        
                        # Marca como cancelado (apareceu no CHURN)
                        if contrato.status_contrato != 'CANCELADO':
                            contrato.status_contrato = 'CANCELADO'
                            contrato.data_cancelamento = pd.to_datetime(row.get('DT_RETIRADA')).date() if pd.notna(row.get('DT_RETIRADA')) else datetime.now().date()
                            contrato.motivo_cancelamento = str(row.get('MOTIVO_RETIRADA', '')) if pd.notna(row.get('MOTIVO_RETIRADA')) else 'CHURN'
                            contrato.elegivel_bonus = False
                            contrato.save()
                            cancelados += 1
                        
                    except Exception:
                        nao_encontrados += 1
                
                except Exception as e:
                    erros += 1
                    continue

            # GAP 6: Reativar apenas contratos que foram cancelados POR CHURN e cuja O.S. não está na planilha atual
            contratos_reativar = ContratoM10.objects.filter(
                status_contrato='CANCELADO',
                motivo_cancelamento__icontains='CHURN',
            ).exclude(ordem_servico__in=ordens_no_churn)
            reativados = contratos_reativar.update(
                status_contrato='ATIVO', data_cancelamento=None, motivo_cancelamento=None
            )

            # Atualizar log
            fim = datetime.now()
            log.finalizado_em = fim
            log.duracao_segundos = int((fim - inicio).total_seconds())
            log.total_processadas = criados_churn + atualizados_churn
            log.total_erros = erros
            log.total_contratos_cancelados = cancelados
            log.total_contratos_reativados = reativados
            log.total_nao_encontrados = nao_encontrados
            log.status = 'PARCIAL' if (cancelados > 0 and nao_encontrados > 0) else 'SUCESSO'
            log.detalhes_json = {
                'ordens_unicas': len(ordens_no_churn),
                'cancelados': cancelados,
                'reativados': reativados,
                'nao_encontrados': nao_encontrados,
                'criados_churn': criados_churn,
                'atualizados_churn': atualizados_churn,
            }
            log.save()

            return Response({
                'message': f'Base CHURN processada! {cancelados} contratos cancelados, {reativados} contratos reativados, {criados_churn + atualizados_churn} registros processados.',
                'total_registros': log.total_linhas,
                'criados': criados_churn,  # Registros criados na ImportacaoChurn
                'atualizados': atualizados_churn,  # Registros atualizados na ImportacaoChurn
                'cancelados': cancelados,
                'reativados': reativados,
                'salvos_churn': criados_churn + atualizados_churn,
                'nao_encontrados': nao_encontrados,
                'log_id': log.id,
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            
            # Atualizar log com erro
            log.status = 'ERRO'
            log.mensagem_erro = str(e)
            log.finalizado_em = datetime.now()
            log.save()
            
            return Response({'error': f'Erro ao processar arquivo: {str(e)}'}, status=500)


class AtualizarFaturasView(APIView):
    """Atualiza múltiplas faturas de uma vez"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Admin', 'BackOffice', 'Diretoria']):
            return Response({'error': 'Sem permissão'}, status=403)

        updates = request.data.get('updates', [])
        
        for update in updates:
            fatura_id = update.get('id')
            field = update.get('field')
            value = update.get('value')

            try:
                fatura = FaturaM10.objects.get(id=fatura_id)
                
                if field == 'status':
                    fatura.status = value
                elif field == 'valor':
                    fatura.valor = float(value) if value else 0
                elif field == 'data_vencimento':
                    fatura.data_vencimento = datetime.strptime(value, '%Y-%m-%d').date() if value else None
                elif field == 'data_pagamento':
                    fatura.data_pagamento = datetime.strptime(value, '%Y-%m-%d').date() if value else None
                elif field == 'numero_fatura_operadora':
                    fatura.numero_fatura_operadora = value

                fatura.save()
                
                # Recalcula elegibilidade do contrato
                fatura.contrato.calcular_elegibilidade()

            except FaturaM10.DoesNotExist:
                continue

        return Response({'message': 'Faturas atualizadas com sucesso!'})


class ExportarM10View(APIView):
    """Exporta dados para Excel"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Cria workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bônus M-10"

        # Cabeçalhos
        headers = ['O.S/PEDIDO', 'Nº Contrato', 'Cliente', 'Vendedor', 'Instalação', 'Plano', 'Status', 'Faturas Pagas', 'Elegível', 'Bônus']
        ws.append(headers)

        # Estilo cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Dados
        contratos = (
            ContratoM10.objects.all()
            .select_related('vendedor')
            .annotate(
                total_faturas=Count('faturas', distinct=True),
                faturas_pagas=Count('faturas', filter=Q(faturas__status='PAGO'), distinct=True),
            )
        )
        for c in contratos:
            total_faturas = c.total_faturas
            faturas_pagas = c.faturas_pagas
            if total_faturas == 0 and c.status_fatura_fpd and str(c.status_fatura_fpd).lower().startswith('paga'):
                total_faturas = 1
                faturas_pagas = 1
            elegivel = (
                c.status_contrato == 'ATIVO'
                and not c.teve_downgrade
                and (total_faturas or 0) > 0
                and faturas_pagas == total_faturas
            )
            bonus = 150 if elegivel else 0
            ws.append([
                c.numero_contrato,
                c.numero_contrato_definitivo or '-',
                c.cliente_nome,
                c.vendedor.username if c.vendedor else '-',
                c.data_instalacao.strftime('%d/%m/%Y'),
                c.plano_atual,
                c.get_status_contrato_display(),
                f"{faturas_pagas}/{total_faturas}",
                'Sim' if elegivel else 'Não',
                f"R$ {bonus:.2f}",
            ])

        # Ajusta largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width


        # Retorna arquivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=bonus_m10_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class ExportarAgendamentosDiaView(APIView):
    """Exporta agendamentos do dia para envio à operadora"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.utils import timezone
        from datetime import date
        
        # Data para filtrar (padrão: hoje)
        data_str = request.query_params.get('data')
        if data_str:
            try:
                data_filtro = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                data_filtro = date.today()
        else:
            data_filtro = date.today()
        
        # Busca vendas agendadas para a data
        vendas = Venda.objects.filter(
            data_agendamento=data_filtro,
            status_esteira__nome__icontains='AGENDADO'
        ).select_related('cliente', 'vendedor', 'plano')
        
        # Cria workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agendamentos"

        # Cabeçalhos conforme modelo
        headers = [
            'Parceiro',
            'Pedido',
            'UF',
            'Cidade',
            'Data/Período de agendamento',
            'Contato parceiro',
            'Contato cliente 1',
            'Contato cliente 2',
            'Contato cliente 3'
        ]
        ws.append(headers)

        # Estilo cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados
        for v in vendas:
            # Formata data e período
            if v.data_agendamento:
                data_fmt = v.data_agendamento.strftime('%d/%m/%Y')
                periodo = v.periodo_agendamento or ''
                data_periodo = f"{data_fmt} - {periodo}" if periodo else data_fmt
            else:
                data_periodo = ''
            
            # Formata telefones (remove formatação, mantém apenas números)
            tel1 = re.sub(r'\D', '', v.telefone1 or '') if v.telefone1 else ''
            tel2 = re.sub(r'\D', '', v.telefone2 or '') if v.telefone2 else ''
            
            ws.append([
                'Record',                          # Parceiro (fixo)
                v.ordem_servico or '',             # Pedido (O.S)
                (v.estado or '').upper()[:2],      # UF
                (v.cidade or '').upper(),          # Cidade
                data_periodo,                      # Data/Período de agendamento
                '3198224-3410',                    # Contato parceiro (fixo)
                tel1,                              # Contato cliente 1
                tel2,                              # Contato cliente 2
                ''                                 # Contato cliente 3 (vazio)
            ])

        # Ajusta largura das colunas
        column_widths = [12, 15, 6, 25, 25, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Borda fina para todas as células com dados
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # Dados (não cabeçalho)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Retorna arquivo
        data_arquivo = data_filtro.strftime("%Y%m%d")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=agendamentos_operadora_{data_arquivo}.xlsx'
        wb.save(response)
        return response


def _normalizar_telefone_chave(telefone):
    """Normaliza telefone para chave de lookup (igual ao webhook: dígitos, sem 55 se len>12)."""
    if not telefone:
        return ""
    tel = re.sub(r'\D', '', str(telefone))
    if tel.startswith('55') and len(tel) > 12:
        tel = tel[2:]
    return tel


# --- Boas práticas WhatsApp (evitar bloqueio) ---
# Ref.: intervalos 30-60s entre mensagens; pausa 1-2 min entre lotes; máx ~20 msg/min; lotes pequenos (5-8).
# Mensagens personalizadas (nome) e ordem aleatória reduzem detecção de padrão.
WHATSAPP_ENVIO_MIN_SEG = 30   # mínimo segundos entre cada mensagem
WHATSAPP_ENVIO_MAX_SEG = 65   # máximo segundos entre cada mensagem
WHATSAPP_LOTE_TAMANHO = 5     # mensagens por lote
WHATSAPP_PAUSA_LOTE_MIN_SEG = 60   # pausa mínima entre lotes (segundos)
WHATSAPP_PAUSA_LOTE_MAX_SEG = 120  # pausa máxima entre lotes (segundos)


class EnviarLembreteInstalacaoView(APIView):
    """Envia lembrete de instalação para clientes agendados na data e turno informados.
    Usa lotes com intervalo aleatório entre mensagens (boas práticas WhatsApp)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        import time
        import random
        from django.utils import timezone
        from datetime import datetime
        from crm_app.whatsapp_service import WhatsAppService

        data_str = request.data.get('data')
        turno = (request.data.get('turno') or '').strip().upper()
        if not data_str or turno not in ('MANHA', 'TARDE'):
            return Response(
                {'detail': 'Envie "data" (YYYY-MM-DD) e "turno" (MANHA ou TARDE).'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            data_filtro = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        offset = max(0, int(request.data.get('offset', 0)))
        limite = max(1, min(10, int(request.data.get('limite', WHATSAPP_LOTE_TAMANHO))))
        min_intervalo = max(15, min(90, int(request.data.get('min_intervalo', WHATSAPP_ENVIO_MIN_SEG))))
        max_intervalo = max(min_intervalo + 5, min(120, int(request.data.get('max_intervalo', WHATSAPP_ENVIO_MAX_SEG))))

        vendas_qs = Venda.objects.filter(
            ativo=True,
            data_agendamento=data_filtro,
            periodo_agendamento=turno,
            status_esteira__nome__icontains='AGENDADO'
        ).exclude(telefone1__isnull=True).exclude(telefone1='').select_related('cliente').order_by('id')
        total_na_data = vendas_qs.count()
        vendas = list(vendas_qs[offset:offset + limite])
        random.shuffle(vendas)

        primeiro_nome = (request.user.first_name or request.user.username or 'Especialista').strip().split()[0] or 'Especialista'
        agora = timezone.now()
        saudacao = 'boa tarde' if agora.hour >= 12 else 'bom dia'
        dd_mm = data_filtro.strftime('%d/%m')
        if turno == 'MANHA':
            periodo_texto = '8h às 12h'
        else:
            periodo_texto = '13h às 18h'

        enviados = 0
        erros = []
        svc = WhatsAppService()

        for i, venda in enumerate(vendas):
            if i > 0:
                delay = random.randint(min_intervalo, max_intervalo)
                time.sleep(delay)
            nome_cliente = (venda.cliente.nome_razao_social if venda.cliente else '').strip() or 'Cliente'
            mensagem = (
                f"Olá, {saudacao} Sr(a). {nome_cliente}\n\n"
                f"Me chamo {primeiro_nome}, sou especialista de qualidade do Record PAP, parceiro Oficial da Nio Fibra.\n\n"
                f"A sua instalação da Nio Fibra está agendada para hoje ({dd_mm}), no período das {periodo_texto}.\n\n"
                "Se você não puder estar presente, é necessário que uma pessoa maior de 18 anos esteja no local.\n\n"
                "Informações sobre sua instalação:\n"
                "A instalação é gratuita.\n"
                "Não realizamos instalações em dias de chuva.\n\n"
                "Para confirmar, digite SIM\n"
                "Para reagendar, envie o dia e o período (manhã/tarde)\n"
                "Para falar com suporte, envie SUPORTE"
            )
            try:
                ok, _ = svc.enviar_mensagem_texto(venda.telefone1, mensagem)
                if ok:
                    enviados += 1
                    tel_chave = _normalizar_telefone_chave(venda.telefone1)
                    if tel_chave:
                        LembreteInstalacaoEnviado.objects.create(
                            telefone=tel_chave,
                            venda=venda,
                            data_agendamento=venda.data_agendamento,
                            periodo_agendamento=venda.periodo_agendamento or turno,
                        )
                else:
                    erros.append(f"Venda #{venda.id} ({venda.telefone1})")
            except Exception as e:
                erros.append(f"Venda #{venda.id}: {str(e)}")

        restantes = max(0, total_na_data - offset - len(vendas))
        proximo_offset = offset + len(vendas) if restantes > 0 else None
        pause_antes_proximo = random.randint(WHATSAPP_PAUSA_LOTE_MIN_SEG, WHATSAPP_PAUSA_LOTE_MAX_SEG) if restantes > 0 else None

        return Response({
            'enviados': enviados,
            'total_na_data': total_na_data,
            'restantes': restantes,
            'proximo_offset': proximo_offset,
            'pause_antes_proximo_seg': pause_antes_proximo,
            'erros': erros[:20],
        }, status=status.HTTP_200_OK)


class EnviarBoasVindasView(APIView):
    """Envia mensagem de boas-vindas para clientes com venda Instalada na data de instalação informada.
    Envio em lotes com intervalo ALEATÓRIO entre mensagens (padrão diferente a cada vez) para evitar bloqueio.
    Parâmetros opcionais: limite (ex.: 5 por vez), offset (para continuar)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        import time
        import random
        from django.utils import timezone
        from datetime import datetime
        from crm_app.whatsapp_service import WhatsAppService
        from crm_app.models import BoasVindasEnviado

        data_str = request.data.get('data')
        if not data_str:
            return Response(
                {'detail': 'Envie "data" (YYYY-MM-DD) - data em que o pedido foi instalado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            data_instalacao = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        offset = max(0, int(request.data.get('offset', 0)))
        limite = max(1, min(10, int(request.data.get('limite', WHATSAPP_LOTE_TAMANHO))))
        min_intervalo = max(15, min(90, int(request.data.get('min_intervalo', WHATSAPP_ENVIO_MIN_SEG))))
        max_intervalo = max(min_intervalo + 5, min(120, int(request.data.get('max_intervalo', WHATSAPP_ENVIO_MAX_SEG))))

        vendas_qs = Venda.objects.filter(
            ativo=True,
            data_instalacao=data_instalacao,
            status_esteira__nome__icontains='INSTALADA',
            boas_vindas_enviado_em__isnull=True,
        ).exclude(telefone1__isnull=True).exclude(telefone1='').select_related('cliente').order_by('id')
        total_na_data = vendas_qs.count()
        vendas = list(vendas_qs[offset:offset + limite])
        random.shuffle(vendas)  # ordem de envio aleatória dentro do lote

        primeiro_nome = (request.user.first_name or request.user.username or 'Especialista').strip().split()[0] or 'Especialista'
        agora = timezone.now()
        saudacao = 'boa tarde' if agora.hour >= 12 else 'bom dia'
        despedida = 'boa tarde!' if agora.hour >= 12 else 'bom dia!'

        msg_base = (
            f"Olá {saudacao}, {{nome_cliente}} tudo bem?\n\n"
            f"Me chamo {primeiro_nome}, sou especialista de qualidade do Record PAP, parceiro Oficial da Nio Fibra.\n\n"
            "Estou entrando em contato para informar que estamos à sua disposição, caso você precise tirar dúvidas sobre seu plano e faturas.\n\n"
            "Sua primeira fatura irá vencer 25 dias após a instalação.\n\n"
            "Você também pode acompanhar sua conta através do app Nio.\n"
            "Instale o aplicativo no seu aparelho celular.\n\n"
            "Disponível para Android e iOS:\n"
            "Google Play Store (Android)\n"
            "https://play.google.com/store/apps/details?id=br.com.niointernet.app\n\n"
            "Apple Store (iOS):\n"
            "https://apps.apple.com/br/app/nio-internet/id6746278488\n\n"
            "Você ainda pode realizar contato pelos canais de comunicação oficiais da Nio:\n"
            "SAC:0800 001 1000\n"
            "WhatsApp: 21-3605-1000\n\n"
            f"Obrigado e tenha um {despedida}"
        )

        enviados = 0
        erros = []
        svc = WhatsAppService()

        for i, venda in enumerate(vendas):
            if i > 0:
                delay = random.randint(min_intervalo, max_intervalo)
                time.sleep(delay)
            nome_cliente = (venda.cliente.nome_razao_social if venda.cliente else '').strip() or 'Cliente'
            mensagem = msg_base.format(nome_cliente=nome_cliente)
            try:
                ok, _ = svc.enviar_mensagem_texto(venda.telefone1, mensagem)
                if ok:
                    enviados += 1
                    venda.boas_vindas_enviado_em = timezone.now()
                    venda.save(update_fields=['boas_vindas_enviado_em'])
                    tel_chave = _normalizar_telefone_chave(venda.telefone1)
                    if tel_chave:
                        BoasVindasEnviado.objects.create(telefone=tel_chave, venda=venda)
                else:
                    erros.append(f"Venda #{venda.id} ({venda.telefone1})")
            except Exception as e:
                erros.append(f"Venda #{venda.id}: {str(e)}")

        restantes = max(0, total_na_data - offset - len(vendas))
        proximo_offset = offset + len(vendas) if restantes > 0 else None
        # Pausa sugerida entre lotes (boas práticas: 1-2 min para não parecer burst)
        pause_antes_proximo = random.randint(WHATSAPP_PAUSA_LOTE_MIN_SEG, WHATSAPP_PAUSA_LOTE_MAX_SEG) if restantes > 0 else None

        return Response({
            'enviados': enviados,
            'total_na_data': total_na_data,
            'restantes': restantes,
            'proximo_offset': proximo_offset,
            'pause_antes_proximo_seg': pause_antes_proximo,
            'erros': erros[:20],
        }, status=status.HTTP_200_OK)


# --- Boas-Vindas Gestão (ferramenta dedicada) ---
# Constantes anti-spam: 1 msg a cada 20-30 min, tudo até 16h
BOAS_VINDAS_INTERVALO_MIN_SEG = 1200   # 20 min
BOAS_VINDAS_INTERVALO_MAX_SEG = 1800   # 30 min
BOAS_VINDAS_HORA_LIMITE = 16            # até 16h


class BoasVindasInstalacoesView(APIView):
    """GET: Lista instalações do dia anterior (ou data informada) para envio de boas-vindas."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        from datetime import datetime, timedelta
        data_str = request.query_params.get('data')
        if data_str:
            try:
                data_filtro = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            data_filtro = (timezone.now() - timedelta(days=1)).date()
        vendas = Venda.objects.filter(
            ativo=True,
            data_instalacao=data_filtro,
            status_esteira__nome__icontains='INSTALADA',
        ).exclude(telefone1__isnull=True).exclude(telefone1='').select_related(
            'cliente', 'vendedor', 'plano', 'status_esteira'
        ).order_by('id')
        pendentes = vendas.filter(boas_vindas_enviado_em__isnull=True)
        enviados = vendas.exclude(boas_vindas_enviado_em__isnull=True)
        items = []
        for v in vendas:
            items.append({
                'id': v.id,
                'cliente': v.cliente.nome_razao_social if v.cliente else '-',
                'telefone': v.telefone1,
                'vendedor': v.vendedor.username if v.vendedor else '-',
                'plano': v.plano.nome if v.plano else '-',
                'data_instalacao': str(v.data_instalacao) if v.data_instalacao else None,
                'boas_vindas_enviado_em': v.boas_vindas_enviado_em.isoformat() if v.boas_vindas_enviado_em else None,
            })
        return Response({
            'data': str(data_filtro),
            'total': vendas.count(),
            'pendentes': pendentes.count(),
            'enviados': enviados.count(),
            'items': items,
        })


class BoasVindasRetornosView(APIView):
    """GET: Lista clientes que receberam boas-vindas e enviaram mensagem (para o BO tratar)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        from datetime import timedelta
        limite_dias = int(request.query_params.get('dias', 30))
        data_min = timezone.now() - timedelta(days=limite_dias)
        bvs = BoasVindasEnviado.objects.filter(
            respondido_em__isnull=False,
            data_envio__gte=data_min,
        ).select_related('venda__cliente', 'venda__vendedor', 'status_boas_vindas').order_by('-respondido_em')
        items = []
        for bv in bvs:
            v = bv.venda
            ultima_msg = (v.cliente_resposta_boas_vindas or '')[:200]
            items.append({
                'id': bv.id,
                'venda_id': v.id,
                'cliente': v.cliente.nome_razao_social if v.cliente else '-',
                'telefone': bv.telefone,
                'vendedor': v.vendedor.username if v.vendedor else '-',
                'data_envio': bv.data_envio.isoformat(),
                'respondido_em': bv.respondido_em.isoformat() if bv.respondido_em else None,
                'ultima_mensagem': ultima_msg[:150] + ('...' if len(ultima_msg) > 150 else ''),
                'status_codigo': bv.status_boas_vindas.codigo if bv.status_boas_vindas else 'PENDENTE',
                'status_nome': bv.status_boas_vindas.nome if bv.status_boas_vindas else 'Pendente',
                'sugestao_ia': bv.sugestao_status_ia,
            })
        return Response({'items': items})


class BoasVindasDetalheView(APIView):
    """GET: Detalhe de um retorno (mensagens, venda, status). POST: Atribuir status."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        bv = BoasVindasEnviado.objects.filter(pk=pk).select_related(
            'venda__cliente', 'venda__vendedor', 'venda__plano', 'status_boas_vindas', 'status_definido_por'
        ).prefetch_related('mensagens').first()
        if not bv:
            return Response({'detail': 'Não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        v = bv.venda
        mensagens = [{'texto': m.texto, 'data_hora': m.data_hora.isoformat(), 'direcao': m.direcao} for m in bv.mensagens.order_by('data_hora')]
        return Response({
            'id': bv.id,
            'venda_id': v.id,
            'cliente': v.cliente.nome_razao_social if v.cliente else '-',
            'telefone': bv.telefone,
            'vendedor': v.vendedor.username if v.vendedor else '-',
            'plano': v.plano.nome if v.plano else '-',
            'data_envio': bv.data_envio.isoformat(),
            'respondido_em': bv.respondido_em.isoformat() if bv.respondido_em else None,
            'mensagens': mensagens,
            'cliente_resposta_boas_vindas': v.cliente_resposta_boas_vindas,
            'status': {'codigo': bv.status_boas_vindas.codigo, 'nome': bv.status_boas_vindas.nome} if bv.status_boas_vindas else None,
            'sugestao_ia': bv.sugestao_status_ia,
            'status_definido_por': bv.status_definido_por.username if bv.status_definido_por else None,
            'status_definido_em': bv.status_definido_em.isoformat() if bv.status_definido_em else None,
        })

    def post(self, request, pk):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        status_codigo = (request.data.get('status_codigo') or '').strip().upper()
        if not status_codigo:
            return Response({'detail': 'Informe status_codigo.'}, status=status.HTTP_400_BAD_REQUEST)
        st = StatusBoasVindas.objects.filter(codigo=status_codigo).first()
        if not st:
            return Response({'detail': f'Status "{status_codigo}" não encontrado.'}, status=status.HTTP_400_BAD_REQUEST)
        bv = BoasVindasEnviado.objects.filter(pk=pk).first()
        if not bv:
            return Response({'detail': 'Não encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        bv.status_boas_vindas = st
        bv.status_definido_por = request.user
        bv.status_definido_em = timezone.now()
        bv.save(update_fields=['status_boas_vindas', 'status_definido_por', 'status_definido_em'])
        return Response({'status': 'ok', 'status_codigo': st.codigo, 'status_nome': st.nome})


class BoasVindasStatusListView(APIView):
    """GET: Lista status disponíveis para atribuição."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        status_list = list(StatusBoasVindas.objects.order_by('ordem').values('id', 'codigo', 'nome', 'cor'))
        return Response({'items': status_list})


class BoasVindasSugestaoIAView(APIView):
    """POST: IA sugere status a partir do texto das mensagens do cliente."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        texto = (request.data.get('texto') or '').strip()
        if not texto:
            return Response({'detail': 'Informe o texto das mensagens.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from crm_app.ai_chat_service import sugerir_status_boas_vindas
            sugestao = sugerir_status_boas_vindas(texto)
            return Response({'sugestao': sugestao})
        except Exception as e:
            logger.warning(f"[BoasVindas] IA sugestão falhou: {e}")
            return Response({'sugestao': 'OUTROS'})


class BoasVindasEnviarGestaoView(APIView):
    """Envia boas-vindas com modelo anti-spam: 1 msg a cada 20-30 min, até 16h.
    Parâmetros: data (YYYY-MM-DD), offset, limite (default 1 para spread máximo)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        import time
        import random
        from datetime import datetime
        from crm_app.whatsapp_service import WhatsAppService

        data_str = request.data.get('data')
        if not data_str:
            return Response(
                {'detail': 'Envie "data" (YYYY-MM-DD) - data em que o pedido foi instalado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            data_instalacao = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        offset = max(0, int(request.data.get('offset', 0)))
        limite = max(1, min(3, int(request.data.get('limite', 1))))  # 1 por vez = spread máximo
        min_intervalo = BOAS_VINDAS_INTERVALO_MIN_SEG
        max_intervalo = BOAS_VINDAS_INTERVALO_MAX_SEG

        vendas_qs = Venda.objects.filter(
            ativo=True,
            data_instalacao=data_instalacao,
            status_esteira__nome__icontains='INSTALADA',
            boas_vindas_enviado_em__isnull=True,
        ).exclude(telefone1__isnull=True).exclude(telefone1='').select_related('cliente').order_by('id')
        total_na_data = vendas_qs.count()
        vendas = list(vendas_qs[offset:offset + limite])
        random.shuffle(vendas)

        primeiro_nome = (request.user.first_name or request.user.username or 'Especialista').strip().split()[0] or 'Especialista'
        agora = timezone.now()
        saudacao = 'boa tarde' if agora.hour >= 12 else 'bom dia'
        despedida = 'boa tarde!' if agora.hour >= 12 else 'bom dia!'

        msg_base = (
            f"Olá {saudacao}, {{nome_cliente}} tudo bem?\n\n"
            f"Me chamo {primeiro_nome}, sou especialista de qualidade do Record PAP, parceiro Oficial da Nio Fibra.\n\n"
            "Estou entrando em contato para informar que estamos à sua disposição, caso você precise tirar dúvidas sobre seu plano e faturas.\n\n"
            "Sua primeira fatura irá vencer 25 dias após a instalação.\n\n"
            "Você também pode acompanhar sua conta através do app Nio.\n"
            "Instale o aplicativo no seu aparelho celular.\n\n"
            "Disponível para Android e iOS:\n"
            "Google Play Store (Android)\n"
            "https://play.google.com/store/apps/details?id=br.com.niointernet.app\n\n"
            "Apple Store (iOS):\n"
            "https://apps.apple.com/br/app/nio-internet/id6746278488\n\n"
            "Você ainda pode realizar contato pelos canais de comunicação oficiais da Nio:\n"
            "SAC:0800 001 1000\n"
            "WhatsApp: 21-3605-1000\n\n"
            f"Obrigado e tenha um {despedida}"
        )

        enviados = 0
        erros = []
        svc = WhatsAppService()

        for i, venda in enumerate(vendas):
            if i > 0:
                delay = random.randint(min_intervalo, max_intervalo)
                time.sleep(delay)
            nome_cliente = (venda.cliente.nome_razao_social if venda.cliente else '').strip() or 'Cliente'
            mensagem = msg_base.format(nome_cliente=nome_cliente)
            try:
                ok, _ = svc.enviar_mensagem_texto(venda.telefone1, mensagem)
                if ok:
                    enviados += 1
                    venda.boas_vindas_enviado_em = timezone.now()
                    venda.save(update_fields=['boas_vindas_enviado_em'])
                    tel_chave = _normalizar_telefone_chave(venda.telefone1)
                    if tel_chave:
                        BoasVindasEnviado.objects.create(telefone=tel_chave, venda=venda)
                else:
                    erros.append(f"Venda #{venda.id} ({venda.telefone1})")
            except Exception as e:
                erros.append(f"Venda #{venda.id}: {str(e)}")

        restantes = max(0, total_na_data - offset - len(vendas))
        proximo_offset = offset + len(vendas) if restantes > 0 else None
        pause_antes_proximo = random.randint(1200, 1800) if restantes > 0 else None  # 20-30 min

        return Response({
            'enviados': enviados,
            'total_na_data': total_na_data,
            'restantes': restantes,
            'proximo_offset': proximo_offset,
            'pause_antes_proximo_seg': pause_antes_proximo,
            'erros': erros[:20],
        }, status=status.HTTP_200_OK)


class BoasVindasAgendarView(APIView):
    """POST: Coloca na fila os envios de boas-vindas. O scheduler processa a cada 5 min.
    Distribui os envios entre 8h e 16h do dia atual (intervalo ~20-30 min entre cada)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        from datetime import datetime, timedelta
        import random

        data_str = request.data.get('data')
        if not data_str:
            return Response(
                {'detail': 'Envie "data" (YYYY-MM-DD) - data das instalações.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            data_instalacao = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'detail': 'Data inválida. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        vendas = list(Venda.objects.filter(
            ativo=True,
            data_instalacao=data_instalacao,
            status_esteira__nome__icontains='INSTALADA',
            boas_vindas_enviado_em__isnull=True,
        ).exclude(telefone1__isnull=True).exclude(telefone1='').values_list('id', flat=True))

        if not vendas:
            return Response({
                'detail': 'Nenhuma instalação pendente de boas-vindas para esta data.',
                'agendados': 0,
            }, status=status.HTTP_200_OK)

        # Evita duplicar na fila
        ja_na_fila = set(
            FilaEnvioBoasVindas.objects.filter(
                data_instalacao=data_instalacao,
                enviado_em__isnull=True,
            ).values_list('venda_id', flat=True)
        )
        vendas = [v for v in vendas if v not in ja_na_fila]
        if not vendas:
            return Response({
                'detail': 'Todas as instalações já estão na fila.',
                'agendados': 0,
            }, status=status.HTTP_200_OK)

        random.shuffle(vendas)
        hoje = timezone.now().date()
        # Janela 8h-16h (até 16h)
        hora_inicio = 8
        hora_fim = 16
        total_min = (hora_fim - hora_inicio) * 60  # 480 min
        n = len(vendas)
        intervalo_medio = max(20, min(30, total_min // max(1, n)))  # 20-30 min entre cada

        criados = 0
        for i, venda_id in enumerate(vendas):
            min_offset = i * intervalo_medio
            variacao = random.randint(-3, 5)  # ± variação
            min_offset = max(0, min_offset + variacao)
            hora = hora_inicio + (min_offset // 60)
            minuto = min_offset % 60
            if hora >= hora_fim:
                hora = hora_fim - 1
                minuto = 59
            agendado_para = timezone.make_aware(
                datetime(hoje.year, hoje.month, hoje.day, hora, minuto, 0)
            )
            if agendado_para <= timezone.now():
                agendado_para = timezone.now() + timedelta(minutes=random.randint(2, 5))
            FilaEnvioBoasVindas.objects.create(
                venda_id=venda_id,
                data_instalacao=data_instalacao,
                agendado_para=agendado_para,
                criado_por=request.user,
            )
            criados += 1

        return Response({
            'detail': f'{criados} envio(s) agendados na fila. O sistema enviará automaticamente até 16h (a cada 5 min verifica).',
            'agendados': criados,
        }, status=status.HTTP_200_OK)


class BoasVindasFilaStatusView(APIView):
    """GET: Status da fila de envios (pendentes, enviados, por data)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice', 'Auditoria', 'Qualidade']):
            return Response({'detail': 'Sem permissão.'}, status=status.HTTP_403_FORBIDDEN)
        data_str = request.query_params.get('data')
        from datetime import timedelta
        if data_str:
            try:
                from datetime import datetime
                data_filtro = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'Data inválida.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            data_filtro = (timezone.now() - timedelta(days=1)).date()

        pendentes = FilaEnvioBoasVindas.objects.filter(
            data_instalacao=data_filtro,
            enviado_em__isnull=True,
        ).count()
        enviados = FilaEnvioBoasVindas.objects.filter(
            data_instalacao=data_filtro,
            enviado_em__isnull=False,
        ).count()
        com_erro = FilaEnvioBoasVindas.objects.filter(
            data_instalacao=data_filtro,
            enviado_em__isnull=True,
            erro__isnull=False,
        ).exclude(erro='').count()

        return Response({
            'data': str(data_filtro),
            'pendentes': pendentes,
            'enviados': enviados,
            'com_erro': com_erro,
            'total_fila': pendentes + enviados,
        })


# --- Antecipar Instalação (solicitação ao GC Nio) ---
def _antecipar_instalacao_queryset_vendas(request):
    """Retorna queryset de Venda permitidas para o usuário: só AGENDADO e com data_agendamento."""
    base = Venda.objects.filter(
        ativo=True,
        status_esteira__nome__iexact='AGENDADO',
        data_agendamento__isnull=False
    ).select_related('cliente', 'status_esteira')
    if is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
        return base
    if is_member(request.user, ['Supervisor']):
        # Supervisor: suas vendas + dos liderados
        liderados_ids = list(request.user.liderados.values_list('id', flat=True))
        return base.filter(Q(vendedor_id=request.user.id) | Q(vendedor_id__in=liderados_ids))
    # Vendedor: só as suas
    return base.filter(vendedor_id=request.user.id)


def _antecipar_instalacao_queryset_vendas_reparo(request):
    """Retorna queryset de Venda INSTALADAS com data_instalacao até 14 dias atrás (para solicitar reparo)."""
    hoje = date.today()
    data_limite = hoje - timedelta(days=14)
    base = Venda.objects.filter(
        ativo=True,
        status_esteira__nome__iexact='INSTALADA',
        data_instalacao__isnull=False,
        data_instalacao__gte=data_limite,
        data_instalacao__lte=hoje,
    ).select_related('cliente', 'status_esteira')
    if is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
        return base
    if is_member(request.user, ['Supervisor']):
        liderados_ids = list(request.user.liderados.values_list('id', flat=True))
        return base.filter(Q(vendedor_id=request.user.id) | Q(vendedor_id__in=liderados_ids))
    return base.filter(vendedor_id=request.user.id)


def _antecipar_instalacao_queryset_vendas_instalacao_fisica(request):
    """Pendência na esteira para sinalização ao GC (com ou sem data física preenchida)."""
    base = Venda.objects.filter(
        ativo=True,
    ).filter(
        Q(status_esteira__nome__icontains='penden') | Q(status_esteira__nome__icontains='pendência')
    ).exclude(
        status_esteira__nome__iexact='INSTALADA'
    ).select_related('cliente', 'status_esteira')
    if is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
        return base
    if is_member(request.user, ['Supervisor']):
        liderados_ids = list(request.user.liderados.values_list('id', flat=True))
        return base.filter(Q(vendedor_id=request.user.id) | Q(vendedor_id__in=liderados_ids))
    return base.filter(vendedor_id=request.user.id)


def _antecipar_instalacao_endereco_completo(v):
    """Formato: Rua X, 123, Sala 1, Centro, São Paulo - SP - CEP 12345-678"""
    partes = []
    if v.logradouro:
        partes.append(v.logradouro)
    if v.numero_residencia:
        partes.append(str(v.numero_residencia))
    if v.complemento:
        partes.append(v.complemento)
    if v.bairro:
        partes.append(v.bairro)
    if v.cidade or v.estado:
        partes.append(f"{v.cidade or ''} - {v.estado or ''}".strip(' -'))
    end = ", ".join(p for p in partes if p)
    if v.cep:
        end = f"{end} - CEP {v.cep}" if end else f"CEP {v.cep}"
    return end or "Endereço não informado"


def _antecipar_instalacao_contato_cliente(v):
    """Retorna o melhor contato disponível do cliente no CRM."""
    telefones = [getattr(v, 'telefone1', None), getattr(v, 'telefone2', None)]
    for telefone in telefones:
        if (telefone or '').strip():
            return telefone.strip()
    return "Não informado"


class BuscarAnteciparInstalacaoView(APIView):
    """GET ?q=CPF ou número da O.S — retorna pedidos agendados, reparo até 14 dias e instalação física em pendência."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        q = (request.query_params.get('q') or '').strip()
        if not q:
            return Response({'detail': 'Informe CPF ou número do pedido (O.S).'}, status=status.HTTP_400_BAD_REQUEST)
        q_limpo = re.sub(r'\D', '', q)
        filters = Q(ordem_servico__icontains=q) | Q(cliente__cpf_cnpj__icontains=q)
        if q_limpo:
            filters |= Q(cliente__cpf_cnpj__icontains=q_limpo) | Q(ordem_servico__icontains=q_limpo)

        results = []
        ids_incluidos = set()

        # Pedidos agendados (antecipação)
        queryset_ag = _antecipar_instalacao_queryset_vendas(request)
        vendas_ag = list(queryset_ag.filter(filters).order_by('-data_agendamento')[:50])
        for v in vendas_ag:
            ids_incluidos.add(v.id)
            data_ag_fmt = v.data_agendamento.strftime('%d/%m/%Y') if v.data_agendamento else ''
            turno = v.get_periodo_agendamento_display() if v.periodo_agendamento else ''
            results.append({
                'id': v.id,
                'ordem_servico': v.ordem_servico or '',
                'endereco_completo': _antecipar_instalacao_endereco_completo(v),
                'telefone_contato': _antecipar_instalacao_contato_cliente(v),
                'data_agendamento': data_ag_fmt,
                'periodo_agendamento': turno,
                'cliente_nome': v.cliente.nome_razao_social if v.cliente else '',
                'tipo_disponivel': 'antecipacao',
                'data_instalacao': None,
                'data_instalacao_fisica': None,
            })

        # Pedidos instalados até 14 dias (reparo)
        queryset_rep = _antecipar_instalacao_queryset_vendas_reparo(request)
        vendas_rep = list(queryset_rep.filter(filters).order_by('-data_instalacao')[:50])
        for v in vendas_rep:
            if v.id in ids_incluidos:
                continue
            ids_incluidos.add(v.id)
            data_inst_fmt = v.data_instalacao.strftime('%d/%m/%Y') if v.data_instalacao else ''
            results.append({
                'id': v.id,
                'ordem_servico': v.ordem_servico or '',
                'endereco_completo': _antecipar_instalacao_endereco_completo(v),
                'telefone_contato': _antecipar_instalacao_contato_cliente(v),
                'data_agendamento': '',
                'periodo_agendamento': '',
                'cliente_nome': v.cliente.nome_razao_social if v.cliente else '',
                'tipo_disponivel': 'reparo',
                'data_instalacao': data_inst_fmt,
                'data_instalacao_fisica': None,
            })

        queryset_if = _antecipar_instalacao_queryset_vendas_instalacao_fisica(request)
        vendas_if = list(queryset_if.filter(filters).order_by('-data_instalacao_fisica')[:50])
        for v in vendas_if:
            if v.id in ids_incluidos:
                continue
            ids_incluidos.add(v.id)
            data_fis = v.data_instalacao_fisica.strftime('%d/%m/%Y') if v.data_instalacao_fisica else ''
            results.append({
                'id': v.id,
                'ordem_servico': v.ordem_servico or '',
                'endereco_completo': _antecipar_instalacao_endereco_completo(v),
                'telefone_contato': _antecipar_instalacao_contato_cliente(v),
                'data_agendamento': '',
                'periodo_agendamento': '',
                'cliente_nome': v.cliente.nome_razao_social if v.cliente else '',
                'tipo_disponivel': 'instalacao_fisica',
                'data_instalacao': v.data_instalacao.strftime('%d/%m/%Y') if v.data_instalacao else '',
                'data_instalacao_fisica': data_fis,
                'aviso_sem_data_instalacao_fisica': not bool(v.data_instalacao_fisica),
            })

        def _sort_key_antecipar(r):
            tipo = r['tipo_disponivel']
            grupo = 0 if tipo == 'antecipacao' else (1 if tipo == 'reparo' else 2)
            ref = r['data_agendamento'] or r.get('data_instalacao_fisica') or r['data_instalacao'] or ''
            return (grupo, ref)

        results.sort(key=_sort_key_antecipar, reverse=True)
        return Response({'results': results})


def _antecipar_instalacao_get_config():
    """Retorna a configuração única (singleton). Cria com valores padrão se não existir."""
    config = AnteciparInstalacaoConfig.objects.first()
    if not config:
        config = AnteciparInstalacaoConfig.objects.create(telefone_gc='', nome_gc='')
    return config


def _antecipar_instalacao_sync_grupos_zapi():
    """
    Sincroniza grupos da Z-API com GrupoDisparo para garantir que o select da configuração
    mostre todos os grupos do número conectado.
    """
    try:
        svc = WhatsAppService()
        grupos_zapi = svc.listar_grupos() or []
    except Exception:
        return
    if not isinstance(grupos_zapi, list):
        return
    for g in grupos_zapi:
        if not isinstance(g, dict):
            continue
        chat_id = str(g.get('id') or '').strip()
        nome = str(g.get('name') or 'Sem Nome').strip()[:100] or 'Sem Nome'
        if not chat_id:
            continue
        obj = GrupoDisparo.objects.filter(chat_id=chat_id).first()
        if obj:
            changed = False
            if obj.nome != nome:
                obj.nome = nome
                changed = True
            if not obj.ativo:
                obj.ativo = True
                changed = True
            if changed:
                obj.save(update_fields=['nome', 'ativo'])
            continue
        GrupoDisparo.objects.create(nome=nome, chat_id=chat_id, ativo=True)


class ConfigAnteciparInstalacaoView(APIView):
    """GET: retorna config (telefone_gc, grupo) + lista de grupos para o select. PATCH: atualiza config (só Admin/Diretoria/BackOffice)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        config = _antecipar_instalacao_get_config()
        _antecipar_instalacao_sync_grupos_zapi()
        grupos = list(
            GrupoDisparo.objects.filter(ativo=True).order_by('nome').values('id', 'nome', 'chat_id')
        )
        return Response({
            'nome_gc': config.nome_gc or '',
            'telefone_gc': config.telefone_gc or '',
            'grupo_id': config.grupo_id,
            'grupo_nome': config.grupo.nome if config.grupo else None,
            'grupos': grupos,
            'pode_editar': is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']),
        })

    def patch(self, request):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({'detail': 'Sem permissão para alterar a configuração.'}, status=status.HTTP_403_FORBIDDEN)
        config = _antecipar_instalacao_get_config()
        if 'telefone_gc' in request.data:
            val = request.data.get('telefone_gc')
            config.telefone_gc = (val if val is not None else '').strip()
        if 'nome_gc' in request.data:
            val = request.data.get('nome_gc')
            config.nome_gc = (val if val is not None else '').strip()[:100]
        if 'grupo_id' in request.data:
            gid = request.data.get('grupo_id')
            if gid is None or gid == '':
                config.grupo_id = None
            else:
                try:
                    gid = int(gid)
                    g = GrupoDisparo.objects.filter(id=gid, ativo=True).first()
                    config.grupo = g
                except (TypeError, ValueError):
                    config.grupo_id = None
        config.atualizado_por = request.user
        config.save()
        return Response({
            'nome_gc': config.nome_gc or '',
            'telefone_gc': config.telefone_gc or '',
            'grupo_id': config.grupo_id,
            'grupo_nome': config.grupo.nome if config.grupo else None,
        })


class HistoricoAnteciparInstalacaoView(APIView):
    """GET: lista solicitações (filtrado por perfil: vendedor=só suas, supervisor=suas+liderados, gestão=todos)."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        base = AnteciparInstalacaoSolicitacao.objects.select_related('usuario', 'venda', 'resposta_gc_por').order_by('-data_solicitacao')
        if is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
            qs = base
        elif is_member(request.user, ['Supervisor']):
            liderados_ids = list(request.user.liderados.values_list('id', flat=True))
            qs = base.filter(Q(usuario_id=request.user.id) | Q(usuario_id__in=liderados_ids))
        else:
            qs = base.filter(usuario_id=request.user.id)
        try:
            page = int(request.query_params.get('page', 1))
        except (TypeError, ValueError):
            page = 1
        if page < 1:
            page = 1
        try:
            page_size = min(int(request.query_params.get('page_size', 20)), 50)
        except (TypeError, ValueError):
            page_size = 20
        start = (page - 1) * page_size
        lista = list(qs[start:start + page_size])
        results = []
        for s in lista:
            results.append({
                'id': s.id,
                'data_solicitacao': s.data_solicitacao.strftime('%d/%m/%Y %H:%M') if s.data_solicitacao else '',
                'usuario_nome': (s.usuario.get_full_name() or s.usuario.username) if s.usuario else '-',
                'ordem_servico': s.ordem_servico or '-',
                'cliente_nome': (s.venda.cliente.nome_razao_social if s.venda and s.venda.cliente else None) or '-',
                'descricao_solicitacao': (s.descricao_solicitacao or '')[:200],
                'tipo_solicitacao': getattr(s, 'tipo_solicitacao', 'antecipacao') or 'antecipacao',
                'observacao_reparo': (getattr(s, 'observacao_reparo', None) or '')[:200],
                'enviado_gc': s.enviado_gc,
                'enviado_grupo': s.enviado_grupo,
                'erros': s.erros or [],
                'sucesso': s.enviado_gc or s.enviado_grupo,
                'resposta_gc': getattr(s, 'resposta_gc', None) or '',
                'resposta_gc_em': s.resposta_gc_em.strftime('%d/%m/%Y %H:%M') if getattr(s, 'resposta_gc_em', None) else '',
                'resposta_gc_por_nome': (s.resposta_gc_por.get_full_name() or s.resposta_gc_por.username) if getattr(s, 'resposta_gc_por', None) else '',
                'resposta_gc_complemento_vendedor': (getattr(s, 'resposta_gc_complemento_vendedor', None) or '')[:500],
            })
        return Response({
            'results': results,
            'total': qs.count(),
            'pode_registrar_resposta_gc': is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']),
        })


def _historico_antecipar_queryset(request):
    """Queryset de solicitações visíveis para o usuário (mesmo filtro do histórico)."""
    base = AnteciparInstalacaoSolicitacao.objects.select_related('usuario', 'venda', 'venda__vendedor', 'resposta_gc_por')
    if is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
        return base
    if is_member(request.user, ['Supervisor']):
        liderados_ids = list(request.user.liderados.values_list('id', flat=True))
        return base.filter(Q(usuario_id=request.user.id) | Q(usuario_id__in=liderados_ids))
    return base.filter(usuario_id=request.user.id)


class RespostaGCAnteciparInstalacaoView(APIView):
    """PATCH: registra resposta do GC e envia mensagem padronizada ao vendedor (só Diretoria/Admin/BackOffice)."""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        if not is_member(request.user, ['Diretoria', 'Admin', 'BackOffice']):
            return Response({'detail': 'Sem permissão para registrar resposta do GC.'}, status=status.HTTP_403_FORBIDDEN)
        resposta_gc = (request.data.get('resposta_gc') or '').strip().lower()
        if resposta_gc not in ('solicitado', 'antecipada', 'nao_antecipada'):
            return Response({'detail': 'resposta_gc deve ser: solicitado, antecipada ou nao_antecipada.'}, status=status.HTTP_400_BAD_REQUEST)
        complemento = (request.data.get('complemento_mensagem_vendedor') or '').strip()[:2000]
        qs = _historico_antecipar_queryset(request)
        try:
            sol = qs.get(id=pk)
        except AnteciparInstalacaoSolicitacao.DoesNotExist:
            return Response({'detail': 'Solicitação não encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        if sol.resposta_gc:
            return Response({'detail': 'Esta solicitação já possui resposta do GC registrada.'}, status=status.HTTP_400_BAD_REQUEST)
        msg = mensagem_resposta_gc_para_vendedor(
            sol.ordem_servico or '', resposta_gc, getattr(sol, 'tipo_solicitacao', None), complemento
        )
        if not msg:
            return Response({'detail': 'Mensagem não definida para esta resposta.'}, status=status.HTTP_400_BAD_REQUEST)
        vendedor = sol.venda.vendedor if sol.venda else None
        telefone = getattr(vendedor, 'tel_whatsapp', None) if vendedor else None
        telefone = (telefone or '').strip()
        enviado_zap = False
        if telefone:
            try:
                svc = WhatsAppService()
                ok, _ = svc.enviar_mensagem_texto(telefone, msg)
                enviado_zap = ok
            except Exception as e:
                logger.exception("Erro ao enviar WhatsApp resposta GC ao vendedor: %s", e)
        from django.utils import timezone
        sol.resposta_gc = resposta_gc
        sol.resposta_gc_em = timezone.now()
        sol.resposta_gc_por = request.user
        sol.resposta_gc_complemento_vendedor = complemento
        sol.save(update_fields=['resposta_gc', 'resposta_gc_em', 'resposta_gc_por', 'resposta_gc_complemento_vendedor'])
        return Response({
            'success': True,
            'message': 'Resposta registrada.' + (' Mensagem enviada ao vendedor por WhatsApp.' if enviado_zap else ' (Vendedor sem telefone ou falha no envio.)' if telefone else ' (Vendedor sem telefone cadastrado.)'),
            'enviado_whatsapp': enviado_zap,
        })


def _mensagem_padrao_reparo(os_num, endereco, data_inst_fmt, observacao_reparo):
    """Mensagem padrão para acionamento de reparo em O.S recém instalada."""
    contato_cliente = (observacao_reparo.get('contato_cliente') or '').strip() or 'NÃO INFORMADO'
    nome_cliente = (observacao_reparo.get('nome_cliente') or '').strip() or 'NÃO INFORMADO'
    gc_nome = (observacao_reparo.get('gc_nome') or '').strip() or 'NÃO INFORMADO'
    data_horario_1 = (observacao_reparo.get('data_horario_1') or '').strip() or 'NÃO INFORMADO'
    data_horario_2 = (observacao_reparo.get('data_horario_2') or '').strip() or 'NÃO INFORMADO'
    texto_livre = (observacao_reparo.get('texto_livre') or '').strip() or 'NÃO INFORMADO'
    base = (
        "*MÁSCARA PADRÃO DE ACIONAMENTO PARA REPAROS DE OSS RECÉM INSTALADAS:*\n\n"
        f"- *OS:* {os_num}\n"
        f"- *NOME DO CLIENTE:* {nome_cliente}\n"
        f"- *ENDEREÇO COMPLETO:* {endereco}\n"
        f"- *CONTATO DO CLIENTE:* {contato_cliente}\n"
        "- *PDV:* 1068561 - RECORD PAP\n"
        f"- *GC:* {gc_nome}\n"
        f"- *DATA INSTALAÇÃO:* {data_inst_fmt}\n"
        "- *DATA E HORÁRIO AGENDADO COM O CLIENTE:*\n"
        f"  1) {data_horario_1}\n"
        f"  2) {data_horario_2}\n"
        f"- *SOLICITAÇÃO:* {texto_livre}"
    )
    return base.upper()


def _mensagem_padrao_instalacao_fisica(os_num, endereco, data_fisica_fmt, descricao):
    """Sinalização de instalação física no cliente com pendência ainda registrada na esteira."""
    base = (
        "*SINALIZAÇÃO - INSTALAÇÃO FÍSICA / PENDÊNCIA NO SISTEMA:*\n\n"
        f"- *OS:* {os_num}\n"
        f"- *ENDEREÇO COMPLETO:* {endereco}\n"
        "- *NOME DO PDV:* 1067100 - RECORD\n"
        f"- *DATA INSTALAÇÃO FÍSICA (NO CLIENTE):* {data_fisica_fmt}\n"
        "- *CONTEXTO:* Houve instalação física; o pedido segue com pendência na esteira.\n"
        f"- *DESCRIÇÃO DETALHADA:* {descricao}"
    )
    return base.upper()


def _antecipar_imagem_para_data_url_e_bytes(uploaded):
    """Retorna (data_url_para_zapi, bytes, nome_arquivo) ou (None, None, None, mensagem_erro)."""
    if not uploaded:
        return None, None, None, None
    raw = uploaded.read()
    if len(raw) > 6 * 1024 * 1024:
        return None, None, None, 'Imagem muito grande (máx. 6 MB).'
    ct = (uploaded.content_type or '').lower().split(';')[0].strip()
    if ct not in ('image/jpeg', 'image/jpg', 'image/png', 'image/webp'):
        return None, None, None, 'Use imagem JPG, PNG ou WEBP.'
    if 'png' in ct:
        mime = 'image/png'
    elif 'webp' in ct:
        mime = 'image/webp'
    else:
        mime = 'image/jpeg'
    b64 = base64.b64encode(raw).decode('ascii')
    data_url = f"data:{mime};base64,{b64}"
    nome = getattr(uploaded, 'name', 'anexo.jpg') or 'anexo.jpg'
    return data_url, raw, nome, None


class SolicitarAnteciparInstalacaoView(APIView):
    """POST JSON ou multipart: antecipação, reparo ou instalacao_fisica (com imagem opcional)."""
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        from django.core.files.base import ContentFile

        venda_id = request.data.get('venda_id')
        tipo = (request.data.get('tipo') or 'antecipacao').strip().lower()
        if tipo not in ('antecipacao', 'reparo', 'instalacao_fisica'):
            tipo = 'antecipacao'

        if not venda_id:
            return Response({'detail': 'venda_id é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)

        imagem_upload = request.FILES.get('imagem')
        if imagem_upload and tipo != 'instalacao_fisica':
            return Response(
                {'detail': 'Anexo de imagem só é permitido para o tipo Instalação física / pendência.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        img_data_url, img_bytes, img_nome, img_err = _antecipar_imagem_para_data_url_e_bytes(
            imagem_upload if tipo == 'instalacao_fisica' else None
        )
        if img_err:
            return Response({'detail': img_err}, status=status.HTTP_400_BAD_REQUEST)

        if tipo == 'reparo':
            queryset = _antecipar_instalacao_queryset_vendas_reparo(request)
            try:
                venda = queryset.get(id=venda_id)
            except Venda.DoesNotExist:
                return Response({'detail': 'Pedido não encontrado ou não está elegível para reparo (instalado há até 14 dias).'}, status=status.HTTP_404_NOT_FOUND)
            turno_map = {'MANHA': 'MANHÃ', 'TARDE': 'TARDE'}
            data_opcao_1 = (request.data.get('data_opcao_1') or '').strip()
            turno_opcao_1 = (request.data.get('turno_opcao_1') or '').strip().upper()
            data_opcao_2 = (request.data.get('data_opcao_2') or '').strip()
            turno_opcao_2 = (request.data.get('turno_opcao_2') or '').strip().upper()
            def _is_date_br(txt):
                try:
                    datetime.strptime(txt, '%d/%m/%Y')
                    return True
                except (TypeError, ValueError):
                    return False
            if not data_opcao_1 or turno_opcao_1 not in turno_map or not data_opcao_2 or turno_opcao_2 not in turno_map:
                return Response(
                    {'detail': 'Informe duas opções de data e turno para retorno do técnico.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not _is_date_br(data_opcao_1) or not _is_date_br(data_opcao_2):
                return Response(
                    {'detail': 'As datas devem estar no formato dd/mm/aaaa.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            texto_livre = (request.data.get('observacao_reparo') or '').strip()[:500]
            data_inst_fmt = venda.data_instalacao.strftime('%d/%m/%Y') if venda.data_instalacao else ''
            endereco = _antecipar_instalacao_endereco_completo(venda)
            os_num = venda.ordem_servico or ''
            contato_cliente = _antecipar_instalacao_contato_cliente(venda)
            nome_cliente = venda.cliente.nome_razao_social if venda.cliente else ''
            data_horario_1 = f"{data_opcao_1} - {turno_map[turno_opcao_1]}"
            data_horario_2 = f"{data_opcao_2} - {turno_map[turno_opcao_2]}"
            gc_nome = (_antecipar_instalacao_get_config().nome_gc or '').strip()
            descricao = (
                "Reparo solicitado para O.S recém instalada (até 14 dias). "
                f"Opções de retorno técnico: {data_horario_1} e {data_horario_2}."
            )
            if texto_livre:
                descricao += f" Solicitação: {texto_livre}"
            mensagem = _mensagem_padrao_reparo(
                os_num,
                endereco,
                data_inst_fmt,
                {
                    'contato_cliente': contato_cliente,
                    'nome_cliente': nome_cliente,
                    'gc_nome': gc_nome,
                    'data_horario_1': data_horario_1,
                    'data_horario_2': data_horario_2,
                    'texto_livre': texto_livre,
                },
            )
        elif tipo == 'instalacao_fisica':
            descricao = (request.data.get('descricao_solicitacao') or '').strip()
            if not descricao:
                return Response(
                    {'detail': 'Descreva a situação (instalação física e pendência no sistema).'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            queryset = _antecipar_instalacao_queryset_vendas_instalacao_fisica(request)
            try:
                venda = queryset.get(id=venda_id)
            except Venda.DoesNotExist:
                return Response(
                    {
                        'detail': 'Pedido não encontrado ou não está elegível (é necessário estar pendente na esteira).'
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )
            data_fis_fmt = venda.data_instalacao_fisica.strftime('%d/%m/%Y') if venda.data_instalacao_fisica else 'NÃO INFORMADA'
            endereco = _antecipar_instalacao_endereco_completo(venda)
            os_num = venda.ordem_servico or ''
            mensagem = _mensagem_padrao_instalacao_fisica(os_num, endereco, data_fis_fmt, descricao)
        else:
            descricao = (request.data.get('descricao_solicitacao') or '').strip()
            if not descricao:
                return Response({'detail': 'Descreva o motivo da solicitação de antecipação.'}, status=status.HTTP_400_BAD_REQUEST)
            queryset = _antecipar_instalacao_queryset_vendas(request)
            try:
                venda = queryset.get(id=venda_id)
            except Venda.DoesNotExist:
                return Response({'detail': 'Pedido não encontrado ou você não tem permissão para solicitar antecipação deste pedido.'}, status=status.HTTP_404_NOT_FOUND)
            data_ag_fmt = venda.data_agendamento.strftime('%d/%m/%Y') if venda.data_agendamento else ''
            turno = venda.get_periodo_agendamento_display() if venda.periodo_agendamento else ''
            endereco = _antecipar_instalacao_endereco_completo(venda)
            os_num = venda.ordem_servico or ''
            mensagem = (
                "*MÁSCARA PADRÃO DE ACIONAMENTO - GRUPO ELITE:*\n\n"
                f"- *OS:* {os_num}\n"
                f"- *ENDEREÇO COMPLETO:* {endereco}\n"
                "- *NOME DO PDV:* 1067100 - RECORD\n"
                f"- *DATA AGENDADA:* {data_ag_fmt} - {turno}\n"
                f"- *DESCRIÇÃO DETALHADA DA SOLICITAÇÃO:* {descricao}"
            )
            mensagem = mensagem.upper()

        config = _antecipar_instalacao_get_config()
        telefone_gc = (config.telefone_gc or '').strip()
        grupo = config.grupo
        enviados = []
        erros = []
        try:
            svc = WhatsAppService()
            if telefone_gc:
                ok1, resp1 = svc.enviar_mensagem_texto(telefone_gc, mensagem)
                if ok1:
                    enviados.append('número do GC')
                else:
                    erros.append(f'GC: {resp1}')
            else:
                erros.append('Telefone do GC não configurado.')
            if grupo and grupo.chat_id:
                ok2, resp2 = svc.enviar_mensagem_texto(grupo.chat_id, mensagem)
                if ok2:
                    enviados.append('grupo')
                else:
                    erros.append(f'Grupo: {resp2}')
            elif not grupo:
                erros.append('Grupo não configurado.')

            caption_img = f"Anexo O.S {os_num} — instalação física / pendência"
            if img_data_url:
                if telefone_gc:
                    r = svc.enviar_imagem_b64(telefone_gc, img_data_url, caption=caption_img)
                    if not r:
                        erros.append('GC: falha no envio da imagem.')
                if grupo and grupo.chat_id:
                    r2 = svc.enviar_imagem_b64(grupo.chat_id, img_data_url, caption=caption_img)
                    if not r2:
                        erros.append('Grupo: falha no envio da imagem.')
        except Exception as e:
            logger.exception("Erro ao enviar WhatsApp antecipar instalação/reparo: %s", e)
            obs_rep = (request.data.get('observacao_reparo') or '').strip()[:500] if tipo == 'reparo' else ''
            create_kw = dict(
                usuario=request.user,
                venda=venda,
                ordem_servico=os_num,
                tipo_solicitacao=tipo,
                descricao_solicitacao=descricao,
                observacao_reparo=obs_rep,
                enviado_gc=False,
                enviado_grupo=False,
                erros=[str(e)],
                mensagem_enviada=mensagem[:2000],
            )
            if img_bytes is not None:
                create_kw['imagem_anexo'] = ContentFile(img_bytes, name=img_nome or 'anexo.jpg')
            AnteciparInstalacaoSolicitacao.objects.create(**create_kw)
            return Response({'detail': f'Erro ao enviar: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        obs_rep_final = (request.data.get('observacao_reparo') or '').strip()[:500] if tipo == 'reparo' else ''
        create_kw = dict(
            usuario=request.user,
            venda=venda,
            ordem_servico=os_num,
            tipo_solicitacao=tipo,
            descricao_solicitacao=descricao,
            observacao_reparo=obs_rep_final,
            enviado_gc=('número do GC' in enviados),
            enviado_grupo=('grupo' in enviados),
            erros=erros,
            mensagem_enviada=mensagem[:2000],
        )
        if img_bytes is not None:
            create_kw['imagem_anexo'] = ContentFile(img_bytes, name=img_nome or 'anexo.jpg')
        AnteciparInstalacaoSolicitacao.objects.create(**create_kw)
        if not enviados:
            return Response({'detail': 'Nenhuma mensagem foi enviada.', 'erros': erros}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({
            'success': True,
            'message': f'Mensagem enviada para {", ".join(enviados)}.',
            'erros': erros if erros else None
        })


class DadosFPDView(APIView):
    """Retorna dados FPD de um contrato por O.S (Ordem de Serviço)"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/dados-fpd/?os=NR_ORDEM"""
        nr_ordem = request.query_params.get('os')
        if not nr_ordem:
            return Response({'error': 'Parâmetro os (Ordem de Serviço) é obrigatório'}, status=400)
        
        try:
            contrato = ContratoM10.objects.get(ordem_servico=nr_ordem)
            
            # Busca registros de importação FPD para esta O.S
            from .models import ImportacaoFPD
            dados_fpd = ImportacaoFPD.objects.filter(nr_ordem=nr_ordem).order_by('-importada_em')
            
            # Dados do contrato
            resultado = {
                'contrato': {
                    'numero_contrato': contrato.numero_contrato,
                    'numero_contrato_definitivo': contrato.numero_contrato_definitivo,
                    'cliente_nome': contrato.cliente_nome,
                    'cpf_cliente': contrato.cpf_cliente,
                    'ordem_servico': contrato.ordem_servico,
                    'vendedor': contrato.vendedor.get_full_name() if contrato.vendedor else None,
                    'data_instalacao': contrato.data_instalacao.isoformat(),
                    'status_contrato': contrato.get_status_contrato_display(),
                },
                'importacoes_fpd': []
            }
            
            # Adiciona dados de cada importação FPD
            for imp in dados_fpd:
                resultado['importacoes_fpd'].append({
                    'id_contrato': imp.id_contrato,
                    'nr_fatura': imp.nr_fatura,
                    'dt_venc_orig': imp.dt_venc_orig.isoformat(),
                    'dt_pagamento': imp.dt_pagamento.isoformat() if imp.dt_pagamento else None,
                    'ds_status_fatura': imp.ds_status_fatura,
                    'vl_fatura': str(imp.vl_fatura),
                    'nr_dias_atraso': imp.nr_dias_atraso,
                    'importada_em': imp.importada_em.isoformat(),
                })
            
            # Dados das faturas M10 vinculadas
            resultado['faturas_m10'] = []
            for fatura in contrato.faturas.all():
                resultado['faturas_m10'].append({
                    'numero_fatura': fatura.numero_fatura,
                    'id_contrato_fpd': fatura.id_contrato_fpd,
                    'dt_pagamento_fpd': fatura.dt_pagamento_fpd.isoformat() if fatura.dt_pagamento_fpd else None,
                    'ds_status_fatura_fpd': fatura.ds_status_fatura_fpd,
                    'status': fatura.status,
                    'valor': str(fatura.valor),
                    'data_vencimento': fatura.data_vencimento.isoformat(),
                    'data_pagamento': fatura.data_pagamento.isoformat() if fatura.data_pagamento else None,
                    'data_importacao_fpd': fatura.data_importacao_fpd.isoformat() if fatura.data_importacao_fpd else None,
                })
            
            return Response(resultado)
            
        except ContratoM10.DoesNotExist:
            return Response({'error': f'Contrato com O.S {nr_ordem} não encontrado'}, status=404)


class BuscarOSFPDView(APIView):
    """Busca por O.S específica na ImportacaoFPD"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/buscar-os-fpd/?os=07309961"""
        nr_ordem = request.query_params.get('os', '').strip()
        
        if not nr_ordem:
            return Response({'error': 'Parâmetro os é obrigatório'}, status=400)
        
        from .models import ImportacaoFPD
        
        # Buscar registros com a O.S
        registros = ImportacaoFPD.objects.filter(nr_ordem__icontains=nr_ordem).order_by('-importada_em')
        
        dados = []
        valor_total = 0
        
        for imp in registros:
            valor_total += float(imp.vl_fatura or 0)
            dados.append({
                'id': imp.id,
                'nr_ordem': imp.nr_ordem,
                'nr_fatura': imp.nr_fatura,
                'dt_venc_orig': imp.dt_venc_orig.isoformat(),
                'dt_pagamento': imp.dt_pagamento.isoformat() if imp.dt_pagamento else None,
                'ds_status_fatura': imp.ds_status_fatura,
                'vl_fatura': str(imp.vl_fatura),
                'nr_dias_atraso': imp.nr_dias_atraso,
                'contrato_m10': f"{imp.contrato_m10.numero_contrato} - {imp.contrato_m10.cliente_nome}" if imp.contrato_m10 else None,
                'importada_em': imp.importada_em.isoformat(),
            })
        
        return Response({
            'total': len(dados),
            'valor_total': str(valor_total),
            'os': nr_ordem,
            'registros': dados,
        })


class BuscarOSChurnView(APIView):
    """Busca por O.S específica na ImportacaoChurn"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/buscar-os-churn/?os=05444203"""
        nr_ordem = request.query_params.get('os', '').strip()
        
        if not nr_ordem:
            return Response({'error': 'Parâmetro os é obrigatório'}, status=400)
        
        from .models import ImportacaoChurn, ContratoM10
        
        # Normalizar O.S para busca (tentar variações)
        os_variants = [nr_ordem]
        if nr_ordem.isdigit():
            os_variants.append(nr_ordem.zfill(8))
            os_variants.append(nr_ordem.lstrip('0') or '0')
        if '-' in nr_ordem:
            part = nr_ordem.split('-', 1)[1].strip()
            if part:
                os_variants.append(part)
                if part.isdigit():
                    os_variants.append(part.zfill(8))
        
        # Buscar registros com a O.S (qualquer variação)
        from django.db.models import Q
        q_filter = Q()
        for variant in os_variants:
            q_filter |= Q(nr_ordem=variant) | Q(numero_pedido=variant)
        
        registros = ImportacaoChurn.objects.filter(q_filter).order_by('-id')
        
        dados = []
        for imp in registros:
            # Buscar ContratoM10 relacionado
            contrato_m10 = None
            try:
                # Tentar encontrar por ordem_servico (com variações)
                for variant in os_variants:
                    c = ContratoM10.objects.filter(ordem_servico=variant).first()
                    if c:
                        contrato_m10 = {
                            'id': c.id,
                            'numero_contrato': c.numero_contrato,
                            'cliente_nome': c.cliente_nome,
                            'status_contrato': c.status_contrato,
                            'data_instalacao': c.data_instalacao.isoformat() if c.data_instalacao else None,
                            'safra': c.safra,
                        }
                        break
            except Exception:
                pass
            
            dados.append({
                'id': imp.id,
                'nr_ordem': imp.nr_ordem,
                'numero_pedido': imp.numero_pedido,
                'uf': imp.uf,
                'municipio': imp.municipio,
                'produto': imp.produto,
                'dt_gross': imp.dt_gross.isoformat() if imp.dt_gross else None,
                'anomes_gross': imp.anomes_gross,
                'dt_retirada': imp.dt_retirada.isoformat() if imp.dt_retirada else None,
                'anomes_retirada': imp.anomes_retirada,
                'tipo_retirada': imp.tipo_retirada,
                'motivo_retirada': imp.motivo_retirada,
                'submotivo_retirada': imp.submotivo_retirada,
                'classificacao': imp.classificacao,
                'contrato_m10': contrato_m10,
            })
        
        return Response({
            'total': len(dados),
            'os': nr_ordem,
            'variantes_tentadas': list(set(os_variants)),
            'registros': dados,
        })


class ListarImportacoesFPDView(APIView):
    """Lista todas as importações FPD com filtros"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/importacoes-fpd/?status=PAGO&mes=2025-01"""
        from .models import ImportacaoFPD
        
        queryset = ImportacaoFPD.objects.all()
        
        # Filtro por status
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(ds_status_fatura__icontains=status)
        
        # Filtro por mês de vencimento
        mes = request.query_params.get('mes')
        if mes:
            try:
                data_inicio = datetime.strptime(mes, '%Y-%m')
                data_fim = (data_inicio.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                queryset = queryset.filter(dt_venc_orig__range=[data_inicio, data_fim])
            except ValueError:
                return Response({'error': 'Formato de mês inválido (use YYYY-MM)'}, status=400)
        
        # Estatísticas
        total_faturas = queryset.count()
        total_valor = queryset.aggregate(Sum('vl_fatura'))['vl_fatura__sum'] or 0
        
        # Paginação
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 100))
        start = (page - 1) * limit
        end = start + limit
        
        dados = []
        for imp in queryset.order_by('-importada_em')[start:end]:
            dados.append({
                'nr_ordem': imp.nr_ordem,
                'id_contrato': imp.id_contrato,
                'nr_fatura': imp.nr_fatura,
                'dt_venc_orig': imp.dt_venc_orig.isoformat(),
                'dt_pagamento': imp.dt_pagamento.isoformat() if imp.dt_pagamento else None,
                'ds_status_fatura': imp.ds_status_fatura,
                'vl_fatura': str(imp.vl_fatura),
                'nr_dias_atraso': imp.nr_dias_atraso,
                'contrato_m10': f"{imp.contrato_m10.numero_contrato} - {imp.contrato_m10.cliente_nome}" if imp.contrato_m10 else None,
                'importada_em': imp.importada_em.isoformat(),
            })
        
        return Response({
            'total': total_faturas,
            'total_valor': str(total_valor),
            'pagina': page,
            'limit': limit,
            'dados': dados,
        })


class ListarLogsImportacaoFPDView(APIView):
    """Lista logs de importação FPD para monitoramento e debug"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/logs-importacao-fpd/?status=ERRO&page=1"""
        from .models import LogImportacaoFPD
        from django.db.models import Sum, Count, Avg, Q
        
        queryset = LogImportacaoFPD.objects.select_related('usuario').all()
        
        # Filtros
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        usuario_id = request.query_params.get('usuario_id')
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)
        
        data_inicio = request.query_params.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(iniciado_em__gte=data_inicio)
        
        data_fim = request.query_params.get('data_fim')
        if data_fim:
            queryset = queryset.filter(iniciado_em__lte=data_fim)
        
        # Paginação
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 50))
        start = (page - 1) * limit
        end = start + limit
        
        total = queryset.count()
        logs = queryset.order_by('-iniciado_em')[start:end]
        
        # Estatísticas gerais
        try:
            stats_geral = LogImportacaoFPD.objects.aggregate(
                total_importacoes=Count('id'),
                total_linhas_processadas=Sum('total_linhas'),
                total_sucesso=Count('id', filter=Q(status='SUCESSO')),
                total_erro=Count('id', filter=Q(status='ERRO')),
                total_parcial=Count('id', filter=Q(status='PARCIAL')),
                media_duracao=Avg('duracao_segundos'),
                total_valor=Sum('total_valor_importado')
            )
            
            total_importacoes = stats_geral['total_importacoes'] or 0
            total_sucesso = stats_geral['total_sucesso'] or 0
            
            resultado = {
                'total': total,
                'page': page,
                'limit': limit,
                'total_pages': (total + limit - 1) // limit if limit > 0 else 1,
                'estatisticas_gerais': {
                    'total_importacoes': total_importacoes,
                    'total_linhas_processadas': int(stats_geral['total_linhas_processadas'] or 0),
                    'total_sucesso': total_sucesso,
                    'total_erro': stats_geral['total_erro'] or 0,
                    'total_parcial': stats_geral['total_parcial'] or 0,
                    'media_duracao_segundos': float(stats_geral['media_duracao'] or 0),
                    'total_valor_importado': str(stats_geral['total_valor'] or 0),
                    'taxa_sucesso': round((total_sucesso / total_importacoes * 100) if total_importacoes > 0 else 0, 2)
                },
                'logs': []
            }
        except Exception as e:
            # Em caso de erro, retornar estrutura vazia mas válida
            resultado = {
                'total': 0,
                'page': page,
                'limit': limit,
                'total_pages': 1,
                'estatisticas_gerais': {
                    'total_importacoes': 0,
                    'total_linhas_processadas': 0,
                    'total_sucesso': 0,
                    'total_erro': 0,
                    'total_parcial': 0,
                    'media_duracao_segundos': 0.0,
                    'total_valor_importado': '0',
                    'taxa_sucesso': 0.0
                },
                'logs': []
            }
        
        for log in logs:
            try:
                log_data = {
                    'id': log.id,
                    'nome_arquivo': log.nome_arquivo or '',
                    'tamanho_arquivo': log.tamanho_arquivo or 0,
                    'usuario': {
                        'id': log.usuario.id if log.usuario else None,
                        'username': log.usuario.username if log.usuario else 'Sistema',
                        'nome_completo': log.usuario.get_full_name() if log.usuario else 'Sistema',
                    },
                    'status': log.status or 'DESCONHECIDO',
                    'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                    'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                    'duracao_segundos': log.duracao_segundos or 0,
                    'total_linhas': log.total_linhas or 0,
                    'total_processadas': log.total_processadas or 0,
                    'total_erros': log.total_erros or 0,
                    'total_contratos_nao_encontrados': log.total_contratos_nao_encontrados or 0,
                    'total_valor_importado': str(log.total_valor_importado) if log.total_valor_importado else '0',
                    'mensagem_erro': log.mensagem_erro or None,
                    'exemplos_nao_encontrados': log.exemplos_nao_encontrados or None,
                }
                
                # Adicionar detalhes completos se requisitado
                if request.query_params.get('detalhes') == 'true':
                    log_data['detalhes_json'] = log.detalhes_json
                
                resultado['logs'].append(log_data)
            except Exception as e:
                # Se houver erro ao processar um log específico, pular para o próximo
                print(f"Erro ao processar log {log.id}: {str(e)}")
                continue
        
        return Response(resultado)


class ListarLogsImportacaoChurnView(APIView):
    """Lista logs de importação CHURN para monitoramento e debug"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        """GET /api/bonus-m10/logs-importacao-churn/?status=ERRO&page=1"""
        from .models import LogImportacaoChurn
        from django.db.models import Sum, Count, Avg, Q
        
        queryset = LogImportacaoChurn.objects.select_related('usuario').all()
        
        # Filtros
        status = request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        usuario_id = request.query_params.get('usuario_id')
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)
        
        data_inicio = request.query_params.get('data_inicio')
        if data_inicio:
            queryset = queryset.filter(iniciado_em__gte=data_inicio)
        
        data_fim = request.query_params.get('data_fim')
        if data_fim:
            queryset = queryset.filter(iniciado_em__lte=data_fim)
        
        # Paginação
        page = int(request.query_params.get('page', 1))
        limit = int(request.query_params.get('limit', 50))
        start = (page - 1) * limit
        end = start + limit
        
        total = queryset.count()
        logs = queryset.order_by('-iniciado_em')[start:end]
        
        # Estatísticas gerais
        try:
            stats_geral = LogImportacaoChurn.objects.aggregate(
                total_importacoes=Count('id'),
                total_linhas_processadas=Sum('total_linhas'),
                total_sucesso=Count('id', filter=Q(status='SUCESSO')),
                total_erro=Count('id', filter=Q(status='ERRO')),
                total_parcial=Count('id', filter=Q(status='PARCIAL')),
                media_duracao=Avg('duracao_segundos'),
                total_cancelados=Sum('total_contratos_cancelados'),
                total_reativados=Sum('total_contratos_reativados')
            )
            
            total_importacoes = stats_geral['total_importacoes'] or 0
            total_sucesso = stats_geral['total_sucesso'] or 0
            
            resultado = {
                'total': total,
                'page': page,
                'limit': limit,
                'total_pages': (total + limit - 1) // limit if limit > 0 else 1,
                'estatisticas_gerais': {
                    'total_importacoes': total_importacoes,
                    'total_linhas_processadas': int(stats_geral['total_linhas_processadas'] or 0),
                    'total_sucesso': total_sucesso,
                    'total_erro': stats_geral['total_erro'] or 0,
                    'total_parcial': stats_geral['total_parcial'] or 0,
                    'media_duracao_segundos': float(stats_geral['media_duracao'] or 0),
                    'total_cancelados': int(stats_geral['total_cancelados'] or 0),
                    'total_reativados': int(stats_geral['total_reativados'] or 0),
                    'taxa_sucesso': round((total_sucesso / total_importacoes * 100) if total_importacoes > 0 else 0, 2)
                },
                'logs': []
            }
        except Exception as e:
            # Em caso de erro, retornar estrutura vazia mas válida
            resultado = {
                'total': 0,
                'page': page,
                'limit': limit,
                'total_pages': 1,
                'estatisticas_gerais': {
                    'total_importacoes': 0,
                    'total_linhas_processadas': 0,
                    'total_sucesso': 0,
                    'total_erro': 0,
                    'total_parcial': 0,
                    'media_duracao_segundos': 0.0,
                    'total_cancelados': 0,
                    'total_reativados': 0,
                    'taxa_sucesso': 0.0
                },
                'logs': []
            }
        
        for log in logs:
            try:
                log_data = {
                    'id': log.id,
                    'nome_arquivo': log.nome_arquivo or '',
                    'tamanho_arquivo': log.tamanho_arquivo or 0,
                    'usuario': {
                        'id': log.usuario.id if log.usuario else None,
                        'username': log.usuario.username if log.usuario else 'Sistema',
                        'nome_completo': log.usuario.get_full_name() if log.usuario else 'Sistema',
                    },
                    'status': log.status or 'DESCONHECIDO',
                    'iniciado_em': log.iniciado_em.isoformat() if log.iniciado_em else None,
                    'finalizado_em': log.finalizado_em.isoformat() if log.finalizado_em else None,
                    'duracao_segundos': log.duracao_segundos or 0,
                    'total_linhas': log.total_linhas or 0,
                    'total_processadas': log.total_processadas or 0,
                    'total_erros': log.total_erros or 0,
                    'total_contratos_cancelados': log.total_contratos_cancelados or 0,
                    'total_contratos_reativados': log.total_contratos_reativados or 0,
                    'total_nao_encontrados': log.total_nao_encontrados or 0,
                    'mensagem_erro': log.mensagem_erro or None,
                }
                
                # Adicionar detalhes completos se requisitado
                if request.query_params.get('detalhes') == 'true':
                    log_data['detalhes_json'] = log.detalhes_json
                
                resultado['logs'].append(log_data)
            except Exception as e:
                # Se houver erro ao processar um log específico, pular para o próximo
                print(f"Erro ao processar log {log.id}: {str(e)}")
                continue
        
        return Response(resultado)


class FaturaM10ListView(generics.ListCreateAPIView):
    """Lista e cria faturas de um contrato específico"""
    serializer_class = FaturaM10Serializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    def create(self, request, *args, **kwargs):
        """Override create para usar update_or_create e evitar erro de duplicação"""
        contrato_id = request.data.get('contrato')
        numero_fatura = request.data.get('numero_fatura')
        
        if not contrato_id or not numero_fatura:
            return Response({'error': 'contrato e numero_fatura são obrigatórios'}, status=400)
        
        try:
            contrato = ContratoM10.objects.get(id=contrato_id)
        except ContratoM10.DoesNotExist:
            return Response({'error': 'Contrato não encontrado'}, status=404)
        
        # Remove campos que não devem ser passados para update_or_create
        dados_fatura = {k: v for k, v in request.data.items() if k not in ['contrato', 'numero_fatura']}
        dados_fatura['contrato'] = contrato
        
        # Usa update_or_create para evitar duplicação
        fatura, created = FaturaM10.objects.update_or_create(
            contrato=contrato,
            numero_fatura=numero_fatura,
            defaults=dados_fatura
        )
        
        serializer = self.get_serializer(fatura)
        status_code = 201 if created else 200
        return Response(serializer.data, status=status_code)

    def _sincronizar_primeira_fatura_com_fpd(self, contrato_id: int):
        """Garante que a fatura 1 reflita os dados importados do FPD (prioridade)."""
        try:
            contrato = ContratoM10.objects.get(id=contrato_id)
        except ContratoM10.DoesNotExist:
            return

        tem_fpd = any([
            contrato.data_vencimento_fpd,
            contrato.status_fatura_fpd,
            contrato.valor_fatura_fpd,
            contrato.data_pagamento_fpd,
        ])
        if not tem_fpd:
            return

        status_fpd = (contrato.status_fatura_fpd or '').upper()
        if 'PAGA' in status_fpd:
            status_m10 = 'PAGO'
        elif 'AGUARDANDO' in status_fpd:
            status_m10 = 'AGUARDANDO'
        elif 'ATRASAD' in status_fpd or 'VENCID' in status_fpd:
            status_m10 = 'ATRASADO'
        else:
            status_m10 = 'NAO_PAGO'

        defaults = {
            'valor': contrato.valor_fatura_fpd or 0,
            'data_vencimento': contrato.data_vencimento_fpd or date.today(),
            'data_pagamento': contrato.data_pagamento_fpd,
            'status': status_m10,
        }
        fatura, _ = FaturaM10.objects.get_or_create(
            contrato=contrato,
            numero_fatura=1,
            defaults=defaults,
        )

        alterado = False
        if contrato.valor_fatura_fpd and fatura.valor != contrato.valor_fatura_fpd:
            fatura.valor = contrato.valor_fatura_fpd
            alterado = True
        if contrato.data_vencimento_fpd and fatura.data_vencimento != contrato.data_vencimento_fpd:
            fatura.data_vencimento = contrato.data_vencimento_fpd
            alterado = True
        if contrato.data_pagamento_fpd and fatura.data_pagamento != contrato.data_pagamento_fpd:
            fatura.data_pagamento = contrato.data_pagamento_fpd
            alterado = True
        if fatura.status != status_m10:
            fatura.status = status_m10
            alterado = True
        if alterado:
            fatura.save(update_fields=['valor', 'data_vencimento', 'data_pagamento', 'status', 'atualizado_em'])
    
    def get_queryset(self):
        contrato_id = self.request.query_params.get('contrato_id')
        if contrato_id:
            self._sincronizar_primeira_fatura_com_fpd(contrato_id)
            return FaturaM10.objects.filter(contrato_id=contrato_id).order_by('numero_fatura')
        return FaturaM10.objects.none()
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class FaturaM10DetailView(generics.RetrieveUpdateAPIView):
    """Detalhes e atualização de uma fatura específica (suporta link ou upload de PDF)"""
    queryset = FaturaM10.objects.all()
    serializer_class = FaturaM10Serializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class BuscarFaturaNioView(APIView):
    """Busca automática de fatura no site da Nio Internet"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        """
        Busca fatura automaticamente e retorna os dados
        
        Body: {
            "cpf": "12345678900",
            "contrato_id": 123,
            "numero_fatura": 1,
            "salvar": false  # Se true, salva automaticamente
        }
        """
        from datetime import date
        
        cpf = request.data.get('cpf')
        contrato_id = request.data.get('contrato_id')
        numero_fatura = request.data.get('numero_fatura')
        salvar = request.data.get('salvar', False)
        
        if not cpf:
            return Response({'error': 'CPF não informado'}, status=400)
        
        # Verificar se a fatura já está disponível
        if contrato_id and numero_fatura:
            try:
                contrato = ContratoM10.objects.get(id=contrato_id)
                fatura = FaturaM10.objects.filter(
                    contrato=contrato,
                    numero_fatura=numero_fatura
                ).first()
                
                if fatura and fatura.data_disponibilidade:
                    hoje = date.today()
                    if fatura.data_disponibilidade > hoje:
                        dias_faltam = (fatura.data_disponibilidade - hoje).days
                        return Response({
                            'error': f'Fatura ainda não disponível. Estará disponível em {dias_faltam} dia(s), a partir de {fatura.data_disponibilidade.strftime("%d/%m/%Y")}.',
                            'data_disponibilidade': fatura.data_disponibilidade,
                            'disponivel': False
                        }, status=400)
            except ContratoM10.DoesNotExist:
                pass
        
        try:
            # Plano A = mesma consulta do WhatsApp: API Nio (consultar_dividas_nio). Sem Playwright.
            import re
            import requests
            from crm_app.nio_api import consultar_dividas_nio, get_invoice_pdf_url

            cpf_limpo = re.sub(r'\D', '', str(cpf))
            if not cpf_limpo or len(cpf_limpo) < 11:
                return Response({'error': 'CPF inválido'}, status=400)

            api_result = consultar_dividas_nio(cpf_limpo, offset=0, limit=50, headless=True)
            invoices = api_result.get('invoices') or []

            if not invoices:
                return Response({
                    'success': True,
                    'sem_dividas': True,
                    'mensagem': 'CPF sem dívidas no momento.'
                }, status=200)

            # Data de vencimento esperada: só preenchemos fatura que corresponda a esta data (evita pegar fatura errada)
            data_vencimento_esperada = None
            if contrato_id and numero_fatura:
                try:
                    contrato = ContratoM10.objects.get(id=contrato_id)
                    fatura = FaturaM10.objects.filter(
                        contrato=contrato,
                        numero_fatura=numero_fatura
                    ).first()
                    if fatura and fatura.data_vencimento:
                        data_vencimento_esperada = fatura.data_vencimento
                    else:
                        data_vencimento_esperada = contrato.calcular_vencimento_fatura_n(numero_fatura)
                except ContratoM10.DoesNotExist:
                    pass
            # Aceitar também do body (valor do formulário quando o usuário alterou a data)
            data_vencimento_form = request.data.get('data_vencimento_esperada')
            if data_vencimento_form:
                try:
                    from datetime import datetime as dt
                    if isinstance(data_vencimento_form, str) and len(data_vencimento_form) >= 10:
                        data_vencimento_esperada = dt.strptime(data_vencimento_form[:10], '%Y-%m-%d').date()
                except Exception:
                    pass

            def _parse_due(inv):
                due = inv.get('due_date_raw') or inv.get('data_vencimento')
                if not due:
                    return None
                if hasattr(due, 'strftime'):
                    return due
                if isinstance(due, str) and len(due) >= 8:
                    try:
                        from datetime import datetime as dt
                        s = due[:10].replace('/', '-')
                        if '-' in s:
                            return dt.strptime(s, '%Y-%m-%d').date()
                        if due[:8].isdigit():
                            return dt.strptime(due[:8], '%Y%m%d').date()
                    except Exception:
                        pass
                return None

            # Matching por data de vencimento: só aceitar fatura Nio cujo vencimento coincida com o esperado (±3 dias)
            TOLERANCIA_DIAS = 3
            inv = None
            data_vencimento = None
            if data_vencimento_esperada:
                melhor_diff = TOLERANCIA_DIAS + 1
                for i in invoices:
                    dv = _parse_due(i)
                    if not dv:
                        continue
                    diff = abs((dv - data_vencimento_esperada).days)
                    if diff <= TOLERANCIA_DIAS and diff < melhor_diff:
                        melhor_diff = diff
                        inv = i
                        data_vencimento = dv
                if inv is None:
                    datas_nio = []
                    for i in invoices:
                        dv = _parse_due(i)
                        if dv:
                            datas_nio.append(dv.strftime('%d/%m/%Y'))
                    return Response({
                        'error': (
                            f'Nenhuma fatura na Nio com vencimento em {data_vencimento_esperada.strftime("%d/%m/%Y")} (±{TOLERANCIA_DIAS} dias). '
                            f'Faturas encontradas na Nio: {", ".join(datas_nio[:10]) or "sem data"}. '
                            'Confira a data de vencimento no formulário ou se a fatura já está disponível na Nio.'
                        ),
                        'data_vencimento_esperada': data_vencimento_esperada.strftime('%Y-%m-%d'),
                        'datas_nio': datas_nio[:10],
                    }, status=400)
            else:
                # Sem data esperada (ex.: contrato não informado): usar primeira fatura e avisar
                inv = invoices[0]
                data_vencimento = _parse_due(inv)

            valor = inv.get('amount')
            codigo_pix = inv.get('pix') or inv.get('codigo_pix')
            codigo_barras = inv.get('barcode') or inv.get('codigo_barras')
            pdf_url = None
            if api_result.get('token') and api_result.get('api_base') and api_result.get('session_id'):
                sess = requests.Session()
                pdf_url = get_invoice_pdf_url(
                    api_result['api_base'],
                    api_result['token'],
                    api_result['session_id'],
                    inv.get('debt_id', ''),
                    str(inv.get('invoice_id', '')),
                    cpf_limpo,
                    inv.get('reference_month', '') or '',
                    sess,
                )

            dados = {
                'valor': valor,
                'codigo_pix': codigo_pix,
                'codigo_barras': codigo_barras,
                'data_vencimento': data_vencimento,
                'pdf_url': pdf_url,
            }

            if not any([dados.get('valor'), dados.get('codigo_pix'), dados.get('codigo_barras')]):
                return Response({
                    'error': 'Fatura encontrada mas sem dados disponíveis. Preencha manualmente.'
                }, status=404)
            
            # Se deve salvar automaticamente
            if salvar and contrato_id and numero_fatura:
                try:
                    contrato = ContratoM10.objects.get(id=contrato_id)
                    fatura, created = FaturaM10.objects.get_or_create(
                        contrato=contrato,
                        numero_fatura=numero_fatura,
                        defaults={
                            'valor': dados['valor'] or 0,
                            'data_vencimento': dados['data_vencimento'] or datetime.now().date(),
                            'status': 'NAO_PAGO'
                        }
                    )
                    
                    # Atualiza campos
                    if dados['valor']:
                        fatura.valor = dados['valor']
                    if dados['data_vencimento']:
                        fatura.data_vencimento = dados['data_vencimento']
                    if dados['codigo_pix']:
                        fatura.codigo_pix = dados['codigo_pix']
                    if dados['codigo_barras']:
                        fatura.codigo_barras = dados['codigo_barras']
                    if dados['pdf_url']:
                        fatura.pdf_url = dados['pdf_url']
                    
                    fatura.save()
                    
                    return Response({
                        'success': True,
                        'message': '✅ Fatura preenchida automaticamente!',
                        'dados': {
                            'valor': float(fatura.valor),
                            'data_vencimento': fatura.data_vencimento.strftime('%Y-%m-%d'),
                            'codigo_pix': fatura.codigo_pix,
                            'codigo_barras': fatura.codigo_barras,
                            'tem_pdf': bool(fatura.pdf_url or fatura.arquivo_pdf),
                            'pdf_url': fatura.pdf_url,
                        }
                    })
                    
                except ContratoM10.DoesNotExist:
                    return Response({'error': 'Contrato não encontrado'}, status=404)
                except Exception as e:
                    return Response({'error': f'Erro ao salvar: {str(e)}'}, status=500)
            
            # Retorna apenas os dados sem salvar
            return Response({
                'success': True,
                'dados': dados
            })
            
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception('BuscarFaturaNioView: erro ao buscar fatura (API Nio)')
            return Response({
                'error': f'Não foi possível buscar a fatura. Verifique o CPF ou tente novamente.'
            }, status=500)


def _run_buscar_faturas_safra_background(historico_id, param, numero_fatura_filtro):
    """
    Executa a busca de faturas da safra em background (thread).
    Atualiza HistoricoBuscaFatura ao final para não bloquear a requisição HTTP.
    """
    import re
    import requests
    from datetime import date, datetime as dt
    from django.utils import timezone
    from django.db.models import Q
    from crm_app.nio_api import consultar_dividas_nio, get_invoice_pdf_url
    from crm_app.models import HistoricoBuscaFatura

    def _api_to_faturas_nio(api_result, cpf_limpo, incluir_pdf=False):
        invoices = api_result.get('invoices') or []
        out = []
        sess = requests.Session() if incluir_pdf else None
        for inv in invoices:
            due = inv.get('due_date_raw') or inv.get('data_vencimento')
            data_vencimento = None
            if due:
                if hasattr(due, 'strftime'):
                    data_vencimento = due
                elif isinstance(due, str) and len(due) >= 8:
                    try:
                        s = due[:10].replace('/', '-')
                        if '-' in s:
                            data_vencimento = dt.strptime(s, '%Y-%m-%d').date()
                        elif due[:8].isdigit():
                            data_vencimento = dt.strptime(due[:8], '%Y%m%d').date()
                    except Exception:
                        pass
            valor = inv.get('amount')
            codigo_pix = inv.get('pix') or inv.get('codigo_pix')
            codigo_barras = inv.get('barcode') or inv.get('codigo_barras')
            pdf_url = None
            if incluir_pdf and api_result.get('token') and api_result.get('api_base') and api_result.get('session_id') and sess:
                pdf_url = get_invoice_pdf_url(
                    api_result['api_base'], api_result['token'], api_result['session_id'],
                    inv.get('debt_id', ''), str(inv.get('invoice_id', '')), cpf_limpo,
                    inv.get('reference_month', '') or '', sess)
            out.append({
                'data_vencimento': data_vencimento,
                'valor': valor,
                'codigo_pix': codigo_pix,
                'codigo_barras': codigo_barras,
                'pdf_url': pdf_url,
            })
        return out

    inicio = timezone.now()
    try:
        historico = HistoricoBuscaFatura.objects.get(id=historico_id)
        data_inicio, data_fim = _safra_to_data_range(param)
        if data_inicio is None or data_fim is None:
            historico.status = 'ERRO'
            historico.mensagem = 'Safra inválida'
            historico.termino_em = timezone.now()
            historico.save()
            return
        contratos = list(ContratoM10.objects.filter(
            data_instalacao__gte=data_inicio,
            data_instalacao__lt=data_fim,
            status_contrato='ATIVO'
        ))
        hoje = date.today()
        resultados = {
            'total_contratos': len(contratos),
            'processados': 0,
            'sucesso': 0,
            'erros': 0,
            'nao_disponiveis': 0,
            'sem_cpf': 0,
            'detalhes': []
        }
        for contrato in contratos:
                if not contrato.cpf_cliente:
                    resultados['sem_cpf'] += 1
                    resultados['detalhes'].append({
                        'contrato': contrato.numero_contrato,
                        'status': 'sem_cpf',
                        'mensagem': 'CPF não cadastrado'
                    })
                    continue
                
                # Busca faturas disponíveis
                from django.db.models import Q
                
                if numero_fatura_filtro:
                    # Busca apenas a fatura especificada
                    fatura = FaturaM10.objects.filter(
                        contrato=contrato,
                        numero_fatura=numero_fatura_filtro
                    ).first()
                    faturas_disponiveis = [fatura] if fatura else []
                else:
                    # Busca apenas faturas não pagas E que já estão disponíveis
                    faturas_disponiveis = FaturaM10.objects.filter(
                        contrato=contrato,
                        status__in=['NAO_PAGO', 'ATRASADO', 'AGUARDANDO']
                    ).filter(
                        Q(data_disponibilidade__isnull=True) | Q(data_disponibilidade__lte=hoje)
                    ).order_by('numero_fatura')
                
                if not faturas_disponiveis:
                    continue
                
                # Busca faturas no Nio via API (mesma do WhatsApp). Matching por vencimento.
                try:
                    cpf_limpo = re.sub(r'\D', '', str(contrato.cpf_cliente))
                    if not cpf_limpo or len(cpf_limpo) < 11:
                        resultados['erros'] += 1
                        resultados['detalhes'].append({
                            'contrato': contrato.numero_contrato,
                            'status': 'cpf_invalido',
                            'mensagem': 'CPF inválido'
                        })
                        continue
                    api_result = consultar_dividas_nio(cpf_limpo, offset=0, limit=50, headless=True)
                    if api_result.get('erro_400'):
                        resultados['erros'] += 1
                        msg = api_result.get('detail') or 'CPF não encontrado ou inválido na base Nio (API 400)'
                        if isinstance(msg, dict):
                            msg = msg.get('message') or msg.get('detail') or str(msg)[:150]
                        resultados['detalhes'].append({
                            'contrato': contrato.numero_contrato,
                            'status': 'erro_400_nio',
                            'mensagem': msg[:200] if isinstance(msg, str) else str(msg)[:200]
                        })
                        continue
                    # PDF desativado na safra: API Nio /invoices/.../download e /pdf retornam 500.
                    # Valor, PIX e código de barras seguem sendo atualizados.
                    faturas_nio = _api_to_faturas_nio(api_result, cpf_limpo, incluir_pdf=False)

                    if not faturas_nio:
                        resultados['detalhes'].append({
                            'contrato': contrato.numero_contrato,
                            'status': 'sem_dividas_nio',
                            'mensagem': 'CPF sem dívidas no Nio no momento'
                        })
                        continue
                    
                    # Faz matching por data de vencimento (com tolerância de ±3 dias)
                    for fatura in faturas_disponiveis:
                        resultados['processados'] += 1
                        match_encontrado = False
                        
                        for fatura_nio in faturas_nio:
                            if not fatura_nio.get('data_vencimento'):
                                continue
                            
                            # Compara datas com tolerância de 3 dias
                            diff_dias = abs((fatura.data_vencimento - fatura_nio['data_vencimento']).days)
                            
                            if diff_dias <= 3:  # Match encontrado!
                                match_encontrado = True
                                
                                # Atualiza a fatura com os dados do Nio
                                alteracoes = []
                                if fatura_nio.get('valor') and fatura.valor != fatura_nio['valor']:
                                    fatura.valor = fatura_nio['valor']
                                    alteracoes.append('valor')
                                
                                if fatura_nio.get('data_vencimento') and fatura.data_vencimento != fatura_nio['data_vencimento']:
                                    fatura.data_vencimento = fatura_nio['data_vencimento']
                                    alteracoes.append('vencimento')
                                
                                if fatura_nio.get('codigo_pix') and fatura.codigo_pix != fatura_nio['codigo_pix']:
                                    fatura.codigo_pix = fatura_nio['codigo_pix']
                                    alteracoes.append('pix')
                                
                                if fatura_nio.get('codigo_barras') and fatura.codigo_barras != fatura_nio['codigo_barras']:
                                    fatura.codigo_barras = fatura_nio['codigo_barras']
                                    alteracoes.append('barcode')
                                
                                if fatura_nio.get('pdf_url') and fatura.pdf_url != fatura_nio['pdf_url']:
                                    fatura.pdf_url = fatura_nio['pdf_url']
                                    alteracoes.append('pdf_url')
                                
                                if alteracoes:
                                    fatura.save()
                                    resultados['sucesso'] += 1
                                    resultados['detalhes'].append({
                                        'contrato': contrato.numero_contrato,
                                        'fatura': fatura.numero_fatura,
                                        'status': 'atualizado',
                                        'mensagem': f'Match por vencimento (diff: {diff_dias} dia(s))',
                                        'alteracoes': alteracoes,
                                        'vencimento_crm': fatura.data_vencimento.strftime('%d/%m/%Y'),
                                        'vencimento_nio': fatura_nio['data_vencimento'].strftime('%d/%m/%Y')
                                    })
                                else:
                                    resultados['detalhes'].append({
                                        'contrato': contrato.numero_contrato,
                                        'fatura': fatura.numero_fatura,
                                        'status': 'sem_alteracao',
                                        'mensagem': 'Match encontrado mas dados já estão atualizados'
                                    })
                                
                                # Remove fatura_nio da lista para não fazer match duplicado
                                faturas_nio.remove(fatura_nio)
                                break
                        
                        if not match_encontrado:
                            resultados['detalhes'].append({
                                'contrato': contrato.numero_contrato,
                                'fatura': fatura.numero_fatura,
                                'status': 'sem_match',
                                'mensagem': f'Fatura vencimento {fatura.data_vencimento.strftime("%d/%m/%Y")} não encontrada no Nio',
                                'vencimento_crm': fatura.data_vencimento.strftime('%d/%m/%Y')
                            })
                    
                except Exception as e:
                    resultados['erros'] += 1
                    resultados['detalhes'].append({
                        'contrato': contrato.numero_contrato,
                        'status': 'erro',
                        'mensagem': str(e)
                    })
        # Atualiza histórico ao concluir
        termino = timezone.now()
        historico.termino_em = termino
        historico.duracao_segundos = (termino - inicio).total_seconds()
        historico.total_faturas = resultados['processados']
        historico.faturas_sucesso = resultados['sucesso']
        historico.faturas_erro = resultados['erros']
        historico.faturas_nao_disponiveis = resultados.get('nao_disponiveis', 0)
        historico.status = 'CONCLUIDA'
        historico.mensagem = f"Processados {resultados['processados']}, sucesso: {resultados['sucesso']}, erros: {resultados['erros']}"
        historico.logs = {'detalhes': resultados['detalhes'][:200]}
        historico.save()
    except Exception as e:
        try:
            historico = HistoricoBuscaFatura.objects.get(id=historico_id)
            historico.status = 'ERRO'
            historico.mensagem = str(e)[:500]
            historico.termino_em = timezone.now()
            historico.save()
        except Exception:
            pass


class BuscarFaturasSafraView(APIView):
    """Busca automática de todas as faturas disponíveis de uma safra (em background)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """
        Inicia busca de faturas em segundo plano. Retorna imediatamente com log_id.
        Body: { "safra": "2025-12" | "safra_id": 1, "numero_fatura": 1 opcional }
        """
        import threading
        from crm_app.models import HistoricoBuscaFatura
        safra_str = request.data.get('safra')
        safra_id = request.data.get('safra_id')
        numero_fatura_filtro = request.data.get('numero_fatura')
        if not safra_str and not safra_id:
            return Response({'error': 'Informe a safra (YYYY-MM) ou safra_id'}, status=400)
        param = safra_id or safra_str
        try:
            data_inicio, data_fim = _safra_to_data_range(param)
            if data_inicio is None or data_fim is None:
                return Response({'error': 'Safra inválida. Use safra_id ou safra (YYYY-MM).'}, status=400)
            contratos = ContratoM10.objects.filter(
                data_instalacao__gte=data_inicio,
                data_instalacao__lt=data_fim,
                status_contrato='ATIVO'
            )
            if not contratos.exists():
                return Response({'error': 'Nenhum contrato encontrado para esta safra'}, status=404)
            safra_display = f"{data_inicio.year}-{data_inicio.month:02d}" if data_inicio else str(param)
            historico = HistoricoBuscaFatura.objects.create(
                tipo_busca='SAFRA',
                safra=safra_display,
                usuario=None,
                status='EM_ANDAMENTO',
                total_contratos=contratos.count(),
            )
            thread = threading.Thread(
                target=_run_buscar_faturas_safra_background,
                args=(historico.id, param, numero_fatura_filtro),
                daemon=True,
            )
            thread.start()
            return Response({
                'success': True,
                'message': 'Busca de faturas iniciada em segundo plano. Atualize a página em alguns minutos para ver os resultados.',
                'log_id': historico.id,
                'background': True,
            }, status=202)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class NioDividasView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        cpf = (request.query_params.get("cpf") or "").strip()
        offset = int(request.query_params.get("offset", 0))
        limit = int(request.query_params.get("limit", 10))

        if not cpf:
            return Response({"detail": "CPF/CNPJ não informado"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Valida CPF/CNPJ mas trata exceção de validação
        try:
            validar_cpf_ou_cnpj(cpf)
        except ValidationError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            result = consultar_dividas_nio(
                cpf=cpf,
                offset=offset,
                limit=limit,
                storage_state=getattr(settings, "NIO_STORAGE_STATE", None),
                headless=True,
            )
        except Exception as exc:
            logger.exception("Erro ao consultar dívidas Nio")
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

        return Response({
            "invoices": result.get("invoices", []),
            "meta": {
                "api_base": result.get("api_base"),
                "session_id": result.get("session_id"),
                "count": len(result.get("invoices", [])),
            },
        })


# ============================================================================
# ENDPOINTS DE DEBUG - SCREENSHOTS
# ============================================================================

from django.http import FileResponse, JsonResponse, Http404
from pathlib import Path
import os
from datetime import datetime
from rest_framework.permissions import AllowAny

@api_view(['GET'])
@permission_classes([AllowAny])  # Temporariamente AllowAny para debug - depois mudar para IsAuthenticated
def listar_screenshots_debug(request):
    """
    Lista todos os screenshots de debug do Plano B (Nio Negocia)
    GET /api/crm/debug/screenshots/
    """
    try:
        # Caminho para a pasta downloads: usar settings.BASE_DIR (em produção Path(__file__) pode ser /)
        base_dir = Path(settings.BASE_DIR)
        downloads_dir = base_dir / 'downloads'
        
        # Debug: informações sobre o diretório
        debug_info = {
            'base_dir': str(base_dir),
            'downloads_dir': str(downloads_dir),
            'downloads_exists': downloads_dir.exists(),
            'downloads_is_dir': downloads_dir.is_dir() if downloads_dir.exists() else False,
        }
        
        screenshots = []
        todos_arquivos = []
        
        if downloads_dir.exists():
            # Listar TODOS os arquivos na pasta downloads para debug
            try:
                todos_arquivos = [f.name for f in downloads_dir.iterdir() if f.is_file()]
                debug_info['total_arquivos_downloads'] = len(todos_arquivos)
                debug_info['arquivos_encontrados'] = todos_arquivos[:20]  # Primeiros 20
            except Exception as e_list:
                debug_info['erro_listar'] = str(e_list)
            
            # Buscar screenshots do Nio Negocia
            try:
                for file in downloads_dir.glob('debug_nio_negocia_*.png'):
                    screenshots.append({
                        'nome': file.name,
                        'tipo': 'screenshot',
                        'tamanho': file.stat().st_size,
                        'data_modificacao': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                        'url': f"/api/crm/debug/screenshots/{file.name}/"
                    })
            except Exception as e_png:
                debug_info['erro_buscar_png'] = str(e_png)
            # Screenshots da automação PAP (produção)
            try:
                for file in downloads_dir.glob('pap_venda_*.png'):
                    screenshots.append({
                        'nome': file.name,
                        'tipo': 'screenshot',
                        'tamanho': file.stat().st_size,
                        'data_modificacao': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                        'url': f"/api/crm/debug/screenshots/{file.name}/"
                    })
            except Exception as e_pap:
                debug_info['erro_buscar_pap'] = str(e_pap)

            # Traces da automação PAP (ver cada clique em https://trace.playwright.dev)
            try:
                for file in downloads_dir.glob('pap_trace_*.zip'):
                    screenshots.append({
                        'nome': file.name,
                        'tipo': 'trace',
                        'tamanho': file.stat().st_size,
                        'data_modificacao': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                        'url': f"/api/crm/debug/screenshots/{file.name}/"
                    })
            except Exception as e_trace:
                debug_info['erro_buscar_trace'] = str(e_trace)

            # Buscar HTMLs de debug também
            try:
                for file in downloads_dir.glob('debug_nio_negocia_*.html'):
                    screenshots.append({
                        'nome': file.name,
                        'tipo': 'html',
                        'tamanho': file.stat().st_size,
                        'data_modificacao': datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
                        'url': f"/api/crm/debug/screenshots/{file.name}/"
                    })
            except Exception as e_html:
                debug_info['erro_buscar_html'] = str(e_html)
        else:
            # Tentar criar a pasta se não existir
            try:
                downloads_dir.mkdir(parents=True, exist_ok=True)
                debug_info['pasta_criada'] = True
            except Exception as e_mkdir:
                debug_info['erro_criar_pasta'] = str(e_mkdir)
        
        # Ordenar por data de modificação (mais recentes primeiro)
        screenshots.sort(key=lambda x: x['data_modificacao'], reverse=True)
        
        return JsonResponse({
            'total': len(screenshots),
            'screenshots': screenshots,
            'debug': debug_info  # Adicionar informações de debug
        })
    
    except Exception as e:
        import traceback
        return JsonResponse({
            'erro': str(e),
            'traceback': traceback.format_exc(),
            'total': 0,
            'screenshots': []
        }, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])  # Temporariamente AllowAny para debug - depois mudar para IsAuthenticated
def baixar_screenshot_debug(request, nome_arquivo):
    """
    Baixa um screenshot específico de debug
    GET /api/crm/debug/screenshots/<nome_arquivo>/
    """
    try:
        # Validar nome do arquivo (segurança: sem path traversal, apenas prefixos permitidos)
        if '/' in nome_arquivo or '\\' in nome_arquivo or '..' in nome_arquivo:
            return JsonResponse({'erro': 'Nome de arquivo inválido.'}, status=400)
        if not (nome_arquivo.startswith('debug_nio_negocia_') or nome_arquivo.startswith('pap_venda_') or nome_arquivo.startswith('pap_trace_')):
            return JsonResponse({
                'erro': 'Nome de arquivo inválido. Apenas screenshots (Nio Negocia, PAP venda) ou traces PAP (pap_trace_*.zip) são permitidos.'
            }, status=400)
        
        # Caminho para a pasta downloads (usar settings.BASE_DIR para produção)
        base_dir = Path(settings.BASE_DIR)
        downloads_dir = base_dir / 'downloads'
        arquivo = downloads_dir / nome_arquivo
        
        if not arquivo.exists():
            return JsonResponse({
                'erro': 'Arquivo não encontrado'
            }, status=404)
        
        # Determinar content-type baseado na extensão
        if nome_arquivo.endswith('.png'):
            content_type = 'image/png'
        elif nome_arquivo.endswith('.html'):
            content_type = 'text/html'
        elif nome_arquivo.endswith('.zip'):
            content_type = 'application/zip'
        else:
            content_type = 'application/octet-stream'
        
        # Retornar arquivo
        return FileResponse(
            open(arquivo, 'rb'),
            content_type=content_type,
            as_attachment=True,
            filename=nome_arquivo
        )
    
    except FileNotFoundError:
        return JsonResponse({
            'erro': 'Arquivo não encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'erro': str(e)
        }, status=500)
