"""Testes do histórico PAP (3 tipos + dedup pela coluna Pedido)."""
from io import BytesIO

from django.test import SimpleTestCase, TestCase
from openpyxl import Workbook
import pandas as pd

from crm_app.historico_pap import (
    extrair_lista_api,
    map_pedido_api,
    montar_xlsx_historico,
    normalizar_pedido,
    normalizar_tipo,
    pedidos_de_dataframe,
    tipos_solicitados,
)


class HistoricoPapNormalizaTest(SimpleTestCase):
    def test_pedido_protocolo(self):
        self.assertEqual(normalizar_pedido("202606292619833371"), "202606292619833371")
        self.assertEqual(normalizar_pedido("Pedido 202606292619833371"), "202606292619833371")
        self.assertEqual(normalizar_pedido("10132079"), "")

    def test_tipos(self):
        self.assertEqual(normalizar_tipo("pré-venda"), "PRE_VENDA")
        self.assertEqual(normalizar_tipo("Interesse"), "INTERESSE")
        self.assertEqual(
            tipos_solicitados(["venda", "interesse", "pré-venda"]),
            ["VENDA", "INTERESSE", "PRE_VENDA"],
        )

    def test_extrair_lista_e_map(self):
        payload = {
            "total": 1,
            "data": [
                {
                    "numeroPedido": "202606292619833371",
                    "status": "Pedido Gerado",
                    "chaveStatusPrimario": "PEDIDO_GERADO",
                    "cliente": "ALINE",
                    "cpf": "12345678901",
                }
            ],
        }
        lista, total = extrair_lista_api(payload)
        self.assertEqual(total, 1)
        row = map_pedido_api(lista[0], "VENDA")
        self.assertEqual(row["pedido"], "202606292619833371")
        self.assertEqual(row["tipo_venda"], "VENDA")
        blob = montar_xlsx_historico([row])
        self.assertGreater(len(blob), 100)

    def test_pedidos_de_excel(self):
        df = pd.DataFrame(
            [
                {"Pedido": "202606292619833371", "Cliente": "A"},
                {"Pedido": "202606292619833371", "Cliente": "dup"},
                {"Pedido": "202607011232833369", "Cliente": "B"},
            ]
        )
        pares = pedidos_de_dataframe(df)
        self.assertEqual(
            [p[0] for p in pares],
            ["202606292619833371", "202607011232833369"],
        )


class HistoricoPapDedupTest(TestCase):
    def test_registrar_exportacao_nao_repete(self):
        from crm_app.historico_pap_service import registrar_exportacao
        from crm_app.models import HistoricoPapPedido

        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        ws["A1"] = "Pedido"
        ws["A2"] = "202606292619833371"
        bio = BytesIO()
        wb.save(bio)
        r1 = registrar_exportacao(None, "exp.xlsx", bio.getvalue())
        r2 = registrar_exportacao(None, "exp.xlsx", bio.getvalue())
        self.assertEqual(r1["novos"], 1)
        self.assertEqual(r2["novos"], 0)
        self.assertEqual(r2["ja_existiam"], 1)
        self.assertEqual(
            HistoricoPapPedido.objects.filter(numero_pedido="202606292619833371").count(),
            1,
        )
        self.assertFalse(r1["grava_venda"])
