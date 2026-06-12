"""Gera Excel da planilha financeira PORTUGA/PAULO maio/2026 (conferência)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

LINHAS = [
    ("ELISANGELA DA SILVA", "11378901690", 150.00, ""),
    ("SAMUEL ARAUJO DA SILVA", "16132705694", 130.00, ""),
    ("VANESSA APARECIDA RODRIGUES DOS SANTOS", "17448367609", 130.00, ""),
    ("MARISA RODRIGUES GUIMARAES", "10053651642", 150.00, ""),
    ("PATRÍCIA GOMES DOS SANTOS", "07761745684", 170.00, ""),
    ("MATEUS SANTOS OLIVEIRA", "07910555660", 170.00, ""),
    ("NILSON VICTOR DA SILVA ARAUJO", "10306217643", 0.00, ""),
    ("POLLYANA GABRIELLE BARBOSA SANTOS", "16190911609", 0.00, ""),
    ("MAKENSON AUGUSTIN", "70320005259", 130.00, ""),
    ("PALOMA RIBEIRO FAGUNDES", "09872667640", 150.00, ""),
    ("DEIVID DE JESUS CARVALHO", "10804843619", 150.00, ""),
    ("SONIA BARROSO DA SILVA", "00536244677", 130.00, ""),
    ("RICHARD KELVIN CRUZ", "14872482611", 0.00, ""),
    ("LUCIA MARIA VIEIRA", "02632574609", 130.00, ""),
    ("FLAVIA HERNANI DA CONCEICAO ARRUDA", "06616753603", 150.00, ""),
    (
        "34.506.166 ANA BEATRIZ DE VASCONCELOS CORREA LEO",
        "34506166000104",
        130.00,
        "MEI",
    ),
    ("ERIKA XAVIER DE ANDRADE BATISTA", "09681821602", 150.00, ""),
    ("FREDERICO DOS SANTOS CASTRO", "11483074609", 130.00, ""),
]

DESCONTOS = [
    ("", "", -25.00, "antecipação de instalação"),
    ("", "", -50.00, "Adiant. CNPJ (1 un.)"),
]

SAIDA = Path(__file__).resolve().parent.parent / "data" / "comissao_export" / "PORTUGA_PAULO_financeiro_maio_2026.xlsx"


def gerar() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maio 2026"

    headers = ("Cliente", "CPF/CNPJ", "valor", "MEI/NMEI")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    red_font = Font(color="FF0000")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for cliente, doc, valor, mei in LINHAS:
        ws.cell(row=row, column=1, value=cliente)
        ws.cell(row=row, column=2, value=doc)
        c_val = ws.cell(row=row, column=3, value=valor)
        c_val.number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=4, value=mei if mei else "-")
        row += 1

    for _cli, _doc, valor, obs in DESCONTOS:
        c_val = ws.cell(row=row, column=3, value=valor)
        c_val.number_format = 'R$ #,##0.00'
        c_val.font = red_font
        ws.cell(row=row, column=4, value=obs).font = red_font
        row += 1

    soma_vendas = sum(x[2] for x in LINHAS)
    soma_descontos = sum(x[2] for x in DESCONTOS)
    total_planilha = soma_vendas + soma_descontos

    ws.cell(row=row, column=1, value="VALOR FINAL PORTUGA MAIO").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = yellow_fill
    total_cell = ws.cell(row=row, column=3, value=total_planilha)
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.font = Font(bold=True)
    total_cell.fill = yellow_fill
    for c in (1, 2, 3, 4):
        ws.cell(row=row, column=c).fill = yellow_fill

    row += 2
    ws.cell(row=row, column=1, value="Conferência (soma automática)")
    ws.cell(row=row, column=3, value=soma_vendas).number_format = 'R$ #,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Descontos")
    ws.cell(row=row, column=3, value=soma_descontos).number_format = 'R$ #,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Total conferido (= sistema R$ 2.075)")
    ws.cell(row=row, column=3, value=total_planilha).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(SAIDA)
    return SAIDA


if __name__ == "__main__":
    path = gerar()
    print(f"Planilha gerada: {path}")
