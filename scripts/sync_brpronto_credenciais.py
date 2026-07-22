"""
Sincroniza login/senha Br Pronto no cadastro a partir de
c:\\BR_PRONTO\\relatorios\\senhas_definitivas_brprontopdv.xlsx

Atualiza apenas usuários cujo matricula_pap (ou brpronto_login) está na lista
de perfis BO (GED_BRPRONTOPDV_NIVEL2_BOPAP) da planilha de solicitação.

NÃO imprime senhas. Uso:
  python manage.py shell < scripts/sync_brpronto_credenciais.py
  ou: python scripts/sync_brpronto_credenciais.py
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "gestao_equipes.settings")

import django

django.setup()

import openpyxl
from django.db.models import Q
from usuarios.models import Usuario

SENHAS_XLSX = Path(r"c:\BR_PRONTO\relatorios\senhas_definitivas_brprontopdv.xlsx")
SOLICITACAO_XLSX = Path(
    r"c:\BR_PRONTO\relatorios\Serasa_BrPDV_Formulário_Solicitação_Acesso_Nio_v1_RECORD-1.xlsx"
)
DOMINIO = "BrPronto"


def _logins_bopap() -> set[str]:
    wb = openpyxl.load_workbook(SOLICITACAO_XLSX, data_only=True)
    ws = wb["Cadastro PAP"]
    out: set[str] = set()
    for r in range(7, 80):
        login = ws.cell(r, 3).value
        perfil = ws.cell(r, 9).value
        if login and perfil and "BOPAP" in str(perfil):
            out.add(str(login).strip().upper())
    return out


def _senhas_por_login() -> dict[str, str]:
    wb = openpyxl.load_workbook(SENHAS_XLSX, data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        login = ws.cell(r, 2).value
        senha = ws.cell(r, 3).value
        if login and senha:
            out[str(login).strip().upper()] = str(senha).strip()
    return out


def main() -> None:
    bos = _logins_bopap()
    senhas = _senhas_por_login()
    print(f"BOs na planilha: {len(bos)} | senhas definitivas: {len(senhas)}")

    atualizados = 0
    sem_user = []
    sem_senha = []

    for login in sorted(bos):
        senha = senhas.get(login)
        if not senha:
            sem_senha.append(login)
            continue
        qs = Usuario.objects.filter(
            Q(matricula_pap__iexact=login) | Q(brpronto_login__iexact=login),
            perfil__cod_perfil__iexact="backoffice",
        )
        if not qs.exists():
            sem_user.append(login)
            continue
        for u in qs:
            # Evita sobrescrever Roger/diretoria com login errado se não for o mat PAP do BO
            if (u.matricula_pap or "").upper() != login and (u.brpronto_login or "").upper() != login:
                continue
            u.brpronto_login = login
            u.brpronto_senha = senha
            u.brpronto_dominio = DOMINIO
            u.brpronto_disponivel_para_automacao = True
            u.save(
                update_fields=[
                    "brpronto_login",
                    "brpronto_senha",
                    "brpronto_dominio",
                    "brpronto_disponivel_para_automacao",
                ]
            )
            atualizados += 1
            print(f"OK user={u.username} id={u.id} login={login} pool=True")

    print(f"\nAtualizados: {atualizados}")
    if sem_user:
        print("Sem usuário no sistema:", ", ".join(sem_user))
    if sem_senha:
        print("Sem senha na planilha definitiva:", ", ".join(sem_senha))


if __name__ == "__main__":
    main()
