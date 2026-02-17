"""
Script para exportar as abas do Excel de Comissão para JSON/CSV no projeto.
Execute no seu PC (onde o Excel está acessível):

  python scripts/export_excel_comissao.py "C:/caminho/para/COMISSÃO_RECORD_AGF.xlsx"

Ou coloque o arquivo em data/ e execute:

  python scripts/export_excel_comissao.py data/COMISSÃO_RECORD_AGF.xlsx

Saída: data/comissao_export/FOLHA_PAGAMENTO.json, REGRAS_FAIXAS.json, REGRAS_VENDEDORES.json (+ .csv)
"""

import csv
import json
import os
import sys
from datetime import date, datetime

def main():
    if len(sys.argv) < 2:
        print("Uso: python export_excel_comissao.py <caminho_do_xlsx>")
        print("Ex:  python export_excel_comissao.py \"C:\\Users\\...\\COMISSÃO_RECORD_AGF.xlsx\"")
        sys.exit(1)

    path_xlsx = os.path.abspath(sys.argv[1])
    if not os.path.isfile(path_xlsx):
        print(f"Arquivo não encontrado: {path_xlsx}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Instale openpyxl: pip install openpyxl")
        sys.exit(1)

    # Abas desejadas (nomes exatos no Excel)
    abas = [
        "FOLHA PAGAMENTO (2)",
        "REGRAS_FAIXAS",
        "REGRAS_VENDEDORES",
    ]

    wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
    out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "comissao_export")
    os.makedirs(out_dir, exist_ok=True)

    for sheet_name in abas:
        if sheet_name not in wb.sheetnames:
            print(f"Aba não encontrada: {sheet_name}. Abas existentes: {wb.sheetnames}")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        def cell_to_json(c):
            if c is None:
                return ""
            if isinstance(c, (datetime, date)):
                return c.isoformat()
            return c

        rows_clean = [[cell_to_json(c) for c in row] for row in rows]

        safe_name = sheet_name.replace(" ", "_").replace("(", "").replace(")", "")
        json_path = os.path.join(out_dir, f"{safe_name}.json")
        csv_path = os.path.join(out_dir, f"{safe_name}.csv")

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(rows_clean, f, ensure_ascii=False, indent=2)

        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=";")
            w.writerows(rows_clean)

        print(f"Exportado: {json_path}")
        print(f"Exportado: {csv_path}")

    wb.close()
    print("\nConcluído. Use os arquivos em data/comissao_export/ para análise.")

if __name__ == "__main__":
    main()
